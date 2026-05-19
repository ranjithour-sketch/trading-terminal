"""
trading.py  —  Indian Intraday & Options Terminal
==================================================
Run:  streamlit run trading.py

Tabs:
  1. 📋 Watchlist          — live prices for all sectors
  2. 🎯 Trade Setup        — 11-factor checklist + entry/exit
  3. 🏦 Smart Money        — institutional activity detector
  4. 🧮 P&L Calculator     — options profit/loss before trade
  5. 📰 News & Events      — market news + daily checklist
"""

import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
import ta
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
from scipy.stats import norm
import math, time, pytz, requests
import xml.etree.ElementTree as ET
from ml_engine import train_model, predict_next_move, approximate_realtime

# ── Zerodha Kite Connect ──────────────────────────────────
try:
    from kiteconnect import KiteConnect
    KITE_AVAILABLE = True
except ImportError:
    KITE_AVAILABLE = False

# ── Load API credentials ──────────────────────────────────
# Works both locally (.env file) and on Streamlit Cloud (secrets)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import os

def _get_secret(key: str, default: str = "") -> str:
    """Load from Streamlit secrets first, then .env, then env vars."""
    try:
        # Streamlit Cloud secrets (production)
        return st.secrets[key]
    except:
        pass
    # Local .env or environment variable
    return os.getenv(key, default)

KITE_API_KEY    = _get_secret("KITE_API_KEY")
KITE_API_SECRET = _get_secret("KITE_API_SECRET")


# ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Intraday & Options Terminal",
    layout="wide", page_icon="🎯",
    initial_sidebar_state="expanded",
)
IST = pytz.timezone("Asia/Kolkata")

# ── Persistent credential storage using JSON file ─────────
import json, os

CREDS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    ".trading_creds.json"
)

def load_creds():
    """
    Load saved credentials.
    Priority: Streamlit secrets > local JSON file > nothing
    """
    # Load Telegram from Streamlit secrets (cloud deployment)
    try:
        if ("tg_token_saved" not in st.session_state
                and st.secrets.get("tg_token")):
            st.session_state["tg_token_saved"] = (
                st.secrets["tg_token"]
            )
        if ("tg_chat_saved" not in st.session_state
                and st.secrets.get("tg_chat")):
            st.session_state["tg_chat_saved"] = (
                st.secrets["tg_chat"]
            )
    except:
        pass

    # Load from local JSON file (local use)
    try:
        if os.path.exists(CREDS_FILE):
            with open(CREDS_FILE, "r") as f:
                data = json.load(f)
            if ("tg_token_saved" not in st.session_state
                    and data.get("tg_token")):
                st.session_state["tg_token_saved"] = (
                    data["tg_token"]
                )
            if ("tg_chat_saved" not in st.session_state
                    and data.get("tg_chat")):
                st.session_state["tg_chat_saved"] = (
                    data["tg_chat"]
                )
    except Exception:
        pass

def save_creds(token: str, chat: str):
    """Save credentials to local JSON file permanently."""
    try:
        with open(CREDS_FILE, "w") as f:
            json.dump({
                "tg_token": token,
                "tg_chat":  chat
            }, f)
        return True
    except Exception:
        return False

# Load credentials on every page load
load_creds()

# ── Background ML Pre-training ────────────────────────────
# Train ML models for top stocks at startup
# Stored in session_state so Diamond scan is instant
def pretrain_ml_models(
    stock_list: list,
    max_stocks: int = 20,
    progress_bar=None,
    status_text=None
):
    """
    Pre-train ML models for stocks in background.
    Results cached in session_state.
    """
    trained = st.session_state.get("ml_pretrained", {})

    _to_train = [
        s for s in stock_list[:max_stocks]
        if s not in trained
    ]

    for _idx, sname in enumerate(_to_train):
        sym = STOCKS.get(sname)
        if not sym:
            continue
        try:
            if status_text:
                status_text.text(
                    f"Training ML: {sname} "
                    f"({_idx+1}/{len(_to_train)})"
                )
            if progress_bar:
                progress_bar.progress(
                    int((_idx+1)/len(_to_train)*100)
                )
            df_ml = candles(sym, "1d")
            if df_ml is not None and len(df_ml) >= 100:
                model = train_model(df_ml)
                if model.get("ok"):
                    pred = predict_next_move(df_ml, model)
                    if pred and pred.get("ok"):
                        trained[sname] = {
                            "direction":   pred["prediction"],
                            "confidence":  pred["confidence"],
                            "reliability": pred["reliability"],
                            "ok": True
                        }
        except Exception:
            continue

    st.session_state["ml_pretrained"]     = trained
    st.session_state["ml_models_trained"] = True


def prefetch_candles_cache(
    stock_list: list,
    timeframes: list = ["1d"],  # Only cache daily — 1h must be fresh
    progress_bar=None,
    status_text=None
):
    """
    Pre-fetch ONLY daily candles for all stocks.
    1h candles are NOT cached — they must be fetched fresh at scan time
    because the morning 1h candle changes significantly after 9:15 AM.
    """
    _cache = st.session_state.get("candle_cache", {})
    _total = len(stock_list) * len(timeframes)
    _done  = 0

    for sname in stock_list:
        sym = STOCKS.get(sname)
        if not sym:
            continue
        for tf in timeframes:
            _key = f"{sym}_{tf}"
            if _key not in _cache:
                try:
                    if status_text:
                        status_text.text(
                            f"Caching {sname} {tf} candles..."
                        )
                    _df = candles(sym, tf)
                    if _df is not None:
                        _cache[_key] = _df
                except Exception:
                    pass
            _done += 1
            if progress_bar:
                progress_bar.progress(
                    int(_done/_total*100)
                )

    st.session_state["candle_cache"]         = _cache
    st.session_state["candle_cache_ready"]   = True
    import time as _ts_time
    st.session_state["prep_timestamp"]       = _ts_time.time()

def get_ml_cached(sname: str) -> dict:
    """Get pre-trained ML result for a stock. Returns dict or None."""
    cache = st.session_state.get("ml_pretrained", {})
    return cache.get(sname)

# ── Kite session management ───────────────────────────────
def get_kite() -> "KiteConnect | None":
    """Returns authenticated KiteConnect instance or None."""
    if not KITE_AVAILABLE or not KITE_API_KEY:
        return None
    token = st.session_state.get("kite_access_token", "")
    if not token:
        # Try loading from saved file
        try:
            kt_file = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                ".kite_token.json"
            )
            if os.path.exists(kt_file):
                import json as _json
                kt_data = _json.load(open(kt_file))
                # Check if token is from today
                from datetime import date as _date
                if kt_data.get("date") == str(_date.today()):
                    token = kt_data.get("token","")
                    if token:
                        st.session_state["kite_access_token"] = token
        except:
            pass
    if not token:
        return None
    try:
        kite = KiteConnect(api_key=KITE_API_KEY)
        kite.set_access_token(token)
        return kite
    except:
        return None

def kite_is_connected() -> bool:
    """True if Kite session is active today."""
    return bool(st.session_state.get("kite_access_token",""))

def save_kite_token(token: str):
    """Save access token to file for today."""
    try:
        import json as _json
        from datetime import date as _date
        kt_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            ".kite_token.json"
        )
        _json.dump({"token": token,
                    "date": str(_date.today())},
                   open(kt_file, "w"))
    except:
        pass

# ── URL-based tab routing ──────────────────────────────────
# Each tab has a URL: ?tab=watchlist, ?tab=setup, etc.
# This allows opening any tab in a separate browser window
TAB_ROUTES = {
    "watchlist": 0,
    "setup":     1,
    "scanner":   2,
    "ml":        3,
    "smart":     4,
    "pulse":     5,
    "options":   6,
    "backtest":  7,
    "hub":       8,
    "manager":   9,
    "paper":     10,
    "orders":    11,
    "evening":   12,
}
TAB_NAMES = [
    "📋 Watchlist",
    "🎯 Trade Setup",
    "🔍 Auto Scanner",
    "🤖 ML Prediction",
    "🏦 Smart Money",
    "📊 Market Pulse",
    "🔗 Options Chain",
    "🧪 Backtest",
    "🎯 Signal Hub",
    "🛡️ Trade Manager",
    "📝 Paper Trading",
    "⚡ Auto Orders",
    "🌙 Evening Scan",
]
TAB_ICONS = ["📋","🎯","🔍","🤖","🏦","📊","🔗","🧪","🎯","🛡️","📝","⚡","🌙"]
TAB_KEYS  = list(TAB_ROUTES.keys())

# Read current tab from URL
_qp = st.query_params
_tab_key = _qp.get("tab", "watchlist")
_default_tab = TAB_ROUTES.get(_tab_key, 0)

# ══════════════════════════════════════════════════════════
# MARKET DATA HELPERS
# ══════════════════════════════════════════════════════════

@st.cache_data(ttl=300)
def get_iv_rank(symbol: str = "NIFTY") -> dict:
    """
    Calculate IV Rank for NIFTY/BANKNIFTY.
    IV Rank = (Current IV - 52w Low IV) / (52w High IV - 52w Low IV) * 100
    Uses VIX as proxy for NIFTY IV.
    IV Rank > 70 = expensive options (avoid buying)
    IV Rank < 30 = cheap options (good time to buy)
    """
    try:
        # Use India VIX as IV proxy
        vix_tk = yf.Ticker("^INDIAVIX")
        hist   = vix_tk.history(period="1y")
        if hist.empty:
            return {"ok": False}

        current_iv = float(hist["Close"].iloc[-1])
        high_52w   = float(hist["Close"].max())
        low_52w    = float(hist["Close"].min())
        iv_range   = high_52w - low_52w

        if iv_range <= 0:
            return {"ok": False}

        iv_rank    = round(
            (current_iv - low_52w) / iv_range * 100, 1
        )
        iv_pct     = round(
            len(hist[hist["Close"] <= current_iv]) /
            len(hist) * 100, 1
        )

        if iv_rank > 70:
            signal = "EXPENSIVE"
            advice = "Options overpriced — avoid buying CE/PE now"
            color  = "#dc2626"
            bg     = "#fef2f2"
        elif iv_rank < 30:
            signal = "CHEAP"
            advice = "Options cheap — good time to buy CE/PE"
            color  = "#16a34a"
            bg     = "#f0fdf4"
        else:
            signal = "MODERATE"
            advice = "Options fairly priced — trade normally"
            color  = "#d97706"
            bg     = "#fffbeb"

        return {
            "ok":         True,
            "current_iv": round(current_iv, 2),
            "iv_rank":    iv_rank,
            "iv_pct":     iv_pct,
            "high_52w":   round(high_52w, 2),
            "low_52w":    round(low_52w, 2),
            "signal":     signal,
            "advice":     advice,
            "color":      color,
            "bg":         bg,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@st.cache_data(ttl=60)
def get_live_greeks(
    symbol: str,
    strike: float,
    opt_type: str = "CE",
    expiry: str = ""
) -> dict:
    """
    Fetch live Greeks for a specific options contract.
    Uses Kite API if connected, else estimates from Black-Scholes.
    opt_type: CE or PE
    """
    try:
        kite = get_kite()
        if kite:
            # Try to get from Kite instruments
            _token = st.session_state.get("kite_access_token","")
            _inst_map = get_kite_instruments(_token)
            # Build option symbol name
            _sym = symbol.replace(" ","").replace("&","")
            _opt_sym = f"{_sym}{expiry}{int(strike)}{opt_type}"

            _inst_token = _inst_map.get(_opt_sym)
            if _inst_token:
                _quote = kite.quote([f"NFO:{_opt_sym}"])
                if _quote:
                    _q = list(_quote.values())[0]
                    return {
                        "ok":     True,
                        "source": "Kite Live",
                        "ltp":    _q.get("last_price", 0),
                        "oi":     _q.get("oi", 0),
                        "volume": _q.get("volume", 0),
                        "iv":     _q.get("implied_volatility", 0),
                        "delta":  _q.get("greeks", {}).get("delta", 0),
                        "theta":  _q.get("greeks", {}).get("theta", 0),
                        "gamma":  _q.get("greeks", {}).get("gamma", 0),
                        "vega":   _q.get("greeks", {}).get("vega", 0),
                    }
    except Exception:
        pass

    # Fallback: Black-Scholes estimation
    try:
        import math as _math
        from scipy.stats import norm as _norm

        # Get current spot price
        _spot_lp = live_price(
            "^NSEI" if "NIFTY" in symbol.upper()
            else "^NSEBANK"
        )
        S = _spot_lp["p"] if _spot_lp["ok"] else strike
        K = strike
        T = max(1, 7) / 365  # assume 7 days to expiry
        r = 0.065  # risk-free rate India
        # Use VIX as proxy for sigma
        _vd = get_india_vix()
        sigma = (_vd["vix"] / 100) if _vd["ok"] else 0.15

        d1 = (
            (_math.log(S/K) + (r + 0.5*sigma**2)*T)
            / (sigma * _math.sqrt(T))
        )
        d2 = d1 - sigma * _math.sqrt(T)

        if opt_type == "CE":
            delta = float(_norm.cdf(d1))
            theta = float(
                -(S * _norm.pdf(d1) * sigma) /
                (2 * _math.sqrt(T)) -
                r * K * _math.exp(-r*T) * _norm.cdf(d2)
            ) / 365
        else:
            delta = float(_norm.cdf(d1) - 1)
            theta = float(
                -(S * _norm.pdf(d1) * sigma) /
                (2 * _math.sqrt(T)) +
                r * K * _math.exp(-r*T) * _norm.cdf(-d2)
            ) / 365

        gamma = float(
            _norm.pdf(d1) / (S * sigma * _math.sqrt(T))
        )
        vega = float(S * _norm.pdf(d1) * _math.sqrt(T)) / 100

        return {
            "ok":     True,
            "source": "Estimated (B-S)",
            "ltp":    0,
            "delta":  round(delta, 4),
            "theta":  round(theta, 4),
            "gamma":  round(gamma, 6),
            "vega":   round(vega, 4),
            "iv":     round(sigma * 100, 2),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@st.cache_data(ttl=120)
def get_oi_change_data(symbol: str = "NIFTY") -> dict:
    """
    Get OI change data from NSE options chain.
    OI change shows where money is flowing in real time.
    Positive OI change at a strike = new positions being built.
    Negative OI change = positions being closed (unwinding).
    """
    try:
        session_ = requests.Session()
        hdrs = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Accept": "application/json",
            "Referer": "https://www.nseindia.com/"
        }
        session_.get(
            "https://www.nseindia.com",
            headers=hdrs, timeout=8
        )
        url = (
            f"https://www.nseindia.com/api/option-chain-indices"
            f"?symbol={symbol}"
        )
        r = session_.get(url, headers=hdrs, timeout=8)

        if r.status_code != 200:
            return {"ok": False}

        data    = r.json()
        records = data.get("records", {})
        spot    = float(records.get("underlyingValue", 0))
        oc_data = records.get("data", [])
        expiries= records.get("expiryDates", [])

        if not oc_data or not expiries:
            return {"ok": False}

        # Use nearest expiry
        nearest_exp = expiries[0]
        exp_data    = [
            d for d in oc_data
            if d.get("expiryDate") == nearest_exp
        ]

        # ATM strike
        step = 50 if symbol == "NIFTY" else 100
        atm  = round(spot / step) * step

        # Build OI change analysis
        strikes_data = []
        total_call_oi_chg = 0
        total_put_oi_chg  = 0

        for item in exp_data:
            strike = item.get("strikePrice", 0)
            ce = item.get("CE", {})
            pe = item.get("PE", {})

            ce_oi     = int(ce.get("openInterest", 0) or 0)
            ce_oi_chg = int(ce.get("changeinOpenInterest", 0) or 0)
            ce_iv     = float(ce.get("impliedVolatility", 0) or 0)
            ce_ltp    = float(ce.get("lastPrice", 0) or 0)

            pe_oi     = int(pe.get("openInterest", 0) or 0)
            pe_oi_chg = int(pe.get("changeinOpenInterest", 0) or 0)
            pe_iv     = float(pe.get("impliedVolatility", 0) or 0)
            pe_ltp    = float(pe.get("lastPrice", 0) or 0)

            total_call_oi_chg += ce_oi_chg
            total_put_oi_chg  += pe_oi_chg

            strikes_data.append({
                "strike":     strike,
                "ce_oi":      ce_oi,
                "ce_oi_chg":  ce_oi_chg,
                "ce_iv":      round(ce_iv, 1),
                "ce_ltp":     round(ce_ltp, 2),
                "pe_oi":      pe_oi,
                "pe_oi_chg":  pe_oi_chg,
                "pe_iv":      round(pe_iv, 1),
                "pe_ltp":     round(pe_ltp, 2),
                "is_atm":     strike == atm,
            })

        # Sort by strike
        strikes_data.sort(key=lambda x: x["strike"])

        # Find max OI and max OI change strikes
        if strikes_data:
            max_call_oi_strike = max(
                strikes_data, key=lambda x: x["ce_oi"]
            )["strike"]
            max_put_oi_strike  = max(
                strikes_data, key=lambda x: x["pe_oi"]
            )["strike"]
            max_call_chg_strike= max(
                strikes_data, key=lambda x: x["ce_oi_chg"]
            )["strike"]
            max_put_chg_strike = max(
                strikes_data, key=lambda x: x["pe_oi_chg"]
            )["strike"]
        else:
            max_call_oi_strike  = atm
            max_put_oi_strike   = atm
            max_call_chg_strike = atm
            max_put_chg_strike  = atm

        # OI change signal
        pcr_chg = round(
            abs(total_put_oi_chg) /
            (abs(total_call_oi_chg) + 1), 2
        )

        if (total_call_oi_chg > 0 and
                total_put_oi_chg > 0 and
                total_put_oi_chg > total_call_oi_chg):
            oi_signal = "BULLISH BUILD"
            oi_color  = "#16a34a"
            oi_advice = (
                "Put writers building positions — "
                "market expected to hold support"
            )
        elif (total_call_oi_chg > 0 and
              total_call_oi_chg > total_put_oi_chg):
            oi_signal = "BEARISH BUILD"
            oi_color  = "#dc2626"
            oi_advice = (
                "Call writers building positions — "
                "market expected to face resistance"
            )
        elif (total_call_oi_chg < 0 and
              total_put_oi_chg < 0):
            oi_signal = "UNWINDING"
            oi_color  = "#f59e0b"
            oi_advice = (
                "Both CE and PE unwinding — "
                "positions being closed, low conviction"
            )
        else:
            oi_signal = "MIXED"
            oi_color  = "#64748b"
            oi_advice = "Mixed OI signals — wait for clarity"

        return {
            "ok":                 True,
            "spot":               spot,
            "atm":                atm,
            "expiry":             nearest_exp,
            "strikes":            strikes_data,
            "total_call_oi_chg":  total_call_oi_chg,
            "total_put_oi_chg":   total_put_oi_chg,
            "pcr_chg":            pcr_chg,
            "max_call_oi_strike": max_call_oi_strike,
            "max_put_oi_strike":  max_put_oi_strike,
            "max_call_chg_strike":max_call_chg_strike,
            "max_put_chg_strike": max_put_chg_strike,
            "oi_signal":          oi_signal,
            "oi_color":           oi_color,
            "oi_advice":          oi_advice,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@st.cache_data(ttl=3600)
def get_economic_calendar() -> list:
    """
    Get upcoming Indian and US economic events.
    Returns list of events for next 7 days.
    Uses a predefined calendar of recurring events
    plus fetches from public sources when available.
    """
    from datetime import datetime as _dt, timedelta as _td
    import pytz as _ptz

    _ist  = _ptz.timezone("Asia/Kolkata")
    _now  = _dt.now(_ist)
    _events = []

    # ── Fixed recurring Indian events ─────────────────────
    # RBI MPC meetings 2025-2026 (approximate dates)
    _rbi_dates = [
        "2026-06-04", "2026-08-06", "2026-10-01",
        "2026-12-03", "2027-02-05"
    ]
    for _rd in _rbi_dates:
        try:
            _rdt = _dt.strptime(_rd, "%Y-%m-%d").date()
            _days = (_rdt - _now.date()).days
            if 0 <= _days <= 30:
                _events.append({
                    "date":    _rd,
                    "event":   "🏦 RBI MPC Policy Decision",
                    "impact":  "HIGH",
                    "country": "India",
                    "note":    "Interest rate decision — major market mover",
                    "days_away": _days,
                })
        except Exception:
            pass

    # ── Try to fetch from investing.com economic calendar ──
    try:
        import requests as _req
        _hdrs = {
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
        }
        # Fallback: use predefined weekly events
        _weekly_events = []

        # US events (affect Indian markets)
        _us_events = [
            ("Monday",    "🇺🇸 US ISM Manufacturing",    "MEDIUM"),
            ("Wednesday", "🇺🇸 US Fed FOMC Minutes",      "HIGH"),
            ("Thursday",  "🇺🇸 US Initial Jobless Claims","LOW"),
            ("Friday",    "🇺🇸 US NFP/Jobs Report",       "HIGH"),
        ]

        # Indian weekly events
        _in_events = [
            ("Monday",   "🇮🇳 India FII/DII Data",        "MEDIUM"),
            ("Thursday", "🇮🇳 NSE F&O Expiry",            "HIGH"),
        ]

        _day_names = [
            "Monday","Tuesday","Wednesday",
            "Thursday","Friday","Saturday","Sunday"
        ]

        for _offset in range(8):
            _dt_check = (_now + _td(days=_offset)).date()
            _day_name = _day_names[_dt_check.weekday()]

            for _day, _evt, _imp in _us_events + _in_events:
                if _day == _day_name:
                    _events.append({
                        "date":    str(_dt_check),
                        "event":   _evt,
                        "impact":  _imp,
                        "country": "US" if "🇺🇸" in _evt else "India",
                        "note":    "",
                        "days_away": _offset,
                    })
    except Exception:
        pass

    # Sort by date
    _events.sort(key=lambda x: x["days_away"])
    return _events[:15]  # next 15 events


@st.cache_data(ttl=300)
def get_india_vix() -> dict:
    """Get India VIX from Yahoo Finance."""
    try:
        vix_tk = yf.Ticker("^INDIAVIX")
        fi = vix_tk.fast_info
        p  = float(fi.last_price)
        pc = float(fi.previous_close) if fi.previous_close else p
        ch = round(((p-pc)/pc)*100, 2) if pc else 0
        level = (
            "🔴 EXTREME FEAR" if p > 25 else
            "🟠 HIGH FEAR"    if p > 20 else
            "🟡 ELEVATED"     if p > 15 else
            "🟢 CALM"         if p > 10 else
            "🟢 VERY CALM"
        )
        advice = (
            "Avoid buying options — premiums very expensive"
            if p > 20 else
            "Good time to buy CE/PE — premiums reasonable"
            if p < 15 else
            "Moderate conditions — trade carefully"
        )
        return {
            "ok": True, "vix": round(p, 2),
            "prev": round(pc, 2), "chg": ch,
            "level": level, "advice": advice
        }
    except Exception as e:
        return {"ok": False, "vix": 0, "level": "Unknown",
                "advice": "VIX data unavailable", "chg": 0}

@st.cache_data(ttl=600)
def get_fii_dii() -> dict:
    """Get FII/DII data from NSE website."""
    try:
        session_ = requests.Session()
        hdrs = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Accept": "application/json",
            "Referer": "https://www.nseindia.com/"
        }
        session_.get("https://www.nseindia.com", headers=hdrs, timeout=8)
        r = session_.get(
            "https://www.nseindia.com/api/fiidiiTradeReact",
            headers=hdrs, timeout=8
        )
        if r.status_code == 200:
            data = r.json()
            if data:
                latest = data[0]
                fii_buy  = float(latest.get("fiiBuy", 0) or 0)
                fii_sell = float(latest.get("fiiSell", 0) or 0)
                dii_buy  = float(latest.get("diiBuy", 0) or 0)
                dii_sell = float(latest.get("diiSell", 0) or 0)
                fii_net  = round(fii_buy - fii_sell, 2)
                dii_net  = round(dii_buy - dii_sell, 2)
                return {
                    "ok": True,
                    "date":     latest.get("date", ""),
                    "fii_buy":  fii_buy,
                    "fii_sell": fii_sell,
                    "fii_net":  fii_net,
                    "dii_buy":  dii_buy,
                    "dii_sell": dii_sell,
                    "dii_net":  dii_net,
                    "data":     data[:10]
                }
        return {"ok": False}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@st.cache_data(ttl=180)
def get_options_chain(symbol: str = "NIFTY") -> dict:
    """Get options chain from NSE."""
    try:
        session_ = requests.Session()
        hdrs = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Accept": "application/json",
            "Referer": "https://www.nseindia.com/"
        }
        session_.get("https://www.nseindia.com", headers=hdrs, timeout=8)
        url = (f"https://www.nseindia.com/api/option-chain-indices"
               f"?symbol={symbol}")
        r = session_.get(url, headers=hdrs, timeout=8)
        if r.status_code == 200:
            data = r.json()
            records = data.get("records", {})
            exp_dates = records.get("expiryDates", [])
            spot = float(records.get("underlyingValue", 0))
            oc_data = records.get("data", [])
            return {
                "ok": True,
                "spot": spot,
                "expiries": exp_dates,
                "data": oc_data,
                "timestamp": records.get("timestamp", "")
            }
        return {"ok": False}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@st.cache_data(ttl=3600)
def get_nifty_history_for_backtest(months: int = 6) -> pd.DataFrame:
    """Get historical data for backtesting."""
    try:
        end   = datetime.now()
        start = end - timedelta(days=months*30)
        df = yf.download(
            "^NSEI", start=start, end=end,
            interval="1d", progress=False,
            auto_adjust=True
        )
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df.dropna(inplace=True)
        return df
    except:
        return pd.DataFrame()

# ══════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* ═══════════════════════════════════════════════
   BASE
═══════════════════════════════════════════════ */
html, body, .stApp, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont,
                 'Segoe UI', sans-serif !important;
    background-color: #f4f6f9 !important;
    color: #1e293b !important;
}

/* ── Layout ───────────────────────────────────────── */
.block-container {
    padding-top: 0.5rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 1400px !important;
}

/* Streamlit header — transparent, no height, no click blocking */
header[data-testid="stHeader"] {
    background: transparent !important;
    border-bottom: none !important;
    box-shadow: none !important;
    pointer-events: none !important;
}

/* Toolbar — allow clicks through */
[data-testid="stToolbar"] {
    pointer-events: all !important;
    z-index: 999 !important;
}

/* Main content area */
.main .block-container {
    padding-top: 0.5rem !important;
}

/* ── Mobile responsive — comprehensive fix ──────── */
@media (max-width: 768px) {
    /* Layout */
    .block-container {
        padding-left: 0.4rem !important;
        padding-right: 0.4rem !important;
        padding-top: 0.3rem !important;
    }

    /* Top navigation bar — make visible on mobile */
    header[data-testid="stHeader"] {
        background: #1e3a5f !important;
        border-bottom: 2px solid #1d4ed8 !important;
    }
    header[data-testid="stHeader"] * {
        color: #ffffff !important;
        fill: #ffffff !important;
    }
    /* Hamburger menu icon */
    button[kind="header"] svg {
        fill: #ffffff !important;
        color: #ffffff !important;
    }
    [data-testid="stToolbar"] {
        background: transparent !important;
    }
    [data-testid="stToolbar"] svg {
        fill: #ffffff !important;
    }
    /* Make all header icons white */
    header svg, header button {
        color: #ffffff !important;
        fill: #ffffff !important;
        stroke: #ffffff !important;
    }

    /* Tabs — horizontal scroll, no wrap */
    .stTabs [data-baseweb="tab-list"] {
        overflow-x: auto !important;
        flex-wrap: nowrap !important;
        -webkit-overflow-scrolling: touch !important;
        scrollbar-width: none !important;
        gap: 2px !important;
        padding: 3px !important;
    }
    .stTabs [data-baseweb="tab-list"]::-webkit-scrollbar {
        display: none !important;
    }
    .stTabs [data-baseweb="tab"] {
        white-space: nowrap !important;
        font-size: 11px !important;
        padding: 5px 8px !important;
        flex-shrink: 0 !important;
        min-width: fit-content !important;
    }
    .stTabs [aria-selected="true"] {
        font-size: 11px !important;
    }

    /* Headings */
    h1 { font-size: 18px !important; }
    h2 { font-size: 16px !important; }
    h3 { font-size: 15px !important; }
    h4 { font-size: 14px !important; }
    p, .stMarkdown p {
        font-size: 13px !important;
        line-height: 1.5 !important;
    }
    caption, .stCaption {
        font-size: 11px !important;
    }

    /* Buttons */
    .stButton button {
        font-size: 12px !important;
        padding: 7px 10px !important;
        border-radius: 8px !important;
    }

    /* Metric cards */
    div[data-testid="metric-container"] {
        padding: 8px 10px !important;
        border-radius: 8px !important;
    }
    div[data-testid="metric-container"] label {
        font-size: 10px !important;
    }
    div[data-testid="metric-container"]
        [data-testid="stMetricValue"] {
        font-size: 16px !important;
        font-weight: 700 !important;
    }

    /* Inputs — prevent iOS zoom */
    .stTextInput input,
    .stNumberInput input,
    .stTextArea textarea,
    .stSelectbox [data-baseweb="select"] div {
        font-size: 16px !important;
        border-radius: 8px !important;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        min-width: 200px !important;
        max-width: 260px !important;
    }
    section[data-testid="stSidebar"] .stButton button {
        font-size: 12px !important;
        padding: 6px 10px !important;
    }

    /* Plotly charts full width */
    .js-plotly-plot {
        width: 100% !important;
    }

    /* Expander */
    .streamlit-expanderHeader {
        font-size: 13px !important;
        padding: 8px 10px !important;
    }

    /* Alert boxes */
    div[data-testid="stSuccess"],
    div[data-testid="stWarning"],
    div[data-testid="stError"],
    div[data-testid="stInfo"] {
        padding: 10px 14px !important;
        font-size: 13px !important;
        border-radius: 8px !important;
    }

    /* Dataframe */
    .stDataFrame {
        font-size: 11px !important;
    }

    /* Columns gap */
    [data-testid="column"] {
        padding: 0 2px !important;
    }
}

/* ═══════════════════════════════════════════════
   SIDEBAR
═══════════════════════════════════════════════ */
section[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #e2e8f0 !important;
    padding: 12px 8px !important;
}
section[data-testid="stSidebar"] * {
    color: #1e293b !important;
    font-family: 'Inter', sans-serif !important;
}
/* Hide sidebar collapse/expand arrow that corrupts text */
[data-testid="stSidebarCollapsedControl"] { display:none !important; }
button[data-testid="collapsedControl"] { display:none !important; }
section[data-testid="stSidebar"] > div > div:first-child {
    padding-top: 2px !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    font-size: 13px !important;
    font-weight: 700 !important;
    color: #374151 !important;
    margin: 12px 0 6px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.5px !important;
}
/* Fix label text corruption in sidebar */
section[data-testid="stSidebar"] label {
    font-size: 12px !important;
    color: #374151 !important;
    font-weight: 500 !important;
}
section[data-testid="stSidebar"] .stTextInput label,
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stCheckbox label {
    display: none !important;
}
section[data-testid="stSidebar"] p {
    font-size: 12px !important;
    color: #64748b !important;
}
section[data-testid="stSidebar"] .stButton button {
    background: #f8fafc !important;
    border: 1px solid #e2e8f0 !important;
    color: #334155 !important;
    border-radius: 8px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    padding: 8px 12px !important;
    width: 100% !important;
    text-align: left !important;
    margin: 1px 0 !important;
    transition: all 0.15s ease !important;
}
section[data-testid="stSidebar"] .stButton button:hover {
    background: #eff6ff !important;
    border-color: #3b82f6 !important;
    color: #1d4ed8 !important;
}
/* Hide expander arrow corruption completely */
section[data-testid="stSidebar"] .streamlit-expanderHeader {
    display: none !important;
}
section[data-testid="stSidebar"] .streamlit-expanderContent {
    border: none !important;
    padding: 0 !important;
}

/* ═══════════════════════════════════════════════
   TABS
═══════════════════════════════════════════════ */
.stTabs [data-baseweb="tab-list"] {
    background: #ffffff !important;
    border-radius: 12px !important;
    padding: 5px !important;
    border: 1px solid #e2e8f0 !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
    gap: 3px !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 8px !important;
    color: #64748b !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    padding: 7px 14px !important;
    border: none !important;
    white-space: nowrap !important;
}
.stTabs [data-baseweb="tab"]:hover {
    background: #f1f5f9 !important;
    color: #334155 !important;
}
.stTabs [aria-selected="true"] {
    background: #3b82f6 !important;
    color: #ffffff !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 4px rgba(59,130,246,0.3) !important;
}
.stTabs [data-baseweb="tab-panel"] {
    padding-top: 1rem !important;
    background: transparent !important;
}

/* ═══════════════════════════════════════════════
   BUTTONS
═══════════════════════════════════════════════ */
.stButton button {
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    border-radius: 8px !important;
    font-size: 13px !important;
    padding: 8px 16px !important;
    transition: all 0.15s ease !important;
    border: 1px solid #e2e8f0 !important;
    color: #374151 !important;
    background: #ffffff !important;
}
.stButton button[kind="primary"] {
    background: #3b82f6 !important;
    color: #ffffff !important;
    border: none !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 6px rgba(59,130,246,0.35) !important;
}
.stButton button[kind="primary"]:hover {
    background: #2563eb !important;
    box-shadow: 0 4px 10px rgba(59,130,246,0.4) !important;
    transform: translateY(-1px) !important;
}
.stButton button:hover {
    border-color: #94a3b8 !important;
    background: #f8fafc !important;
}

/* ═══════════════════════════════════════════════
   INPUTS & SELECTS
═══════════════════════════════════════════════ */
.stTextInput input,
.stNumberInput input,
.stTextArea textarea {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 8px !important;
    color: #1e293b !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 13px !important;
}
.stTextInput input:focus,
.stNumberInput input:focus {
    border-color: #3b82f6 !important;
    box-shadow: 0 0 0 3px rgba(59,130,246,0.12) !important;
}
.stSelectbox [data-baseweb="select"] > div {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 8px !important;
    color: #1e293b !important;
    font-size: 13px !important;
}
[data-baseweb="popover"] [data-baseweb="menu"] {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 8px !important;
}
[data-baseweb="option"] {
    background: #ffffff !important;
    color: #1e293b !important;
    font-size: 13px !important;
}
[data-baseweb="option"]:hover {
    background: #eff6ff !important;
}

/* ═══════════════════════════════════════════════
   METRIC CARDS
═══════════════════════════════════════════════ */
div[data-testid="metric-container"] {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important;
    padding: 14px 18px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
}
div[data-testid="metric-container"] label {
    color: #64748b !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.4px !important;
}
div[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #1e293b !important;
    font-weight: 700 !important;
    font-size: 22px !important;
}

/* ═══════════════════════════════════════════════
   ALERTS
═══════════════════════════════════════════════ */
div[data-testid="stSuccess"] {
    background: #f0fdf4 !important;
    border: 1px solid #86efac !important;
    border-radius: 10px !important;
    color: #166534 !important;
}
div[data-testid="stWarning"] {
    background: #fffbeb !important;
    border: 1px solid #fcd34d !important;
    border-radius: 10px !important;
    color: #92400e !important;
}
div[data-testid="stError"] {
    background: #fef2f2 !important;
    border: 1px solid #fca5a5 !important;
    border-radius: 10px !important;
    color: #991b1b !important;
}
div[data-testid="stInfo"] {
    background: #eff6ff !important;
    border: 1px solid #93c5fd !important;
    border-radius: 10px !important;
    color: #1e40af !important;
}
div[data-testid="stSuccess"] p,
div[data-testid="stWarning"] p,
div[data-testid="stError"] p,
div[data-testid="stInfo"] p {
    font-size: 14px !important;
    line-height: 1.6 !important;
}

/* ═══════════════════════════════════════════════
   EXPANDER — fix arrow corruption
═══════════════════════════════════════════════ */
.streamlit-expanderHeader {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important;
    color: #374151 !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    padding: 10px 14px !important;
}
.streamlit-expanderHeader svg {
    display: none !important;
}
.streamlit-expanderHeader p,
.streamlit-expanderHeader span {
    color: #374151 !important;
    font-size: 13px !important;
    font-weight: 600 !important;
}
/* Remove the ::before arrow that causes corruption */
.streamlit-expanderHeader::before {
    content: none !important;
}
details summary::marker,
details summary::-webkit-details-marker {
    display: none !important;
    content: '' !important;
}
.streamlit-expanderContent {
    background: #f8fafc !important;
    border: 1px solid #e2e8f0 !important;
    border-top: none !important;
    border-radius: 0 0 10px 10px !important;
    padding: 12px !important;
}

/* ═══════════════════════════════════════════════
   DATAFRAME / TABLE
═══════════════════════════════════════════════ */
.stDataFrame {
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}
.stDataFrame table {
    color: #1e293b !important;
    font-size: 13px !important;
}
.stDataFrame thead th {
    background: #f8fafc !important;
    color: #64748b !important;
    font-weight: 600 !important;
    font-size: 12px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.3px !important;
    border-bottom: 1px solid #e2e8f0 !important;
}
.stDataFrame tbody tr:hover {
    background: #f0f9ff !important;
}

/* ═══════════════════════════════════════════════
   SCANNER CARDS — ensure stock name is visible
═══════════════════════════════════════════════ */
.scanner-card {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important;
    padding: 14px 16px !important;
    color: #1e293b !important;
}
.scanner-card .stock-name {
    font-size: 15px !important;
    font-weight: 600 !important;
    color: #1e293b !important;
}

/* ═══════════════════════════════════════════════
   MARKDOWN & HEADINGS
═══════════════════════════════════════════════ */
h1, h2, h3, h4, h5, h6 {
    font-family: 'Inter', sans-serif !important;
    color: #1e293b !important;
    font-weight: 700 !important;
    letter-spacing: -0.3px !important;
}
.stMarkdown p {
    color: #374151 !important;
    font-size: 14px !important;
    line-height: 1.6 !important;
}
.stMarkdown strong { color: #1e293b !important; }
caption, .stCaption {
    color: #64748b !important;
    font-size: 12px !important;
}

/* ═══════════════════════════════════════════════
   PROGRESS BAR
═══════════════════════════════════════════════ */
.stProgress > div > div {
    background: #3b82f6 !important;
    border-radius: 4px !important;
}

/* ═══════════════════════════════════════════════
   CHECKBOX & TOGGLE
═══════════════════════════════════════════════ */
.stCheckbox label,
.stToggle label { color: #374151 !important; font-size: 13px !important; }

/* ═══════════════════════════════════════════════
   SCROLLBAR
═══════════════════════════════════════════════ */
::-webkit-scrollbar { width: 5px; height: 5px }
::-webkit-scrollbar-track { background: #f1f5f9 }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px }
::-webkit-scrollbar-thumb:hover { background: #94a3b8 }

/* ═══════════════════════════════════════════════
   PLOTLY CHARTS — light background
═══════════════════════════════════════════════ */
.js-plotly-plot .plotly .main-svg {
    background: #ffffff !important;
    border-radius: 10px !important;
}

/* ═══════════════════════════════════════════════
   SLIDER
═══════════════════════════════════════════════ */
.stSlider [data-baseweb="slider"] div[role="slider"] {
    background: #3b82f6 !important;
    border-color: #3b82f6 !important;
}
/* ── Sidebar ALL arrow/corruption fixes ────────────── */
/* Hide collapse arrow buttons */
[data-testid="stSidebarCollapsedControl"] { display:none !important; }
[data-testid="stSidebarCollapseButton"]   { display:none !important; }
button[data-testid="collapsedControl"]    { display:none !important; }

/* Hide expander arrow icon in sidebar — this was causing _arrow_right */
section[data-testid="stSidebar"] details summary {
    list-style: none !important;
}
section[data-testid="stSidebar"] details summary::-webkit-details-marker {
    display: none !important;
}
section[data-testid="stSidebar"] .streamlit-expanderHeader svg {
    display: none !important;
}

/* Hide text input labels in sidebar */
section[data-testid="stSidebar"] .stTextInput label { 
    display: none !important; 
}

/* Hide selectbox labels in sidebar */
section[data-testid="stSidebar"] .stSelectbox label { 
    display: none !important; 
}

/* Remove top padding */
section[data-testid="stSidebar"] > div:first-child { 
    padding-top: 0 !important; 
}

</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# STOCK & SECTOR DATABASE
# ══════════════════════════════════════════════════════════
STOCKS = {
    # ── Indices ───────────────────────────────────────
    "NIFTY 50":         "^NSEI",
    "BANK NIFTY":       "^NSEBANK",
    "SENSEX":           "^BSESN",
    "NIFTY IT":         "^CNXIT",
    "NIFTY AUTO":       "^CNXAUTO",
    "NIFTY PHARMA":     "^CNXPHARMA",
    "NIFTY METAL":      "^CNXMETAL",
    "NIFTY FMCG":       "^CNXFMCG",
    "NIFTY MIDCAP":     "^NSEMDCP50",
    "INDIA VIX":        "^INDIAVIX",
    # ── Banking & Finance ─────────────────────────────
    "HDFC Bank":        "HDFCBANK.NS",
    "ICICI Bank":       "ICICIBANK.NS",
    "SBI":              "SBIN.NS",
    "Kotak Bank":       "KOTAKBANK.NS",
    "Axis Bank":        "AXISBANK.NS",
    "IndusInd Bank":    "INDUSINDBK.NS",
    "Yes Bank":         "YESBANK.NS",
    "PNB":              "PNB.NS",
    "Bank of Baroda":   "BANKBARODA.NS",
    "Canara Bank":      "CANBK.NS",
    "Federal Bank":     "FEDERALBNK.NS",
    "IDFC First":       "IDFCFIRSTB.NS",
    "Bajaj Finance":    "BAJFINANCE.NS",
    "Bajaj Finserv":    "BAJAJFINSV.NS",
    "Shriram Finance":  "SHRIRAMFIN.NS",
    "Muthoot Finance":  "MUTHOOTFIN.NS",
    "LIC Housing":      "LICHSGFIN.NS",
    "HDFC Life":        "HDFCLIFE.NS",
    "SBI Life":         "SBILIFE.NS",
    "ICICI Lombard":    "ICICIGI.NS",
    # ── IT & Technology ──────────────────────────────
    "TCS":              "TCS.NS",
    "Infosys":          "INFY.NS",
    "Wipro":            "WIPRO.NS",
    "HCL Tech":         "HCLTECH.NS",
    "Tech Mahindra":    "TECHM.NS",
    "Persistent":       "PERSISTENT.NS",
    "Coforge":          "COFORGE.NS",
    "LTIMindtree":      "LTIM.NS",
    "Mphasis":          "MPHASIS.NS",
    "Tata Elxsi":       "TATAELXSI.NS",
    "KPIT Tech":        "KPITTECH.NS",
    "Cyient":           "CYIENT.NS",
    # ── Energy & Power ───────────────────────────────
    "Reliance":         "RELIANCE.NS",
    "ONGC":             "ONGC.NS",
    "Indian Oil":       "IOC.NS",
    "BPCL":             "BPCL.NS",
    "NTPC":             "NTPC.NS",
    "Power Grid":       "POWERGRID.NS",
    "Adani Green":      "ADANIGREEN.NS",
    "Adani Power":      "ADANIPOWER.NS",
    "Adani Energy":     "ADANIENSOL.NS",
    "Tata Power":       "TATAPOWER.NS",
    "Gail":             "GAIL.NS",
    "Coal India":       "COALINDIA.NS",
    "NHPC":             "NHPC.NS",
    "SJVN":             "SJVN.NS",
    "Torrent Power":    "TORNTPOWER.NS",
    "JSW Energy":       "JSWENERGY.NS",
    # ── Auto & Auto Ancillary ─────────────────────────
    "Maruti":           "MARUTI.NS",
    "M&M":              "M&M.NS",
    "Hero MotoCorp":    "HEROMOTOCO.NS",
    "Bajaj Auto":       "BAJAJ-AUTO.NS",
    "TVS Motor":        "TVSMOTOR.NS",
    "Eicher Motors":    "EICHERMOT.NS",
    "Ashok Leyland":    "ASHOKLEY.NS",
    "Tata Motors":      "TATAMOTORS.NS",
    "Samvardhana":      "MOTHERSON.NS",
    "Bosch":            "BOSCHLTD.NS",
    "Balkrishna Ind":   "BALKRISIND.NS",
    "MRF":              "MRF.NS",
    "Apollo Tyres":     "APOLLOTYRE.NS",
    # ── Pharma & Healthcare ───────────────────────────
    "Sun Pharma":       "SUNPHARMA.NS",
    "Dr Reddy":         "DRREDDY.NS",
    "Cipla":            "CIPLA.NS",
    "Divi's Lab":       "DIVISLAB.NS",
    "Lupin":            "LUPIN.NS",
    "Apollo Hosp":      "APOLLOHOSP.NS",
    "Mankind Pharma":   "MANKIND.NS",
    "Zydus Life":       "ZYDUSLIFE.NS",
    "Alkem Lab":        "ALKEM.NS",
    "Torrent Pharma":   "TORNTPHARM.NS",
    "Aurobindo":        "AUROPHARMA.NS",
    "Max Healthcare":   "MAXHEALTH.NS",
    # ── FMCG & Consumer ──────────────────────────────
    "HUL":              "HINDUNILVR.NS",
    "ITC":              "ITC.NS",
    "Nestle":           "NESTLEIND.NS",
    "Britannia":        "BRITANNIA.NS",
    "Dabur":            "DABUR.NS",
    "Tata Consumer":    "TATACONSUM.NS",
    "Marico":           "MARICO.NS",
    "Colgate":          "COLPAL.NS",
    "Emami":            "EMAMILTD.NS",
    "Godrej Consumer":  "GODREJCP.NS",
    "United Spirits":   "MCDOWELL-N.NS",
    # ── Metals & Mining ──────────────────────────────
    "Tata Steel":       "TATASTEEL.NS",
    "JSW Steel":        "JSWSTEEL.NS",
    "Hindalco":         "HINDALCO.NS",
    "Vedanta":          "VEDL.NS",
    "SAIL":             "SAIL.NS",
    "NMDC":             "NMDC.NS",
    "Jindal Steel":     "JSPL.NS",
    "APL Apollo":       "APLAPOLLO.NS",
    "Hindustan Zinc":   "HINDZINC.NS",
    "National Aluminium":"NATIONALUM.NS",
    # ── Infrastructure & Realty ───────────────────────
    "L&T":              "LT.NS",
    "UltraTech":        "ULTRACEMCO.NS",
    "DLF":              "DLF.NS",
    "Godrej Properties":"GODREJPROP.NS",
    "Prestige Estate":  "PRESTIGE.NS",
    "Oberoi Realty":    "OBEROIRLTY.NS",
    "Macrotech":        "LODHA.NS",
    "ACC":              "ACC.NS",
    "Ambuja Cement":    "AMBUJACEM.NS",
    "Shree Cement":     "SHREECEM.NS",
    "Dalmia Bharat":    "DALBHARAT.NS",
    # ── Adani Group ──────────────────────────────────
    "Adani Ports":      "ADANIPORTS.NS",
    "Adani Enterprises":"ADANIENT.NS",
    "Adani Total Gas":  "ATGL.NS",
    "Adani Wilmar":     "AWL.NS",
    "NDTV":             "NDTV.NS",
    # ── Tata Group ───────────────────────────────────
    "Tata Chemicals":   "TATACHEM.NS",
    "Titan":            "TITAN.NS",
    "Trent":            "TRENT.NS",
    "Tata Comm":        "TATACOMM.NS",
    "Indian Hotels":    "INDHOTEL.NS",
    "Voltas":           "VOLTAS.NS",
    # ── Defence & Railways ───────────────────────────
    "HAL":              "HAL.NS",
    "BEL":              "BEL.NS",
    "Mazagon Dock":     "MAZDOCK.NS",
    "RVNL":             "RVNL.NS",
    "IRFC":             "IRFC.NS",
    "IRCTC":            "IRCTC.NS",
    "IRCON":            "IRCON.NS",
    "BEML":             "BEML.NS",
    "Cochin Shipyard":  "COCHINSHIP.NS",
    "Garden Reach":     "GRSE.NS",
    "BDL":              "BDL.NS",
    "Data Patterns":    "DATAPATTNS.NS",
    # ── Telecom ──────────────────────────────────────
    "Bharti Airtel":    "BHARTIARTL.NS",
    "Vodafone Idea":    "IDEA.NS",
    "Indus Towers":     "INDUSTOWER.NS",
    # ── New Age & Consumer Tech ───────────────────────
    "Zomato":           "ZOMATO.NS",
    "Nykaa":            "NYKAA.NS",
    "Paytm":            "PAYTM.NS",
    "DMart":            "DMART.NS",
    "PB Fintech":       "POLICYBZR.NS",
    "CarTrade":         "CARTRADE.NS",
    # ── Chemicals & Specialty ─────────────────────────
    "Asian Paints":     "ASIANPAINT.NS",
    "Pidilite":         "PIDILITIND.NS",
    "SRF":              "SRF.NS",
    "Navin Fluorine":   "NAVINFLUOR.NS",
    "Deepak Nitrite":   "DEEPAKNTR.NS",
    "Atul Ltd":         "ATUL.NS",
    # ── Capital Goods & Engineering ───────────────────
    "Siemens":          "SIEMENS.NS",
    "ABB India":        "ABB.NS",
    "Havells":          "HAVELLS.NS",
    "Crompton":         "CROMPTON.NS",
    "Thermax":          "THERMAX.NS",
    "Cummins":          "CUMMINSIND.NS",
    "Bharat Forge":     "BHARATFORG.NS",
    "Schaeffler":       "SCHAEFFLER.NS",
    # ── Media & Entertainment ─────────────────────────
    "Sun TV":           "SUNTV.NS",
    "Zee Entertainment":"ZEEL.NS",
    "PVR Inox":         "PVRINOX.NS",
    # ── Additional Banking & NBFC ─────────────────────
    "RBL Bank":         "RBLBANK.NS",
    "Karnataka Bank":   "KTKBANK.NS",
    "DCB Bank":         "DCBBANK.NS",
    "Ujjivan Small":    "UJJIVANSFB.NS",
    "Equitas Small":    "EQUITASBNK.NS",
    "AU Small Finance": "AUBANK.NS",
    "Cholamandalam":    "CHOLAFIN.NS",
    "Mahindra Finance": "M&MFIN.NS",
    "Piramal Enterprises":"PIRAMALENT.NS",
    "Aditya Birla Cap": "ABCAPITAL.NS",
    "Manappuram":       "MANAPPURAM.NS",
    "IIFL Finance":     "IIFL.NS",
    # ── Additional IT ─────────────────────────────────
    "Oracle India":     "OFSS.NS",
    "Wipro":            "WIPRO.NS",
    "Hexaware":         "HEXAWARE.NS",
    "Sonata Software":  "SONATSOFTW.NS",
    "Mastek":           "MASTEK.NS",
    "Tanla Platforms":  "TANLA.NS",
    "Route Mobile":     "ROUTE.NS",
    "Intellect Design": "INTELLECT.NS",
    # ── Additional Pharma ─────────────────────────────
    "Biocon":           "BIOCON.NS",
    "Ipca Labs":        "IPCALAB.NS",
    "Ajanta Pharma":    "AJANTPHARM.NS",
    "Alembic Pharma":   "APLLTD.NS",
    "Gland Pharma":     "GLAND.NS",
    "Syngene":          "SYNGENE.NS",
    "Aarti Drugs":      "AARTIDRUGS.NS",
    "Laurus Labs":      "LAURUSLABS.NS",
    "Granules India":   "GRANULES.NS",
    "Suven Pharma":     "SUVENPHAR.NS",
    # ── Additional FMCG & Consumer ────────────────────
    "Varun Beverages":  "VBL.NS",
    "United Breweries": "UBL.NS",
    "Jubilant Food":    "JUBLFOOD.NS",
    "Westlife Food":    "WESTLIFE.NS",
    "Restaurant Brands":"RBASKETSS.NS",
    "Devyani Intl":     "DEVYANI.NS",
    "Bikaji Foods":     "BIKAJI.NS",
    "Mrs Bectors":      "BECTORFOOD.NS",
    "Radico Khaitan":   "RADICO.NS",
    "Globus Spirits":   "GLOBUSSPR.NS",
    # ── Additional Auto ───────────────────────────────
    "Mahindra CIE":     "MAHINDCIE.NS",
    "Endurance Tech":   "ENDURANCE.NS",
    "Minda Corp":       "MINDACORP.NS",
    "Suprajit Engg":    "SUPRAJIT.NS",
    "Craftsman Auto":   "CRAFTSMAN.NS",
    "Sona BLW":         "SONACOMS.NS",
    "Uno Minda":        "UNOMINDA.NS",
    # ── Additional Metals ─────────────────────────────
    "Welspun Corp":     "WELCORP.NS",
    "Ratnamani Metals": "RATNAMANI.NS",
    "Mishra Dhatu":     "MIDHANI.NS",
    "Gravita India":    "GRAVITA.NS",
    "Hindustan Copper": "HINDCOPPER.NS",
    # ── Textiles ──────────────────────────────────────
    "Page Industries":  "PAGEIND.NS",
    "Trident":          "TRIDENT.NS",
    "Vardhman Textile": "VTL.NS",
    "GHCL":             "GHCL.NS",
    "KPR Mill":         "KPRMILL.NS",
    "Welspun India":    "WELSPUNLTD.NS",
    # ── Real Estate & Construction ────────────────────
    "NCC Ltd":          "NCC.NS",
    "KNR Constructions":"KNRCON.NS",
    "PNC Infratech":    "PNCINFRA.NS",
    "IRB Infra":        "IRB.NS",
    "Ashoka Buildcon":  "ASHOKA.NS",
    "PSP Projects":     "PSPPROJECT.NS",
    "Brigade Enterprises":"BRIGADE.NS",
    "Sobha":            "SOBHA.NS",
    "Signature Global": "SIGNATURE.NS",
    # ── Power & Utilities ─────────────────────────────
    "Tata Consultancy": "CESC.NS",
    "CESC":             "CESC.NS",
    "Torrent Power":    "TORNTPOWER.NS",
    "GIPCL":            "GIPCL.NS",
    "GE Vernova":       "GEVERNOVA.NS",
    "Hitachi Energy":   "POWERINDIA.NS",
    "KPI Green":        "KPIGREEN.NS",
    "Waaree Energies":  "WAAREEENER.NS",
    "Premier Energies": "PREMIERENS.NS",
    # ── Agrochemicals & Fertilisers ───────────────────
    "UPL":              "UPL.NS",
    "PI Industries":    "PIIND.NS",
    "Coromandel Intl":  "COROMANDEL.NS",
    "Bayer Cropscience":"BAYERCROP.NS",
    "Chambal Fert":     "CHAMBLFERT.NS",
    "Rallis India":     "RALLIS.NS",
    "Astec Lifesciences":"ASTEC.NS",
    "Dhanuka Agritech": "DHANUKA.NS",
    # ── Logistics & Shipping ──────────────────────────
    "Container Corp":   "CONCOR.NS",
    "VRL Logistics":    "VRLLOG.NS",
    "TCI Express":      "TCIEXP.NS",
    "Delhivery":        "DELHIVERY.NS",
    "Blue Dart":        "BLUEDART.NS",
    "Mahindra Logistics":"MAHLOG.NS",
    "Gateway Distriparks":"GDL.NS",
    "Shipping Corp":    "SCI.NS",
    # ── Healthcare Devices & Hospitals ────────────────
    "Narayana Hrudayalaya":"NH.NS",
    "Fortis Healthcare":"FORTIS.NS",
    "Krishna Institute":"KIMS.NS",
    "Global Health":    "MEDANTA.NS",
    "Aster DM":         "ASTERDM.NS",
    "Poly Medicure":    "POLYMED.NS",
    "Vijaya Diagnostic":"VIJAYA.NS",
    # ── Specialty Finance ─────────────────────────────
    "BSE Ltd":          "BSE.NS",
    "CDSL":             "CDSL.NS",
    "CAMS":             "CAMS.NS",
    "KFintech":         "KFINTECH.NS",
    "Angel One":        "ANGELONE.NS",
    "5Paisa Capital":   "5PAISA.NS",
    "MCX":              "MCX.NS",
    "Multi Comm Exch":  "MCX.NS",
    # ── Hotels & Tourism ──────────────────────────────
    "EIH Ltd":          "EIHOTEL.NS",
    "Lemon Tree":       "LEMONTREE.NS",
    "Chalet Hotels":    "CHALET.NS",
    "Mahindra Holidays":"MHRIL.NS",
    "Thomas Cook":      "THOMASCOOK.NS",
    # ── Retail ────────────────────────────────────────
    "Titan":            "TITAN.NS",
    "Kalyan Jewellers": "KALYANKJIL.NS",
    "Senco Gold":       "SENCO.NS",
    "PC Jeweller":      "PCJEWELLER.NS",
    "Shopper Stop":     "SHOPERSTOP.NS",
    "V-Mart Retail":    "VMART.NS",
    "Aditya Birla Fash":"ABFRL.NS",
    "Vedant Fashions":  "MANYAVAR.NS",
    # ── Miscellaneous ─────────────────────────────────
    "3M India":         "3MINDIA.NS",
    "Honeywell Auto":   "HONAUT.NS",
    "ABB India":        "ABB.NS",
    "Thermax":          "THERMAX.NS",
    "Triveni Turbine":  "TRITURBINE.NS",
    "Lakshmi Machine":  "LMWLTD.NS",
    "TD Power Systems": "TDPOWERSYS.NS",
    "Elecon Engg":      "ELECON.NS",
    "Rexnord Elect":    "REXNORD.NS",
    "Kaynes Tech":      "KAYNES.NS",
    "Syrma SGS":        "SYRMA.NS",
    "Avalon Tech":      "AVALON.NS",

    # ══════════════════════════════════════════════════════
    # COMMODITIES — MCX (Multi Commodity Exchange)
    # Yahoo Finance tickers for Indian commodities
    # ══════════════════════════════════════════════════════

    # ── Precious Metals ───────────────────────────────────
    "Gold (MCX)":           "GC=F",     # Gold Futures (USD) - global
    "Silver (MCX)":         "SI=F",     # Silver Futures (USD) - global
    "Gold Spot":            "GLD",      # SPDR Gold ETF proxy
    "Silver Spot":          "SLV",      # iShares Silver ETF proxy

    # ── India Gold/Silver ETFs on NSE ─────────────────────
    "Nippon Gold ETF":      "GOLDBEES.NS",
    "SBI Gold ETF":         "SETFGOLD.NS",
    "Nippon Silver ETF":    "SILVERBEES.NS",

    # ── Energy ────────────────────────────────────────────
    "Crude Oil (MCX)":      "CL=F",     # WTI Crude Futures
    "Brent Crude":          "BZ=F",     # Brent Crude Futures
    "Natural Gas (MCX)":    "NG=F",     # Natural Gas Futures

    # ── Base Metals ───────────────────────────────────────
    "Copper (MCX)":         "HG=F",     # Copper Futures
    "Aluminium":            "ALI=F",    # Aluminium Futures
    "Nickel":               "NI=F",     # Nickel Futures
    "Zinc":                 "ZNC=F",    # Zinc Futures
    "Lead":                 "LE=F",     # Lead Futures

    # ── Agricultural ──────────────────────────────────────
    "Cotton (MCX)":         "CT=F",     # Cotton Futures
    "Mentha Oil":           "MO=F",     # Mentha Oil
    "Cardamom":             "CD=F",     # Cardamom
    "Castor Seed":          "CS=F",     # Castor Seed
    "Crude Palm Oil":       "KO=F",     # Palm Oil

    # ── Global Commodity ETFs (tradeable via NSE) ─────────
    "Mirae Commodity ETF":  "MCXCMMDTY.NS",

    # ── Currency & Forex ──────────────────────────────────
    "USD/INR":              "USDINR=X",
    "EUR/INR":              "EURINR=X",
    "GBP/INR":              "GBPINR=X",
    "JPY/INR":              "JPYINR=X",

    # ── Global Indices (for context) ──────────────────────
    "Dow Jones":            "^DJI",
    "S&P 500":              "^GSPC",
    "NASDAQ":               "^IXIC",
    "Nikkei 225":           "^N225",
    "Hang Seng":            "^HSI",
    "FTSE 100":             "^FTSE",
    "Gift Nifty":           "^NSEMDCP50", # proxy
}

SECTORS = {
    "🏆 Top 30 F&O": [
        "NIFTY 50","BANK NIFTY","Reliance","HDFC Bank",
        "ICICI Bank","TCS","Infosys","SBI","Wipro",
        "Bajaj Finance","ITC","Sun Pharma","L&T","Maruti",
        "Coal India","NTPC","Bharti Airtel","Tata Steel",
        "Axis Bank","HCL Tech","Power Grid","Adani Ports",
        "Hindalco","ONGC","Bajaj Auto","Titan",
        "JSW Steel","UltraTech","BEL","Adani Enterprises",
    ],
    "🏦 Banking": [
        "HDFC Bank","ICICI Bank","SBI","Kotak Bank",
        "Axis Bank","IndusInd Bank","Bajaj Finance",
        "PNB","Bank of Baroda","Canara Bank",
        "Federal Bank","IDFC First","Bajaj Finserv",
        "Yes Bank","Shriram Finance","Muthoot Finance",
    ],
    "💻 IT": [
        "TCS","Infosys","Wipro","HCL Tech",
        "Tech Mahindra","Persistent","Coforge",
        "LTIMindtree","Mphasis","Tata Elxsi","KPIT Tech",
    ],
    "🛢️ Energy & Power": [
        "Reliance","ONGC","Indian Oil","BPCL",
        "NTPC","Power Grid","Adani Green","Adani Power",
        "Tata Power","Gail","NHPC","JSW Energy",
        "Torrent Power","Coal India",
    ],
    "🚗 Auto": [
        "Maruti","M&M","Hero MotoCorp","Bajaj Auto",
        "TVS Motor","Eicher Motors","Ashok Leyland",
        "Tata Motors","Bosch","MRF","Apollo Tyres",
    ],
    "💊 Pharma": [
        "Sun Pharma","Dr Reddy","Cipla","Divi's Lab",
        "Lupin","Apollo Hosp","Mankind Pharma",
        "Zydus Life","Torrent Pharma","Aurobindo",
    ],
    "🛒 FMCG": [
        "HUL","ITC","Nestle","Britannia","Dabur",
        "Tata Consumer","Marico","Colgate","Emami",
    ],
    "⚙️ Metals": [
        "Tata Steel","JSW Steel","Hindalco","Vedanta",
        "SAIL","NMDC","Jindal Steel","Hindustan Zinc",
    ],
    "🛡️ Defence": [
        "HAL","BEL","Mazagon Dock","RVNL","IRFC",
        "IRCTC","BEML","Cochin Shipyard","Garden Reach","BDL",
    ],
    "🏗️ Infra & Cement": [
        "L&T","UltraTech","DLF","Adani Ports",
        "Godrej Properties","ACC","Ambuja Cement",
        "Shree Cement","Prestige Estate",
    ],
    "⚡ Adani Group": [
        "Adani Ports","Adani Enterprises","Adani Green",
        "Adani Power","Adani Total Gas","Adani Energy",
    ],
    "🔬 Chemicals": [
        "Asian Paints","Pidilite","SRF","Navin Fluorine",
        "Deepak Nitrite","Atul Ltd",
    ],
    "🏭 Capital Goods": [
        "Siemens","ABB India","Havells","Thermax",
        "Cummins","Bharat Forge","Crompton",
        "Kaynes Tech","Syrma SGS","Triveni Turbine",
    ],
    "🌾 Agro & Fertilisers": [
        "UPL","PI Industries","Coromandel Intl",
        "Chambal Fert","Rallis India","Dhanuka Agritech",
    ],
    "🚚 Logistics": [
        "Container Corp","Delhivery","Blue Dart",
        "VRL Logistics","TCI Express","Gateway Distriparks",
    ],
    "🏨 Hotels & Retail": [
        "Indian Hotels","EIH Ltd","Lemon Tree",
        "Titan","Kalyan Jewellers","Senco Gold",
        "Vedant Fashions","DMart","Trent",
    ],
    "🏥 Healthcare": [
        "Apollo Hosp","Narayana Hrudayalaya","Fortis Healthcare",
        "Global Health","Aster DM","Max Healthcare",
        "Krishna Institute","Vijaya Diagnostic",
    ],
    "💹 Broking & Exchanges": [
        "BSE Ltd","CDSL","CAMS","Angel One","MCX",
        "KFintech","5Paisa Capital",
    ],
    "🧵 Textiles": [
        "Page Industries","Trident","Vardhman Textile",
        "KPR Mill","Welspun India","GHCL",
    ],
    "☀️ Renewable Energy": [
        "Adani Green","Adani Power","Tata Power",
        "NHPC","JSW Energy","KPI Green",
        "Waaree Energies","Premier Energies",
    ],
    "🏗️ Construction": [
        "NCC Ltd","KNR Constructions","IRB Infra",
        "Ashoka Buildcon","PSP Projects","RVNL",
        "PNC Infratech","L&T",
    ],

    # ── Commodity Sectors ─────────────────────────────────
    "🥇 Precious Metals": [
        "Gold (MCX)","Silver (MCX)",
        "Nippon Gold ETF","SBI Gold ETF",
        "HDFC Gold ETF","Nippon Silver ETF",
    ],
    "🛢️ Energy Commodities": [
        "Crude Oil (MCX)","Brent Crude","Natural Gas (MCX)",
    ],
    "⚙️ Base Metals MCX": [
        "Copper (MCX)","Aluminium","Lead",
    ],
    "🌾 Agricultural MCX": [
        "Cotton (MCX)",
    ],
    "💱 Currency": [
        "USD/INR","EUR/INR","GBP/INR","JPY/INR",
    ],
    "🌍 Global Indices": [
        "Dow Jones","S&P 500","NASDAQ",
        "Nikkei 225","Hang Seng","FTSE 100",
    ],
}

LOT_SIZES = {
    "^NSEI":250,"^NSEBANK":15,
    "RELIANCE.NS":250,"TCS.NS":150,
    "HDFCBANK.NS":550,"ICICIBANK.NS":700,
    "INFY.NS":300,"SBIN.NS":1500,
    "BAJFINANCE.NS":125,"ASHOKLEY.NS":1500,
    "AXISBANK.NS":1200,"ITC.NS":3200,
    "WIPRO.NS":3000,"KOTAKBANK.NS":400,
    "HCLTECH.NS":350,"TECHM.NS":600,
}

# ══════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════
def now_ist():
    return datetime.now(IST)

def market_open():
    n = now_ist()
    if n.weekday() >= 5:
        return False
    return (n.replace(hour=9,  minute=15, second=0) <= n <=
            n.replace(hour=15, minute=30, second=0))

def best_trading_time():
    n   = now_ist()
    hr  = n.hour
    mn  = n.minute
    t   = hr * 60 + mn
    if t < 9*60+15:  return "pre_market"
    if t < 9*60+30:  return "avoid"       # opening 15 min
    if t < 10*60+30: return "best"        # 9:30–10:30
    if t < 12*60:    return "good"        # 10:30–12:00
    if t < 13*60+30: return "ok"          # lunch
    if t < 14*60+30: return "good"        # afternoon
    if t < 15*60:    return "caution"     # 2:30–3:00
    if t <= 15*60+30:return "avoid"       # last 30 min
    return "closed"

def live_price(sym: str) -> dict:
    """
    Gets live price — uses Kite if connected, else Yahoo Finance.
    Kite = real-time tick data (zero delay).
    Yahoo = ~15 second delay via fast_info.
    No cache when Kite is connected — always fresh.
    """
    # ── Try Kite first (real-time) ────────────────────────
    kite = get_kite()
    if kite and not sym.startswith("^"):
        try:
            # Convert Yahoo symbol to NSE symbol
            # e.g. HDFCBANK.NS -> NSE:HDFCBANK
            nse_sym = sym.replace(".NS","").replace(".BO","")
            exchange = "BSE" if ".BO" in sym else "NSE"
            quote_key = f"{exchange}:{nse_sym}"
            quote = kite.quote([quote_key])
            q     = quote[quote_key]
            p     = float(q["last_price"])
            pv    = float(q["ohlc"]["close"])
            ch    = round(((p-pv)/pv)*100, 2) if pv else 0
            return {
                "ok":True,
                "p":round(p,2),
                "prev":round(pv,2),
                "chg":ch,
                "chg_abs":round(p-pv,2),
                "high":round(float(q["ohlc"]["high"]),2),
                "low": round(float(q["ohlc"]["low"]),2),
                "source":"kite"
            }
        except Exception as _kite_err:
            # Store error for debugging
            st.session_state["kite_candle_error"] = str(_kite_err)
            pass  # Fall through to Yahoo

    # ── Yahoo Finance fallback ────────────────────────────
    @st.cache_data(ttl=15)
    def _yahoo_price(sym_: str) -> dict:
        try:
            tk = yf.Ticker(sym_)
            fi = tk.fast_info
            p  = float(fi.last_price)
            pv = float(fi.previous_close)
            if not p or p <= 0:
                hist = tk.history(period="2d", interval="1d")
                if not hist.empty:
                    p  = float(hist["Close"].iloc[-1])
                    pv = float(hist["Close"].iloc[-2]) if len(hist)>1 else p
            ch = round(((p-pv)/pv)*100, 2) if pv else 0
            return {
                "ok":True, "p":round(p,2), "prev":round(pv,2),
                "chg":ch, "chg_abs":round(p-pv,2),
                "high":round(float(fi.day_high or p),2),
                "low": round(float(fi.day_low  or p),2),
                "source":"yahoo"
            }
        except:
            return {"ok":False,"p":0,"prev":0,"chg":0,
                    "chg_abs":0,"high":0,"low":0,
                    "source":"yahoo"}
    return _yahoo_price(sym)

# ── Persistent Trade Storage ──────────────────────────────
import json as _json
import os as _os

_TRADES_FILE  = "active_trades.json"
_JOURNAL_FILE = "trade_journal.json"

def save_paper_trades(trades: list):
    """Save paper trades to JSON file."""
    try:
        with open("paper_trades.json", "w") as _f:
            _json.dump(trades, _f, indent=2, default=str)
    except Exception:
        pass

def load_paper_trades() -> list:
    """Load paper trades from JSON file."""
    try:
        if _os.path.exists("paper_trades.json"):
            with open("paper_trades.json", "r") as _f:
                data = _json.load(_f)
                return data if isinstance(data, list) else []
    except Exception:
        pass
    return []


def save_trades(trades: list):
    """Save trades to JSON file — persists across page refresh."""
    try:
        with open(_TRADES_FILE, "w") as _f:
            _json.dump(trades, _f, indent=2, default=str)
    except Exception:
        pass

def load_trades() -> list:
    """Load trades from JSON file on startup."""
    try:
        if _os.path.exists(_TRADES_FILE):
            with open(_TRADES_FILE, "r") as _f:
                data = _json.load(_f)
                return data if isinstance(data, list) else []
    except Exception:
        pass
    return []

def save_journal(journal: list):
    """Save trade journal to JSON file."""
    try:
        with open(_JOURNAL_FILE, "w") as _f:
            _json.dump(journal, _f, indent=2, default=str)
    except Exception:
        pass

def load_journal() -> list:
    """Load trade journal from JSON file."""
    try:
        if _os.path.exists(_JOURNAL_FILE):
            with open(_JOURNAL_FILE, "r") as _f:
                data = _json.load(_f)
                return data if isinstance(data, list) else []
    except Exception:
        pass
    return []


def get_kite_instruments(_token: str = "") -> dict:
    """
    Cache NSE instruments in st.session_state.
    session_state persists across Streamlit reruns unlike globals.
    Returns dict: {tradingsymbol: instrument_token}
    """
    if not _token or not KITE_AVAILABLE or not KITE_API_KEY:
        return {}

    # Check if we already have instruments for this token
    cached_token = st.session_state.get("kite_inst_token", "")
    cached_map   = st.session_state.get("kite_inst_map", {})

    if cached_token == _token and cached_map:
        return cached_map

    # Fetch fresh from Kite
    try:
        kite_inst = KiteConnect(api_key=KITE_API_KEY)
        kite_inst.set_access_token(_token)
        instruments = kite_inst.instruments("NSE")
        result = {
            inst["tradingsymbol"]: inst["instrument_token"]
            for inst in instruments
        }
        # Save to session state — persists across reruns
        st.session_state["kite_inst_map"]   = result
        st.session_state["kite_inst_token"] = _token
        st.session_state["kite_inst_count"] = len(result)
        return result
    except Exception as e:
        st.session_state["kite_inst_error"] = str(e)
        return {}

def candles(sym: str, interval: str) -> pd.DataFrame:
    """
    Fetch OHLCV candles — Kite if connected (real-time),
    else Yahoo Finance (~15 min delay for intraday).
    """
    # ── Try Kite first ────────────────────────────────────
    kite = get_kite()

    # Skip Kite for symbols that are Yahoo Finance only
    # These include: indices (^), currencies (=X), commodities
    _yahoo_only = (
        sym.startswith("^") or
        sym.endswith("=X") or
        sym.endswith("=F") or
        "GC" in sym or "CL" in sym or "SI" in sym
    )

    if kite and not _yahoo_only:
        try:
            nse_sym = sym.replace(".NS","").replace(".BO","")

            # Kite interval mapping
            kite_interval = {
                "1m":  "minute",
                "5m":  "5minute",
                "15m": "15minute",
                "30m": "30minute",
                "1h":  "60minute",
                "1d":  "day",
            }.get(interval, "15minute")

            # Date range
            days_ = {
                "minute":2,"5minute":20,
                "15minute":60,"30minute":60,
                "60minute":200,"day":2000
            }.get(kite_interval, 60)

            from_dt = datetime.now() - timedelta(days=days_)
            to_dt   = datetime.now()

            # Use cached instruments lookup
            # Pass token so cache busts when new login happens
            _cur_token = st.session_state.get(
                "kite_access_token", ""
            )
            instruments_map = get_kite_instruments(_cur_token)
            inst_token = instruments_map.get(nse_sym)

            # Debug: store which path was taken
            if inst_token:
                st.session_state["kite_data_source"] = (
                    f"Kite: {nse_sym} token {inst_token}"
                )
            else:
                st.session_state["kite_data_source"] = (
                    f"Kite: {nse_sym} NOT FOUND in instruments"
                )

            if inst_token:
                hist = kite.historical_data(
                    inst_token,
                    from_date=from_dt,
                    to_date=to_dt,
                    interval=kite_interval,
                    continuous=False
                )
                if hist:
                    df_k = pd.DataFrame(hist)
                    df_k.rename(columns={
                        "date":"Date",
                        "open":"Open","high":"High",
                        "low":"Low","close":"Close",
                        "volume":"Volume"
                    }, inplace=True)
                    df_k.set_index("Date", inplace=True)
                    if df_k.index.tzinfo is None:
                        df_k.index = df_k.index.tz_localize(IST)
                    else:
                        df_k.index = df_k.index.tz_convert(IST)
                    df_k.dropna(inplace=True)
                    return df_k
        except Exception as _kite_err:
            # Store error for debugging
            st.session_state["kite_candle_error"] = str(_kite_err)
            pass  # Fall through to Yahoo

    # ── Yahoo Finance fallback (cached) ───────────────────
    @st.cache_data(ttl=60)
    def _yahoo_candles(sym_: str, interval_: str) -> pd.DataFrame:
        days = {"1m":1,"5m":4,"15m":20,"30m":40,
                "1h":59,"1d":300}.get(interval_, 20)
        end   = datetime.now()
        start = end - timedelta(days=days)
        try:
            df = yf.download(sym_,
                start=start.strftime("%Y-%m-%d"),
                end=(end+timedelta(days=1)).strftime("%Y-%m-%d"),
                interval=interval_,
                auto_adjust=True, progress=False,
                threads=False)
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            df.dropna(inplace=True)
            if not df.empty:
                df.index = (df.index.tz_convert(IST)
                            if df.index.tzinfo
                            else df.index.tz_localize("UTC")
                                         .tz_convert(IST))
            return df
        except:
            return pd.DataFrame()
    return _yahoo_candles(sym, interval)

@st.cache_data(ttl=300)
def bulk_prices(names: list) -> pd.DataFrame:
    syms = [STOCKS[n] for n in names if n in STOCKS]
    nmap = {STOCKS[n]:n for n in names if n in STOCKS}
    if not syms: return pd.DataFrame()
    try:
        import warnings
        warnings.filterwarnings("ignore")
        raw = yf.download(syms, period="2d", interval="1d",
                          auto_adjust=True, progress=False,
                          group_by="ticker")
        rows = []
        for s in syms:
            nm = nmap.get(s, s)
            try:
                # Handle both single and multi ticker downloads
                if len(syms) == 1:
                    cc = raw["Close"].dropna()
                    hh = raw["High"].dropna()
                    ll = raw["Low"].dropna()
                elif s in raw.columns.get_level_values(1):
                    cc = raw["Close"][s].dropna()
                    hh = raw["High"][s].dropna()
                    ll = raw["Low"][s].dropna()
                elif hasattr(raw, 'columns') and s in raw:
                    cc = raw[s]["Close"].dropna()
                    hh = raw[s]["High"].dropna()
                    ll = raw[s]["Low"].dropna()
                else:
                    raise ValueError("ticker not in data")
                if len(cc) < 2: raise ValueError("not enough rows")
                pr = float(cc.iloc[-1])
                pv = float(cc.iloc[-2])
                ch = round(((pr-pv)/pv)*100, 2)
                rows.append({"Name":nm,"Sym":s,
                    "Price":round(pr,2),"Chg%":ch,
                    "Chg₹":round(pr-pv,2),
                    "High":round(float(hh.iloc[-1]),2),
                    "Low": round(float(ll.iloc[-1]),2)})
            except:
                rows.append({"Name":nm,"Sym":s,"Price":None,
                             "Chg%":None,"Chg₹":None,
                             "High":None,"Low":None})
        return pd.DataFrame(rows)
    except:
        return pd.DataFrame()

@st.cache_data(ttl=600)
def get_news(q: str) -> list:
    try:
        url  = (f"https://news.google.com/rss/search"
                f"?q={q.replace(' ','+')}"
                f"&hl=en-IN&gl=IN&ceid=IN:en")
        resp = requests.get(
            url, headers={"User-Agent":"Mozilla/5.0"},
            timeout=6)
        root = ET.fromstring(resp.content)
        out  = []
        for item in root.findall(".//item")[:10]:
            ttl = item.findtext("title","")
            lnk = item.findtext("link","")
            dt  = item.findtext("pubDate","")[:22]
            src = ""
            if " - " in ttl:
                ttl, src = ttl.rsplit(" - ",1)
            out.append({"title":ttl.strip(),"link":lnk,
                        "date":dt,"src":src.strip()})
        return out
    except:
        return []


# ══════════════════════════════════════════════════════════
# TELEGRAM HELPER  (global — used by scanner + test button)
# ══════════════════════════════════════════════════════════
def send_telegram(token: str, chat_id: str,
                  message: str) -> bool:
    """Sends message via official Telegram Bot API."""
    try:
        resp = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={
                "chat_id":    str(chat_id).strip(),
                "text":       message,
                "parse_mode": "HTML"
            },
            timeout=10
        )
        return resp.status_code == 200
    except Exception:
        return False

def tg_configured() -> bool:
    """True if token + chat_id are saved in session state."""
    return (
        bool(st.session_state.get("tg_token_saved", "")) and
        bool(st.session_state.get("tg_chat_saved",  ""))
    )

# ══════════════════════════════════════════════════════════
# SIGNAL ENGINE  (indicators + 11-factor checklist)
# ══════════════════════════════════════════════════════════
def compute_all(df: pd.DataFrame, lp: dict) -> dict | None:
    if df is None or len(df) < 55: return None
    try:
        c = df["Close"].squeeze().astype(float)
        h = df["High"].squeeze().astype(float)
        l = df["Low"].squeeze().astype(float)
        v = df["Volume"].squeeze().astype(float)
        o = df["Open"].squeeze().astype(float)

        # ── Indicators ────────────────────────────────────
        e9   = ta.trend.ema_indicator(c, window=9)
        e21  = ta.trend.ema_indicator(c, window=21)
        e50  = ta.trend.ema_indicator(c, window=50)
        rsi  = ta.momentum.rsi(c, window=14)
        macd = ta.trend.macd(c)
        msig = ta.trend.macd_signal(c)
        adx  = ta.trend.adx(h, l, c, window=14)
        vwap = (c * v).cumsum() / v.cumsum()
        bbu  = ta.volatility.bollinger_hband(c, window=20)
        bbl  = ta.volatility.bollinger_lband(c, window=20)
        atr  = ta.volatility.average_true_range(
                   h, l, c, window=14)
        cmf  = ta.volume.chaikin_money_flow(
                   h, l, c, v, window=20)
        obv  = ta.volume.on_balance_volume(c, v)

        cp   = lp["p"] if lp["ok"] else float(c.iloc[-1])
        e9v  = float(e9.iloc[-1])
        e21v = float(e21.iloc[-1])
        e50v = float(e50.iloc[-1])
        rv   = float(rsi.iloc[-1])
        rv1  = float(rsi.iloc[-2])
        mv   = float(macd.iloc[-1])
        msv  = float(msig.iloc[-1])
        mv1  = float(macd.iloc[-2])
        msv1 = float(msig.iloc[-2])
        adxv = float(adx.iloc[-1])
        vwv  = float(vwap.iloc[-1])

        # VWAP Standard Deviation Bands
        # +1 SD = overbought, -1 SD = oversold
        try:
            _vwap_std  = float(c.tail(20).std())
            vwap_upper = round(vwv + _vwap_std, 2)
            vwap_lower = round(vwv - _vwap_std, 2)
            vwap_u2    = round(vwv + 2 * _vwap_std, 2)
            vwap_l2    = round(vwv - 2 * _vwap_std, 2)
            # Position relative to VWAP bands
            if cp >= vwap_u2:
                vwap_zone = "EXTREME_OB"  # extremely overbought
            elif cp >= vwap_upper:
                vwap_zone = "OVERBOUGHT"
            elif cp <= vwap_l2:
                vwap_zone = "EXTREME_OS"  # extremely oversold
            elif cp <= vwap_lower:
                vwap_zone = "OVERSOLD"
            else:
                vwap_zone = "FAIR_VALUE"
        except Exception:
            vwap_upper = round(vwv * 1.005, 2)
            vwap_lower = round(vwv * 0.995, 2)
            vwap_u2    = round(vwv * 1.010, 2)
            vwap_l2    = round(vwv * 0.990, 2)
            vwap_zone  = "FAIR_VALUE"

        atrv = float(atr.iloc[-1])
        # Validate ATR — must be positive and reasonable
        # If NaN or zero or negative — use 1% of price as fallback
        if not atrv or atrv <= 0 or atrv != atrv:  # NaN check
            atrv = round(cp * 0.01, 2)
        # ATR should not exceed 5% of price (sanity check)
        if atrv > cp * 0.05:
            atrv = round(cp * 0.015, 2)
        cmfv = float(cmf.iloc[-1])
        bbup = float(bbu.iloc[-1])
        bblw = float(bbl.iloc[-1])

        vol_avg   = float(v.tail(20).mean())
        vol_ratio = round(float(v.iloc[-1])/(vol_avg+1e-9),2)
        vsurge    = vol_ratio >= 1.2

        # OBV trend
        obv_bull = float(obv.iloc[-1]) > float(obv.iloc[-5])

        # ── CPR (Central Pivot Range) ─────────────────────
        # MUST use previous FULL DAY data for CPR
        # Using previous candle is wrong for intraday charts
        try:
            # Group candles by date and get previous day OHLC
            df_daily = df.copy()
            df_daily.index = pd.to_datetime(df_daily.index)
            daily_grp = df_daily.resample("1D").agg({
                "High":  "max",
                "Low":   "min",
                "Close": "last",
                "Open":  "first"
            }).dropna()

            if len(daily_grp) >= 2:
                # Use yesterday's full day data
                prev_h = float(daily_grp["High"].iloc[-2])
                prev_l = float(daily_grp["Low"].iloc[-2])
                prev_c = float(daily_grp["Close"].iloc[-2])
            else:
                # Fallback: use last 20 candles range
                prev_h = float(h.tail(20).max())
                prev_l = float(l.tail(20).min())
                prev_c = float(c.iloc[-1])
        except Exception:
            prev_h = float(h.tail(20).max())
            prev_l = float(l.tail(20).min())
            prev_c = float(c.iloc[-1])

        cpr_pivot = round((prev_h + prev_l + prev_c) / 3, 2)
        cpr_bc    = round((prev_h + prev_l) / 2, 2)
        cpr_tc    = round(cpr_pivot - cpr_bc + cpr_pivot, 2)

        # Make sure TC > BC
        if cpr_tc < cpr_bc:
            cpr_tc, cpr_bc = cpr_bc, cpr_tc

        cpr_width    = round(cpr_tc - cpr_bc, 2)
        cpr_width_pct= round(cpr_width / cpr_pivot * 100, 2)

        # CPR width classification
        # Narrow = trending day, Wide = sideways day
        if cpr_width_pct < 0.25:
            cpr_type = "Narrow (Trending day expected)"
        elif cpr_width_pct < 0.5:
            cpr_type = "Moderate"
        else:
            cpr_type = "Wide (Sideways day expected)"

        # Price position relative to CPR
        if cp > cpr_tc:
            cpr_position = "ABOVE"   # bullish
            cpr_bias     = "Bullish"
        elif cp < cpr_bc:
            cpr_position = "BELOW"   # bearish
            cpr_bias     = "Bearish"
        else:
            cpr_position = "INSIDE"  # sideways
            cpr_bias     = "Sideways"

        # Virgin CPR — price never touched yesterday's CPR
        # (strong magnet for price today)
        virgin_cpr = not (
            float(l.iloc[-1]) <= cpr_tc and
            float(h.iloc[-1]) >= cpr_bc
        )

        # ── Supertrend Indicator ─────────────────────────
        # Supertrend = ATR-based trend following indicator
        # Period=7, Multiplier=3 (standard settings)
        st_period = 7
        st_mult   = 3.0
        try:
            st_atr = ta.volatility.AverageTrueRange(
                h, l, c, window=st_period
            ).average_true_range()

            hl2 = (h + l) / 2
            upper_band = hl2 + (st_mult * st_atr)
            lower_band = hl2 - (st_mult * st_atr)

            # Calculate Supertrend
            st_vals  = [0.0] * len(c)
            st_trend = [1]   * len(c)  # 1=uptrend, -1=downtrend

            for idx in range(1, len(c)):
                # Upper band
                if (upper_band.iloc[idx] < upper_band.iloc[idx-1] or
                        float(c.iloc[idx-1]) > upper_band.iloc[idx-1]):
                    ub = float(upper_band.iloc[idx])
                else:
                    ub = float(upper_band.iloc[idx-1])

                # Lower band
                if (lower_band.iloc[idx] > lower_band.iloc[idx-1] or
                        float(c.iloc[idx-1]) < lower_band.iloc[idx-1]):
                    lb = float(lower_band.iloc[idx])
                else:
                    lb = float(lower_band.iloc[idx-1])

                # Trend direction
                if st_trend[idx-1] == -1:
                    if float(c.iloc[idx]) > ub:
                        st_trend[idx] = 1
                        st_vals[idx]  = lb
                    else:
                        st_trend[idx] = -1
                        st_vals[idx]  = ub
                else:
                    if float(c.iloc[idx]) < lb:
                        st_trend[idx] = -1
                        st_vals[idx]  = ub
                    else:
                        st_trend[idx] = 1
                        st_vals[idx]  = lb

            st_value    = round(st_vals[-1], 2)
            st_dir      = st_trend[-1]   # 1=buy, -1=sell
            st_bull     = st_dir == 1
            st_crossed  = st_trend[-1] != st_trend[-2]  # fresh crossover
            st_signal   = ("BUY" if st_bull else "SELL")
            st_label    = (f"Supertrend BUY ₹{st_value:,.2f}"
                           if st_bull
                           else f"Supertrend SELL ₹{st_value:,.2f}")
        except Exception:
            st_value   = 0.0
            st_bull    = cp > float(c.mean())
            st_crossed = False
            st_signal  = "BUY" if st_bull else "SELL"
            st_label   = st_signal
            st_dir     = 1 if st_bull else -1

        # ── Fibonacci Retracement Levels ──────────────────
        # Calculate from recent swing high and swing low (last 50 candles)
        fib_period = min(50, len(c))
        fib_high   = float(h.tail(fib_period).max())
        fib_low    = float(l.tail(fib_period).min())
        fib_range  = fib_high - fib_low

        # Key Fibonacci levels
        fib_236 = round(fib_high - 0.236 * fib_range, 2)
        fib_382 = round(fib_high - 0.382 * fib_range, 2)
        fib_500 = round(fib_high - 0.500 * fib_range, 2)
        fib_618 = round(fib_high - 0.618 * fib_range, 2)
        fib_786 = round(fib_high - 0.786 * fib_range, 2)

        # Find nearest Fibonacci level to current price
        fib_levels = {
            "23.6%": fib_236,
            "38.2%": fib_382,
            "50.0%": fib_500,
            "61.8%": fib_618,
            "78.6%": fib_786,
        }
        nearest_fib = min(
            fib_levels.items(),
            key=lambda x: abs(x[1] - cp)
        )
        fib_nearest_name = nearest_fib[0]
        fib_nearest_val  = nearest_fib[1]
        fib_distance_pct = round(
            abs(cp - fib_nearest_val) / cp * 100, 2
        )

        # Is price near a key Fibonacci level (within 0.5%)?
        near_fib = fib_distance_pct < 0.5
        # Is price bouncing from Fibonacci support?
        fib_support = (near_fib and
                       cp > fib_nearest_val and
                       float(l.iloc[-1]) <= fib_nearest_val * 1.005)

        # Support / Resistance
        window = 5
        s_lvls, r_lvls = [], []
        for i in range(window, len(df)-window):
            if float(l.iloc[i]) == float(
                    l.iloc[i-window:i+window+1].min()):
                s_lvls.append(float(l.iloc[i]))
            if float(h.iloc[i]) == float(
                    h.iloc[i-window:i+window+1].max()):
                r_lvls.append(float(h.iloc[i]))
        sup = round(max([s for s in s_lvls if s < cp],
                        default=cp*0.98), 2)
        res = round(min([r for r in r_lvls if r > cp],
                        default=cp*1.02), 2)

        risk_pts   = round(cp - sup, 2)
        reward_pts = round(res - cp, 2)
        rr_ratio   = round(reward_pts/(risk_pts+0.001), 2)

        # ── Weekly & Monthly S/R + Pivots ─────────────────
        # Fetch weekly and monthly candles for key levels
        try:
            df_w = df.copy()
            df_w.index = pd.to_datetime(df_w.index)
            # Weekly OHLC
            w_grp  = df_w.resample("1W").agg({
                "High":"max","Low":"min","Close":"last","Open":"first"
            }).dropna().tail(8)  # last 8 weeks
            # Monthly OHLC
            m_grp  = df_w.resample("1ME").agg({
                "High":"max","Low":"min","Close":"last","Open":"first"
            }).dropna().tail(3)  # last 3 months

            # Weekly pivot
            if len(w_grp) >= 2:
                pw_h = float(w_grp["High"].iloc[-2])
                pw_l = float(w_grp["Low"].iloc[-2])
                pw_c = float(w_grp["Close"].iloc[-2])
                w_pivot = round((pw_h + pw_l + pw_c) / 3, 2)
                w_r1 = round(2 * w_pivot - pw_l, 2)
                w_s1 = round(2 * w_pivot - pw_h, 2)
                w_r2 = round(w_pivot + (pw_h - pw_l), 2)
                w_s2 = round(w_pivot - (pw_h - pw_l), 2)
            else:
                w_pivot = w_r1 = w_s1 = w_r2 = w_s2 = cp

            # Monthly pivot
            if len(m_grp) >= 2:
                pm_h = float(m_grp["High"].iloc[-2])
                pm_l = float(m_grp["Low"].iloc[-2])
                pm_c = float(m_grp["Close"].iloc[-2])
                m_pivot = round((pm_h + pm_l + pm_c) / 3, 2)
                m_r1 = round(2 * m_pivot - pm_l, 2)
                m_s1 = round(2 * m_pivot - pm_h, 2)
                m_r2 = round(m_pivot + (pm_h - pm_l), 2)
                m_s2 = round(m_pivot - (pm_h - pm_l), 2)
            else:
                m_pivot = m_r1 = m_s1 = m_r2 = m_s2 = cp

            # Weekly S/R from swing highs/lows
            w_highs = [float(w_grp["High"].iloc[i])
                       for i in range(len(w_grp))]
            w_lows  = [float(w_grp["Low"].iloc[i])
                       for i in range(len(w_grp))]
            w_sup = round(max(
                [l for l in w_lows if l < cp], default=cp*0.95
            ), 2)
            w_res = round(min(
                [h for h in w_highs if h > cp], default=cp*1.05
            ), 2)

            # Monthly S/R
            m_highs = [float(m_grp["High"].iloc[i])
                       for i in range(len(m_grp))]
            m_lows  = [float(m_grp["Low"].iloc[i])
                       for i in range(len(m_grp))]
            m_sup = round(max(
                [l for l in m_lows if l < cp], default=cp*0.92
            ), 2)
            m_res = round(min(
                [h for h in m_highs if h > cp], default=cp*1.08
            ), 2)

        except Exception:
            w_pivot = w_r1 = w_s1 = w_r2 = w_s2 = cp
            m_pivot = m_r1 = m_s1 = m_r2 = m_s2 = cp
            w_sup   = round(cp * 0.97, 2)
            w_res   = round(cp * 1.03, 2)
            m_sup   = round(cp * 0.94, 2)
            m_res   = round(cp * 1.06, 2)

        # ── SL and Targets ────────────────────────────────
        # Entry logic:
        # If current price is ABOVE EMA9 → entry = current price (already above EMA9)
        # If current price is BELOW EMA9 → entry = EMA9 (wait for pullback)
        # For PE: opposite logic

        # CE entry
        if cp >= e9v:
            entry_long = round(cp, 2)   # Price already above EMA9 — enter now
        else:
            entry_long = round(e9v, 2)  # Wait for price to reach EMA9

        # PE entry
        if cp <= e9v:
            entry_short = round(cp, 2)  # Price already below EMA9 — enter now
        else:
            entry_short = round(e9v, 2) # Wait for price to reach EMA9

        # ── SL and Targets — always validated ─────────────
        # Use ATR for SL distance
        sl_dist   = max(atrv * 1.0, cp * 0.005)  # min 0.5% of price

        # CE: SL MUST be below entry, Targets MUST be above entry
        sl_long   = round(entry_long - sl_dist, 2)
        tgt1      = round(entry_long + sl_dist * 1.5, 2)  # R:R 1.5:1
        tgt2      = round(entry_long + sl_dist * 2.5, 2)  # R:R 2.5:1
        tgt3      = round(entry_long + sl_dist * 4.0, 2)  # R:R 4.0:1

        # PE: SL MUST be above entry, Targets MUST be below entry
        sl_short  = round(entry_short + sl_dist, 2)
        tgt1s     = round(entry_short - sl_dist * 1.5, 2)
        tgt2s     = round(entry_short - sl_dist * 2.5, 2)
        tgt3s     = round(entry_short - sl_dist * 4.0, 2)

        # ── Final hard validation ────────────────────────
        # CE checks
        assert_ce_ok = sl_long < entry_long < tgt1
        if not assert_ce_ok:
            # Force correct values using percentage
            sl_long = round(entry_long * 0.985, 2)
            tgt1    = round(entry_long * 1.015, 2)
            tgt2    = round(entry_long * 1.025, 2)
            tgt3    = round(entry_long * 1.040, 2)

        # PE checks
        assert_pe_ok = sl_short > entry_short > tgt1s
        if not assert_pe_ok:
            # Force correct values using percentage
            sl_short = round(entry_short * 1.015, 2)
            tgt1s    = round(entry_short * 0.985, 2)
            tgt2s    = round(entry_short * 0.975, 2)
            tgt3s    = round(entry_short * 0.960, 2)

        # Liquidity sweep
        rh20 = float(h.tail(20).max())
        rl20 = float(l.tail(20).min())
        sweep_low  = (float(l.iloc[-1]) < rl20*1.001 and
                      cp > rl20*1.002)
        sweep_high = (float(h.iloc[-1]) > rh20*0.999 and
                      cp < rh20*0.998)

        # BOS
        bos_bull = (float(h.iloc[-1]) > float(h.iloc[-2]) >
                    float(h.iloc[-3]) and
                    float(v.iloc[-1]) > vol_avg*1.3)
        bos_bear = (float(l.iloc[-1]) < float(l.iloc[-2]) <
                    float(l.iloc[-3]) and
                    float(v.iloc[-1]) > vol_avg*1.3)

        # Candlestick patterns
        body0  = abs(cp - float(o.iloc[-1]))
        avg_b  = float(abs(c - o).rolling(10).mean().iloc[-1])
        uw     = float(h.iloc[-1]) - max(cp, float(o.iloc[-1]))
        lw     = min(cp, float(o.iloc[-1])) - float(l.iloc[-1])
        c1v    = float(c.iloc[-2])
        o1v    = float(o.iloc[-2])
        h1v    = float(h.iloc[-2])
        l1v    = float(l.iloc[-2])
        c2v    = float(c.iloc[-3])
        o2v    = float(o.iloc[-3])

        patterns = []
        if body0 < avg_b*0.2:
            patterns.append(("⚡ Doji","neutral",
                             "Indecision — big move coming"))
        if lw > body0*2 and uw < body0*0.5 and c1v < o1v:
            patterns.append(("🔨 Hammer","bullish",
                             "Buyers rejected lows — reversal"))
        if uw > body0*2 and lw < body0*0.5 and c1v > o1v:
            patterns.append(("⭐ Shooting Star","bearish",
                             "Sellers rejected highs — reversal"))
        if (cp>float(o.iloc[-1]) and c1v<o1v and
                float(o.iloc[-1])<c1v and cp>o1v):
            patterns.append(("🟢 Bullish Engulfing","bullish",
                             "Bulls overwhelmed bears — strong buy"))
        if (cp<float(o.iloc[-1]) and c1v>o1v and
                float(o.iloc[-1])>c1v and cp<o1v):
            patterns.append(("🔴 Bearish Engulfing","bearish",
                             "Bears overwhelmed bulls — strong sell"))
        if (c2v<o2v and abs(c1v-o1v)<avg_b*0.4 and
                cp>float(o.iloc[-1]) and cp>(o2v+c2v)/2):
            patterns.append(("🌅 Morning Star","bullish",
                             "3-candle reversal — trend change"))
        if (c2v>o2v and abs(c1v-o1v)<avg_b*0.4 and
                cp<float(o.iloc[-1]) and cp<(o2v+c2v)/2):
            patterns.append(("🌆 Evening Star","bearish",
                             "3-candle reversal — trend change"))
        if (body0>avg_b*2.5 and uw<body0*0.1 and lw<body0*0.1):
            bias = "bullish" if cp>float(o.iloc[-1]) else "bearish"
            patterns.append((
                f"💪 {'Bull' if bias=='bullish' else 'Bear'} Marubozu",
                bias, "Strong momentum — continuation likely"))

        # ── Additional patterns ────────────────────────────
        # Three White Soldiers — 3 consecutive bullish candles
        try:
            c3v = float(c.iloc[-4]); o3v = float(o.iloc[-4])
            if (c2v>o2v and c1v>o1v and cp>float(o.iloc[-1]) and
                    c2v>c3v and c1v>c2v and cp>c1v and
                    abs(c2v-o2v)>avg_b*0.7 and
                    abs(c1v-o1v)>avg_b*0.7):
                patterns.append((
                    "🪖 Three White Soldiers", "bullish",
                    "3 strong bullish candles — powerful uptrend"
                ))
        except Exception:
            pass

        # Three Black Crows — 3 consecutive bearish candles
        try:
            if (c2v<o2v and c1v<o1v and cp<float(o.iloc[-1]) and
                    c2v<c3v and c1v<c2v and cp<c1v and
                    abs(c2v-o2v)>avg_b*0.7 and
                    abs(c1v-o1v)>avg_b*0.7):
                patterns.append((
                    "🐦 Three Black Crows", "bearish",
                    "3 strong bearish candles — powerful downtrend"
                ))
        except Exception:
            pass

        # Inside Bar (NR7) — candle inside previous candle
        if (float(h.iloc[-1]) < h1v and
                float(l.iloc[-1]) > l1v):
            patterns.append((
                "📦 Inside Bar", "neutral",
                "Price consolidating — breakout imminent. "
                "Buy direction of breakout."
            ))

        # Pin Bar — long wick rejection candle
        if lw > body0 * 3 and uw < body0 * 0.5:
            patterns.append((
                "📌 Bullish Pin Bar", "bullish",
                "Strong rejection of lows — buyers in control"
            ))
        if uw > body0 * 3 and lw < body0 * 0.5:
            patterns.append((
                "📌 Bearish Pin Bar", "bearish",
                "Strong rejection of highs — sellers in control"
            ))

        # ── 11-FACTOR CHECKLIST ───────────────────────────
        # Each factor: (label, passed, value_str, why)

        # Uptrend score (10 conditions)
        up_conds = {
            "EMA9>EMA21":  e9v > e21v,
            "EMA21>EMA50": e21v > e50v,
            "Price>EMA9":  cp > e9v,
            "Price>VWAP":  cp > vwv,
            "RSI 55-75":   55 < rv < 75,
            "RSI Rising":  rv > rv1,
            "MACD>Signal": mv > msv,
            "MACD Cross":  (mv>msv and mv1<=msv1),
            "ADX>20":      adxv > 20,
            "Vol Surge":   vsurge,
        }
        up_score = sum(up_conds.values())

        dn_conds = {
            "EMA9<EMA21":  e9v < e21v,
            "EMA21<EMA50": e21v < e50v,
            "Price<EMA9":  cp < e9v,
            "Price<VWAP":  cp < vwv,
            "RSI 25-45":   25 < rv < 45,
            "RSI Falling": rv < rv1,
            "MACD<Signal": mv < msv,
            "MACD Cross↓": (mv<msv and mv1>=msv1),
            "ADX>20":      adxv > 20,
            "Vol Surge":   vsurge,
        }
        dn_score = sum(dn_conds.values())

        direction = (
            "UPTREND"   if up_score >= 6 else
            "DOWNTREND" if dn_score >= 6 else
            "SIDEWAYS"
        )
        score = up_score if direction != "DOWNTREND" else dn_score

        # The 11 checklist items for CE trade
        tt = best_trading_time()
        good_time = tt in ("best","good","ok")

        ce_checklist = [
            ("1. Uptrend Score ≥ 7",
             up_score >= 7,
             f"{up_score}/10",
             "Core trend condition"),
            ("2. RSI 55–68 (momentum zone)",
             55 < rv < 68,
             f"{rv:.1f}",
             "Bullish but not overbought"),
            ("3. Price above VWAP",
             cp > vwv,
             f"Price ₹{cp:,.0f} | VWAP ₹{vwv:,.0f}",
             "Buyers in control today"),
            ("4. Volume Surge ≥ 1.2×",
             vsurge,
             f"{vol_ratio:.2f}× avg",
             "Institutional participation"),
            ("5. MACD above Signal",
             mv > msv,
             f"MACD {mv:.2f} | Sig {msv:.2f}",
             "Bullish momentum confirmed"),
            ("6. ADX > 20 (trending)",
             adxv > 20,
             f"ADX {adxv:.1f}",
             "Market is trending, not sideways"),
            ("7. CMF > 0 (money flowing in)",
             cmfv > 0,
             f"CMF {cmfv:+.3f}",
             "Institutional money entering"),
            ("8. OBV Rising",
             obv_bull,
             "Yes" if obv_bull else "No",
             "Volume confirming price"),
            ("8b. Bollinger Band — Price above midline",
             cp > (bbup + bblw) / 2,
             f"Mid ₹{(bbup+bblw)/2:,.0f} | "
             f"Upper ₹{bbup:,.0f} | Lower ₹{bblw:,.0f}",
             "Price above BB midline confirms uptrend"),
            ("9. Risk-Reward ≥ 1.5",
             rr_ratio >= 1.5,
             f"{rr_ratio}:1",
             "Reward must justify risk"),
            ("9b. CPR — Price above CPR",
             cpr_position == "ABOVE",
             f"{cpr_bias} | {cpr_type[:8]}",
             "Central Pivot Range bias must be bullish"),
            ("9c. Supertrend BUY signal",
             st_bull,
             f"{st_signal} ₹{st_value:,.0f}"
             + (" FRESH!" if st_crossed else ""),
             "Supertrend must be in BUY zone"),
            ("10. No Bearish Pattern",
             not any(p[1]=="bearish" for p in patterns),
             ", ".join(p[0] for p in patterns) or "None",
             "Candle pattern must not contradict"),
            ("11. Good Trading Time",
             good_time,
             tt.replace("_"," ").upper(),
             "Avoid first 15 min and last 30 min"),
        ]

        # PE checklist (mirror)
        pe_checklist = [
            ("1. Downtrend Score ≥ 7",
             dn_score >= 7,
             f"{dn_score}/10",
             "Core trend condition"),
            ("2. RSI 32–45 (bearish zone)",
             32 < rv < 45,
             f"{rv:.1f}",
             "Bearish but not oversold"),
            ("3. Price below VWAP",
             cp < vwv,
             f"Price ₹{cp:,.0f} | VWAP ₹{vwv:,.0f}",
             "Sellers in control today"),
            ("4. Volume Surge ≥ 1.2×",
             vsurge,
             f"{vol_ratio:.2f}× avg",
             "Institutional participation"),
            ("5. MACD below Signal",
             mv < msv,
             f"MACD {mv:.2f} | Sig {msv:.2f}",
             "Bearish momentum confirmed"),
            ("6. ADX > 20 (trending)",
             adxv > 20,
             f"ADX {adxv:.1f}",
             "Market is trending, not sideways"),
            ("7. CMF < 0 (money flowing out)",
             cmfv < 0,
             f"CMF {cmfv:+.3f}",
             "Institutional money exiting"),
            ("8. OBV Falling",
             not obv_bull,
             "Yes" if not obv_bull else "No",
             "Volume confirming bearish move"),
            ("8b. Bollinger Band — Price below midline",
             cp < (bbup + bblw) / 2,
             f"Mid ₹{(bbup+bblw)/2:,.0f} | "
             f"Upper ₹{bbup:,.0f} | Lower ₹{bblw:,.0f}",
             "Price below BB midline confirms downtrend"),
            ("9. Risk-Reward ≥ 1.5",
             rr_ratio >= 1.5,
             f"{rr_ratio}:1",
             "Reward must justify risk"),
            ("9b. CPR — Price below CPR",
             cpr_position == "BELOW",
             f"{cpr_bias} | {cpr_type[:8]}",
             "Central Pivot Range bias must be bearish"),
            ("9c. Supertrend SELL signal",
             not st_bull,
             f"{st_signal} ₹{st_value:,.0f}"
             + (" FRESH!" if st_crossed else ""),
             "Supertrend must be in SELL zone"),
            ("10. No Bullish Pattern",
             not any(p[1]=="bullish" for p in patterns),
             ", ".join(p[0] for p in patterns) or "None",
             "Candle pattern must not contradict"),
            ("11. Good Trading Time",
             good_time,
             tt.replace("_"," ").upper(),
             "Avoid first 15 min and last 30 min"),
        ]

        ce_pass = sum(1 for c in ce_checklist if c[1])
        pe_pass = sum(1 for c in pe_checklist if c[1])

        return dict(
            cp=cp, direction=direction, score=score,
            up_score=up_score, dn_score=dn_score,
            up_conds=up_conds, dn_conds=dn_conds,
            rv=rv, adxv=adxv, atrv=atrv,
            e9v=e9v, e21v=e21v, e50v=e50v, vwv=vwv,
            vwap_upper=vwap_upper, vwap_lower=vwap_lower,
            vwap_u2=vwap_u2, vwap_l2=vwap_l2,
            vwap_zone=vwap_zone,
            bbup=bbup, bblw=bblw, cmfv=cmfv,
            vol_ratio=vol_ratio, vsurge=vsurge,
            obv_bull=obv_bull, rr_ratio=rr_ratio,
            sup=sup, res=res,
            # Weekly levels
            w_pivot=w_pivot, w_r1=w_r1, w_s1=w_s1,
            w_r2=w_r2, w_s2=w_s2, w_sup=w_sup, w_res=w_res,
            # Monthly levels
            m_pivot=m_pivot, m_r1=m_r1, m_s1=m_s1,
            m_r2=m_r2, m_s2=m_s2, m_sup=m_sup, m_res=m_res,
            sl_long=sl_long, sl_short=sl_short,
            tgt1=tgt1, tgt2=tgt2, tgt3=tgt3,
            tgt1s=tgt1s, tgt2s=tgt2s, tgt3s=tgt3s,
            sweep_low=sweep_low, sweep_high=sweep_high,
            bos_bull=bos_bull, bos_bear=bos_bear,
            patterns=patterns,
            ce_checklist=ce_checklist, pe_checklist=pe_checklist,
            ce_pass=ce_pass, pe_pass=pe_pass,
            good_time=good_time, time_state=tt,
            entry_long=entry_long, entry_short=entry_short,
            sl_dist=sl_dist,
            # CPR
            cpr_pivot=cpr_pivot, cpr_tc=cpr_tc, cpr_bc=cpr_bc,
            cpr_width=cpr_width, cpr_width_pct=cpr_width_pct,
            cpr_type=cpr_type, cpr_bias=cpr_bias,
            cpr_position=cpr_position, virgin_cpr=virgin_cpr,
            # Supertrend
            st_value=st_value, st_bull=st_bull,
            st_crossed=st_crossed, st_signal=st_signal,
            st_label=st_label, st_dir=st_dir,
            # Fibonacci
            fib_high=fib_high, fib_low=fib_low,
            fib_236=fib_236, fib_382=fib_382,
            fib_500=fib_500, fib_618=fib_618, fib_786=fib_786,
            fib_nearest_name=fib_nearest_name,
            fib_nearest_val=fib_nearest_val,
            fib_distance_pct=fib_distance_pct,
            near_fib=near_fib, fib_support=fib_support,
            # MACD values
            mv=mv, msv=msv, mv1=mv1, msv1=msv1,
            macd_bull=(mv > msv),
            # series
            e9s=e9, e21s=e21, e50s=e50,
            vwaps=vwap, rsis=rsi,
            macds=macd, msigs=msig,
            bbus=bbu, bbls=bbl,
            cmfs=cmf, obvs=obv,
        )
    except Exception as e:
        return None

# ══════════════════════════════════════════════════════════
# BLACK-SCHOLES
# ══════════════════════════════════════════════════════════
def bs(S,K,T,r,sig,t="CE"):
    if T<=0: return max(S-K,0) if t=="CE" else max(K-S,0)
    d1=(math.log(S/K)+(r+.5*sig**2)*T)/(sig*math.sqrt(T))
    d2=d1-sig*math.sqrt(T)
    if t=="CE":
        return max(round(S*norm.cdf(d1)-K*math.exp(-r*T)*norm.cdf(d2),2),0)
    return max(round(K*math.exp(-r*T)*norm.cdf(-d2)-S*norm.cdf(-d1),2),0)

def greeks(S,K,T,r,sig,t="CE"):
    if T<=0: return {"d":0,"g":0,"th":0,"v":0}
    d1=(math.log(S/K)+(r+.5*sig**2)*T)/(sig*math.sqrt(T))
    d2=d1-sig*math.sqrt(T)
    g =norm.pdf(d1)/(S*sig*math.sqrt(T))
    vg=S*norm.pdf(d1)*math.sqrt(T)/100
    if t=="CE":
        dlt=norm.cdf(d1)
        th=(-(S*norm.pdf(d1)*sig)/(2*math.sqrt(T))
            -r*K*math.exp(-r*T)*norm.cdf(d2))/365
    else:
        dlt=norm.cdf(d1)-1
        th=(-(S*norm.pdf(d1)*sig)/(2*math.sqrt(T))
            +r*K*math.exp(-r*T)*norm.cdf(-d2))/365
    return {"d":round(dlt,4),"g":round(g,6),
            "th":round(th,2),"v":round(vg,2)}

# ══════════════════════════════════════════════════════════
# HEADER  (status bar + index row)
# ══════════════════════════════════════════════════════════
n      = now_ist()
mopen  = market_open()
tstate = best_trading_time()
tclr   = {"best":"#16a34a","good":"#15803d",
           "ok":"#d97706","caution":"#ea580c",
           "avoid":"#dc2626","closed":"#94a3b8",
           "pre_market":"#64748b"}.get(tstate,"#64748b")
tmsg   = {"best":"✅ BEST TIME — 9:30–10:30 AM",
           "good":"🟡 GOOD TIME — steady market",
           "ok":"🟡 OK — lunch hours",
           "caution":"⚠️ CAUTION — choppy",
           "avoid":"❌ AVOID — too volatile",
           "closed":"⛔ MARKET CLOSED",
           "pre_market":"🕐 PRE-MARKET"
          }.get(tstate,"—")

st.markdown(f"""
<div style='background:linear-gradient(135deg,#1e3a5f,#1d4ed8);
     border-radius:12px;padding:12px 16px;
     margin-bottom:10px;
     box-shadow:0 4px 12px rgba(29,78,216,0.25)'>
  <div style='display:flex;justify-content:space-between;
       align-items:center;flex-wrap:wrap;gap:6px'>
    <span style='font-size:16px;font-weight:700;
                 color:#ffffff;letter-spacing:-0.3px'>
        🎯 Intraday &amp; Options Terminal
    </span>
    <span style='background:{"rgba(22,163,74,0.3)" if mopen else "rgba(220,38,38,0.3)"};
                 color:{"#86efac" if mopen else "#fca5a5"};
                 font-weight:700;padding:3px 10px;
                 border-radius:20px;font-size:12px'>
        {"🟢 OPEN" if mopen else "🔴 CLOSED"}
    </span>
    <span style='color:#93c5fd;font-weight:600;
                 font-size:12px'>{tmsg}</span>
    <span style='color:#bfdbfe;font-size:11px'>
        🕐 {n.strftime("%d %b  %H:%M IST")}
    </span>
  </div>
</div>
""", unsafe_allow_html=True)

# Index cards
idx = {"NIFTY":"^NSEI","BANKNIFTY":"^NSEBANK",
       "SENSEX":"^BSESN","NIFTY IT":"^CNXIT"}
ic  = st.columns(len(idx))
for i,(nm,sy) in enumerate(idx.items()):
    ip = live_price(sy)
    if ip["ok"]:
        col = "#00ff88" if ip["chg"]>=0 else "#ff4455"
        arr = "▲" if ip["chg"]>=0 else "▼"
        ic[i].markdown(f"""
        <div style='background:#ffffff;border:1px solid #e2e8f0;
             border-radius:12px;padding:12px;text-align:center;
             box-shadow:0 1px 3px rgba(0,0,0,0.06)'>
          <div style='color:#64748b;font-size:11px;font-weight:500;
                      letter-spacing:0.5px'>{nm}</div>
          <div style='color:#1e293b;font-size:22px;font-weight:700;
                      margin:4px 0'>
              {ip['p']:,.0f}</div>
          <div style='color:{col};font-size:13px;font-weight:600'>
              {arr}{abs(ip['chg']):.2f}%
              <span style='color:#94a3b8;font-size:11px;
                           font-weight:400'>
                  {ip['chg_abs']:+.1f}
              </span>
          </div>
        </div>""", unsafe_allow_html=True)

st.markdown("<div style='margin:6px 0'></div>",
            unsafe_allow_html=True)

# ── Open in new window panel ───────────────────────────────
# Auto-detect base URL (before expander)
import socket as _socket
try:
    _hn = _socket.gethostname()
    _is_local = (_hn.startswith("DESKTOP") or
                 _hn.startswith("LAPTOP") or
                 "ranjith" in _hn.lower() or
                 _hn == "localhost")
except:
    _is_local = False

_base = ("http://localhost:8501" if _is_local
         else "https://ranjithour-sketch-trading-terminal-trading.streamlit.app")

with st.expander("Open any tab in a separate browser window"):
    st.caption(
        "Click any link below to open that tab in a new browser "
        "window. You can have multiple tabs open side by side."
    )
    link_cols = st.columns(5)
    link_data = [
        ("📋 Watchlist",    "watchlist", "#3b82f6"),
        ("🎯 Trade Setup",  "setup",     "#16a34a"),
        ("🔍 Scanner",      "scanner",   "#9333ea"),
        ("🤖 ML",           "ml",        "#0891b2"),
        ("🏦 Smart Money",  "smart",     "#d97706"),
        ("📊 Market Pulse", "pulse",     "#0f766e"),
        ("🔗 Options",      "options",   "#7c3aed"),
        ("🧪 Backtest",     "backtest",  "#1d4ed8"),
        ("🎯 Signal Hub",   "hub",       "#0f766e"),
        ("🛡️ Trade Manager","manager",   "#1e3a5f"),
        ("📝 Paper Trading", "paper",     "#0f766e"),
        ("⚡ Auto Orders",   "orders",    "#dc2626"),
        ("🌙 Evening Scan",  "evening",   "#1e1b4b"),
    ]
    for idx, (lname, lkey, lcolor) in enumerate(link_data):
        col_idx = idx % 5
        with link_cols[col_idx]:
            st.markdown(
                f"<a href='{_base}/?tab={lkey}' "
                f"target='_blank' "
                f"style='display:block;"
                f"background:{lcolor};"
                f"color:white;"
                f"text-decoration:none;"
                f"padding:8px 12px;"
                f"border-radius:8px;"
                f"font-size:13px;"
                f"font-weight:600;"
                f"text-align:center;"
                f"margin:3px 0;"
                f"font-family:Inter,sans-serif'>"
                f"{lname}</a>",
                unsafe_allow_html=True
            )
    st.markdown(
        "<div style='margin-top:8px;font-size:12px;"
        "color:#64748b'>"
        "💡 Tip: Bookmark your most used tabs for instant access. "
        "Works on same WiFi network too — use your Network URL "
        "from the terminal instead of localhost.</div>",
        unsafe_allow_html=True
    )

# ══════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════
# ── Zerodha Kite Connection ───────────────────────────────
if KITE_AVAILABLE and KITE_API_KEY:
    if kite_is_connected():
        st.sidebar.success("Kite LIVE — Real-time data")
        if st.sidebar.button("Disconnect", key="kite_disc"):
            st.session_state.pop("kite_access_token", None)
            try:
                _kt = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    ".kite_token.json"
                )
                if os.path.exists(_kt):
                    os.remove(_kt)
            except:
                pass
            st.rerun()
    else:
        st.sidebar.warning("Yahoo data (15-min delay)")
        # Check if redirected back from Kite with request_token
        _qp = st.query_params
        if "request_token" in _qp:
            _req = _qp["request_token"]
            try:
                _k2 = KiteConnect(api_key=KITE_API_KEY)
                _sess = _k2.generate_session(
                    _req,
                    api_secret=KITE_API_SECRET
                )
                _tok = _sess["access_token"]
                st.session_state["kite_access_token"] = _tok
                save_kite_token(_tok)
                st.query_params.clear()
                st.sidebar.success("Kite connected!")
                st.rerun()
            except Exception as _e:
                st.sidebar.error(f"Login failed: {_e}")


        # Login URL - v=3 is required by Kite Connect
        _login_url = (
            f"https://kite.trade/connect/login"
            f"?api_key={KITE_API_KEY}&v=3"
        )

        st.sidebar.link_button(
            "Login with Zerodha Kite",
            url=_login_url,
            use_container_width=True,
            type="primary"
        )
        st.sidebar.caption(
            "After login, Zerodha redirects back here automatically."
        )
        st.sidebar.caption(
            "Ensure kite.trade Redirect URL is set to your app URL."
        )
else:
    st.sidebar.info("Yahoo Finance (delayed)")

st.sidebar.markdown("---")

# ── Prepare for Trading button ────────────────────────────
ml_cache        = st.session_state.get("ml_pretrained", {})
candle_ready    = st.session_state.get("candle_cache_ready", False)
ml_trained      = st.session_state.get("ml_models_trained", False)

_prep_done = ml_trained and candle_ready
_ml_count  = len(ml_cache)
_cc_count  = len(st.session_state.get("candle_cache", {}))

# Status display
if _prep_done:
    # Check how old the cache is
    import time as _prep_time
    _cache_ts  = st.session_state.get("prep_timestamp", 0)
    _cache_age = round((_prep_time.time() - _cache_ts) / 60)
    _age_warn  = _cache_age > 60  # warn if older than 60 min

    if _age_warn:
        st.sidebar.warning(
            f"⚠️ Cache is {_cache_age} min old. "
            f"Re-run Prepare for fresh data."
        )
    else:
        st.sidebar.success(
            f"✅ Ready to Trade! "
            f"ML: {_ml_count} stocks cached "
            f"({_cache_age} min ago)"
        )
elif ml_trained:
    st.sidebar.warning(
        f"⚠️ ML cached ({_ml_count} stocks) "
        f"but candles not pre-fetched"
    )
else:
    st.sidebar.info("⚡ Click Prepare before 9:30 AM scan")

# Prepare for Trading button
if st.sidebar.button(
    "⚡ Prepare for Trading",
    key="prepare_trading_btn",
    type="primary",
    help="Pre-trains ML + caches candles for all scan stocks. "
         "Run at 9:00 AM for fastest 9:30 scan."
):
    # Collect all stocks to prepare
    # Use SECTORS (globally defined) instead of SCANNER_UNIVERSE
    # which is only defined inside the scanner tab
    _all_scan_stocks = []
    for _grp, _stks in SECTORS.items():
        _all_scan_stocks.extend(_stks)
    # Also add top F&O stocks
    _top_fo = [
        "NIFTY 50","BANK NIFTY","Reliance","HDFC Bank",
        "ICICI Bank","TCS","Infosys","SBI","Wipro",
        "Bajaj Finance","ITC","Sun Pharma","L&T","Maruti",
        "Axis Bank","HCL Tech","ONGC","Bharti Airtel",
        "Tata Steel","JSW Steel","Kotak Bank","Titan Company",
        "Asian Paints","Nestle India","Power Grid","NTPC",
        "Bajaj Auto","Eicher Motors","UltraTech Cement",
        "Britannia","Cipla","Dr Reddys","Divis Lab"
    ]
    _all_scan_stocks.extend(_top_fo)
    # Remove duplicates preserve order
    _seen = set()
    _unique_stocks = []
    for _s in _all_scan_stocks:
        if _s not in _seen and _s in STOCKS:
            _seen.add(_s)
            _unique_stocks.append(_s)

    _total_stocks = len(_unique_stocks)

    st.sidebar.markdown(
        f"Preparing {_total_stocks} stocks... "
        f"This takes 3-4 minutes. "
        f"Do this at 9:00 AM."
    )

    # Phase 1: Pre-fetch candles
    with st.sidebar:
        st.markdown("**Phase 1: Caching candles...**")
        _p1_bar  = st.progress(0)
        _p1_text = st.empty()
        prefetch_candles_cache(
            _unique_stocks,
            timeframes=["1h","1d"],
            progress_bar=_p1_bar,
            status_text=_p1_text
        )
        _p1_text.text("✅ Candles cached!")

        # Phase 2: Pre-train ML
        st.markdown("**Phase 2: Training ML models...**")
        _p2_bar  = st.progress(0)
        _p2_text = st.empty()
        pretrain_ml_models(
            _unique_stocks,
            max_stocks=_total_stocks,
            progress_bar=_p2_bar,
            status_text=_p2_text
        )
        _p2_text.text(
            f"✅ ML trained for "
            f"{len(st.session_state.get('ml_pretrained',{}))} stocks!"
        )

    st.rerun()

# Clear cache button
if _prep_done:
    if st.sidebar.button(
        "🗑️ Clear Cache & Retrain",
        key="clear_prep_cache"
    ):
        st.session_state.pop("ml_pretrained", None)
        st.session_state.pop("ml_models_trained", None)
        st.session_state.pop("candle_cache", None)
        st.session_state.pop("candle_cache_ready", None)
        st.rerun()

st.sidebar.markdown("---")

# Stock search
# Use empty string "" as label — avoids _arrow_right corruption
# Do NOT use label_visibility="collapsed" — that causes the bug
with st.sidebar:
    st.markdown(
        "<p style='font-size:12px;color:#64748b;"
        "margin:0 0 4px 0'>Search stock</p>",
        unsafe_allow_html=True
    )
    srch = st.text_input(
        "Search stock",
        placeholder="Reliance, TCS, Gold...",
        key="sidebar_search_main",
        label_visibility="hidden"
    )
    if srch:
        q    = srch.strip().lower()
        hits = {k:v for k,v in STOCKS.items()
                if q in k.lower() or q in v.lower()}
        if hits:
            st.markdown(
                "<p style='font-size:11px;color:#94a3b8;"
                "margin:4px 0 2px 0'>Results</p>",
                unsafe_allow_html=True
            )
            pk = st.selectbox(
                "Select stock",
                list(hits.keys()),
                key="sidebar_search_result",
                label_visibility="hidden"
            )
            if st.button(
                "Load", type="primary", key="sb_load",
                use_container_width=True
            ):
                st.session_state["sn"] = pk
                st.session_state["st"] = hits[pk]
                st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown("<b>Quick pick</b>", unsafe_allow_html=True)

# Flat list — no expanders (expanders corrupt label text in Streamlit)
SIDEBAR_PICKS = {
    "Top F&O": [
        "NIFTY 50","BANK NIFTY","Reliance","HDFC Bank",
        "ICICI Bank","TCS","Infosys","SBI",
    ],
    "Banking": [
        "HDFC Bank","ICICI Bank","SBI","Axis Bank",
        "Kotak Bank","Bajaj Finance",
    ],
    "IT": [
        "TCS","Infosys","Wipro","HCL Tech",
        "Tech Mahindra","Persistent",
    ],
    "Energy": [
        "Reliance","ONGC","NTPC","Power Grid",
        "Tata Power","Gail",
    ],
    "Commodities": [
        "Gold (MCX)","Silver (MCX)",
        "Crude Oil (MCX)","Copper (MCX)",
        "USD/INR","Natural Gas (MCX)",
    ],
    "Global": [
        "Dow Jones","S&P 500","NASDAQ",
        "Nikkei 225","Hang Seng",
    ],
}

# Show as a selectbox for sector then buttons for stocks
sidebar_sector = st.sidebar.selectbox(
    "Sector",
    list(SIDEBAR_PICKS.keys()),
    key="sidebar_sector_pick",
    label_visibility="collapsed"
)
for ni, nm in enumerate(SIDEBAR_PICKS[sidebar_sector]):
    sk = STOCKS.get(nm, "")
    if st.sidebar.button(
        nm,
        key=f"sb_{sidebar_sector[:3]}_{ni}",
        use_container_width=True
    ):
        st.session_state["sn"] = nm
        st.session_state["st"] = sk

st.sidebar.markdown("---")
tf = st.sidebar.selectbox(
    "Timeframe",
    ["1m","5m","15m","30m","1h","1d"],
    index=2,
    key="global_tf"
)
# Sync scanner timeframe with global timeframe
if "scan_tf" not in st.session_state:
    st.session_state["scan_tf"] = tf

auto_rf = st.sidebar.checkbox("Auto Refresh (2 min)", False)

st.sidebar.markdown("---")
st.sidebar.markdown("<b>Open in new window</b>", unsafe_allow_html=True)
_base_url = _base
for _lname, _lkey in [
    ("📋 Watchlist",    "watchlist"),
    ("🎯 Trade Setup",  "setup"),
    ("🔍 Scanner",      "scanner"),
    ("🤖 ML",           "ml"),
    ("🏦 Smart Money",  "smart"),
    ("🧮 P&L Calc",     "calc"),
]:
    st.sidebar.markdown(
        f"<a href='{_base_url}/?tab={_lkey}' "
        f"target='_blank' "
        f"style='display:block;"
        f"background:#f8fafc;"
        f"border:1px solid #e2e8f0;"
        f"color:#374151;"
        f"text-decoration:none;"
        f"padding:7px 12px;"
        f"border-radius:7px;"
        f"font-size:12px;"
        f"font-weight:500;"
        f"margin:2px 0;"
        f"font-family:Inter,sans-serif'>"
        f"{_lname} ↗</a>",
        unsafe_allow_html=True
    )

st.sidebar.markdown("---")
st.sidebar.markdown("""
### Entry Rules Summary
✅ CE (Buy Call):
- Score ≥ 7 + RSI 55–68
- Price > VWAP
- Volume surge ✅
- MACD > Signal
- 9 or more checks GREEN

✅ PE (Buy Put):
- Score ≥ 7 + RSI 32–45
- Price < VWAP
- Volume surge ✅
- MACD < Signal
- 9 or more checks GREEN

⛔ Never trade when:
- Score < 6
- RSI > 70 or < 30
- Time shows ❌ AVOID
""")

if "sn" not in st.session_state:
    st.session_state["sn"] = "NIFTY 50"
    st.session_state["st"] = "^NSEI"

sname = st.session_state["sn"]
stick = st.session_state["st"]

# ══════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════
T1,T2,T3,T4,T5,T6,T7,T8,T9,T10,T11,T12,T13 = st.tabs(TAB_NAMES)

# ── Reusable inline stock search widget ──────────────────
def inline_stock_search(tab_key: str):
    """
    Shows a compact search box + sector picker inline.
    Returns True if a new stock was selected.
    """
    changed = False
    sc1, sc2, sc3 = st.columns([3, 2, 1])
    with sc1:
        q = st.text_input(
            "Search stock",
            placeholder="Adani Ports, TCS, NIFTY...",
            key=f"inline_srch_{tab_key}",
            label_visibility="collapsed"
        )
        if q:
            hits = {k:v for k,v in STOCKS.items()
                    if q.strip().lower() in k.lower()
                    or q.strip().lower() in v.lower()}
            if hits:
                pick = st.selectbox(
                    "Results",
                    list(hits.keys()),
                    key=f"inline_pick_{tab_key}",
                    label_visibility="collapsed"
                )
                if st.button(
                    "Load",
                    key=f"inline_load_{tab_key}",
                    type="primary"
                ):
                    st.session_state["sn"] = pick
                    st.session_state["st"] = hits[pick]
                    changed = True
                    st.rerun()
            elif q:
                st.caption("No results found")
    with sc2:
        sector_opts = [""] + list(SECTORS.keys())
        chosen_sec  = st.selectbox(
            "Sector",
            sector_opts,
            key=f"inline_sec_{tab_key}",
            label_visibility="collapsed",
            format_func=lambda x: "Browse by sector..." if x=="" else x
        )
        if chosen_sec:
            sec_stocks = SECTORS[chosen_sec]
            chosen_st  = st.selectbox(
                "Stock",
                sec_stocks,
                key=f"inline_sec_st_{tab_key}",
                label_visibility="collapsed"
            )
            if st.button(
                "Load",
                key=f"inline_sec_load_{tab_key}",
                type="primary"
            ):
                st.session_state["sn"] = chosen_st
                st.session_state["st"] = STOCKS.get(
                    chosen_st, ""
                )
                changed = True
                st.rerun()
    with sc3:
        st.markdown(
            f"<div style='padding:6px 0;font-size:13px;"
            f"color:#64748b'>"
            f"📍 <b style='color:#1e293b'>"
            f"{st.session_state.get('sn','NIFTY 50')}"
            f"</b></div>",
            unsafe_allow_html=True
        )
    return changed

# ╔══════════════════════════════════════════════════════╗
# ║  TAB 1 — WATCHLIST                                  ║
# ╚══════════════════════════════════════════════════════╝
with T1:
    st.markdown("### 📋 Live Stock Prices")
    wc1,wc2,wc3 = st.columns([2,1,1])
    with wc1:
        wsec = st.selectbox("Sector",list(SECTORS.keys()),
                            key="wsec_t1")
    with wc2:
        if st.button("🔄 Refresh",type="primary",key="wrf_t1"):
            st.cache_data.clear()
    with wc3:
        showall = st.checkbox("All stocks", key="showall_wl")

    slist = list(STOCKS.keys()) if showall else SECTORS[wsec]
    with st.spinner("Fetching live prices..."):
        pdf = bulk_prices(slist)

    if pdf.empty:
        st.error("Could not fetch. Check internet.")
    else:
        valid = pdf[pdf["Price"].notna()].copy()

        # ── Grid of cards ─────────────────────────────────
        n_cols = 4
        btn_idx = 0
        for chunk_start in range(0,len(valid),n_cols):
            chunk = valid.iloc[chunk_start:chunk_start+n_cols]
            cols  = st.columns(n_cols)
            for ci,(_, row) in enumerate(chunk.iterrows()):
                chg = row["Chg%"] or 0
                col = "#00ff88" if chg>=0 else "#ff4455"
                arr = "▲" if chg>=0 else "▼"
                with cols[ci]:
                    if st.button(
                        f"**{row['Name']}**\n"
                        f"₹{row['Price']:,.2f}  "
                        f"{arr}{abs(chg):.2f}%",
                        key=f"wlb_{btn_idx}",
                        width="stretch"):
                        st.session_state["sn"] = row["Name"]
                        st.session_state["st"] = row["Sym"]
                        st.rerun()
                btn_idx += 1

        # ── Full table ────────────────────────────────────
        st.markdown("---")
        tbl = valid[["Name","Price","Chg%","Chg₹",
                     "High","Low"]].copy()
        tbl.columns = ["Stock","Price ₹","Change %",
                       "Change ₹","Day High","Day Low"]
        tbl["Price ₹"]  = tbl["Price ₹"].apply(
            lambda x: f"₹{x:,.2f}")
        tbl["Day High"] = tbl["Day High"].apply(
            lambda x: f"₹{x:,.2f}")
        tbl["Day Low"]  = tbl["Day Low"].apply(
            lambda x: f"₹{x:,.2f}")
        tbl["Change %"] = tbl["Change %"].apply(
            lambda x: f"{x:+.2f}%")
        tbl["Change ₹"] = tbl["Change ₹"].apply(
            lambda x: f"{x:+.2f}")
        st.dataframe(tbl, width="stretch", hide_index=True)

# ╔══════════════════════════════════════════════════════╗
# ║  TAB 2 — TRADE SETUP                                ║
# ╚══════════════════════════════════════════════════════╝
with T2:
    # ── Inline stock search ───────────────────────────────
    st.markdown(
        "<div style='background:#f0f9ff;border:1px solid "
        "#bae6fd;border-radius:10px;padding:10px 14px;"
        "margin-bottom:12px'>"
        "<span style='font-size:13px;color:#0369a1;"
        "font-weight:600'>🔍 Search or browse stocks</span>"
        "</div>",
        unsafe_allow_html=True
    )
    inline_stock_search("t2")

    # Show which sector this stock is in
    _stock_sector = None
    for _sname, _stocks in SECTORS.items():
        if sname in _stocks:
            _stock_sector = _sname
            break
    if _stock_sector:
        st.caption(
            f"📂 {sname} is in **{_stock_sector}** sector. "
            f"To find it in Auto Scanner — select '{_stock_sector}' "
            f"and scan on same timeframe ({tf})."
        )

    st.markdown("---")

    # ── Live price card ───────────────────────────────────
    lp = live_price(stick)

    # Show data source indicator
    _kite_on = kite_is_connected()
    _kite_src = st.session_state.get("kite_data_source","")
    _kite_err = st.session_state.get("kite_candle_error","")
    _src_col  = "#f0fdf4" if _kite_on else "#fffbeb"
    _src_bdr  = "#86efac" if _kite_on else "#fcd34d"
    _src_txt  = "#166534" if _kite_on else "#92400e"

    if _kite_on and _kite_src.startswith("Kite:") and "token" in _kite_src:
        _src_msg = (
            "⚡ <b>Kite LIVE</b> — Real-time candles active. "
            "All signals are live."
        )
    elif _kite_on:
        _src_msg = (
            "⚡ Kite connected but loading candle data... "
            "First load takes ~30 seconds to fetch instruments."
            + (f" Error: {_kite_err}" if _kite_err else "")
        )
        _src_col = "#fffbeb"
        _src_bdr = "#fcd34d"
        _src_txt = "#92400e"
    else:
        _src_msg = (
            "📊 <b>Yahoo Finance</b> — 15 min delay. "
            "Login with Kite in sidebar for live signals."
        )
    st.markdown(
        f"<div style='background:{_src_col};border:1px solid {_src_bdr};"
        f"border-radius:8px;padding:8px 14px;margin-bottom:8px;"
        f"font-size:12px;color:{_src_txt}'>{_src_msg}</div>",
        unsafe_allow_html=True
    )
    if lp["ok"]:
        pc  = "#00ff88" if lp["chg"]>=0 else "#ff4455"
        arr = "▲" if lp["chg"]>=0 else "▼"
        st.markdown(f"""
        <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)' style='display:flex;
             justify-content:space-between;
             align-items:center;flex-wrap:wrap;gap:12px'>
          <div>
            <div style='color:#64748b;font-size:13px'>
                {sname}
                <span style='color:#94a3b8;
                             margin-left:8px'>{stick}</span>
            </div>
            <div style='font-size:34px;font-weight:700;line-height:1.1' style='color:#1e293b'>
                ₹{lp['p']:,.2f}
            </div>
            <div style='color:{pc};font-size:17px;
                        margin-top:4px'>
                {arr} {abs(lp['chg']):.2f}%
                <span style='color:#94a3b8;font-size:13px'>
                    ({lp['chg_abs']:+.2f})
                </span>
            </div>
          </div>
          <div style='color:#64748b;font-size:13px;
                      line-height:2.2;text-align:right'>
            High <b style='color:#1e293b'>
                ₹{lp['high']:,}</b><br>
            Low  <b style='color:#1e293b'>
                ₹{lp['low']:,}</b><br>
            Prev <b style='color:#1e293b'>
                ₹{lp['prev']:,}</b>
          </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.warning("Could not fetch live price")

    # Fetch + compute
    with st.spinner("Analysing..."):
        df = candles(stick, tf)

    col_info, col_ref = st.columns([4,1])
    with col_info:
        if not df.empty:
            st.caption(
                f"✅ {len(df)} candles | "
                f"Last: {df.index[-1].strftime('%d %b %H:%M')} IST"
                f" | ⚠️ ~15 min delay"
            )
    with col_ref:
        if st.button("🔄 Refresh",type="primary",
                     key="t2_ref",width="stretch"):
            st.cache_data.clear(); st.rerun()

    if df.empty or len(df) < 55:
        st.error("Not enough data. Try 1d timeframe.")
        st.stop()

    sig = compute_all(df, lp)
    if not sig:
        st.error("Could not calculate signals.")
        st.stop()

    cp = sig["cp"]

    # ── Trade direction cards ──────────────────────────────
    st.markdown("---")
    d1c, d2c = st.columns(2)

    def score_bar(score, max_score=10):
        pct  = int(score/max_score*100)
        col  = ("#00ff88" if pct>=70
                else "#ffcc00" if pct>=50
                else "#ff4455")
        return (f"<div style='background:#f1f5f9;border-radius:4px;"
                f"height:8px;width:100%;margin-top:6px'>"
                f"<div style='background:{col};width:{pct}%;"
                f"height:8px;border-radius:4px'></div></div>")

    with d1c:
        uc = ("#00ff88" if sig['up_score']>=7
              else "#ffcc00" if sig['up_score']>=5
              else "#ff4455")
        st.markdown(f"""
        <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)'>
          <div style='font-size:11px;color:#64748b;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px'>UPTREND SCORE</div>
          <div style='font-size:52px;font-weight:700;
                      color:{uc};line-height:1'>
              {sig['up_score']}/10
          </div>
          <div style='color:{uc};font-size:13px;
                      margin-top:4px'>
              {'🔥 BUY CE' if sig['up_score']>=7
               else '⏳ WAIT' if sig['up_score']>=5
               else '❌ NO CE TRADE'}
          </div>
          {score_bar(sig['up_score'])}
          <div style='margin-top:10px;font-size:13px;
                      color:#555'>
              CE checklist: {sig['ce_pass']}/11 passed
          </div>
        </div>
        """, unsafe_allow_html=True)

    with d2c:
        dc = ("#ff4455" if sig['dn_score']>=7
              else "#ffcc00" if sig['dn_score']>=5
              else "#555")
        st.markdown(f"""
        <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)'>
          <div style='font-size:11px;color:#64748b;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px'>DOWNTREND SCORE</div>
          <div style='font-size:52px;font-weight:700;
                      color:{dc};line-height:1'>
              {sig['dn_score']}/10
          </div>
          <div style='color:{dc};font-size:13px;
                      margin-top:4px'>
              {'🔥 BUY PE' if sig['dn_score']>=7
               else '⏳ WAIT' if sig['dn_score']>=5
               else '❌ NO PE TRADE'}
          </div>
          {score_bar(sig['dn_score'])}
          <div style='margin-top:10px;font-size:13px;
                      color:#555'>
              PE checklist: {sig['pe_pass']}/11 passed
          </div>
        </div>
        """, unsafe_allow_html=True)

    # ── Multi-Timeframe Confirmation ──────────────────────
    st.markdown("---")
    st.markdown("### 🕐 Multi-Timeframe Confirmation")
    st.caption(
        "Checks 15m, 1h and 1d simultaneously. "
        "All 3 agreeing = highest confidence trade."
    )

    mtf_tfs   = [("15m","Intraday"),("1h","Short-term"),("1d","Medium-term")]
    mtf_cols  = st.columns(3)
    mtf_results = []

    for (mtf_tf, mtf_label), mtf_col in zip(mtf_tfs, mtf_cols):
        with st.spinner(f"Loading {mtf_tf}..."):
            mtf_df  = candles(stick, mtf_tf)
            mtf_sig = None
            if mtf_df is not None and len(mtf_df) >= 55:
                try:
                    mtf_sig = compute_all(mtf_df, live_price(stick))
                except:
                    pass

        mtf_dir = mtf_sig["direction"] if mtf_sig else "UNKNOWN"
        mtf_results.append(mtf_dir)
        mtf_col_hex = (
            "#16a34a" if mtf_dir=="UPTREND"
            else "#dc2626" if mtf_dir=="DOWNTREND"
            else "#f59e0b"
        )
        mtf_bg = (
            "#f0fdf4" if mtf_dir=="UPTREND"
            else "#fef2f2" if mtf_dir=="DOWNTREND"
            else "#fffbeb"
        )
        with mtf_col:
            if mtf_sig:
                st.markdown(
                    f"<div style='background:{mtf_bg};"
                    f"border:1.5px solid {mtf_col_hex};"
                    f"border-radius:10px;padding:14px;"
                    f"text-align:center'>"
                    f"<div style='font-size:11px;color:#64748b'>"
                    f"{mtf_tf} — {mtf_label}</div>"
                    f"<div style='font-size:18px;font-weight:700;"
                    f"color:{mtf_col_hex};margin:4px 0'>{mtf_dir}</div>"
                    f"<div style='font-size:12px;color:#475569'>"
                    f"Score {max(mtf_sig['up_score'],mtf_sig['dn_score'])}/10</div>"
                    f"<div style='font-size:11px;color:#64748b'>"
                    f"ST: {'BUY' if mtf_sig['st_bull'] else 'SELL'} | "
                    f"RSI: {mtf_sig['rv']:.0f}</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f"<div style='background:#f8fafc;"
                    f"border:1px solid #e2e8f0;border-radius:10px;"
                    f"padding:14px;text-align:center'>"
                    f"<div style='font-size:11px;color:#64748b'>"
                    f"{mtf_tf}</div>"
                    f"<div style='color:#94a3b8'>No data</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )

    # MTF verdict
    bull_c = mtf_results.count("UPTREND")
    bear_c = mtf_results.count("DOWNTREND")

    if bull_c == 3:
        st.success(
            "🔥 ALL 3 TIMEFRAMES BULLISH — "
            "Highest confidence BUY CE setup!"
        )
    elif bear_c == 3:
        st.error(
            "🔥 ALL 3 TIMEFRAMES BEARISH — "
            "Highest confidence BUY PE setup!"
        )
    elif bull_c == 2:
        st.warning(
            "⚡ 2/3 timeframes BULLISH — "
            "Good CE setup. Wait for 15m to confirm."
        )
    elif bear_c == 2:
        st.warning(
            "⚡ 2/3 timeframes BEARISH — "
            "Good PE setup. Wait for 15m to confirm."
        )
    else:
        st.info(
            "⚠️ Timeframes conflicting — "
            "mixed signals. Wait for alignment."
        )

    # ── ML Prediction + Diamond Verdict ──────────────────
    st.markdown("---")
    st.markdown("### 🤖 ML Confirmation")
    st.caption(
        "ML is trained on this stock's own historical data. "
        "When ML agrees with technical signal — "
        "confidence is significantly higher."
    )

    with st.spinner("Running ML prediction..."):
        try:
            _ml_df = candles(stick, "1d")
            _ml_pred = None
            _ml_dir  = "UNKNOWN"
            _ml_conf = 0
            _ml_ok   = False
            if _ml_df is not None and len(_ml_df) >= 100:
                _ml_model = train_model(_ml_df)
                if _ml_model.get("ok"):
                    _ml_pred = predict_next_move(_ml_df, _ml_model)
                    if _ml_pred and _ml_pred.get("ok"):
                        _ml_dir  = _ml_pred["prediction"]
                        _ml_conf = _ml_pred["confidence"]
                        _ml_ok   = True
        except Exception:
            _ml_ok = False

    # Current signal direction from Trade Setup
    _t2_dir = sig["direction"] if sig else "UNKNOWN"
    _ml_agrees = _ml_ok and _ml_dir == _t2_dir
    _mtf_all   = bull_c == 3 or bear_c == 3

    # Show ML result
    ml_col1, ml_col2 = st.columns(2)
    with ml_col1:
        if _ml_ok:
            _mc = "#16a34a" if _ml_dir=="UPTREND" else "#dc2626" if _ml_dir=="DOWNTREND" else "#f59e0b"
            _mb = "#f0fdf4" if _ml_dir=="UPTREND" else "#fef2f2" if _ml_dir=="DOWNTREND" else "#fffbeb"
            st.markdown(
                f"<div style='background:{_mb};"
                f"border:1.5px solid {_mc};"
                f"border-radius:12px;padding:16px;"
                f"text-align:center'>"
                f"<div style='font-size:11px;color:#64748b;"
                f"text-transform:uppercase;letter-spacing:1px'>"
                f"ML Prediction (Daily)</div>"
                f"<div style='font-size:28px;font-weight:700;"
                f"color:{_mc};margin:6px 0'>{_ml_dir}</div>"
                f"<div style='font-size:14px;color:{_mc};"
                f"font-weight:600'>{_ml_conf}% confidence</div>"
                f"<div style='font-size:12px;color:#64748b;"
                f"margin-top:4px'>{_ml_pred.get('reliability','')}</div>"
                f"</div>",
                unsafe_allow_html=True
            )
        else:
            st.info("ML needs 100+ daily candles for this stock.")

    with ml_col2:
        # Agreement summary
        _agree_count = sum([
            _ml_agrees,
            _mtf_all,
            sig["up_score"] >= 7 if _t2_dir=="UPTREND"
            else sig["dn_score"] >= 7
        ])
        st.markdown(
            f"<div style='background:#f8fafc;"
            f"border:1px solid #e2e8f0;"
            f"border-radius:12px;padding:16px'>"
            f"<div style='font-size:13px;font-weight:700;"
            f"color:#374151;margin-bottom:12px'>"
            f"Confirmation Checklist</div>"
            f"<div style='font-size:13px;line-height:2.2'>"
            f"{'✅' if sig['up_score']>=7 or sig['dn_score']>=7 else '❌'} "
            f"Technical score 7+ "
            f"({max(sig['up_score'],sig['dn_score'])}/10)<br>"
            f"{'✅' if _ml_agrees else '❌'} "
            f"ML prediction agrees "
            f"({'Yes' if _ml_agrees else 'No — ' + _ml_dir})<br>"
            f"{'✅' if _mtf_all else '❌'} "
            f"All 3 timeframes agree "
            f"({'Yes' if _mtf_all else str(max(bull_c,bear_c)) + '/3 agree'})<br>"
            f"{'✅' if sig.get('st_bull') == (_t2_dir=='UPTREND') else '❌'} "
            f"Supertrend confirms</div>"
            f"</div>",
            unsafe_allow_html=True
        )

    # Diamond verdict for manually searched stock
    st.markdown("---")
    _is_diamond = (
        _ml_agrees and _mtf_all and
        (sig["up_score"] >= 7 or sig["dn_score"] >= 7)
    )

    if _is_diamond:
        _d_action = "BUY CE" if _t2_dir == "UPTREND" else "BUY PE"
        st.markdown(
            f"<div style='background:linear-gradient("
            f"135deg,#1e1b4b,#3730a3);"
            f"border-radius:14px;padding:20px 24px;"
            f"text-align:center'>"
            f"<div style='font-size:28px;font-weight:700;"
            f"color:#ffffff'>💎 DIAMOND SIGNAL</div>"
            f"<div style='font-size:16px;color:#c7d2fe;"
            f"margin-top:8px'>{sname} — {_d_action}</div>"
            f"<div style='font-size:13px;color:#a5b4fc;"
            f"margin-top:6px'>"
            f"Technical ✅  ML ✅  All 3 Timeframes ✅</div>"
            f"<div style='font-size:13px;color:#818cf8;"
            f"margin-top:4px'>"
            f"ML: {_ml_dir} ({_ml_conf}%) | "
            f"Score: {max(sig['up_score'],sig['dn_score'])}/10"
            f"</div></div>",
            unsafe_allow_html=True
        )
        # Send to Telegram
        if tg_configured():
            if st.button(
                "📱 Send Diamond Signal to Telegram",
                key="t2_send_diamond",
                type="primary",
                use_container_width=True
            ):
                _tok = st.session_state.get("tg_token_saved","")
                _cid = st.session_state.get("tg_chat_saved","")
                _msg = (
                    f"💎 DIAMOND — {sname}\n"
                    f"{_d_action} | "
                    f"Score {max(sig['up_score'],sig['dn_score'])}/10\n"
                    f"ML: {_ml_dir} ({_ml_conf}%)\n"
                    f"All 3 timeframes confirmed\n"
                    f"Price: Rs {sig['cp']:,.2f}\n"
                    f"Entry: Rs {sig['e9v']:,.2f} | "
                    f"SL: Rs {sig['sl_long'] if _t2_dir=='UPTREND' else sig['sl_short']:,.2f}\n"
                    f"T1: Rs {sig['tgt1'] if _t2_dir=='UPTREND' else sig['tgt1s']:,.2f}"
                )
                if send_telegram(_tok, _cid, _msg):
                    st.success("✅ Diamond signal sent to Telegram!")
                else:
                    st.error("❌ Failed to send")
    elif _ml_agrees:
        st.success(
            f"⚡ ML agrees with technical signal ({_ml_dir}) — "
            f"good setup. Waiting for all 3 timeframes to align "
            f"for Diamond confirmation."
        )
    elif _mtf_all:
        st.warning(
            f"⚡ All timeframes agree but ML says {_ml_dir}. "
            f"Technical signal is strong. "
            f"Trade with caution."
        )
    else:
        st.info(
            "No Diamond Signal yet for this stock. "
            "Technical score, ML and all timeframes need to "
            "agree simultaneously."
        )

    # ── 11-FACTOR CHECKLISTS ───────────────────────────────
    st.markdown("---")
    st.markdown("### ✅ Pre-Trade Checklist")
    st.caption(
        "All 11 factors must be green before entering. "
        "Even one red factor means wait."
    )

    cl1, cl2 = st.columns(2)

    def render_checklist(checklist, title, color):
        passed = sum(1 for c in checklist if c[1])
        total  = len(checklist)
        pct    = int(passed/total*100)
        hdr_col= ("#00ff88" if pct>=82
                  else "#ffcc00" if pct>=60
                  else "#ff4455")
        st.markdown(f"""
        <div style='font-size:15px;font-weight:600;
                    color:{hdr_col};margin-bottom:8px'>
            {title} — {passed}/{total} passed
            <span style='font-size:12px;color:#64748b;
                         margin-left:8px'>
                ({pct}% ready)
            </span>
        </div>
        """, unsafe_allow_html=True)
        for label, passed_, value, why in checklist:
            css_style = ("background:#071407;border-left:3px solid #00ff88" if passed_ 
                     else "background:#140707;border-left:3px solid #ff4455")
            ico = "✅" if passed_ else "❌"
            st.markdown(f"""
            <div style='{css_style};border-radius:6px;padding:9px 14px;margin:3px 0;font-size:14px;color:#ccc'>
              <span style='font-weight:600'>{ico} {label}</span>
              <span style='float:right;color:#6b7280;
                           font-size:12px'>{value}</span>
              <div style='font-size:11px;color:#64748b;
                          margin-top:2px'>{why}</div>
            </div>
            """, unsafe_allow_html=True)

    with cl1:
        render_checklist(
            sig["ce_checklist"],
            "📈 CE (CALL) — Bullish",
            "#00ff88"
        )
    with cl2:
        render_checklist(
            sig["pe_checklist"],
            "📉 PE (PUT) — Bearish",
            "#ff4455"
        )

    # ── ENTRY / EXIT BOXES ────────────────────────────────
    st.markdown("---")
    st.markdown("### 🎯 Entry & Exit Points")

    # Smart money alerts first
    if sig["sweep_low"]:
        st.success(
            "🚀 **LIQUIDITY SWEEP OF LOWS** — "
            "Institutions just triggered stop losses and bought. "
            "Highest probability CE setup right now!"
        )
    if sig["sweep_high"]:
        st.error(
            "⚠️ **LIQUIDITY SWEEP OF HIGHS** — "
            "Institutions just triggered stop losses and sold. "
            "Highest probability PE setup right now!"
        )
    if sig["bos_bull"]:
        st.success(
            "📊 **BULLISH BREAK OF STRUCTURE** — "
            "Institutions confirmed uptrend with high volume."
        )
    if sig["bos_bear"]:
        st.error(
            "📊 **BEARISH BREAK OF STRUCTURE** — "
            "Institutions confirmed downtrend with high volume."
        )

    # Candlestick patterns
    if sig["patterns"]:
        for pname, pbias, pmeaning in sig["patterns"]:
            pc_ = ("#00ff88" if pbias=="bullish"
                   else "#ff4455" if pbias=="bearish"
                   else "#ffcc00")
            st.markdown(f"""
            <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)' style='border-left:3px solid {pc_}'>
              <span style='color:{pc_};font-weight:600'>
                  {pname}
              </span>
              <span style='color:#666;font-size:13px;
                           margin-left:10px'>
                  {pmeaning}
              </span>
            </div>
            """, unsafe_allow_html=True)

    # Determine recommended action
    ce_ready = sig["ce_pass"] >= 9 and sig["up_score"] >= 7
    pe_ready = sig["pe_pass"] >= 9 and sig["dn_score"] >= 7

    if ce_ready:
        # ── BUY CE entry box ──────────────────────────────
        st.success(
            f"🟢 BUY CE (CALL OPTION) — "
            f"{sig['ce_pass']}/11 checks passed"
        )

        ec1, ec2, ec3 = st.columns(3)
        with ec1:
            st.markdown(
                "<div style='background:#f0fdf4;border-radius:8px;"
                "padding:16px;text-align:center;height:140px'>"
                "<div style='font-size:11px;color:#64748b;letter-spacing:1px;"
                "text-transform:uppercase;margin-bottom:8px'>Entry Zone</div>"
                f"<div style='color:#00ff88;font-size:22px;font-weight:700'>"
                f"₹{sig.get('entry_long', sig['e9v']):,.2f}</div>"
                "<div style='color:#64748b;font-size:12px;margin-top:8px'>"
                f"{'Enter at market price' if sig.get('entry_long', sig['e9v']) >= sig['cp']*0.999 else 'Wait for pullback to EMA9'}<br>then enter on next green candle"
                "</div></div>",
                unsafe_allow_html=True
            )
        with ec2:
            st.markdown(
                "<div style='background:#fef2f2;border-radius:8px;"
                "padding:16px;text-align:center;height:140px'>"
                "<div style='font-size:11px;color:#64748b;letter-spacing:1px;"
                "text-transform:uppercase;margin-bottom:8px'>Stop Loss</div>"
                f"<div style='color:#ff4455;font-size:22px;font-weight:700'>"
                f"₹{sig['sl_long']:,.2f}</div>"
                "<div style='color:#64748b;font-size:12px;margin-top:8px'>"
                "ATR-based SL<br>"
                f"Exit if price closes below EMA21<br>(₹{sig['e21v']:,.2f})"
                "</div></div>",
                unsafe_allow_html=True
            )
        with ec3:
            st.markdown(
                "<div style='background:#eff6ff;border-radius:8px;"
                "padding:16px;text-align:center;height:140px'>"
                "<div style='font-size:11px;color:#64748b;letter-spacing:1px;"
                "text-transform:uppercase;margin-bottom:8px'>Targets</div>"
                f"<div style='color:#88aaff;font-size:15px;font-weight:600;"
                f"line-height:1.8'>T1 ₹{sig['tgt1']:,.2f}<br>"
                f"T2 ₹{sig['tgt2']:,.2f}<br>"
                f"T3 ₹{sig['tgt3']:,.2f}</div>"
                "</div>",
                unsafe_allow_html=True
            )

        _ce_entry = sig.get("entry_long", sig["e9v"])
        _ce_sl    = sig["sl_long"]
        _ce_t1    = sig["tgt1"]
        _ce_risk  = abs(_ce_entry - _ce_sl)
        _ce_rew   = abs(_ce_t1   - _ce_entry)
        _ce_rr    = round(_ce_rew / (_ce_risk + 0.001), 2)
        st.markdown(
            f"**Risk-Reward:** {_ce_rr}:1 &nbsp;|&nbsp; "
            f"**ATR:** ₹{sig['atrv']:,.2f} &nbsp;|&nbsp; "
            f"**Support:** ₹{sig['sup']:,} &nbsp;|&nbsp; "
            f"**Resistance:** ₹{sig['res']:,}"
        )

        st.markdown("#### ⚠️ Exit Rules — Follow without exception")
        st.warning(
            f"🔴 **Stop loss hit** → Price closes below "
            f"₹{sig['sl_long']:,.2f} → Exit immediately  \n"
            f"📊 **RSI overbought** → RSI crosses above 72 → Book profit  \n"
            f"🎯 **Target reached** → At T1 (₹{sig['tgt1']:,.2f}) book 50% qty  \n"
            f"&nbsp;&nbsp; At T2 (₹{sig['tgt2']:,.2f}) book another 30%  \n"
            f"⏰ **Time exit** → Exit ALL positions by 2:45 PM  \n"
            f"📉 **EMA21 break** → Price closes below ₹{sig['e21v']:,.2f} → Exit"
        )

    elif pe_ready:
        # ── BUY PE entry box ──────────────────────────────
        st.error(
            f"🔴 BUY PE (PUT OPTION) — "
            f"{sig['pe_pass']}/11 checks passed"
        )

        pc1, pc2, pc3 = st.columns(3)
        with pc1:
            st.markdown(
                "<div style='background:#fef2f2;border-radius:8px;"
                "padding:16px;text-align:center;height:140px'>"
                "<div style='font-size:11px;color:#64748b;letter-spacing:1px;"
                "text-transform:uppercase;margin-bottom:8px'>Entry Zone</div>"
                f"<div style='color:#ff4455;font-size:22px;font-weight:700'>"
                f"₹{sig.get('entry_short', sig['e9v']):,.2f}</div>"
                "<div style='color:#64748b;font-size:12px;margin-top:8px'>"
                f"{'Enter at market price' if sig.get('entry_short', sig['e9v']) <= sig['cp']*1.001 else 'Wait for bounce to EMA9'}<br>then enter on next red candle"
                "</div></div>",
                unsafe_allow_html=True
            )
        with pc2:
            st.markdown(
                "<div style='background:#1a0800;border-radius:8px;"
                "padding:16px;text-align:center;height:140px'>"
                "<div style='font-size:11px;color:#64748b;letter-spacing:1px;"
                "text-transform:uppercase;margin-bottom:8px'>Stop Loss</div>"
                f"<div style='color:#ff8844;font-size:22px;font-weight:700'>"
                f"₹{sig['sl_short']:,.2f}</div>"
                "<div style='color:#64748b;font-size:12px;margin-top:8px'>"
                "ATR-based SL<br>"
                f"Exit if price closes above EMA21<br>(₹{sig['e21v']:,.2f})"
                "</div></div>",
                unsafe_allow_html=True
            )
        with pc3:
            st.markdown(
                "<div style='background:#eff6ff;border-radius:8px;"
                "padding:16px;text-align:center;height:140px'>"
                "<div style='font-size:11px;color:#64748b;letter-spacing:1px;"
                "text-transform:uppercase;margin-bottom:8px'>Targets</div>"
                f"<div style='color:#88aaff;font-size:15px;font-weight:600;"
                f"line-height:1.8'>T1 ₹{sig['tgt1s']:,.2f}<br>"
                f"T2 ₹{sig['tgt2s']:,.2f}<br>"
                f"T3 ₹{sig['tgt3s']:,.2f}</div>"
                "</div>",
                unsafe_allow_html=True
            )

        _pe_entry = sig.get("entry_short", sig["e9v"])
        _pe_sl    = sig["sl_short"]
        _pe_t1    = sig["tgt1s"]
        _pe_risk  = abs(_pe_sl    - _pe_entry)
        _pe_rew   = abs(_pe_entry - _pe_t1)
        _pe_rr    = round(_pe_rew / (_pe_risk + 0.001), 2)
        st.markdown(
            f"**Risk-Reward:** {_pe_rr}:1 &nbsp;|&nbsp; "
            f"**ATR:** ₹{sig['atrv']:,.2f} &nbsp;|&nbsp; "
            f"**Support:** ₹{sig['sup']:,} &nbsp;|&nbsp; "
            f"**Resistance:** ₹{sig['res']:,}"
        )

        st.markdown("#### ⚠️ Exit Rules — Follow without exception")
        st.warning(
            f"🔴 **Stop loss hit** → Price closes above "
            f"₹{sig['sl_short']:,.2f} → Exit immediately  \n"
            f"📊 **RSI oversold** → RSI crosses below 28 → Book profit  \n"
            f"🎯 **Target reached** → At T1 (₹{sig['tgt1s']:,.2f}) book 50%  \n"
            f"&nbsp;&nbsp; At T2 (₹{sig['tgt2s']:,.2f}) book another 30%  \n"
            f"⏰ **Time exit** → Exit ALL positions by 2:45 PM  \n"
            f"📈 **EMA21 break** → Price closes above ₹{sig['e21v']:,.2f} → Exit"
        )

    else:
        # ── No trade ready ────────────────────────────────
        missing_ce = [c[0] for c in sig["ce_checklist"] if not c[1]]
        missing_pe = [c[0] for c in sig["pe_checklist"] if not c[1]]

        st.warning(
            f"⏳ **NO TRADE YET** — "
            f"CE {sig['ce_pass']}/11 | PE {sig['pe_pass']}/11  \n\n"
            f"**Missing for CE:** {', '.join(missing_ce[:4])}  \n"
            f"**Missing for PE:** {', '.join(missing_pe[:4])}  \n\n"
            "Wait for score to reach 7+ and 9+ checks green before entering."
        )

    # ── CHART ─────────────────────────────────────────────
    st.markdown("---")
    st.markdown(f"### 🕯️ {sname} — {tf}")

    plot_df = df.tail(100).copy()
    fig = make_subplots(
        rows=4, cols=1,
        shared_xaxes=True,
        row_heights=[0.55,0.15,0.15,0.15],
        vertical_spacing=0.02,
    )

    # Candles
    fig.add_trace(go.Candlestick(
        x=plot_df.index,
        open=plot_df["Open"],  high=plot_df["High"],
        low=plot_df["Low"],    close=plot_df["Close"],
        name="Price",
        increasing_line_color="lime",
        decreasing_line_color="#ff4455",
    ), row=1, col=1)

    # EMAs
    for ser,col_,nm in [
        (sig["e9s"],"yellow","EMA9"),
        (sig["e21s"],"orange","EMA21"),
        (sig["e50s"],"cyan","EMA50"),
    ]:
        fig.add_trace(go.Scatter(
            x=plot_df.index, y=ser.tail(100),
            line=dict(color=col_,width=1.2),name=nm
        ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=plot_df.index, y=sig["vwaps"].tail(100),
        line=dict(color="white",width=1.2,dash="dash"),
        name="VWAP"
    ), row=1, col=1)

    # VWAP Bands on chart
    if sig and sig.get("vwap_upper"):
        _vw_vals = sig["vwaps"].tail(100)
        _std_ser = sig["rsis"].tail(100) * 0  # zero series for shift
        try:
            _vw_std = float(sig["vwaps"].tail(20).std())
            _vw_up  = _vw_vals + _vw_std
            _vw_dn  = _vw_vals - _vw_std
            fig.add_trace(go.Scatter(
                x=plot_df.index[-100:], y=_vw_up,
                line=dict(color="rgba(251,191,36,0.5)", width=1, dash="dot"),
                name="VWAP +1SD", showlegend=True
            ), row=1, col=1)
            fig.add_trace(go.Scatter(
                x=plot_df.index[-100:], y=_vw_dn,
                line=dict(color="rgba(251,191,36,0.5)", width=1, dash="dot"),
                name="VWAP -1SD",
                fill="tonexty",
                fillcolor="rgba(251,191,36,0.03)",
                showlegend=True
            ), row=1, col=1)
        except Exception:
            pass

    # Bollinger Bands on chart
    if sig:
        # Calculate BB series for chart
        _bb_mid = plot_df["Close"].rolling(20).mean()
        _bb_std = plot_df["Close"].rolling(20).std()
        _bb_up  = _bb_mid + 2 * _bb_std
        _bb_low = _bb_mid - 2 * _bb_std

        fig.add_trace(go.Scatter(
            x=plot_df.index, y=_bb_up.tail(100),
            line=dict(color="rgba(147,197,253,0.6)",
                      width=1, dash="dot"),
            name="BB Upper",
            showlegend=True
        ), row=1, col=1)
        fig.add_trace(go.Scatter(
            x=plot_df.index, y=_bb_low.tail(100),
            line=dict(color="rgba(147,197,253,0.6)",
                      width=1, dash="dot"),
            name="BB Lower",
            fill="tonexty",
            fillcolor="rgba(147,197,253,0.05)",
            showlegend=True
        ), row=1, col=1)
        fig.add_trace(go.Scatter(
            x=plot_df.index, y=_bb_mid.tail(100),
            line=dict(color="rgba(147,197,253,0.4)",
                      width=1),
            name="BB Mid",
            showlegend=False
        ), row=1, col=1)

    # Supertrend line on chart
    if sig and sig["st_value"] > 0:
        st_color = "#16a34a" if sig["st_bull"] else "#dc2626"
        st_series = [sig["st_value"]] * len(plot_df)
        fig.add_trace(go.Scatter(
            x=plot_df.index,
            y=st_series,
            line=dict(color=st_color, width=2, dash="dot"),
            name=f"Supertrend {sig['st_signal']}",
            opacity=0.8
        ), row=1, col=1)

        # Fibonacci levels on chart
        fib_data = [
            (sig["fib_236"], "Fib 23.6%", "#94a3b8"),
            (sig["fib_382"], "Fib 38.2%", "#f59e0b"),
            (sig["fib_500"], "Fib 50.0%", "#f59e0b"),
            (sig["fib_618"], "Fib 61.8%", "#ef4444"),
            (sig["fib_786"], "Fib 78.6%", "#ef4444"),
        ]
        for fib_val, fib_name, fib_col in fib_data:
            if sig["fib_low"] < fib_val < sig["fib_high"]:
                fig.add_hline(
                    y=fib_val,
                    line_dash="dot",
                    line_color=fib_col,
                    line_width=1,
                    opacity=0.5,
                    annotation_text=fib_name,
                    annotation_position="right",
                    row=1, col=1
                )

    # Weekly Pivot lines on chart
    if sig and sig.get("w_pivot"):
        _wp_levels = [
            (sig["w_pivot"], "W.Pivot", "#a78bfa", "dash"),
            (sig["w_r1"],    "W.R1",    "#f87171", "dot"),
            (sig["w_s1"],    "W.S1",    "#4ade80", "dot"),
        ]
        for _wv, _wn, _wc, _wd in _wp_levels:
            if abs(_wv - sig["cp"]) / sig["cp"] < 0.05:
                fig.add_hline(
                    y=_wv, line_dash=_wd,
                    line_color=_wc, line_width=1,
                    opacity=0.6,
                    annotation_text=_wn,
                    annotation_position="left",
                    row=1, col=1
                )

    # Monthly Pivot lines on chart
    if sig and sig.get("m_pivot"):
        _mp_lvls = [
            (sig["m_pivot"], "M.Pivot", "#818cf8", "longdash"),
            (sig["m_r1"],    "M.R1",    "#f43f5e", "longdashdot"),
            (sig["m_s1"],    "M.S1",    "#34d399", "longdashdot"),
        ]
        for _mv, _mn, _mc, _md in _mp_lvls:
            if abs(_mv - sig["cp"]) / sig["cp"] < 0.08:
                fig.add_hline(
                    y=_mv, line_dash=_md,
                    line_color=_mc, line_width=1.5,
                    opacity=0.5,
                    annotation_text=_mn,
                    annotation_position="right",
                    row=1, col=1
                )

    # CPR lines on chart
    if sig:
        # CPR TC (top)
        fig.add_hline(
            y=sig["cpr_tc"],
            line_dash="dot", line_color="#f59e0b",
            line_width=1.5, opacity=0.8,
            annotation_text=f"CPR TC {sig['cpr_tc']:,}",
            annotation_position="left",
            row=1, col=1
        )
        # CPR Pivot
        fig.add_hline(
            y=sig["cpr_pivot"],
            line_dash="solid", line_color="#f59e0b",
            line_width=2, opacity=0.9,
            annotation_text=f"Pivot {sig['cpr_pivot']:,}",
            annotation_position="left",
            row=1, col=1
        )
        # CPR BC (bottom)
        fig.add_hline(
            y=sig["cpr_bc"],
            line_dash="dot", line_color="#f59e0b",
            line_width=1.5, opacity=0.8,
            annotation_text=f"CPR BC {sig['cpr_bc']:,}",
            annotation_position="left",
            row=1, col=1
        )

    fig.add_trace(go.Scatter(
        x=plot_df.index, y=sig["bbus"].tail(100),
        line=dict(color="rgba(60,200,100,0.4)",
                  width=1,dash="dot"),name="BB Upper"
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=plot_df.index, y=sig["bbls"].tail(100),
        line=dict(color="rgba(200,60,60,0.4)",
                  width=1,dash="dot"),name="BB Lower",
        fill="tonexty",
        fillcolor="rgba(255,255,255,0.02)"
    ), row=1, col=1)

    # SL lines on chart
    if ce_ready:
        fig.add_hline(
            y=sig["sl_long"],
            line_dash="dash",line_color="#ff4455",
            opacity=0.7,
            annotation_text=f"SL ₹{sig['sl_long']:,}",
            annotation_position="right",
            row=1, col=1)
        for i,(tv,tn) in enumerate([
                (sig["tgt1"],"T1"),
                (sig["tgt2"],"T2"),
                (sig["tgt3"],"T3")]):
            fig.add_hline(
                y=tv,
                line_dash="dot",line_color="#00ff88",
                opacity=0.5,
                annotation_text=f"{tn} ₹{tv:,}",
                annotation_position="right",
                row=1, col=1)

    elif pe_ready:
        fig.add_hline(
            y=sig["sl_short"],
            line_dash="dash",line_color="#ff8844",
            opacity=0.7,
            annotation_text=f"SL ₹{sig['sl_short']:,}",
            annotation_position="right",
            row=1, col=1)
        for tv,tn in [
                (sig["tgt1s"],"T1"),
                (sig["tgt2s"],"T2"),
                (sig["tgt3s"],"T3")]:
            fig.add_hline(
                y=tv,
                line_dash="dot",line_color="#ff4455",
                opacity=0.5,
                annotation_text=f"{tn} ₹{tv:,}",
                annotation_position="right",
                row=1, col=1)

    # Volume
    vc = ["lime" if float(plot_df["Close"].iloc[i]) >=
                    float(plot_df["Open"].iloc[i])
          else "#ff4455"
          for i in range(len(plot_df))]
    fig.add_trace(go.Bar(
        x=plot_df.index, y=plot_df["Volume"],
        marker_color=vc, name="Volume", opacity=0.7
    ), row=2, col=1)

    # RSI
    fig.add_trace(go.Scatter(
        x=plot_df.index, y=sig["rsis"].tail(100),
        line=dict(color="violet",width=1.5), name="RSI"
    ), row=3, col=1)
    for lvl,col_ in [(70,"red"),(68,"lime"),
                     (45,"orange"),(30,"red")]:
        fig.add_hline(y=lvl, line_dash="dash",
                      line_color=col_,opacity=0.4,
                      row=3, col=1)
    fig.add_hrect(y0=55,y1=68,
                  fillcolor="rgba(0,255,100,0.05)",
                  line_width=0, row=3, col=1)
    fig.add_hrect(y0=32,y1=45,
                  fillcolor="rgba(255,60,60,0.05)",
                  line_width=0, row=3, col=1)

    # MACD
    fig.add_trace(go.Scatter(
        x=plot_df.index, y=sig["macds"].tail(100),
        line=dict(color="yellow",width=1.3), name="MACD"
    ), row=4, col=1)
    fig.add_trace(go.Scatter(
        x=plot_df.index, y=sig["msigs"].tail(100),
        line=dict(color="#ff4455",width=1.3), name="Signal"
    ), row=4, col=1)
    hist = (sig["macds"]-sig["msigs"]).tail(100)
    fig.add_trace(go.Bar(
        x=plot_df.index, y=hist,
        marker_color=["lime" if v>=0 else "#ff4455"
                      for v in hist],
        name="Hist", opacity=0.6
    ), row=4, col=1)
    fig.add_hline(y=0, line_dash="dot",
                  line_color="#333", row=4, col=1)

    fig.update_layout(
        template="plotly_white",
        xaxis_rangeslider_visible=False,
        height=740,
        margin=dict(l=10,r=10,t=30,b=10),
        legend=dict(orientation="h",y=1.01,x=0,
                    font=dict(size=11)),
        title=(f"{sname} | {tf} | "
               "🟡EMA9 🟠EMA21 🔵EMA50 ⬜VWAP | "
               "Green zone=Buy RSI | Red zone=Sell RSI")
    )
    fig.update_yaxes(title_text="₹",     row=1, col=1)
    fig.update_yaxes(title_text="Vol",   row=2, col=1)
    fig.update_yaxes(title_text="RSI",
                     range=[0,100],      row=3, col=1)
    fig.update_yaxes(title_text="MACD",  row=4, col=1)
    st.plotly_chart(fig, width="stretch")

    # Key levels summary
    st.markdown("### 📌 Key Levels Summary")
    kl1,kl2,kl3,kl4,kl5,kl6 = st.columns(6)
    kl1.metric("Support",    f"₹{sig['sup']:,}")
    kl2.metric("Resistance", f"₹{sig['res']:,}")
    kl3.metric("VWAP",       f"₹{sig['vwv']:,.0f}")
    kl4.metric("EMA9",       f"₹{sig['e9v']:,.0f}")
    kl5.metric("EMA21",      f"₹{sig['e21v']:,.0f}")
    kl6.metric("ATR",        f"₹{sig['atrv']:,.1f}")

    # ── Weekly & Monthly Pivot Levels ─────────────────────
    st.markdown("#### 📅 Weekly & Monthly Key Levels")
    st.caption(
        "These levels are used by institutions for swing trades. "
        "Weekly for 1-3 day trades, Monthly for monthly options."
    )

    wm1, wm2 = st.columns(2)
    with wm1:
        st.markdown("**📅 Weekly Levels**")
        w_items = [
            ("W.Pivot", sig.get("w_pivot", 0), "#7c3aed"),
            ("W.R1 (Resistance)", sig.get("w_r1", 0), "#dc2626"),
            ("W.R2 (Resistance)", sig.get("w_r2", 0), "#dc2626"),
            ("W.S1 (Support)",    sig.get("w_s1", 0), "#16a34a"),
            ("W.S2 (Support)",    sig.get("w_s2", 0), "#16a34a"),
            ("Weekly High (Res)", sig.get("w_res", 0), "#dc2626"),
            ("Weekly Low (Sup)",  sig.get("w_sup", 0), "#16a34a"),
        ]
        for _wn, _wv, _wc in w_items:
            if _wv > 0:
                _is_near = abs(_wv - sig["cp"]) / sig["cp"] < 0.01
                st.markdown(
                    f"<div style='display:flex;justify-content:"
                    f"space-between;padding:5px 8px;"
                    f"background:{'#faf5ff' if _is_near else '#f8fafc'};"
                    f"border-radius:6px;margin:2px 0;"
                    f"border:{'1.5px solid '+_wc if _is_near else 'none'}'>"
                    f"<span style='font-size:12px;color:#64748b'>"
                    f"{_wn}{'  ◀ NEAR' if _is_near else ''}</span>"
                    f"<span style='font-size:13px;font-weight:700;"
                    f"color:{_wc}'>₹{_wv:,.0f}</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
    with wm2:
        st.markdown("**🗓️ Monthly Levels**")
        m_items = [
            ("M.Pivot", sig.get("m_pivot", 0), "#7c3aed"),
            ("M.R1 (Resistance)", sig.get("m_r1", 0), "#dc2626"),
            ("M.R2 (Resistance)", sig.get("m_r2", 0), "#dc2626"),
            ("M.S1 (Support)",    sig.get("m_s1", 0), "#16a34a"),
            ("M.S2 (Support)",    sig.get("m_s2", 0), "#16a34a"),
            ("Monthly High (Res)",sig.get("m_res", 0), "#dc2626"),
            ("Monthly Low (Sup)", sig.get("m_sup", 0), "#16a34a"),
        ]
        for _mn, _mv, _mc in m_items:
            if _mv > 0:
                _is_near = abs(_mv - sig["cp"]) / sig["cp"] < 0.02
                st.markdown(
                    f"<div style='display:flex;justify-content:"
                    f"space-between;padding:5px 8px;"
                    f"background:{'#eff6ff' if _is_near else '#f8fafc'};"
                    f"border-radius:6px;margin:2px 0;"
                    f"border:{'1.5px solid '+_mc if _is_near else 'none'}'>"
                    f"<span style='font-size:12px;color:#64748b'>"
                    f"{_mn}{'  ◀ NEAR' if _is_near else ''}</span>"
                    f"<span style='font-size:13px;font-weight:700;"
                    f"color:{_mc}'>₹{_mv:,.0f}</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )

    # VWAP Bands display
    _vz = sig.get("vwap_zone", "FAIR_VALUE")
    _vz_map = {
        "EXTREME_OB": ("🔴 EXTREME OVERBOUGHT — +2SD above VWAP", "#dc2626", "#fef2f2"),
        "OVERBOUGHT": ("🟠 OVERBOUGHT — +1SD above VWAP", "#ea580c", "#fff7ed"),
        "FAIR_VALUE": ("🟢 FAIR VALUE — Inside VWAP bands", "#16a34a", "#f0fdf4"),
        "OVERSOLD":   ("🟢 OVERSOLD — -1SD below VWAP. CE entry zone!", "#16a34a", "#f0fdf4"),
        "EXTREME_OS": ("🔥 EXTREME OVERSOLD — -2SD below VWAP. Strong CE entry!", "#7c3aed", "#faf5ff"),
    }
    _vz_label, _vz_col, _vz_bg = _vz_map.get(_vz, ("Normal", "#374151", "#f8fafc"))
    vb1, vb2, vb3, vb4 = st.columns(4)
    vb1.metric("VWAP", f"₹{sig['vwv']:,.0f}")
    vb2.metric("VWAP +1SD", f"₹{sig.get('vwap_upper', sig['vwv']):,.0f}")
    vb3.metric("VWAP -1SD", f"₹{sig.get('vwap_lower', sig['vwv']):,.0f}")
    vb4.metric("Zone", _vz.replace("_"," "))
    st.markdown(
        f"<div style='background:{_vz_bg};"
        f"border:1px solid {_vz_col};"
        f"border-radius:8px;padding:8px 14px;"
        f"font-size:13px;color:{_vz_col};margin-bottom:8px'>"
        f"<b>VWAP Bands:</b> {_vz_label}</div>",
        unsafe_allow_html=True
    )

    # Bollinger Bands display
    _bb_mid_v = round((sig["bbup"] + sig["bblw"]) / 2, 2)
    _bb_width = round(
        (sig["bbup"] - sig["bblw"]) / _bb_mid_v * 100, 2
    )
    _cp_vs_bb = (
        "🔴 Near Upper Band — Overbought"
        if sig["cp"] >= sig["bbup"] * 0.99
        else "🟢 Near Lower Band — Oversold / Entry zone"
        if sig["cp"] <= sig["bblw"] * 1.01
        else "🟡 Inside Bands — Normal range"
    )
    bb1, bb2, bb3, bb4 = st.columns(4)
    bb1.metric("BB Upper",  f"₹{sig['bbup']:,.0f}")
    bb2.metric("BB Mid",    f"₹{_bb_mid_v:,.0f}")
    bb3.metric("BB Lower",  f"₹{sig['bblw']:,.0f}")
    bb4.metric("BB Width",  f"{_bb_width:.1f}%",
               help="Low width = squeeze = breakout coming")
    st.markdown(
        f"<div style='background:#f0f9ff;"
        f"border:1px solid #bae6fd;border-radius:8px;"
        f"padding:8px 14px;font-size:13px;color:#0369a1'>"
        f"<b>Bollinger Bands:</b> {_cp_vs_bb} &nbsp;|&nbsp; "
        f"{'Squeeze — big move coming soon!' if _bb_width < 2 else 'Normal band width'}"
        f"</div>",
        unsafe_allow_html=True
    )

    # CPR section
    # ── Live Options Greeks ───────────────────────────────────
    if sig:
        st.markdown("### 🔢 Options Greeks")
        st.caption(
            "Greeks show how your option price will behave. "
            "Kite live if connected, else Black-Scholes estimate."
        )
        _gk1, _gk2 = st.columns([1,2])
        with _gk1:
            _g_strike = st.number_input(
                "Option strike",
                value=float(round(sig["cp"]/50)*50),
                step=50.0, key="g_strike"
            )
            _g_type = st.selectbox(
                "Option type",
                ["CE","PE"],
                index=0 if sig["direction"]=="UPTREND" else 1,
                key="g_type"
            )
        with _gk2:
            if st.button(
                "Calculate Greeks",
                key="calc_greeks",
                type="primary",
                use_container_width=True
            ):
                with st.spinner("Fetching Greeks..."):
                    _g_sym = (
                        "BANKNIFTY" if "BANK" in sname.upper()
                        else "NIFTY"
                    )
                    _greeks = get_live_greeks(
                        _g_sym, _g_strike, _g_type
                    )
                if _greeks["ok"]:
                    _gc1,_gc2,_gc3,_gc4 = st.columns(4)
                    _gc1.metric("Delta", f"{_greeks['delta']:.3f}",
                        help=f"₹1 move = ₹{abs(_greeks['delta']):.2f} option change")
                    _gc2.metric("Theta", f"₹{_greeks['theta']:.2f}/day",
                        delta="decay", delta_color="inverse")
                    _gc3.metric("Gamma", f"{_greeks['gamma']:.5f}")
                    _gc4.metric("Vega",  f"{_greeks['vega']:.3f}")
                    st.caption(f"Source: {_greeks['source']}")

                    _smove = abs(
                        sig["tgt1"] - sig["cp"]
                        if sig["direction"]=="UPTREND"
                        else sig["cp"] - sig["tgt1s"]
                    )
                    _ogain = round(abs(_greeks["delta"]) * _smove, 2)
                    _ddays = round(abs(_ogain / (_greeks["theta"] - 0.001))) if _greeks["theta"] != 0 else 99
                    st.info(
                        f"Stock moves ₹{_smove:.0f} to T1 → "
                        f"option gains ~₹{_ogain:.0f} | "
                        f"Theta neutralizes gain in ~{_ddays} days if flat"
                    )
                else:
                    st.warning("Greeks unavailable.")


    # ── Supertrend Display ────────────────────────────────
    if sig:
        st_col = "#16a34a" if sig["st_bull"] else "#dc2626"
        st_bg  = "#f0fdf4" if sig["st_bull"] else "#fef2f2"
        st_crossed_badge = (
            "<span style='background:#f59e0b;color:#fff;"
            "padding:2px 8px;border-radius:10px;"
            "font-size:11px;margin-left:8px'>FRESH CROSSOVER</span>"
            if sig["st_crossed"] else ""
        )
        st.markdown(
            f"<div style='background:{st_bg};border:1.5px solid "
            f"{st_col};border-radius:10px;padding:12px 18px;"
            f"margin-bottom:10px;display:flex;"
            f"justify-content:space-between;align-items:center'>"
            f"<div>"
            f"<span style='font-size:13px;font-weight:700;"
            f"color:{st_col}'>⚡ SUPERTREND — {sig['st_signal']}"
            f"</span>{st_crossed_badge}</div>"
            f"<div style='font-size:14px;font-weight:700;"
            f"color:{st_col}'>₹{sig['st_value']:,.2f}</div>"
            f"<div style='font-size:12px;color:#64748b'>"
            f"{'Price above Supertrend — Bullish' if sig['st_bull'] else 'Price below Supertrend — Bearish'}"
            f"</div></div>",
            unsafe_allow_html=True
        )

    # ── Fibonacci Levels ───────────────────────────────────
    if sig:
        st.markdown("#### 📐 Fibonacci Retracement Levels")
        st.caption(
            f"Swing High ₹{sig['fib_high']:,.2f} → "
            f"Swing Low ₹{sig['fib_low']:,.2f} "
            f"| Nearest: {sig['fib_nearest_name']} "
            f"₹{sig['fib_nearest_val']:,.2f} "
            f"({sig['fib_distance_pct']:.2f}% away)"
            + (" ← Price near Fibonacci!" if sig["near_fib"] else "")
        )

        fc1,fc2,fc3,fc4,fc5 = st.columns(5)
        fib_display = [
            (fc1, "23.6%", sig["fib_236"], "#94a3b8", "Weak support"),
            (fc2, "38.2%", sig["fib_382"], "#f59e0b", "Key support"),
            (fc3, "50.0%", sig["fib_500"], "#f59e0b", "Mid point"),
            (fc4, "61.8%", sig["fib_618"], "#ef4444", "Golden ratio"),
            (fc5, "78.6%", sig["fib_786"], "#ef4444", "Deep support"),
        ]
        cur = sig["cp"]
        for col, label, val, col_hex, desc in fib_display:
            diff = round(((cur - val) / val) * 100, 2)
            is_near = abs(diff) < 0.5
            bg = "#fffbeb" if is_near else "#f8fafc"
            border = f"2px solid {col_hex}" if is_near else "1px solid #e2e8f0"
            col.markdown(
                f"<div style='background:{bg};border:{border};"
                f"border-radius:8px;padding:10px;text-align:center'>"
                f"<div style='font-size:11px;color:#64748b'>{label}</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:{col_hex}'>₹{val:,.0f}</div>"
                f"<div style='font-size:10px;color:#94a3b8'>{desc}</div>"
                f"<div style='font-size:10px;color:"
                f"{'#16a34a' if diff>=0 else '#dc2626'}'>"
                f"{diff:+.1f}%</div>"
                + ('<div style="font-size:9px;color:#f59e0b">◀ NEAR</div>' if is_near else '')
                + "</div>",
                unsafe_allow_html=True
            )

    st.markdown("#### 🔄 CPR — Central Pivot Range")
    cpr_col = (
        "#16a34a" if sig["cpr_position"] == "ABOVE"
        else "#dc2626" if sig["cpr_position"] == "BELOW"
        else "#f59e0b"
    )
    cpr_bg = (
        "#f0fdf4" if sig["cpr_position"] == "ABOVE"
        else "#fef2f2" if sig["cpr_position"] == "BELOW"
        else "#fffbeb"
    )
    cp1,cp2,cp3,cp4,cp5 = st.columns(5)
    cp1.metric("CPR Pivot",  f"₹{sig['cpr_pivot']:,}")
    cp2.metric("CPR Top (TC)",f"₹{sig['cpr_tc']:,}")
    cp3.metric("CPR Bot (BC)",f"₹{sig['cpr_bc']:,}")
    cp4.metric("CPR Width",  f"{sig['cpr_width_pct']:.2f}%")
    cp5.metric("Bias",       sig["cpr_bias"])

    st.markdown(
        f"<div style='background:{cpr_bg};"
        f"border:1.5px solid {cpr_col};"
        f"border-radius:10px;padding:12px 18px;"
        f"margin:8px 0;display:flex;gap:20px;"
        f"flex-wrap:wrap;align-items:center'>"
        f"<span style='font-size:16px;font-weight:700;"
        f"color:{cpr_col}'>Price is {sig['cpr_position']} CPR</span>"
        f"<span style='font-size:13px;color:#475569'>"
        f"{sig['cpr_type']}</span>"
        + ("<span style='background:#7c3aed;color:white;"
           "padding:3px 10px;border-radius:12px;"
           "font-size:12px'>✨ Virgin CPR</span>"
           if sig['virgin_cpr'] else "")
        + "</div>",
        unsafe_allow_html=True
    )

    with st.expander("📖 How to use CPR"):
        st.markdown("""
        **CPR (Central Pivot Range)** is calculated from
        yesterday's High, Low and Close.

        | Position | Meaning | Trade |
        |----------|---------|-------|
        | Price **above** TC | Strong bullish | Buy CE |
        | Price **inside** CPR | Sideways — avoid | Wait |
        | Price **below** BC | Strong bearish | Buy PE |

        **CPR Width tells you what kind of day to expect:**
        - **Narrow CPR (< 0.25%)** — Strong trending day.
          Big moves expected. Best for directional trades.
        - **Wide CPR (> 0.5%)** — Choppy sideways day.
          Avoid options. Wait for breakout.

        **Virgin CPR** — Price never touched yesterday's CPR.
        This acts as a strong magnet — price is likely to
        come back and test it today. High probability level.
        """)

    with st.expander("📋 Last 20 candles"):
        t20 = df.tail(20)[["Open","High","Low",
                            "Close","Volume"]].copy()
        t20["Chg%"] = (t20["Close"].pct_change()*100).round(2)
        t20.index   = t20.index.strftime("%d-%b %H:%M")
        st.dataframe(t20.round(2), width="stretch")

    # ══════════════════════════════════════════════════════
    # DURATION-BASED SIGNAL ENGINE
    # ══════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ⏱️ Duration Signal — Hold or Exit?")
    st.caption(
        "Enter your trade details below. "
        "The terminal will tell you every day whether "
        "to hold your position or exit early."
    )

    # ── Trade details input ────────────────────────────────
    with st.expander(
        "Enter your trade details",
        expanded="dur_expiry" not in st.session_state
    ):
        di1, di2, di3 = st.columns(3)
        with di1:
            dur_type = st.selectbox(
                "Option type",
                ["CE (Call — Bullish)",
                 "PE (Put — Bearish)"],
                key="dur_type"
            )
            dur_duration = st.selectbox(
                "Holding duration",
                ["Intraday (exit by 2:45 PM)",
                 "Weekly (next Thursday expiry)",
                 "Monthly (last Thursday expiry)",
                 "Custom expiry date"],
                key="dur_duration"
            )
        with di2:
            dur_entry = st.number_input(
                "Your entry price (Rs)",
                value=0.0, step=0.5,
                min_value=0.0,
                key="dur_entry"
            )
            dur_sl = st.number_input(
                "Your stop loss (Rs)",
                value=sig["sl_long"]
                if "CE" in dur_type else sig["sl_short"]
                if sig else 0.0,
                step=0.5,
                key="dur_sl"
            )
        with di3:
            dur_target = st.number_input(
                "Your target price (Rs)",
                value=sig["tgt1"]
                if "CE" in dur_type else sig["tgt1s"]
                if sig else 0.0,
                step=0.5,
                key="dur_target"
            )
            if "Custom" in dur_duration:
                dur_expiry = st.date_input(
                    "Custom expiry date",
                    key="dur_custom_expiry"
                )
            else:
                # Auto-calculate next Thursday
                today   = now_ist().date()
                days_to_thu = (3 - today.weekday()) % 7
                if days_to_thu == 0:
                    days_to_thu = 7
                if "Monthly" in dur_duration:
                    # Last Thursday of current month
                    import calendar
                    last_day = calendar.monthrange(
                        today.year, today.month
                    )[1]
                    from datetime import date as date_type
                    last_thu = max(
                        date_type(today.year, today.month, d)
                        for d in range(1, last_day+1)
                        if date_type(
                            today.year, today.month, d
                        ).weekday() == 3
                    )
                    dur_expiry = last_thu
                elif "Intraday" in dur_duration:
                    dur_expiry = today
                else:
                    from datetime import date as date_type
                    dur_expiry = (
                        today +
                        timedelta(days=days_to_thu)
                    )
                st.date_input(
                    "Expiry date",
                    value=dur_expiry,
                    disabled=True,
                    key="dur_expiry_display"
                )

        if st.button(
            "Generate Hold/Exit Analysis",
            type="primary",
            key="dur_analyse",
            use_container_width=True
        ):
            st.session_state["_dur_expiry"]  = str(dur_expiry)
            st.session_state["_dur_entry"]   = dur_entry
            st.session_state["_dur_sl"]      = dur_sl
            st.session_state["_dur_target"]  = dur_target
            st.session_state["_dur_type"]    = dur_type
            st.session_state["_dur_active"]  = True

    # ── Show analysis if active ────────────────────────────
    if st.session_state.get("_dur_active") and sig:
        from datetime import date as date_type, datetime as dt_type
        import math

        # Load saved values
        try:
            expiry_date = date_type.fromisoformat(
                st.session_state.get("dur_expiry", "")
            )
        except:
            expiry_date = now_ist().date()

        s_entry  = st.session_state.get("_dur_entry",  0)
        s_sl     = st.session_state.get("_dur_sl",     0)
        s_target = st.session_state.get("_dur_target", 0)
        s_type   = st.session_state.get("_dur_type",   "CE")
        today    = now_ist().date()

        days_left  = (expiry_date - today).days
        days_total = max(
            (expiry_date - (
                expiry_date - timedelta(days=7)
            )).days, 1
        )
        is_ce = "CE" in s_type

        # ── Current price & P&L ───────────────────────────
        cur_price = lp["p"] if lp["ok"] else sig["cp"]
        pnl_pts   = (
            cur_price - s_entry if is_ce
            else s_entry - cur_price
        )
        pnl_pct   = round(
            pnl_pts / (s_entry + 0.001) * 100, 2
        ) if s_entry > 0 else 0

        # ── Hold/Exit decision engine ──────────────────────
        hold_signals   = []
        exit_signals   = []
        warning_signals= []

        # 1. Price vs EMA21
        if is_ce:
            if cur_price > sig["e21v"]:
                hold_signals.append(
                    f"✅ Price ₹{cur_price:,} is above EMA21 "
                    f"₹{sig['e21v']:,} — trend intact"
                )
            else:
                exit_signals.append(
                    f"❌ Price ₹{cur_price:,} broke below "
                    f"EMA21 ₹{sig['e21v']:,} — exit now"
                )
        else:
            if cur_price < sig["e21v"]:
                hold_signals.append(
                    f"✅ Price ₹{cur_price:,} is below EMA21 "
                    f"₹{sig['e21v']:,} — trend intact"
                )
            else:
                exit_signals.append(
                    f"❌ Price ₹{cur_price:,} broke above "
                    f"EMA21 ₹{sig['e21v']:,} — exit now"
                )

        # 2. Stop loss check
        if s_sl > 0:
            if is_ce and cur_price <= s_sl:
                exit_signals.append(
                    f"🚨 Stop loss hit — price ₹{cur_price:,} "
                    f"at or below SL ₹{s_sl:,} — EXIT IMMEDIATELY"
                )
            elif not is_ce and cur_price >= s_sl:
                exit_signals.append(
                    f"🚨 Stop loss hit — price ₹{cur_price:,} "
                    f"at or above SL ₹{s_sl:,} — EXIT IMMEDIATELY"
                )
            else:
                dist = abs(cur_price - s_sl)
                dist_pct = round(dist / cur_price * 100, 2)
                hold_signals.append(
                    f"✅ Stop loss safe — "
                    f"₹{dist:,.0f} ({dist_pct}%) away"
                )

        # 3. Target check
        if s_target > 0:
            if is_ce and cur_price >= s_target:
                warning_signals.append(
                    f"🎯 Target ₹{s_target:,} reached — "
                    f"book at least 50% quantity now"
                )
            elif not is_ce and cur_price <= s_target:
                warning_signals.append(
                    f"🎯 Target ₹{s_target:,} reached — "
                    f"book at least 50% quantity now"
                )

        # 4. CPR bias
        if sig["cpr_bias"] == "Bullish" and is_ce:
            hold_signals.append(
                "✅ CPR bias Bullish — supports CE position"
            )
        elif sig["cpr_bias"] == "Bearish" and not is_ce:
            hold_signals.append(
                "✅ CPR bias Bearish — supports PE position"
            )
        elif sig["cpr_position"] == "INSIDE":
            warning_signals.append(
                "⚠️ Price inside CPR — choppy zone, "
                "reduce position size"
            )
        else:
            exit_signals.append(
                f"❌ CPR bias {sig['cpr_bias']} "
                f"is against your position"
            )

        # 5. RSI check
        rsi_v = sig["rv"]
        if is_ce:
            if rsi_v > 75:
                warning_signals.append(
                    f"⚠️ RSI {rsi_v:.0f} is overbought — "
                    "consider booking partial profit"
                )
            elif 50 <= rsi_v <= 75:
                hold_signals.append(
                    f"✅ RSI {rsi_v:.0f} in bullish zone — "
                    "momentum intact"
                )
            else:
                warning_signals.append(
                    f"⚠️ RSI {rsi_v:.0f} weakening — "
                    "watch closely"
                )
        else:
            if rsi_v < 25:
                warning_signals.append(
                    f"⚠️ RSI {rsi_v:.0f} is oversold — "
                    "consider booking partial profit"
                )
            elif 25 <= rsi_v <= 50:
                hold_signals.append(
                    f"✅ RSI {rsi_v:.0f} in bearish zone — "
                    "momentum intact"
                )
            else:
                warning_signals.append(
                    f"⚠️ RSI {rsi_v:.0f} weakening — "
                    "watch closely"
                )

        # 6. Time decay warning
        if days_left <= 1:
            exit_signals.append(
                "🚨 Expiry tomorrow — exit today by 2:45 PM "
                "unless deeply profitable"
            )
        elif days_left <= 2:
            warning_signals.append(
                f"⚠️ Only {days_left} days to expiry — "
                "theta decay is very high now, "
                "consider exiting if not in profit"
            )
        elif days_left <= 4:
            warning_signals.append(
                f"⚠️ {days_left} days to expiry — "
                "time decay accelerating"
            )
        else:
            hold_signals.append(
                f"✅ {days_left} days to expiry — "
                "enough time for the move to develop"
            )

        # 7. Volume check
        if sig["vsurge"]:
            hold_signals.append(
                "✅ Volume surge confirms price move"
            )

        # 8. Signal score
        tech_score = (
            sig["up_score"] if is_ce else sig["dn_score"]
        )
        if tech_score >= 7:
            hold_signals.append(
                f"✅ Technical score {tech_score}/10 — "
                "strong signal, hold"
            )
        elif tech_score >= 5:
            warning_signals.append(
                f"⚠️ Technical score {tech_score}/10 — "
                "weakening, watch"
            )
        else:
            exit_signals.append(
                f"❌ Technical score {tech_score}/10 — "
                "signal too weak, consider exiting"
            )

        # ── Final verdict ──────────────────────────────────
        if exit_signals:
            verdict      = "EXIT NOW"
            verdict_col  = "#dc2626"
            verdict_bg   = "#fef2f2"
            verdict_icon = "🔴"
        elif len(warning_signals) >= 2:
            verdict      = "REDUCE POSITION"
            verdict_col  = "#d97706"
            verdict_bg   = "#fffbeb"
            verdict_icon = "🟡"
        elif len(hold_signals) >= 3:
            verdict      = "HOLD"
            verdict_col  = "#16a34a"
            verdict_bg   = "#f0fdf4"
            verdict_icon = "🟢"
        else:
            verdict      = "MONITOR CLOSELY"
            verdict_col  = "#d97706"
            verdict_bg   = "#fffbeb"
            verdict_icon = "🟡"

        # ── Display verdict ────────────────────────────────
        st.markdown(
            f"<div style='background:{verdict_bg};"
            f"border:2px solid {verdict_col};"
            f"border-radius:14px;padding:20px 24px;"
            f"margin:12px 0;text-align:center'>"
            f"<div style='font-size:13px;color:#64748b;"
            f"letter-spacing:2px;text-transform:uppercase'>"
            f"Today's Recommendation — {today.strftime('%d %b %Y')}"
            f"</div>"
            f"<div style='font-size:48px;font-weight:700;"
            f"color:{verdict_col};margin:8px 0;line-height:1'>"
            f"{verdict_icon} {verdict}</div>"
            f"<div style='font-size:14px;color:#475569'>"
            f"{sname} {s_type} | "
            f"Entry ₹{s_entry:,} | "
            f"Expiry {expiry_date.strftime('%d %b %Y')} "
            f"({days_left} days left)"
            f"</div></div>",
            unsafe_allow_html=True
        )

        # ── P&L display ────────────────────────────────────
        pl1, pl2, pl3, pl4 = st.columns(4)
        pnl_col_m = "normal" if pnl_pts >= 0 else "inverse"
        pl1.metric(
            "Current Price",
            f"₹{cur_price:,.2f}",
            delta=f"{pnl_pts:+.2f} pts",
            delta_color=pnl_col_m
        )
        pl2.metric("Entry", f"₹{s_entry:,}")
        pl3.metric("Stop Loss", f"₹{s_sl:,}")
        pl4.metric("Target", f"₹{s_target:,}")

        # P&L %
        pnl_c = "#16a34a" if pnl_pct >= 0 else "#dc2626"
        st.markdown(
            f"<div style='background:#f8fafc;"
            f"border-radius:8px;padding:10px 16px;"
            f"font-size:14px;margin:8px 0'>"
            f"Unrealised P&L: "
            f"<b style='color:{pnl_c}'>{pnl_pct:+.2f}%"
            f" ({pnl_pts:+.2f} pts)</b>"
            f"</div>",
            unsafe_allow_html=True
        )

        # ── Signal breakdown ───────────────────────────────
        st.markdown("#### Signal breakdown")

        col_hold, col_warn, col_exit = st.columns(3)
        with col_hold:
            st.markdown(
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#16a34a;margin-bottom:8px'>"
                f"✅ Hold signals ({len(hold_signals)})</div>",
                unsafe_allow_html=True
            )
            for hs in hold_signals:
                st.markdown(
                    f"<div style='background:#f0fdf4;"
                    f"border-left:3px solid #86efac;"
                    f"border-radius:0 6px 6px 0;"
                    f"padding:8px 12px;margin:4px 0;"
                    f"font-size:12px;color:#166534'>"
                    f"{hs}</div>",
                    unsafe_allow_html=True
                )

        with col_warn:
            st.markdown(
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#d97706;margin-bottom:8px'>"
                f"⚠️ Warnings ({len(warning_signals)})</div>",
                unsafe_allow_html=True
            )
            for ws in warning_signals:
                st.markdown(
                    f"<div style='background:#fffbeb;"
                    f"border-left:3px solid #fcd34d;"
                    f"border-radius:0 6px 6px 0;"
                    f"padding:8px 12px;margin:4px 0;"
                    f"font-size:12px;color:#92400e'>"
                    f"{ws}</div>",
                    unsafe_allow_html=True
                )

        with col_exit:
            st.markdown(
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#dc2626;margin-bottom:8px'>"
                f"❌ Exit signals ({len(exit_signals)})</div>",
                unsafe_allow_html=True
            )
            for es in exit_signals:
                st.markdown(
                    f"<div style='background:#fef2f2;"
                    f"border-left:3px solid #fca5a5;"
                    f"border-radius:0 6px 6px 0;"
                    f"padding:8px 12px;margin:4px 0;"
                    f"font-size:12px;color:#991b1b'>"
                    f"{es}</div>",
                    unsafe_allow_html=True
                )

        # ── Daily check reminder ───────────────────────────
        st.markdown("---")
        st.markdown("#### 📅 Daily holding checklist")
        _dir_word  = "above" if is_ce else "below"
        _bias_word = "Bullish" if is_ce else "Bearish"
        st.info(
            f"Check these every morning at 9:30 AM "
            f"while holding {sname} {s_type}:\n\n"
            f"**1.** Is price still {_dir_word} "
            f"EMA21 (₹{sig['e21v']:,})?\n"
            f"**2.** Is CPR bias still {_bias_word}?\n"
            f"**3.** Is RSI still in the right zone?\n"
            f"**4.** Did any bad news come overnight?\n"
            f"**5.** Days to expiry: **{days_left}** — "
            f"exit by 2:45 PM on expiry day no matter what."
        )

        if st.button(
            "Clear — start new trade",
            key="dur_clear",
            type="secondary"
        ):
            for k in ["dur_active","dur_expiry","dur_entry",
                      "dur_sl","dur_target","dur_type"]:
                st.session_state.pop(k, None)
            st.rerun()

# ╔══════════════════════════════════════════════════════╗
# ║  TAB 3 — SMART MONEY                                ║
# ╚══════════════════════════════════════════════════════╝

with T3:
    st.markdown("### 🔍 Auto Scanner — 30 Stocks Live")

    # ── Quick stock search inside scanner ────────────────
    with st.expander("🔍 Search a specific stock", expanded=False):
        sc_search = st.text_input(
            "Type stock name",
            placeholder="e.g. Adani Ports, TCS, HDFC...",
            key="scanner_search_box"
        )
        if sc_search:
            hits = {k:v for k,v in STOCKS.items()
                    if sc_search.strip().lower() in k.lower()}
            if hits:
                picked = st.selectbox(
                    "Select stock",
                    list(hits.keys()),
                    key="scanner_search_pick",
                    label_visibility="collapsed"
                )
                sc1, sc2 = st.columns(2)
                with sc1:
                    if st.button(
                        f"📊 Analyse {picked}",
                        key="scanner_search_analyse",
                        type="primary",
                        use_container_width=True
                    ):
                        st.session_state["sn"] = picked
                        st.session_state["st"] = hits[picked]
                        st.rerun()
                with sc2:
                    if tg_configured() and st.button(
                        "📱 Quick Signal",
                        key="scanner_search_tg",
                        use_container_width=True
                    ):
                        sym_q = hits[picked]
                        lp_q  = live_price(sym_q)
                        if lp_q["ok"]:
                            msg_q = (
                                f"<b>Quick Signal: {picked}</b>\n"
                                f"Price: Rs {lp_q['p']:,.2f}\n"
                                f"Change: {lp_q['chg']:+.2f}%\n"
                                f"Check Trade Setup tab for full analysis."
                            )
                            if send_telegram(
                                st.session_state["tg_token_saved"],
                                st.session_state["tg_chat_saved"],
                                msg_q
                            ):
                                st.success("Sent ✅")
                            else:
                                st.error("Failed ❌")
            elif sc_search:
                st.caption("No results found")

    _kite_scan = kite_is_connected()
    _cur_tok   = st.session_state.get("kite_access_token","")
    _inst_map  = st.session_state.get("kite_inst_map",{})
    _inst_err  = st.session_state.get("kite_inst_error","")

    if _kite_scan:
        if _inst_map:
            st.success(
                f"⚡ Kite LIVE — Real-time candles active. "
                f"{len(_inst_map):,} instruments loaded."
            )
        else:
            st.warning(
                "⚡ Kite connected but instruments not loaded yet. "
                "Click the button below to load instruments first."
                + (f" Error: {_inst_err}" if _inst_err else "")
            )
            if st.button(
                "Load Kite Instruments (required for live data)",
                key="load_instruments",
                type="primary"
            ):
                with st.spinner(
                    "Loading NSE instruments from Kite (~20 seconds)..."
                ):
                    result = get_kite_instruments(_cur_tok)
                if result:
                    st.success(
                        f"✅ {len(result):,} instruments loaded! "
                        "Scanner will now use live data."
                    )
                    st.rerun()
                else:
                    st.error(
                        "Failed to load instruments. "
                        "Check Kite connection and try again."
                    )
    else:
        st.warning(
            "📊 Yahoo Finance data (15-min delay) — "
            "Login with Zerodha Kite in sidebar for live signals"
        )
    st.caption(
        "Scans stocks simultaneously across sectors. "
        "Click Scan, review results, then manually send signals to Telegram."
    )

    # ── Scanner stock universe — uses same SECTORS as watchlist ──
    SCANNER_UNIVERSE = {
        k: v for k, v in SECTORS.items()
    }
    # Add a "All Sectors" option
    SCANNER_UNIVERSE = {
        "🌐 All Sectors (top 5 each)": [
            stock
            for sector_stocks in list(SECTORS.values())
            for stock in sector_stocks[:5]
        ],
        **SECTORS
    }

    st.markdown("---")

    # ── Telegram setup ────────────────────────────────────
    # Show configured badge if already set
    if tg_configured():
        st.success(
            "📱 Telegram configured ✅ — "
            "Use the Send buttons on each signal card to "
            "send alerts manually. Expand below to change settings."
        )
    st.markdown("#### 📱 Telegram Alert Setup")
    with st.expander(
        "Configure Telegram alerts — click to setup",
        expanded=not tg_configured()
    ):
        st.markdown("""
        **Your own private Telegram bot — 100% free, no third party.**

        **One-time setup (3 minutes):**

        **Step 1 — Create your bot:**
        - Open Telegram on your phone
        - Search for **@BotFather**
        - Send: `/newbot`
        - Give it a name e.g. `My Trading Bot`
        - Give it a username e.g. `ranjith_trading_bot`
        - BotFather sends you a **Token** — copy it
          (looks like: `7123456789:AAHxxxxxxxxxxxxxxx`)

        **Step 2 — Get your Chat ID:**
        - Search for **@userinfobot** in Telegram
        - Send any message to it
        - It replies with your **Chat ID** (a number like `987654321`)

        **Step 3 — Enter both below and test:**
        """)

        # ── Input fields (pre-filled if saved) ────────
        saved_token = st.session_state.get(
            "tg_token_saved", ""
        )
        saved_chat  = st.session_state.get(
            "tg_chat_saved", ""
        )

        if saved_token and saved_chat:
            st.success(
                "✅ Telegram credentials loaded from saved file. "
                "You don't need to re-enter them."
            )

        tg_token = st.text_input(
            "Step 1 — Bot Token",
            value=saved_token,
            placeholder="7123456789:AAHxxxxxxxxxxx",
            type="password",
            key="tg_token",
            help="Get this from @BotFather on Telegram"
        )
        tg_chat = st.text_input(
            "Step 2 — Chat ID",
            value=saved_chat,
            placeholder="987654321",
            key="tg_chat",
            help="Get this from @userinfobot on Telegram"
        )

        # ── Save credentials whenever typed ───────────
        if tg_token:
            st.session_state["tg_token_saved"] = tg_token
        if tg_chat:
            st.session_state["tg_chat_saved"] = tg_chat
        # Auto-save both if both are filled
        if tg_token and tg_chat:
            save_creds(tg_token, tg_chat)

        st.markdown("---")

        # ── Test button — always visible ───────────────
        st.markdown(
            "<div style='font-size:13px;color:#374151;"
            "margin-bottom:8px'>"
            "Step 3 — Click the button below to verify "
            "your connection:</div>",
            unsafe_allow_html=True
        )

        if st.button(
            "📲 Send Test Message to Telegram",
            key="tg_test",
            type="primary",
            use_container_width=True
        ):
            tok = (tg_token or
                   st.session_state.get("tg_token_saved",""))
            cid = (tg_chat or
                   st.session_state.get("tg_chat_saved",""))

            if not tok or not cid:
                st.error(
                    "❌ Please enter both Bot Token "
                    "and Chat ID above first."
                )
            else:
                with st.spinner("Sending test message..."):
                    ok = send_telegram(
                        tok, cid,
                        "🎯 <b>Trading Terminal Connected!</b>\n"
                        "You will now receive trade signals here.\n"
                        "Use Send buttons in the scanner to send signals manually."
                    )
                if ok:
                    st.session_state["tg_token_saved"] = tok
                    st.session_state["tg_chat_saved"]  = cid
                    # Save permanently to file
                    saved = save_creds(tok, cid)
                    st.success(
                        "✅ Test message sent! "
                        "Check your Telegram. "
                        "Credentials saved permanently — "
                        "you won't need to enter them again."
                        + (" 💾" if saved else "")
                    )
                    st.balloons()
                else:
                    try:
                        resp = requests.post(
                            f"https://api.telegram.org/bot{tok}"
                            f"/sendMessage",
                            json={"chat_id": cid,
                                  "text": "test"},
                            timeout=10
                        )
                        err = resp.json().get(
                            "description", resp.text
                        )
                    except Exception as ex:
                        err = str(ex)
                    st.error(
                        f"❌ Failed: {err}  \n"
                        "Fix: Open Telegram → find your bot "
                        "→ press START first."
                    )

        # Show current status
        if tg_configured():
            st.success(
                "✅ Telegram is configured and active. "
                "You will receive alerts when scanner "
                "finds score 8+ signals."
            )

    st.markdown("---")

    # ── Scanner controls ───────────────────────────────────
    sc1, sc2, sc3 = st.columns([2, 1, 1])
    with sc1:
        scan_group = st.selectbox(
            "Stock group to scan",
            list(SCANNER_UNIVERSE.keys()),
            key="scan_group"
        )
    with sc2:
        # Default to same timeframe as Trade Setup sidebar
        _global_tf  = st.session_state.get("global_tf","15m")
        _tf_options = ["15m","30m","1h","1d"]
        _tf_default = _tf_options.index(_global_tf) if _global_tf in _tf_options else 0
        scan_tf = st.selectbox(
            "Timeframe",
            _tf_options,
            index=_tf_default,
            key="scan_tf",
            help="Matches Trade Setup timeframe automatically"
        )
    with sc3:
        min_rr_scan = st.selectbox(
            "Min R:R",
            ["1.0","1.5","2.0"],
            index=0,
            key="min_rr_scan",
            help="Minimum Risk:Reward ratio"
        )

    # Show timeframe sync status
    _gtf = st.session_state.get("global_tf","15m")
    if scan_tf == _gtf:
        st.success(
            f"✅ Scanner timeframe ({scan_tf}) matches "
            f"Trade Setup timeframe ({_gtf}) — "
            f"scores will be consistent."
        )
    else:
        st.warning(
            f"⚠️ Scanner uses {scan_tf} but Trade Setup uses {_gtf}. "
            f"Scores may differ. Change sidebar timeframe to {scan_tf} "
            f"to match."
        )

    # ── Time Session Filter ───────────────────────────────
    from datetime import datetime as _dtnow
    import pytz as _pytz
    _ist = _pytz.timezone("Asia/Kolkata")
    _now_ist = _dtnow.now(_ist)
    _hour = _now_ist.hour
    _min  = _now_ist.minute
    _time_dec = _hour + _min / 60.0  # decimal time

    if 9.25 <= _time_dec < 10.0:
        st.error(
            "⏰ **OPENING SESSION WARNING (9:15 - 10:00 AM)** — "
            "Market is highly volatile right now. "
            "Signals are unreliable during first 45 minutes. "
            "Wait until 10:00 AM before acting on any signal."
        )
        _session_ok = False
        _session_label = "🔴 OPENING — Wait until 10:00 AM"
    elif 10.0 <= _time_dec < 13.0:
        st.success(
            "✅ **MID SESSION (10:00 AM - 1:00 PM)** — "
            "Best time to trade. Signals are most reliable now."
        )
        _session_ok = True
        _session_label = "🟢 MID SESSION — Best time to trade"
    elif 13.0 <= _time_dec < 14.5:
        st.info(
            "ℹ️ **AFTERNOON SESSION (1:00 - 2:30 PM)** — "
            "Good for continuation trades. "
            "Avoid new entries after 2:00 PM."
        )
        _session_ok = True
        _session_label = "🟡 AFTERNOON — Continuation trades only"
    elif 14.5 <= _time_dec < 15.3:
        st.warning(
            "⚠️ **PRE-CLOSE (2:30 - 3:15 PM)** — "
            "Exit existing positions only. "
            "Do NOT enter new trades."
        )
        _session_ok = False
        _session_label = "🔴 PRE-CLOSE — Exit only, no new entries"
    elif _time_dec >= 15.3 or _time_dec < 9.25:
        st.info(
            "🕐 **MARKET CLOSED** — "
            "You can scan to prepare for tomorrow. "
            "Do not place any orders."
        )
        _session_ok = True  # Allow scanning for prep
        _session_label = "⚫ MARKET CLOSED — Prep for tomorrow"
    else:
        _session_ok = True
        _session_label = "🟢 Market open"

    # IV Rank quick check before scanning
    _iv_quick = get_iv_rank()
    if _iv_quick["ok"]:
        _iv_col = _iv_quick["color"]
        _iv_bg  = _iv_quick["bg"]
        st.markdown(
            f"<div style='background:{_iv_bg};"
            f"border:1px solid {_iv_col};"
            f"border-radius:8px;padding:8px 16px;"
            f"font-size:13px;margin-bottom:8px'>"
            f"<b style='color:{_iv_col}'>IV Rank: "
            f"{_iv_quick['iv_rank']} — {_iv_quick['signal']}</b>"
            f" &nbsp;|&nbsp; {_iv_quick['advice']}"
            f"</div>",
            unsafe_allow_html=True
        )
    st.caption(
        "Scanner uses the exact same scoring logic as Trade Setup. "
        "If a stock scores 7+ in Trade Setup on 15m — "
        "select its sector and scan on 15m to find it here."
    )

    sl1, sl2 = st.columns(2)
    with sl1:
        min_score_scan = st.slider(
            "Minimum score to show",
            min_value=0,
            max_value=10,
            value=6,
            key="min_score_scan",
            help="Only show stocks with score above this"
        )
    with sl2:
        min_combined_scan = st.slider(
            "Minimum combined score",
            min_value=0,
            max_value=10,
            value=5,
            key="min_combined_scan",
            help=(
                "Combined = 60% technical + 40% historical. "
                "Higher = more consistent signal."
            )
        )

    alert_score = min_score_scan  # used for display only

    # Mobile-friendly controls - stack vertically on small screens
    # Fetch NIFTY correlation signal for filter
    if "nifty_corr_sig" not in st.session_state:
        with st.spinner("Loading NIFTY trend for correlation filter..."):
            try:
                _nf_df = candles("^NSEI", scan_tf)
                _nf_lp = live_price("^NSEI")
                if _nf_df is not None and len(_nf_df) >= 55:
                    _nf_sig = compute_all(_nf_df, _nf_lp)
                    st.session_state["nifty_corr_sig"] = _nf_sig
            except Exception:
                pass

    _nf_s = st.session_state.get("nifty_corr_sig")
    if _nf_s:
        _nfd = _nf_s.get("direction","SIDEWAYS")
        _nfc = "#16a34a" if _nfd=="UPTREND" else "#dc2626" if _nfd=="DOWNTREND" else "#f59e0b"
        st.markdown(
            f"<div style='background:{_nfc}22;border:1px solid {_nfc};"
            f"border-radius:8px;padding:8px 14px;margin-bottom:8px;"
            f"font-size:13px;color:{_nfc}'>"
            f"<b>NIFTY Correlation:</b> {_nfd} "
            f"(Score {max(_nf_s['up_score'],_nf_s['dn_score'])}/10) — "
            f"{'CE signals shown only' if _nfd=='UPTREND' else 'PE signals shown only' if _nfd=='DOWNTREND' else 'Both CE and PE signals shown'}"
            f"</div>",
            unsafe_allow_html=True
        )

    sc_b1, sc_b2 = st.columns([3,1])
    with sc_b1:
        run_scanner = st.button(
            "🚀 Scan All Stocks Now",
            type="primary",
            key="run_scanner",
            use_container_width=True
        )
    with sc_b2:
        if st.button(
            "🗑️ Clear & Rescan",
            key="scan_clear_cache",
            use_container_width=True,
            help="Clears old results and fetches fresh data"
        ):
            # Clear all cached results
            for _k in ["scan_results","scan_group_used",
                       "scan_tf_used","scan_time"]:
                st.session_state.pop(_k, None)
            st.cache_data.clear()
            st.rerun()
    st.markdown(
        "<div style='margin:6px 0;padding:10px 14px;"
        "background:#f0f9ff;border:1px solid #bae6fd;"
        "border-radius:8px;font-size:13px;color:#0369a1'>"
        "🔄 <b>Auto scan every 5 min</b> — "
        "Toggle below to enable automatic scanning</div>",
        unsafe_allow_html=True
    )
    auto_scan = st.checkbox(
        "Enable auto scan every 5 minutes",
        value=False,
        key="auto_scan"
    )

    # send_telegram is now a global function

    # ── Run scanner ────────────────────────────────────────
    def get_option_levels(price: float, direction: str) -> dict:
        """Calculate ATM / ITM / OTM strike levels."""
        if price > 30000: step = 500
        elif price > 10000: step = 100
        elif price > 3000:  step = 50
        elif price > 500:   step = 20
        else:               step = 10

        atm = round(price / step) * step

        if direction == "UPTREND":
            return {
                "opt_type": "CE (Call)",
                "ITM": atm - step,
                "ATM": atm,
                "OTM": atm + step,
                "recommend": atm,
                "step": step,
            }
        else:
            return {
                "opt_type": "PE (Put)",
                "ITM": atm + step,
                "ATM": atm,
                "OTM": atm - step,
                "recommend": atm,
                "step": step,
            }

    def historical_consistency(sdf: pd.DataFrame,
                               direction: str) -> dict:
        """
        Validates signal across last 3, 5, 10 candles.
        Ensures results dont flip every 15 minutes.
        Returns a consistency score 0-10.
        """
        try:
            c   = sdf["Close"].squeeze().astype(float)
            v   = sdf["Volume"].squeeze().astype(float)
            e9  = ta.trend.ema_indicator(c, 9)
            e21 = ta.trend.ema_indicator(c, 21)
            rsi = ta.momentum.rsi(c, 14)
            vol_avg = v.rolling(20).mean()

            # 1. Last 3 candles price vs EMA9 and EMA21
            c3 = 0
            for i in range(1, 4):
                if direction == "UPTREND":
                    if (float(c.iloc[-i]) > float(e9.iloc[-i])
                            and float(e9.iloc[-i]) >
                            float(e21.iloc[-i])):
                        c3 += 1
                else:
                    if (float(c.iloc[-i]) < float(e9.iloc[-i])
                            and float(e9.iloc[-i]) <
                            float(e21.iloc[-i])):
                        c3 += 1

            # 2. Last 5 candles price vs EMA21
            c5 = 0
            for i in range(1, 6):
                if direction == "UPTREND":
                    if float(c.iloc[-i]) > float(e21.iloc[-i]):
                        c5 += 1
                else:
                    if float(c.iloc[-i]) < float(e21.iloc[-i]):
                        c5 += 1

            # 3. RSI in correct zone last 5 candles
            rsi_ok = 0
            for i in range(1, 6):
                rv = float(rsi.iloc[-i])
                if direction == "UPTREND" and 45 < rv < 78:
                    rsi_ok += 1
                elif direction == "DOWNTREND" and 22 < rv < 55:
                    rsi_ok += 1

            # 4. Volume above average last 3 candles
            vol_ok = sum(
                1 for i in range(1, 4)
                if float(v.iloc[-i]) > float(vol_avg.iloc[-i])
            )

            # Weighted consistency score
            score = round(
                (c3/3)*4 + (c5/5)*3 +
                (rsi_ok/5)*2 + (vol_ok/3)*1, 1
            )

            reliability = (
                "🔥 Very High" if score >= 8 else
                "✅ High"      if score >= 6 else
                "📈 Moderate"  if score >= 4 else
                "⚠️ Low — wait"
            )
            return {
                "score":       score,
                "reliability": reliability,
                "candles_3":   c3,
                "candles_5":   c5,
                "rsi_ok":      rsi_ok,
                "vol_ok":      vol_ok,
            }
        except:
            return {"score":0, "reliability":"—",
                    "candles_3":0, "candles_5":0,
                    "rsi_ok":0, "vol_ok":0}

    def run_scan_engine(stocks_to_scan, timeframe,
                        min_sc, alert_sc):
        results = []
        alerted = []
        prog    = st.progress(0, text="Starting scan...")
        total   = len(stocks_to_scan)

        for i, sname in enumerate(stocks_to_scan):
            sym = STOCKS.get(sname)
            if not sym:
                prog.progress(int((i+1)/total*100),
                              text=f"Skipping {sname}...")
                continue

            prog.progress(
                int((i+1)/total*100),
                text=f"Scanning {sname}... ({i+1}/{total})"
            )

            try:
                sdf = candles(sym, timeframe)
                if sdf.empty or len(sdf) < 55:
                    continue

                slp = live_price(sym)
                sig = compute_all(sdf, slp)
                if sig is None:
                    continue

                up_s = sig["up_score"]
                dn_s = sig["dn_score"]
                best = max(up_s, dn_s)

                if best < min_sc:
                    continue

                direction = (
                    "UPTREND"   if up_s > dn_s else
                    "DOWNTREND" if dn_s > up_s else
                    "MIXED"
                )
                if direction == "MIXED":
                    continue

                action = ("BUY CE" if direction == "UPTREND"
                          else "BUY PE")

                # Historical consistency (key fix for
                # results changing every 15 min)
                hist = historical_consistency(sdf, direction)

                # Skip low consistency signals
                if hist["score"] < 3:
                    continue

                # Option levels
                opt = get_option_levels(sig["cp"], direction)

                # SL and targets
                sl_v = (sig["sl_long"]
                        if direction == "UPTREND"
                        else sig["sl_short"])
                t1_v = (sig["tgt1"]
                        if direction == "UPTREND"
                        else sig["tgt1s"])
                t2_v = (sig["tgt2"]
                        if direction == "UPTREND"
                        else sig["tgt2s"])
                t3_v = (sig["tgt3"]
                        if direction == "UPTREND"
                        else sig["tgt3s"])

                # Risk reward — calculated from ENTRY price
                _entry_used = (sig["entry_long"]
                               if direction == "UPTREND"
                               else sig["entry_short"])
                risk   = abs(_entry_used - sl_v)
                reward = abs(t1_v - _entry_used)
                rr     = round(reward / (risk + 0.001), 2)

                # ── Fix 1: Skip if R:R below threshold ─────
                _min_rr = float(st.session_state.get("min_rr_scan","1.0"))
                if rr < _min_rr:
                    continue

                # ── Fix 2: Skip if price too far from entry ──
                _entry    = sig.get("entry_long", sig["e9v"])
                _cp       = sig["cp"]
                _atr      = sig["atrv"]
                _max_dist = _atr * 1.5

                if direction == "UPTREND":
                    if _cp < (_entry - _max_dist):
                        continue
                elif direction == "DOWNTREND":
                    if _cp > (_entry + _max_dist):
                        continue

                # ── Fix 3: NIFTY Correlation Filter ──────────
                # Skip CE signals if NIFTY itself is in downtrend
                # Skip PE signals if NIFTY itself is in uptrend
                _nifty_sig = st.session_state.get("nifty_corr_sig")
                if _nifty_sig:
                    _nifty_dir = _nifty_sig.get("direction","SIDEWAYS")
                    if (direction == "UPTREND" and
                            _nifty_dir == "DOWNTREND"):
                        # CE signal but NIFTY is falling — skip
                        continue
                    elif (direction == "DOWNTREND" and
                            _nifty_dir == "UPTREND"):
                        # PE signal but NIFTY is rising — skip
                        continue

                # CPR is informational only — not used as filter
                # CPR from previous day may not match intraday moves
                cpr_pos  = sig.get("cpr_position","INSIDE")
                cpr_bias = sig.get("cpr_bias","Sideways")

                # ── Fix 3: Confidence scoring ──────────────
                conflicts = 0
                confirms  = 0

                # Check Supertrend agreement
                st_bull = sig.get("st_bull", True)
                if direction == "UPTREND":
                    if st_bull:
                        confirms += 1
                    else:
                        conflicts += 1
                else:
                    if not st_bull:
                        confirms += 1
                    else:
                        conflicts += 1

                # Check RSI agreement
                rsi_v = sig["rv"]
                if direction == "UPTREND":
                    if 50 <= rsi_v <= 75:
                        confirms += 1
                    elif rsi_v < 40 or rsi_v > 80:
                        conflicts += 1
                else:
                    if 25 <= rsi_v <= 50:
                        confirms += 1
                    elif rsi_v > 60 or rsi_v < 20:
                        conflicts += 1

                # Check MACD agreement
                if direction == "UPTREND":
                    if sig.get("macd_bull", False):
                        confirms += 1
                    else:
                        conflicts += 1
                else:
                    if not sig.get("macd_bull", True):
                        confirms += 1
                    else:
                        conflicts += 1

                # Skip if too many conflicts
                if conflicts >= 2:
                    continue

                # Confidence label
                if confirms >= 3 and conflicts == 0:
                    confidence = "HIGH"
                elif confirms >= 2 and conflicts <= 1:
                    confidence = "MEDIUM"
                else:
                    confidence = "LOW"

                # ── Signal Validity Timer ──────────────────
                # Estimate how many candles the signal will stay valid
                # Based on RSI distance from zone edge, volume, momentum

                _rsi_now = sig["rv"]
                _rsi_prev= float(sig["rsis"].iloc[-2]) if hasattr(sig.get("rsis",""), "iloc") else _rsi_now

                # RSI momentum — is it moving toward or away from zone?
                _rsi_slope = _rsi_now - _rsi_prev

                if direction == "UPTREND":
                    # CE signal — RSI should be 55-68
                    # Distance from danger zone (75 = overbought)
                    _rsi_dist = 75 - _rsi_now
                    _rsi_moving_ok = _rsi_slope > 0 and _rsi_now < 68
                else:
                    # PE signal — RSI should be 32-45
                    # Distance from danger zone (25 = oversold)
                    _rsi_dist = _rsi_now - 25
                    _rsi_moving_ok = _rsi_slope < 0 and _rsi_now > 32

                # Volume momentum
                _vol_ratio = sig.get("vol_ratio", 1.0)
                _vol_ok    = _vol_ratio >= 1.2

                # Price position relative to EMA9
                _e9v = sig.get("e9v", sig["cp"])
                _price_dist_pct = abs(sig["cp"] - _e9v) / _e9v * 100

                # Calculate validity candles
                if _rsi_dist > 15 and _vol_ok and _rsi_moving_ok:
                    _valid_candles = 6  # ~90 min on 15m
                    _validity = "🟢 Strong — Valid ~6 candles (90 min)"
                    _validity_col = "#16a34a"
                    _validity_bg  = "#f0fdf4"
                elif _rsi_dist > 8 and (_vol_ok or _rsi_moving_ok):
                    _valid_candles = 4  # ~60 min on 15m
                    _validity = "🟡 Moderate — Valid ~4 candles (60 min)"
                    _validity_col = "#d97706"
                    _validity_bg  = "#fffbeb"
                elif _rsi_dist > 4:
                    _valid_candles = 2  # ~30 min on 15m
                    _validity = "🟠 Short — Valid ~2 candles (30 min). Enter now!"
                    _validity_col = "#ea580c"
                    _validity_bg  = "#fff7ed"
                else:
                    _valid_candles = 1
                    _validity = "🔴 Expiring — Enter immediately or skip!"
                    _validity_col = "#dc2626"
                    _validity_bg  = "#fef2f2"

                # Combined score
                combined = round(
                    best*0.6 + hist["score"]*0.4, 1
                )

                # Price change
                try:
                    chg = round(
                        ((sig["cp"] -
                          float(sdf["Close"].iloc[-2])) /
                         float(sdf["Close"].iloc[-2]))*100, 2
                    )
                except:
                    chg = 0

                result = {
                    "Stock":       sname,
                    "Sym":         sym,
                    "Score":       best,
                    "HistScore":   hist["score"],
                    "Combined":    combined,
                    "Reliability": hist["reliability"],
                    "Direction":   direction,
                    "Action":      action,
                    "Price":       sig["cp"],
                    "Change%":     chg,
                    "RSI":         sig["rv"],
                    "ADX":         sig["adxv"],
                    "VolSurge":    sig["vsurge"],
                    "Entry":       (sig.get("entry_long", sig["e9v"])
                                   if direction == "UPTREND"
                                   else sig.get("entry_short", sig["e9v"])),
                    "SL":          sl_v,
                    "T1":          t1_v,
                    "T2":          t2_v,
                    "T3":          t3_v,
                    "RR":          rr,
                    "OptType":     opt["opt_type"],
                    "ATM":         opt["ATM"],
                    "ITM":         opt["ITM"],
                    "OTM":         opt["OTM"],
                    "Consist3":    hist["candles_3"],
                    "Consist5":    hist["candles_5"],
                    # CPR
                    "CPR_Bias":    sig.get("cpr_bias","—"),
                    "CPR_Pos":     sig.get("cpr_position","—"),
                    "CPR_Type":    sig.get("cpr_type","—"),
                    "CPR_Pivot":   sig.get("cpr_pivot",0),
                    "Virgin_CPR":  sig.get("virgin_cpr",False),
                    # Confidence
                    "Confidence":   confidence,
                    "Confirms":     confirms,
                    "Conflicts":    conflicts,
                    "Validity":     _validity,
                    "ValidCandles": _valid_candles,
                    "ValidityCol":  _validity_col,
                    "ValidityBg":   _validity_bg,
                }
                results.append(result)

                # ── ML + Multi-timeframe confirmation ─────────
                # Use pre-trained cache if available (faster)
                ml_direction  = "UNKNOWN"
                ml_confidence = 0
                ml_agrees     = False
                mtf_1h_dir    = "UNKNOWN"
                mtf_1d_dir    = "UNKNOWN"

                _cached_ml = get_ml_cached(sname)
                if _cached_ml and _cached_ml.get("ok"):
                    ml_direction  = _cached_ml["direction"]
                    ml_confidence = _cached_ml["confidence"]
                    ml_agrees     = (ml_direction == direction)
                else:
                    # Train fresh — use candle cache if available
                    try:
                        _cc = st.session_state.get("candle_cache",{})
                        df_ml = _cc.get(f"{sym}_1d") or candles(sym,"1d")
                        if df_ml is not None and len(df_ml) >= 100:
                            ml_model = train_model(df_ml)
                            if ml_model.get("ok"):
                                ml_pred = predict_next_move(
                                    df_ml, ml_model
                                )
                                if ml_pred and ml_pred.get("ok"):
                                    ml_direction  = ml_pred["prediction"]
                                    ml_confidence = ml_pred["confidence"]
                                    ml_agrees = (ml_direction == direction)
                    except Exception:
                        pass

                try:
                    # Always fetch 1h fresh — morning candle changes
                    # after 9:15 AM open, cached data is stale
                    df_1h = candles(sym, "1h")
                    if df_1h is not None and len(df_1h) >= 55:
                        sig_1h = compute_all(df_1h, slp)
                        if sig_1h:
                            mtf_1h_dir = sig_1h["direction"]
                except Exception:
                    pass

                try:
                    df_1d = _cc.get(f"{sym}_1d") or candles(sym, "1d")
                    if df_1d is not None and len(df_1d) >= 55:
                        sig_1d = compute_all(df_1d, slp)
                        if sig_1d:
                            mtf_1d_dir = sig_1d["direction"]
                except Exception:
                    pass

                mtf_all_ok = (
                    mtf_1h_dir == direction and
                    mtf_1d_dir == direction
                )
                is_diamond = (
                    ml_agrees and mtf_all_ok and
                    confidence in ["HIGH","MEDIUM"] and
                    rr >= 1.5
                )

                result["ML_Direction"]  = ml_direction
                result["ML_Confidence"] = ml_confidence
                result["ML_Agrees"]     = ml_agrees
                result["MTF_1H"]        = mtf_1h_dir
                result["MTF_1D"]        = mtf_1d_dir
                result["MTF_All_OK"]    = mtf_all_ok
                result["Is_Diamond"]    = is_diamond


                # No auto alerts — user sends manually

            except Exception:
                continue

        prog.empty()
        return results, alerted
    # ── Display results ────────────────────────────────────
    # ── Show previous scan results if available ───────
    # This keeps results visible when buttons are clicked
    if "scan_results" in st.session_state and not run_scanner:
        results   = st.session_state["scan_results"]
        scan_time = st.session_state.get("scan_time","")
        grp_used  = st.session_state.get("scan_group_used","")
        tf_used   = st.session_state.get("scan_tf_used","")
        st.warning(
            f"⚠️ Showing CACHED results from {scan_time} — "
            f"{grp_used} | {tf_used} | "
            f"Click **Scan All Stocks Now** or **🗑️ Clear & Rescan** "
            f"to get fresh signals."
        )
        # Re-apply stale signal filter on cached results
        # Prices may have changed since last scan
        filtered_results = []
        for _r in results:
            _lp_chk = live_price(_r["Sym"])
            if _lp_chk["ok"]:
                _cur = _lp_chk["p"]
                _ent = _r["Entry"]
                _atr_chk = abs(_ent - _r["SL"])  # approximate ATR from SL distance
                _max_d   = _atr_chk * 2.0
                if _r["Direction"] == "UPTREND":
                    if _cur < (_ent - _max_d):
                        continue  # Price moved too far below entry
                else:
                    if _cur > (_ent + _max_d):
                        continue  # Price moved too far above entry
                # Update price in cached result
                _r["Price"] = _cur
                _r["Change%"] = _lp_chk["chg"]
            filtered_results.append(_r)
        results = filtered_results

        total_scanned = len(SCANNER_UNIVERSE.get(grp_used,[]))

        # ── Summary metrics ───────────────────────────────
        results_sorted = sorted(
            results, key=lambda x: x["Score"], reverse=True
        )
        strong_r = [r for r in results_sorted if r["Score"] >= 8]
        good_r   = [r for r in results_sorted if 6 <= r["Score"] < 8]

        # ── Export Excel button — shown right after scan ───
        if results_sorted:
            _xb1, _xb2 = st.columns([3,1])
            with _xb1:
                st.success(
                    f"✅ {len(results_sorted)} signals found — "
                    f"click Export to download for paper trading"
                )
            with _xb2:
                _do_export = st.button(
                    "📥 Export Excel",
                    key="scan_export_top",
                    type="primary",
                    use_container_width=True
                )

            if _do_export:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import (
                    PatternFill, Font, Alignment, Border, Side
                )
                from datetime import datetime as _xdt
                from datetime import date as _xdate

                wb  = Workbook()
                ws  = wb.active
                ws.title = "Scanner Signals"

                _hf    = PatternFill("solid", fgColor="1e3a5f")
                _hfont = Font(color="FFFFFF", bold=True, size=11)
                _gf    = PatternFill("solid", fgColor="d1fae5")
                _rf    = PatternFill("solid", fgColor="fee2e2")
                _pf    = PatternFill("solid", fgColor="ede9fe")
                _ca    = Alignment(horizontal="center")
                _bd    = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )

                ws.merge_cells("A1:R1")
                ws["A1"] = (
                    f"Trading Terminal | "
                    f"{_xdt.now().strftime('%d %b %Y %H:%M IST')}"
                )
                ws["A1"].fill = PatternFill("solid", fgColor="1e3a5f")
                ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
                ws["A1"].alignment = _ca

                _hdrs = [
                    "Stock","Signal","Score","Combined","R:R",
                    "Price","Change%","Entry","Stop Loss",
                    "Target 1","Target 2","ATM","ITM","OTM",
                    "RSI","CPR","Confidence","Diamond"
                ]
                for _ci,_h in enumerate(_hdrs,1):
                    _c = ws.cell(row=2, column=_ci, value=_h)
                    _c.fill=_hf; _c.font=_hfont
                    _c.alignment=_ca; _c.border=_bd
                    ws.column_dimensions[_c.column_letter].width = max(12,len(_h)+2)

                _diamond_r = [r for r in results_sorted if r.get("Is_Diamond",False)]
                _all_sigs  = (
                    [dict(r,_d=True) for r in _diamond_r] +
                    [dict(r,_d=False) for r in results_sorted if not r.get("Is_Diamond",False)]
                )

                for _ri,_r in enumerate(_all_sigs,3):
                    _is_d  = _r.get("_d",False)
                    _is_ce = _r["Direction"]=="UPTREND"
                    _rfl   = _pf if _is_d else _gf if _is_ce else _rf
                    _row   = [
                        _r["Stock"],_r["Action"],_r["Score"],_r["Combined"],
                        _r["RR"],_r["Price"],f"{_r.get('Change%',0):+.2f}%",
                        _r["Entry"],_r["SL"],_r["T1"],_r["T2"],
                        _r.get("ATM",""),_r.get("ITM",""),_r.get("OTM",""),
                        round(_r.get("RSI",0),1),_r.get("CPR_Pos",""),
                        _r.get("Confidence",""),"💎 YES" if _is_d else "No"
                    ]
                    for _ci,_v in enumerate(_row,1):
                        _c=ws.cell(row=_ri,column=_ci,value=_v)
                        _c.fill=_rfl; _c.alignment=_ca; _c.border=_bd

                # Paper Trading sheet
                ws2 = wb.create_sheet("Paper Trading")
                _pt_hdrs = [
                    "Date","Stock","Signal","Entry Price","Lots",
                    "ATM Strike","Premium Paid","Stop Loss",
                    "Target 1","Target 2","Exit Price",
                    "P&L Points","P&L ₹","Result","Notes"
                ]
                _ptf = PatternFill("solid", fgColor="0f766e")
                for _ci,_h in enumerate(_pt_hdrs,1):
                    _c=ws2.cell(row=1,column=_ci,value=_h)
                    _c.fill=_ptf
                    _c.font=Font(color="FFFFFF",bold=True)
                    _c.alignment=_ca; _c.border=_bd
                    ws2.column_dimensions[_c.column_letter].width=max(14,len(_h)+2)

                for _ri,_r in enumerate(_all_sigs,2):
                    _is_ce=_r["Direction"]=="UPTREND"
                    _ptrow=[
                        _xdate.today().strftime("%d %b %Y"),
                        _r["Stock"],_r["Action"],_r["Entry"],
                        1,_r.get("ATM",""),"",_r["SL"],
                        _r["T1"],_r["T2"],"","","","",""
                    ]
                    _prf=_gf if _is_ce else _rf
                    for _ci,_v in enumerate(_ptrow,1):
                        _c=ws2.cell(row=_ri,column=_ci,value=_v)
                        _c.fill=_prf; _c.alignment=_ca; _c.border=_bd

                _buf=io.BytesIO()
                wb.save(_buf); _buf.seek(0)
                _fname=f"signals_{_xdt.now().strftime('%d%b%Y_%H%M')}.xlsx"
                st.download_button(
                    label="📥 Download Excel Now",
                    data=_buf.getvalue(),
                    file_name=_fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="scan_dl_top"
                )
                st.success(f"✅ {len(_all_sigs)} signals exported!")
        ce_list  = [r for r in results_sorted if r["Direction"]=="UPTREND"]
        pe_list  = [r for r in results_sorted if r["Direction"]=="DOWNTREND"]

        sm1,sm2,sm3,sm4,sm5 = st.columns(5)
        sm1.metric("Scanned",   total_scanned if total_scanned else len(results))
        sm2.metric("Signals",   len(results))
        sm3.metric("Strong 8+", len(strong_r))
        sm4.metric("BUY CE",    len(ce_list))
        sm5.metric("BUY PE",    len(pe_list))

        # ── Strong signals ─────────────────────────────────
        # ── 💎 DIAMOND SIGNALS ──────────────────────────
        diamond_r = [
            r for r in results_sorted
            if r.get("Is_Diamond", False)
        ]

        if diamond_r:
                st.markdown("---")
                _d_count = len(diamond_r)
                st.markdown(
                    f"<div style='background:linear-gradient("
                    f"135deg,#1e1b4b,#3730a3);"
                    f"border-radius:14px;padding:16px 20px;"
                    f"margin-bottom:16px'>"
                    f"<div style='font-size:22px;font-weight:700;"
                    f"color:#ffffff'>💎 DIAMOND SIGNALS"
                    f" — {_d_count} found</div>"
                    f"<div style='font-size:13px;color:#c7d2fe;"
                    f"margin-top:4px'>"
                    f"Technical + ML + All 3 Timeframes confirmed."
                    f" Highest possible confidence.</div></div>",
                    unsafe_allow_html=True
                )

                for idx_d, r in enumerate(diamond_r):
                    _dc  = "#16a34a" if r["Direction"]=="UPTREND" else "#dc2626"
                    _dbg = "#f0fdf4" if r["Direction"]=="UPTREND" else "#fef2f2"
                    _ml_dir  = r.get("ML_Direction","—")
                    _ml_conf = r.get("ML_Confidence", 0)
                    st.markdown(
                        f"<div style='background:linear-gradient("
                        f"135deg,#faf5ff,#ede9fe);"
                        f"border:2px solid #7c3aed;"
                        f"border-radius:14px;padding:18px;"
                        f"margin-bottom:12px'>"
                        f"<div style='display:flex;justify-content:"
                        f"space-between;align-items:center;"
                        f"flex-wrap:wrap;gap:8px;margin-bottom:10px'>"
                        f"<div><span style='font-size:20px;"
                        f"font-weight:700;color:#1e293b'>"
                        f"{r['Stock']}</span>"
                        f"<span style='background:#7c3aed;color:white;"
                        f"padding:3px 10px;border-radius:12px;"
                        f"font-size:11px;margin-left:8px'>"
                        f"💎 DIAMOND</span></div>"
                        f"<span style='background:{_dbg};color:{_dc};"
                        f"padding:4px 16px;border-radius:20px;"
                        f"font-size:14px;font-weight:700'>"
                        f"{r['Action']}</span></div>"
                        f"<div style='display:grid;"
                        f"grid-template-columns:repeat(4,1fr);"
                        f"gap:8px;margin-bottom:10px'>"
                        f"<div style='background:white;"
                        f"border-radius:8px;padding:10px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"Technical</div>"
                        f"<div style='font-size:20px;font-weight:700;"
                        f"color:{_dc}'>{r['Score']}/10</div></div>"
                        f"<div style='background:white;"
                        f"border-radius:8px;padding:10px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"ML</div>"
                        f"<div style='font-size:13px;font-weight:700;"
                        f"color:#7c3aed'>{_ml_dir}</div>"
                        f"<div style='font-size:11px;color:#64748b'>"
                        f"{_ml_conf}%</div></div>"
                        f"<div style='background:white;"
                        f"border-radius:8px;padding:10px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"Timeframes</div>"
                        f"<div style='font-size:11px;color:#16a34a;"
                        f"font-weight:700'>15m ✅ 1h ✅ 1d ✅"
                        f"</div></div>"
                        f"<div style='background:white;"
                        f"border-radius:8px;padding:10px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"R:R</div>"
                        f"<div style='font-size:20px;font-weight:700;"
                        f"color:#1e293b'>{r['RR']}:1</div></div></div>"
                        f"<div style='display:grid;"
                        f"grid-template-columns:repeat(4,1fr);"
                        f"gap:6px;margin-bottom:8px'>"
                        f"<div style='background:#f0fdf4;"
                        f"border-radius:6px;padding:8px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"Entry</div>"
                        f"<div style='font-size:14px;font-weight:700;"
                        f"color:#16a34a'>₹{r['Entry']:,.0f}</div></div>"
                        f"<div style='background:#fef2f2;"
                        f"border-radius:6px;padding:8px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"SL</div>"
                        f"<div style='font-size:14px;font-weight:700;"
                        f"color:#dc2626'>₹{r['SL']:,.0f}</div></div>"
                        f"<div style='background:#eff6ff;"
                        f"border-radius:6px;padding:8px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"T1</div>"
                        f"<div style='font-size:14px;font-weight:700;"
                        f"color:#1d4ed8'>₹{r['T1']:,.0f}</div></div>"
                        f"<div style='background:#eff6ff;"
                        f"border-radius:6px;padding:8px;"
                        f"text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>"
                        f"T2</div>"
                        f"<div style='font-size:14px;font-weight:700;"
                        f"color:#1d4ed8'>₹{r['T2']:,.0f}</div></div>"
                        f"</div>"
                        f"<div style='font-size:12px;color:#7c3aed'>"
                        f"<b>{r['OptType']}</b> | ATM ✅ {r['ATM']} "
                        f"| ITM {r['ITM']} | OTM {r['OTM']}"
                        f"</div></div>",
                        unsafe_allow_html=True
                    )
                    _dc1, _dc2, _dc3, _dc4 = st.columns(4)
                    with _dc1:
                        if st.button(
                            "📊 Trade Setup",
                            key=f"dia_ts_{idx_d}",
                            type="primary",
                            use_container_width=True
                        ):
                            st.session_state["sn"] = r["Stock"]
                            st.session_state["st"] = r["Sym"]
                            st.rerun()
                    with _dc2:
                        if st.button(
                            "🛡️ Add to TM",
                            key=f"dia_tm_{idx_d}",
                            use_container_width=True,
                            help="Add to Trade Manager instantly"
                        ):
                            import datetime as _dtmdt
                            if "active_trades" not in st.session_state:
                                st.session_state["active_trades"] = load_trades()
                            _dia_tm = {
                                "id":          len(st.session_state["active_trades"]) + 1,
                                "stock":       r["Stock"],
                                "sym":         r["Sym"],
                                "type":        r["Action"],
                                "entry":       r["Entry"],
                                "sl":          r["SL"],
                                "target":      r["T1"],
                                "lots":        1,
                                "lots_rem":    1,
                                "style":       "Intraday (exit 2:45 PM)",
                                "tf":          scan_tf,
                                "opt_price":   0.0,
                                "added_at":    _dtmdt.datetime.now().strftime("%d %b %H:%M"),
                                "status":      "ACTIVE",
                                "last_action": "Added from 💎 Diamond Scanner",
                                "atm_strike":  r.get("ATM",""),
                            }
                            st.session_state["active_trades"].append(_dia_tm)
                            save_trades(st.session_state["active_trades"])
                            st.success(
                                f"✅ 💎 {r['Stock']} added to Trade Manager!"
                            )
                    with _dc3:
                        if st.button(
                            "🔬 Full Analysis",
                            key=f"dia_fa_{idx_d}",
                            use_container_width=True
                        ):
                            _dk = f"expand_dia_{idx_d}"
                            st.session_state[_dk] = (
                                not st.session_state.get(_dk, False)
                            )
                    with _dc4:
                        if tg_configured():
                            if st.button(
                                "📱 Send Signal",
                                key=f"dia_tg_{idx_d}",
                                use_container_width=True
                            ):
                                _tok = st.session_state.get(
                                    "tg_token_saved", ""
                                )
                                _cid = st.session_state.get(
                                    "tg_chat_saved", ""
                                )
                                _msg = (
                                    f"💎 DIAMOND — {r['Stock']}\n"
                                    f"{r['Action']} | "
                                    f"Score {r['Score']}/10\n"
                                    f"ML: {_ml_dir} ({_ml_conf}%)\n"
                                    f"All 3 timeframes confirmed\n"
                                    f"Entry Rs{r['Entry']:,.0f} | "
                                    f"SL Rs{r['SL']:,.0f}\n"
                                    f"T1 Rs{r['T1']:,.0f} | "
                                    f"T2 Rs{r['T2']:,.0f}\n"
                                    f"R:R {r['RR']}:1 | ATM {r['ATM']}"
                                )
                                if send_telegram(_tok, _cid, _msg):
                                    st.success("Diamond signal sent!")
                                else:
                                    st.error("Failed to send")
                    st.markdown(
                        "<div style='margin:6px 0'></div>",
                        unsafe_allow_html=True
                    )


        if strong_r:
            st.markdown("---")
            sh1, sh2, sh3 = st.columns([3,1,1])
            with sh1:
                st.markdown(f"### 🔥 STRONG SIGNALS — {len(strong_r)} found")
                st.caption("Confirmed by technical score AND historical consistency.")
            with sh2:
                if tg_configured():
                    if st.button("📱 Send All", key="scan_tg_all",
                                 type="primary", use_container_width=True):
                        tok = st.session_state.get("tg_token_saved","")
                        cid = st.session_state.get("tg_chat_saved","")
                        sent = 0
                        for r_s in strong_r:
                            msg_s = (
                                f"<b>{r_s['Stock']}</b> — {r_s['Action']}\n"
                                f"Score: {r_s['Score']}/10 | Combined: {r_s['Combined']}/10\n"
                                f"Price: Rs {r_s['Price']:,.2f}\n"
                                f"Entry: Rs {r_s['Entry']:,.2f} | SL: Rs {r_s['SL']:,.2f}\n"
                                f"T1: Rs {r_s['T1']:,.2f} | T2: Rs {r_s['T2']:,.2f}\n"
                                f"ATM: {r_s['ATM']} | RSI: {r_s['RSI']:.0f} | R:R {r_s['RR']}:1"
                            )
                            if send_telegram(tok, cid, msg_s):
                                sent += 1
                        if sent > 0:
                            st.success(f"✅ Sent {sent}/{len(strong_r)} signals!")
                        else:
                            st.error("❌ Failed. Check Telegram setup.")

            # Excel Export button
            with sh3:
                if st.button(
                    "📥 Export Excel",
                    key="scan_export_excel",
                    use_container_width=True,
                    help="Export all signals to Excel for paper trading"
                ):
                    import io
                    from openpyxl import Workbook
                    from openpyxl.styles import (
                        PatternFill, Font, Alignment, Border, Side
                    )
                    from datetime import datetime as _xdt

                    wb  = Workbook()
                    ws  = wb.active
                    ws.title = "Scanner Signals"

                    # ── Styles ─────────────────────────────
                    hdr_fill = PatternFill(
                        "solid", fgColor="1e3a5f"
                    )
                    hdr_font = Font(
                        color="FFFFFF", bold=True, size=11
                    )
                    green_fill = PatternFill(
                        "solid", fgColor="d1fae5"
                    )
                    red_fill   = PatternFill(
                        "solid", fgColor="fee2e2"
                    )
                    purple_fill= PatternFill(
                        "solid", fgColor="ede9fe"
                    )
                    center  = Alignment(horizontal="center")
                    thin    = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )

                    # ── Title row ──────────────────────────
                    ws.merge_cells("A1:P1")
                    ws["A1"] = (
                        f"Trading Terminal — Scanner Signals | "
                        f"{_xdt.now().strftime('%d %b %Y %H:%M IST')}"
                    )
                    ws["A1"].font      = Font(bold=True, size=13)
                    ws["A1"].fill      = PatternFill("solid", fgColor="1e3a5f")
                    ws["A1"].font      = Font(color="FFFFFF", bold=True, size=13)
                    ws["A1"].alignment = center

                    # ── Headers ────────────────────────────
                    headers = [
                        "Stock","Signal","Score","Combined",
                        "R:R","Price","Change%","Entry","Stop Loss",
                        "Target 1","Target 2","ATM Strike",
                        "ITM Strike","OTM Strike","RSI","CPR",
                        "Confidence","Diamond","Sector"
                    ]
                    for ci, hdr in enumerate(headers, 1):
                        cell = ws.cell(row=2, column=ci, value=hdr)
                        cell.fill      = hdr_fill
                        cell.font      = hdr_font
                        cell.alignment = center
                        cell.border    = thin
                        ws.column_dimensions[
                            cell.column_letter
                        ].width = max(12, len(hdr)+2)

                    # ── All signals (Diamond + Strong) ─────
                    all_signals = (
                        [dict(r, _is_d=True)
                         for r in diamond_r] +
                        [dict(r, _is_d=False)
                         for r in strong_r]
                    )

                    for ri, r in enumerate(all_signals, 3):
                        is_d   = r.get("_is_d", False)
                        is_ce  = r["Direction"] == "UPTREND"
                        row_fill = (
                            purple_fill if is_d
                            else green_fill if is_ce
                            else red_fill
                        )
                        row_data = [
                            r["Stock"],
                            r["Action"],
                            r["Score"],
                            r["Combined"],
                            r["RR"],
                            r["Price"],
                            f"{r['Change%']:+.2f}%",
                            r["Entry"],
                            r["SL"],
                            r["T1"],
                            r["T2"],
                            r.get("ATM", ""),
                            r.get("ITM", ""),
                            r.get("OTM", ""),
                            round(r.get("RSI", 0), 1),
                            r.get("CPR_Pos", ""),
                            r.get("Confidence", ""),
                            "💎 YES" if is_d else "No",
                            r.get("Sector", ""),
                        ]
                        for ci, val in enumerate(row_data, 1):
                            cell = ws.cell(row=ri, column=ci, value=val)
                            cell.fill      = row_fill
                            cell.alignment = center
                            cell.border    = thin

                    # ── Paper Trading Sheet ────────────────
                    ws2 = wb.create_sheet("Paper Trading")
                    pt_headers = [
                        "Date","Stock","Signal","Entry Price",
                        "Lots","Option Strike","Premium Paid",
                        "Stop Loss","Target 1","Target 2",
                        "Exit Price","P&L Points","P&L ₹",
                        "Result","Notes"
                    ]
                    pt_fill = PatternFill("solid", fgColor="0f766e")
                    for ci, hdr in enumerate(pt_headers, 1):
                        cell = ws2.cell(row=1, column=ci, value=hdr)
                        cell.fill      = pt_fill
                        cell.font      = Font(
                            color="FFFFFF", bold=True
                        )
                        cell.alignment = center
                        cell.border    = thin
                        ws2.column_dimensions[
                            cell.column_letter
                        ].width = max(14, len(hdr)+2)

                    # Pre-fill paper trading rows from signals
                    from datetime import date as _date
                    for ri, r in enumerate(all_signals, 2):
                        pt_data = [
                            _date.today().strftime("%d %b %Y"),
                            r["Stock"],
                            r["Action"],
                            r["Entry"],
                            1,  # default 1 lot
                            r.get("ATM", ""),
                            "",  # premium paid (fill manually)
                            r["SL"],
                            r["T1"],
                            r["T2"],
                            "",  # exit price
                            "",  # P&L points
                            "",  # P&L ₹
                            "",  # result
                            "",  # notes
                        ]
                        row_fill2 = (
                            green_fill
                            if r["Direction"] == "UPTREND"
                            else red_fill
                        )
                        for ci, val in enumerate(pt_data, 1):
                            cell = ws2.cell(
                                row=ri, column=ci, value=val
                            )
                            cell.fill      = row_fill2
                            cell.alignment = center
                            cell.border    = thin

                    # ── Market Context Sheet ───────────────
                    ws3 = wb.create_sheet("Market Context")
                    ws3["A1"] = "Market Context"
                    ws3["A1"].font = Font(bold=True, size=12)
                    _ctx_data = [
                        ["Date", _xdt.now().strftime("%d %b %Y")],
                        ["Time", _xdt.now().strftime("%H:%M IST")],
                        ["Scan Group", grp_used],
                        ["Timeframe", tf_used],
                        ["Total Signals", len(all_signals)],
                        ["Diamond Signals", len(diamond_r)],
                        ["Strong Signals", len(strong_r)],
                    ]
                    for ri, (k, v) in enumerate(_ctx_data, 2):
                        ws3.cell(row=ri, column=1, value=k).font = Font(bold=True)
                        ws3.cell(row=ri, column=2, value=str(v))

                    # ── Save and download ──────────────────
                    _buf = io.BytesIO()
                    wb.save(_buf)
                    _buf.seek(0)
                    _fname = (
                        f"signals_{_xdt.now().strftime('%d%b%Y_%H%M')}.xlsx"
                    )
                    st.download_button(
                        label="📥 Download Excel File",
                        data=_buf.getvalue(),
                        file_name=_fname,
                        mime="application/vnd.openxmlformats-"
                             "officedocument.spreadsheetml.sheet",
                        key="scan_dl_excel"
                    )
                    st.success(
                        f"✅ Excel ready — {len(all_signals)} signals "
                        f"exported with paper trading sheet!"
                    )

            for idx_r, r in enumerate(strong_r):
                dir_col  = "#16a34a" if r["Direction"]=="UPTREND" else "#dc2626"
                bg_light = "#f0fdf4" if r["Direction"]=="UPTREND" else "#fef2f2"
                chg_col  = "#16a34a" if r["Change%"] >= 0 else "#dc2626"
                arr      = "▲" if r["Change%"] >= 0 else "▼"
                border_col = "#86efac" if r["Direction"]=="UPTREND" else "#fca5a5"

                st.markdown(
                    f"<div style='background:#ffffff;border:1.5px solid {border_col};"
                    f"border-radius:12px;padding:16px 18px;margin-bottom:10px'>"
                    f"<div style='display:flex;justify-content:space-between;"
                    f"align-items:center;flex-wrap:wrap;gap:6px;margin-bottom:8px'>"
                    f"<span style='font-size:17px;font-weight:700;color:#1e293b'>{r['Stock']}</span>"
                    f"<span style='background:{bg_light};color:{dir_col};padding:4px 14px;"
                    f"border-radius:20px;font-size:13px;font-weight:700'>{r['Action']}</span>"
                    f"<span style='font-size:12px;color:#64748b'>{r['Reliability']}</span>"
                    f"</div>"
                    f"<div style='display:flex;gap:16px;flex-wrap:wrap;margin-bottom:10px'>"
                    f"<div><div style='font-size:10px;color:#94a3b8'>Signal</div>"
                    f"<div style='font-size:20px;font-weight:700;color:{dir_col}'>{r['Score']}/10</div></div>"
                    f"<div><div style='font-size:10px;color:#94a3b8'>Combined</div>"
                    f"<div style='font-size:20px;font-weight:700;color:#1e293b'>{r['Combined']}/10</div></div>"
                    f"<div><div style='font-size:10px;color:#94a3b8'>R:R</div>"
                    f"<div style='font-size:14px;font-weight:600;color:#374151'>{r['RR']}:1</div></div>"
                    f"<div><div style='font-size:10px;color:#94a3b8'>Price</div>"
                    f"<div style='font-size:14px;font-weight:600;color:#1e293b'>"
                    f"₹{r['Price']:,.2f} <span style='color:{chg_col}'>{arr}{abs(r['Change%']):.1f}%</span></div></div>"
                    f"</div>"
                    f"<div style='display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:10px'>"
                    f"<div style='background:#f0fdf4;border-radius:6px;padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>Entry</div>"
                    f"<div style='font-size:13px;font-weight:700;color:#16a34a'>₹{r['Entry']:,.0f}</div></div>"
                    f"<div style='background:#fef2f2;border-radius:6px;padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>Stop Loss</div>"
                    f"<div style='font-size:13px;font-weight:700;color:#dc2626'>₹{r['SL']:,.0f}</div></div>"
                    f"<div style='background:#eff6ff;border-radius:6px;padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>Target 1</div>"
                    f"<div style='font-size:13px;font-weight:700;color:#1d4ed8'>₹{r['T1']:,.0f}</div></div>"
                    f"<div style='background:#eff6ff;border-radius:6px;padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>Target 2</div>"
                    f"<div style='font-size:13px;font-weight:700;color:#1d4ed8'>₹{r['T2']:,.0f}</div></div>"
                    f"</div>"
                    f"<div style='background:#faf5ff;border-radius:6px;padding:8px 12px;margin-bottom:8px;"
                    f"font-size:12px;color:#7c3aed'>"
                    f"<b>{r['OptType']}</b> | ATM ✅ {r['ATM']} | ITM {r['ITM']} | OTM {r['OTM']}"
                    f"{'  ✨ Virgin CPR' if r.get('Virgin_CPR') else ''}"
                    f" | CPR: Price {r.get('CPR_Pos','—')}</div>"
                    + (
                        "<div style='background:#f0fdf4;border-left:3px solid #16a34a;"
                        "border-radius:0 6px 6px 0;padding:6px 12px;"
                        "font-size:11px;color:#166534;margin-bottom:6px'>"
                        "✅ CPR confirms — trade with full confidence</div>"
                        if (
                            (r.get("CPR_Pos")=="ABOVE" and r["Direction"]=="UPTREND") or
                            (r.get("CPR_Pos")=="BELOW" and r["Direction"]=="DOWNTREND")
                        ) else
                        "<div style='background:#fef2f2;border-left:3px solid #dc2626;"
                        "border-radius:0 6px 6px 0;padding:6px 12px;"
                        "font-size:11px;color:#991b1b;margin-bottom:6px'>"
                        "⚠️ CPR conflicts with signal — consider skipping or reduce size</div>"
                        if (
                            (r.get("CPR_Pos")=="BELOW" and r["Direction"]=="UPTREND") or
                            (r.get("CPR_Pos")=="ABOVE" and r["Direction"]=="DOWNTREND")
                        ) else
                        "<div style='background:#fffbeb;border-left:3px solid #f59e0b;"
                        "border-radius:0 6px 6px 0;padding:6px 12px;"
                        "font-size:11px;color:#92400e;margin-bottom:6px'>"
                        "⚠️ Price inside CPR — choppy zone, trade with smaller size</div>"
                    )
                    + f"</div>",
                    unsafe_allow_html=True
                )
                bc1, bc2, bc3 = st.columns(3)
                with bc1:
                    if st.button(
                        "🔬 Full Analysis",
                        key=f"scan_full_{idx_r}",
                        type="primary",
                        use_container_width=True
                    ):
                        # Toggle expanded analysis for this stock
                        key = f"expand_{idx_r}"
                        st.session_state[key] = not st.session_state.get(key, False)
                with bc2:
                    if st.button(
                        "📊 Trade Setup →",
                        key=f"scan_an_{idx_r}",
                        use_container_width=True
                    ):
                        st.session_state["sn"] = r["Stock"]
                        st.session_state["st"] = r["Sym"]
                        st.rerun()
                with bc3:
                    if tg_configured():
                        if st.button(
                            "📱 Send Signal",
                            key=f"scan_tg_{idx_r}",
                            use_container_width=True
                        ):
                            tok = st.session_state.get("tg_token_saved","")
                            cid = st.session_state.get("tg_chat_saved","")
                            msg = (
                                f"<b>{r['Stock']}</b> — {r['Action']}\n"
                                f"Score: {r['Score']}/10 | Combined: {r['Combined']}/10\n"
                                f"Reliability: {r['Reliability']}\n"
                                f"Price: Rs {r['Price']:,.2f}\n"
                                f"Entry: Rs {r['Entry']:,.2f} | SL: Rs {r['SL']:,.2f}\n"
                                f"T1: Rs {r['T1']:,.2f} | T2: Rs {r['T2']:,.2f}\n"
                                f"ATM: {r['ATM']} | ITM: {r['ITM']} | OTM: {r['OTM']}\n"
                                f"RSI: {r['RSI']:.0f} | R:R {r['RR']}:1"
                            )
                            if send_telegram(tok, cid, msg):
                                st.success(f"✅ {r['Stock']} sent!")
                            else:
                                st.error("❌ Failed — check Telegram setup")

                # ── COMBINED ANALYSIS PANEL ────────────────────
                if st.session_state.get(f"expand_{idx_r}", False):
                    with st.container():
                        st.markdown(
                            f"<div style='background:#f0f9ff;"
                            f"border:2px solid #3b82f6;"
                            f"border-radius:14px;padding:20px;"
                            f"margin:8px 0'>"
                            f"<div style='font-size:16px;font-weight:700;"
                            f"color:#1e40af;margin-bottom:16px'>"
                            f"🔬 Full Analysis — {r['Stock']}"
                            f"</div></div>",
                            unsafe_allow_html=True
                        )

                        # Load data for this stock
                        _sym  = r["Sym"]
                        _name = r["Stock"]
                        _lp   = live_price(_sym)

                        with st.spinner(
                            f"Loading full analysis for {_name}..."
                        ):
                            _df  = candles(_sym, scan_tf)
                            _sig = compute_all(_df, _lp) if (
                                _df is not None and
                                len(_df) >= 55
                            ) else None
                            # ML
                            _df_ml = candles(_sym, "1d")
                            _model = None
                            _pred  = None
                            if _df_ml is not None and len(_df_ml) >= 100:
                                _model = train_model(_df_ml)
                                if _model.get("ok"):
                                    _pred = predict_next_move(
                                        _df_ml, _model
                                    )
                            # Real-time approximation
                            _rt = approximate_realtime(
                                _df,
                                _lp["p"] if _lp["ok"] else 0
                            ) if _df is not None else None

                        # ── TOP ROW: Scanner + ML side by side ─────
                        col_sc, col_ml = st.columns(2)

                        with col_sc:
                            st.markdown(
                                "<div style='background:#ffffff;"
                                "border:1px solid #e2e8f0;"
                                "border-radius:12px;padding:16px'>"
                                "<div style='font-size:13px;font-weight:700;"
                                "color:#64748b;margin-bottom:12px;"
                                "text-transform:uppercase;letter-spacing:1px'>"
                                "📊 Scanner Signal</div>",
                                unsafe_allow_html=True
                            )
                            dir_c = (
                                "#16a34a"
                                if r["Direction"] == "UPTREND"
                                else "#dc2626"
                            )
                            st.markdown(
                                f"<div style='font-size:28px;"
                                f"font-weight:700;color:{dir_c}'>"
                                f"{r['Action']}</div>"
                                f"<div style='font-size:13px;"
                                f"color:#475569;margin-top:4px'>"
                                f"Score: <b>{r['Score']}/10</b> | "
                                f"Combined: <b>{r['Combined']}/10</b>"
                                f"</div>"
                                f"<div style='font-size:12px;"
                                f"color:#64748b;margin-top:8px'>"
                                f"Reliability: {r['Reliability']}</div>"
                                f"<hr style='border-color:#f1f5f9'>"
                                f"<table style='width:100%;font-size:13px'>"
                                f"<tr><td style='color:#64748b'>Entry</td>"
                                f"<td style='text-align:right;color:#16a34a;"
                                f"font-weight:700'>₹{r['Entry']:,.2f}</td></tr>"
                                f"<tr><td style='color:#64748b'>Stop Loss</td>"
                                f"<td style='text-align:right;color:#dc2626;"
                                f"font-weight:700'>₹{r['SL']:,.2f}</td></tr>"
                                f"<tr><td style='color:#64748b'>Target 1</td>"
                                f"<td style='text-align:right;color:#1d4ed8;"
                                f"font-weight:700'>₹{r['T1']:,.2f}</td></tr>"
                                f"<tr><td style='color:#64748b'>Target 2</td>"
                                f"<td style='text-align:right;color:#1d4ed8;"
                                f"font-weight:700'>₹{r['T2']:,.2f}</td></tr>"
                                f"<tr><td style='color:#64748b'>R:R Ratio</td>"
                                f"<td style='text-align:right;font-weight:700;"
                                f"color:#1e293b'>{r['RR']}:1</td></tr>"
                                f"<tr><td style='color:#64748b'>RSI</td>"
                                f"<td style='text-align:right;"
                                f"color:#1e293b'>{r['RSI']:.0f}</td></tr>"
                                f"<tr><td style='color:#64748b'>ADX</td>"
                                f"<td style='text-align:right;"
                                f"color:#1e293b'>{r['ADX']:.0f}</td></tr>"
                                f"</table>"
                                f"<hr style='border-color:#f1f5f9'>"
                                f"<div style='font-size:12px;color:#7c3aed;"
                                f"font-weight:600'>OPTIONS</div>"
                                f"<div style='font-size:12px;color:#475569;"
                                f"margin-top:4px'>"
                                f"{r['OptType']}<br>"
                                f"✅ ATM {r['ATM']} (recommended)<br>"
                                f"ITM {r['ITM']} | OTM {r['OTM']}"
                                f"</div>"
                                f"<div style='margin-top:8px;font-size:12px;"
                                f"color:#475569'>"
                                f"CPR: Price <b>{r.get('CPR_Pos','—')}</b> | "
                                f"{r.get('CPR_Bias','—')}"
                                f"{'  ✨ Virgin CPR' if r.get('Virgin_CPR') else ''}"
                                f"</div></div>",
                                unsafe_allow_html=True
                            )

                        with col_ml:
                            st.markdown(
                                "<div style='background:#ffffff;"
                                "border:1px solid #e2e8f0;"
                                "border-radius:12px;padding:16px'>"
                                "<div style='font-size:13px;font-weight:700;"
                                "color:#64748b;margin-bottom:12px;"
                                "text-transform:uppercase;letter-spacing:1px'>"
                                "🤖 ML Prediction (1d)</div>",
                                unsafe_allow_html=True
                            )

                            if _pred and _pred.get("ok"):
                                pc_ = _pred["sig_color"]
                                pred_bg = (
                                    "#f0fdf4"
                                    if _pred["prediction"]=="UPTREND"
                                    else "#fef2f2"
                                    if _pred["prediction"]=="DOWNTREND"
                                    else "#fffbeb"
                                )
                                st.markdown(
                                    f"<div style='background:{pred_bg};"
                                    f"border-radius:8px;padding:14px;"
                                    f"text-align:center;margin-bottom:12px'>"
                                    f"<div style='font-size:32px;font-weight:700;"
                                    f"color:{pc_}'>{_pred['prediction']}</div>"
                                    f"<div style='font-size:16px;color:{pc_};"
                                    f"font-weight:600'>{_pred['signal']}</div>"
                                    f"<div style='font-size:13px;color:#64748b;"
                                    f"margin-top:6px'>"
                                    f"{_pred['reliability']} "
                                    f"({_pred['confidence']}%)"
                                    f"</div></div>",
                                    unsafe_allow_html=True
                                )

                                # Probability bars
                                probs = _pred["probabilities"]
                                for pname, pval in probs.items():
                                    bc_ = (
                                        "#16a34a" if pname=="UPTREND"
                                        else "#dc2626" if pname=="DOWNTREND"
                                        else "#f59e0b"
                                    )
                                    st.markdown(
                                        f"<div style='margin:4px 0'>"
                                        f"<div style='display:flex;"
                                        f"justify-content:space-between;"
                                        f"font-size:12px;color:#475569'>"
                                        f"<span>{pname}</span>"
                                        f"<span style='font-weight:600;"
                                        f"color:{bc_}'>{pval:.1f}%</span></div>"
                                        f"<div style='background:#f1f5f9;"
                                        f"height:6px;border-radius:3px;margin-top:2px'>"
                                        f"<div style='background:{bc_};"
                                        f"width:{pval:.0f}%;height:6px;"
                                        f"border-radius:3px'></div></div></div>",
                                        unsafe_allow_html=True
                                    )

                                # Model accuracy
                                st.markdown(
                                    f"<div style='margin-top:10px;font-size:12px;"
                                    f"color:#64748b'>"
                                    f"Model accuracy: "
                                    f"<b>{_pred['model_accuracy']}%</b> | "
                                    f"Trained on "
                                    f"<b>{_pred['n_trained']}</b> candles"
                                    f"</div>",
                                    unsafe_allow_html=True
                                )

                                # Agreement check
                                scan_dir = r["Direction"]
                                ml_dir   = _pred["prediction"]
                                if scan_dir == "UPTREND" and ml_dir == "UPTREND":
                                    st.success(
                                        "🔥 Scanner + ML both agree BULLISH — "
                                        "highest confidence CE setup!"
                                    )
                                elif scan_dir == "DOWNTREND" and ml_dir == "DOWNTREND":
                                    st.error(
                                        "🔥 Scanner + ML both agree BEARISH — "
                                        "highest confidence PE setup!"
                                    )
                                else:
                                    st.warning(
                                        f"⚠️ Scanner says {scan_dir} but "
                                        f"ML says {ml_dir}. "
                                        "Wait for both to agree."
                                    )
                            else:
                                st.info(
                                    "ML needs 100+ daily candles. "
                                    "Not enough data for this stock."
                                )

                            # Live bias from real-time approximation
                            if _rt and _rt.get("ok"):
                                st.markdown("---")
                                bc_rt = _rt["bias_color"]
                                st.markdown(
                                    f"<div style='background:#f8fafc;"
                                    f"border-radius:8px;padding:10px 14px'>"
                                    f"<div style='font-size:11px;color:#64748b;"
                                    f"text-transform:uppercase;letter-spacing:1px'>"
                                    f"Live Approximation</div>"
                                    f"<div style='font-size:20px;font-weight:700;"
                                    f"color:{bc_rt};margin:4px 0'>"
                                    f"{_rt['live_bias']}</div>"
                                    f"<div style='font-size:12px;color:#64748b'>"
                                    f"RSI: {_rt['rsi_live']} | "
                                    f"Since last candle: "
                                    f"{_rt['since_close']:+.2f}% | "
                                    f"Micro: {_rt['micro_trend']}"
                                    f"</div></div>",
                                    unsafe_allow_html=True
                                )
                            st.markdown("</div>", unsafe_allow_html=True)

                        # ── MINI CHART ─────────────────────────────
                        if _df is not None and len(_df) >= 30:
                            st.markdown(
                                f"#### 🕯️ {_name} — {scan_tf} chart"
                            )
                            plot_mini = _df.tail(60).copy()
                            fig_mini  = make_subplots(
                                rows=2, cols=1,
                                shared_xaxes=True,
                                row_heights=[0.75, 0.25],
                                vertical_spacing=0.03
                            )
                            # Candles
                            fig_mini.add_trace(go.Candlestick(
                                x=plot_mini.index,
                                open=plot_mini["Open"],
                                high=plot_mini["High"],
                                low=plot_mini["Low"],
                                close=plot_mini["Close"],
                                name="Price",
                                increasing_line_color="#16a34a",
                                decreasing_line_color="#dc2626",
                            ), row=1, col=1)

                            # EMAs
                            if _sig:
                                for ser, col_, nm in [
                                    (_sig["e9s"],   "#f59e0b", "EMA9"),
                                    (_sig["e21s"],  "#ea580c", "EMA21"),
                                    (_sig["vwaps"], "#64748b", "VWAP"),
                                ]:
                                    fig_mini.add_trace(go.Scatter(
                                        x=plot_mini.index,
                                        y=ser.tail(60),
                                        line=dict(color=col_, width=1.5),
                                        name=nm
                                    ), row=1, col=1)

                                # CPR lines
                                for cpr_val, cpr_nm, cpr_cl in [
                                    (_sig["cpr_tc"],    "TC",    "#f59e0b"),
                                    (_sig["cpr_pivot"], "Pivot", "#f59e0b"),
                                    (_sig["cpr_bc"],    "BC",    "#f59e0b"),
                                ]:
                                    fig_mini.add_hline(
                                        y=cpr_val,
                                        line_dash="dot",
                                        line_color=cpr_cl,
                                        line_width=1,
                                        opacity=0.7,
                                        annotation_text=cpr_nm,
                                        row=1, col=1
                                    )

                                # SL line
                                sl_plot = (
                                    _sig["sl_long"]
                                    if r["Direction"] == "UPTREND"
                                    else _sig["sl_short"]
                                )
                                fig_mini.add_hline(
                                    y=sl_plot,
                                    line_dash="dash",
                                    line_color="#dc2626",
                                    line_width=1.5,
                                    opacity=0.8,
                                    annotation_text=f"SL ₹{sl_plot:,}",
                                    row=1, col=1
                                )

                            # Volume
                            vc = [
                                "#16a34a"
                                if float(plot_mini["Close"].iloc[i]) >=
                                   float(plot_mini["Open"].iloc[i])
                                else "#dc2626"
                                for i in range(len(plot_mini))
                            ]
                            fig_mini.add_trace(go.Bar(
                                x=plot_mini.index,
                                y=plot_mini["Volume"],
                                marker_color=vc,
                                name="Vol", opacity=0.7
                            ), row=2, col=1)

                            fig_mini.update_layout(
                                template="plotly_white",
                                height=420,
                                xaxis_rangeslider_visible=False,
                                margin=dict(l=10,r=10,t=20,b=10),
                                legend=dict(
                                    orientation="h", y=1.02,
                                    font=dict(size=10)
                                ),
                            )
                            st.plotly_chart(
                                fig_mini, use_container_width=True
                            )

                st.markdown(
                    "<div style='margin:4px 0'></div>",
                    unsafe_allow_html=True
                )

        # ── Good signals ───────────────────────────────────
        if good_r:
            st.markdown("---")
            st.markdown(f"### 📈 GOOD SIGNALS — {len(good_r)} found")
            st.caption("Watch these — enter when score reaches 8+")
            for gi in range(0, len(good_r), 2):
                chunk = good_r[gi:gi+2]
                gcols = st.columns(2)
                for ci, r in enumerate(chunk):
                    dc = "#16a34a" if r["Direction"]=="UPTREND" else "#dc2626"
                    with gcols[ci]:
                        st.markdown(
                            f"<div style='background:#ffffff;border:1px solid #e2e8f0;"
                            f"border-radius:10px;padding:14px;margin-bottom:6px'>"
                            f"<div style='display:flex;justify-content:space-between'>"
                            f"<b style='color:#1e293b;font-size:15px'>{r['Stock']}</b>"
                            f"<span style='color:{dc};font-weight:700'>{r['Score']}/10</span></div>"
                            f"<div style='font-size:12px;color:#64748b;margin-top:6px'>"
                            f"{r['Action']} | ₹{r['Price']:,.0f} | RSI {r['RSI']:.0f}</div>"
                            f"<div style='font-size:11px;color:#94a3b8;margin-top:4px'>"
                            f"Entry ₹{r['Entry']:,.0f} | SL ₹{r['SL']:,.0f}</div>"
                            f"<div style='font-size:11px;color:#7c3aed;margin-top:4px'>"
                            f"{r['OptType']} ATM {r['ATM']}</div></div>",
                            unsafe_allow_html=True
                        )
                        if st.button("View", key=f"scan_vw_{gi}_{ci}",
                                     use_container_width=True):
                            st.session_state["sn"] = r["Stock"]
                            st.session_state["st"] = r["Sym"]
                            st.rerun()

        # ── Score chart ────────────────────────────────────
        if results_sorted:
            st.markdown("---")
            import plotly.graph_objects as go_scan
            fig_scan = go_scan.Figure(go_scan.Bar(
                x=[r["Stock"] for r in results_sorted],
                y=[r["Combined"] for r in results_sorted],
                marker_color=["#16a34a" if r["Direction"]=="UPTREND"
                              else "#dc2626" for r in results_sorted],
                text=[f"{r['Combined']}/10" for r in results_sorted],
                textposition="outside",
            ))
            fig_scan.add_hline(y=8, line_dash="dash", line_color="#16a34a",
                               annotation_text="Strong (8+)")
            fig_scan.add_hline(y=6, line_dash="dot", line_color="#f59e0b",
                               annotation_text="Good (6+)")
            fig_scan.update_layout(
                template="plotly_white", height=300,
                yaxis_range=[0,11],
                margin=dict(l=10,r=10,t=20,b=60),
                xaxis_tickangle=-35, showlegend=False,
                title="Combined Score (60% Technical + 40% Historical)"
            )
            st.plotly_chart(fig_scan, use_container_width=True)

    if run_scanner or auto_scan:
        if auto_scan and not run_scanner:
            st.info("Auto scan active — running every 5 minutes")

        stocks_to_scan = SCANNER_UNIVERSE[scan_group]
        st.markdown(f"**Scanning {len(stocks_to_scan)} stocks "
                f"on {scan_tf} timeframe...**")

        with st.spinner("Scanning... please wait"):
            results, alerted = run_scan_engine(
                stocks_to_scan, scan_tf,
                min_score_scan, alert_score
            )
        # Filter by combined score
        min_comb = st.session_state.get("min_combined_scan", 5)
        results = [
        r for r in results
        if r.get("Combined", 0) >= min_comb
        ]

        # ── Save results to session state ──────────────
        # This ensures results persist when buttons are clicked
        if results:
            st.session_state["scan_results"]     = results
            st.session_state["scan_group_used"]  = scan_group
            st.session_state["scan_tf_used"]     = scan_tf
            st.session_state["scan_time"]        = (
                now_ist().strftime("%H:%M:%S IST")
            )

        if not results:
            st.warning(
            "No stocks met the minimum score. "
            "Try lowering the min score or use 1d timeframe."
        )
        else:
            total_scanned = len(stocks_to_scan)

            # ── Summary metrics ───────────────────────────────
            results_sorted = sorted(
                results, key=lambda x: x["Score"], reverse=True
            )
            strong_r = [r for r in results_sorted if r["Score"] >= 8]
            good_r   = [r for r in results_sorted if 6 <= r["Score"] < 8]
            ce_list  = [r for r in results_sorted if r["Direction"]=="UPTREND"]
            pe_list  = [r for r in results_sorted if r["Direction"]=="DOWNTREND"]

            sm1,sm2,sm3,sm4,sm5 = st.columns(5)
            sm1.metric("Scanned",   total_scanned if total_scanned else len(results))
            sm2.metric("Signals",   len(results))
            sm3.metric("Strong 8+", len(strong_r))
            sm4.metric("BUY CE",    len(ce_list))
            sm5.metric("BUY PE",    len(pe_list))

            # ── Strong signals ─────────────────────────────────
            if strong_r:
                st.markdown("---")
                sh1, sh2 = st.columns([3,1])
                with sh1:
                    st.markdown(f"### 🔥 STRONG SIGNALS — {len(strong_r)} found")
                    st.caption("Confirmed by technical score AND historical consistency.")
                with sh2:
                    if tg_configured():
                        if st.button("📱 Send All", key="scan_tg_all",
                                     type="primary", use_container_width=True):
                            tok = st.session_state.get("tg_token_saved","")
                            cid = st.session_state.get("tg_chat_saved","")
                            sent = 0
                            for r_s in strong_r:
                                msg_s = (
                                    f"<b>{r_s['Stock']}</b> — {r_s['Action']}\n"
                                    f"Score: {r_s['Score']}/10 | Combined: {r_s['Combined']}/10\n"
                                    f"Price: Rs {r_s['Price']:,.2f}\n"
                                    f"Entry: Rs {r_s['Entry']:,.2f} | SL: Rs {r_s['SL']:,.2f}\n"
                                    f"T1: Rs {r_s['T1']:,.2f} | T2: Rs {r_s['T2']:,.2f}\n"
                                    f"ATM: {r_s['ATM']} | RSI: {r_s['RSI']:.0f} | R:R {r_s['RR']}:1"
                                )
                                if send_telegram(tok, cid, msg_s):
                                    sent += 1
                            if sent > 0:
                                st.success(f"✅ Sent {sent}/{len(strong_r)} signals!")
                            else:
                                st.error("❌ Failed. Check Telegram setup.")

                for idx_r, r in enumerate(strong_r):
                    dir_col  = "#16a34a" if r["Direction"]=="UPTREND" else "#dc2626"
                    bg_light = "#f0fdf4" if r["Direction"]=="UPTREND" else "#fef2f2"
                    chg_col  = "#16a34a" if r["Change%"] >= 0 else "#dc2626"
                    arr      = "▲" if r["Change%"] >= 0 else "▼"
                    border_col = "#86efac" if r["Direction"]=="UPTREND" else "#fca5a5"

                    # Confidence badge
                    conf = r.get("Confidence","MEDIUM")
                    conf_col = (
                        "#16a34a" if conf=="HIGH"
                        else "#d97706" if conf=="MEDIUM"
                        else "#dc2626"
                    )
                    conf_bg = (
                        "#f0fdf4" if conf=="HIGH"
                        else "#fffbeb" if conf=="MEDIUM"
                        else "#fef2f2"
                    )
                    conf_icon = (
                        "🔥" if conf=="HIGH"
                        else "⚡" if conf=="MEDIUM"
                        else "⚠️"
                    )

                    st.markdown(
                        f"<div style='background:#ffffff;border:1.5px solid {border_col};"
                        f"border-radius:12px;padding:16px 18px;margin-bottom:10px'>"
                        f"<div style='display:flex;justify-content:space-between;"
                        f"align-items:center;flex-wrap:wrap;gap:6px;margin-bottom:8px'>"
                        f"<span style='font-size:17px;font-weight:700;color:#1e293b'>{r['Stock']}</span>"
                        f"<span style='background:{bg_light};color:{dir_col};padding:4px 14px;"
                        f"border-radius:20px;font-size:13px;font-weight:700'>{r['Action']}</span>"
                        f"<span style='background:{conf_bg};color:{conf_col};"
                        f"padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600'>"
                        f"{conf_icon} {conf} CONFIDENCE</span>"
                        f"<span style='font-size:12px;color:#64748b'>{r['Reliability']}</span>"
                        f"</div>"
                        f"<div style='display:flex;gap:16px;flex-wrap:wrap;margin-bottom:10px'>"
                        f"<div><div style='font-size:10px;color:#94a3b8'>Signal</div>"
                        f"<div style='font-size:20px;font-weight:700;color:{dir_col}'>{r['Score']}/10</div></div>"
                        f"<div><div style='font-size:10px;color:#94a3b8'>Combined</div>"
                        f"<div style='font-size:20px;font-weight:700;color:#1e293b'>{r['Combined']}/10</div></div>"
                        f"<div><div style='font-size:10px;color:#94a3b8'>R:R</div>"
                        f"<div style='font-size:14px;font-weight:600;color:#374151'>{r['RR']}:1</div></div>"
                        f"<div><div style='font-size:10px;color:#94a3b8'>Price</div>"
                        f"<div style='font-size:14px;font-weight:600;color:#1e293b'>"
                        f"₹{r['Price']:,.2f} <span style='color:{chg_col}'>{arr}{abs(r['Change%']):.1f}%</span></div></div>"
                        f"</div>"
                        f"<div style='display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:10px'>"
                        f"<div style='background:#f0fdf4;border-radius:6px;padding:8px;text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>Entry</div>"
                        f"<div style='font-size:13px;font-weight:700;color:#16a34a'>₹{r['Entry']:,.0f}</div></div>"
                        f"<div style='background:#fef2f2;border-radius:6px;padding:8px;text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>Stop Loss</div>"
                        f"<div style='font-size:13px;font-weight:700;color:#dc2626'>₹{r['SL']:,.0f}</div></div>"
                        f"<div style='background:#eff6ff;border-radius:6px;padding:8px;text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>Target 1</div>"
                        f"<div style='font-size:13px;font-weight:700;color:#1d4ed8'>₹{r['T1']:,.0f}</div></div>"
                        f"<div style='background:#eff6ff;border-radius:6px;padding:8px;text-align:center'>"
                        f"<div style='font-size:10px;color:#64748b'>Target 2</div>"
                        f"<div style='font-size:13px;font-weight:700;color:#1d4ed8'>₹{r['T2']:,.0f}</div></div>"
                        f"</div>"
                        f"<div style='background:#faf5ff;border-radius:6px;padding:8px 12px;margin-bottom:8px;"
                        f"font-size:12px;color:#7c3aed'>"
                        f"<b>{r['OptType']}</b> | ATM ✅ {r['ATM']} | ITM {r['ITM']} | OTM {r['OTM']}"
                        f"{'  ✨ Virgin CPR' if r.get('Virgin_CPR') else ''}"
                        f" | CPR: Price {r.get('CPR_Pos','—')}</div>"
                        + f"<div style='background:{r.get('ValidityBg','#f8fafc')};"
                          f"border-left:3px solid {r.get('ValidityCol','#94a3b8')};"
                          f"border-radius:0 6px 6px 0;padding:6px 12px;"
                          f"margin-top:4px;font-size:12px;"
                          f"color:{r.get('ValidityCol','#64748b')}'>"
                          f"⏱️ <b>Signal Validity:</b> "
                          f"{r.get('Validity','Checking...')}</div>"
                        + "</div>",
                        unsafe_allow_html=True
                    )
                    bc1, bc2, bc3 = st.columns(3)
                    with bc1:
                        if st.button(f"📊 Analyse", key=f"scan_an_{idx_r}",
                                     type="primary", use_container_width=True):
                            st.session_state["sn"] = r["Stock"]
                            st.session_state["st"] = r["Sym"]
                            st.rerun()
                    with bc2:
                        if st.button(
                            f"🛡️ Add to Trade Manager",
                            key=f"scan_tm_{idx_r}",
                            use_container_width=True,
                            help="Add this signal directly to Trade Manager"
                        ):
                            import datetime as _tmdt
                            # Pre-fill all details from signal
                            if "active_trades" not in st.session_state:
                                st.session_state["active_trades"] = load_trades()
                            _new_tm = {
                                "id":        len(st.session_state["active_trades"]) + 1,
                                "stock":     r["Stock"],
                                "sym":       r["Sym"],
                                "type":      r["Action"],
                                "entry":     r["Entry"],
                                "sl":        r["SL"],
                                "target":    r["T1"],
                                "lots":      1,
                                "lots_rem":  1,
                                "style":     "Intraday (exit 2:45 PM)",
                                "tf":        scan_tf,
                                "opt_price": 0.0,
                                "added_at":  _tmdt.datetime.now().strftime("%d %b %H:%M"),
                                "status":    "ACTIVE",
                                "last_action": "Added from Scanner",
                                "atm_strike": r.get("ATM",""),
                            }
                            st.session_state["active_trades"].append(_new_tm)
                            save_trades(st.session_state["active_trades"])
                            st.success(
                                f"✅ {r['Stock']} {r['Action']} "
                                f"added to Trade Manager!"
                            )
                    with bc3:
                        if tg_configured():
                            if st.button(f"📱 Send Signal", key=f"scan_tg_{idx_r}",
                                         use_container_width=True):
                                tok = st.session_state.get("tg_token_saved","")
                                cid = st.session_state.get("tg_chat_saved","")
                                msg = (
                                    f"<b>{r['Stock']}</b> — {r['Action']}\n"
                                    f"Score: {r['Score']}/10 | Combined: {r['Combined']}/10\n"
                                    f"Reliability: {r['Reliability']}\n"
                                    f"Price: Rs {r['Price']:,.2f}\n"
                                    f"Entry: Rs {r['Entry']:,.2f} | SL: Rs {r['SL']:,.2f}\n"
                                    f"T1: Rs {r['T1']:,.2f} | T2: Rs {r['T2']:,.2f}\n"
                                    f"ATM: {r['ATM']} | ITM: {r['ITM']} | OTM: {r['OTM']}\n"
                                    f"RSI: {r['RSI']:.0f} | R:R {r['RR']}:1"
                                )
                                if send_telegram(tok, cid, msg):
                                    st.success(f"✅ {r['Stock']} sent!")
                                else:
                                    st.error("❌ Failed — check Telegram setup")
                    st.markdown("<div style='margin:4px 0'></div>", unsafe_allow_html=True)

            # ── Good signals ───────────────────────────────────
            if good_r:
                st.markdown("---")
                st.markdown(f"### 📈 GOOD SIGNALS — {len(good_r)} found")
                st.caption("Watch these — enter when score reaches 8+")
                for gi in range(0, len(good_r), 2):
                    chunk = good_r[gi:gi+2]
                    gcols = st.columns(2)
                    for ci, r in enumerate(chunk):
                        dc = "#16a34a" if r["Direction"]=="UPTREND" else "#dc2626"
                        with gcols[ci]:
                            st.markdown(
                                f"<div style='background:#ffffff;border:1px solid #e2e8f0;"
                                f"border-radius:10px;padding:14px;margin-bottom:6px'>"
                                f"<div style='display:flex;justify-content:space-between'>"
                                f"<b style='color:#1e293b;font-size:15px'>{r['Stock']}</b>"
                                f"<span style='color:{dc};font-weight:700'>{r['Score']}/10</span></div>"
                                f"<div style='font-size:12px;color:#64748b;margin-top:6px'>"
                                f"{r['Action']} | ₹{r['Price']:,.0f} | RSI {r['RSI']:.0f}</div>"
                                f"<div style='font-size:11px;color:#94a3b8;margin-top:4px'>"
                                f"Entry ₹{r['Entry']:,.0f} | SL ₹{r['SL']:,.0f}</div>"
                                f"<div style='font-size:11px;color:#7c3aed;margin-top:4px'>"
                                f"{r['OptType']} ATM {r['ATM']}</div></div>",
                                unsafe_allow_html=True
                            )
                            if st.button("View", key=f"scan_vw_{gi}_{ci}",
                                         use_container_width=True):
                                st.session_state["sn"] = r["Stock"]
                                st.session_state["st"] = r["Sym"]
                                st.rerun()

            # ── Score chart ────────────────────────────────────

            # ── Store stocks for Signal Monitor ───────────────
            if results_sorted:
                st.session_state["monitor_stocks"] = [
                    {
                        "stock": r["Stock"],
                        "sym":   r["Sym"],
                        "dir":   r["Direction"],
                        "score": r["Score"],
                        "entry": r["Entry"],
                        "sl":    r["SL"],
                        "t1":    r["T1"],
                        "valid": r.get("ValidCandles",4),
                    }
                    for r in results_sorted
                ]
                st.session_state["monitor_scan_tf"]   = scan_tf
                st.session_state["monitor_scan_time"] = datetime.now().strftime("%H:%M")

            # ── Signal Monitor ─────────────────────────────────
            _mlist = st.session_state.get("monitor_stocks",[])
            if _mlist:
                st.markdown("---")
                st.markdown("### 🔭 Signal Monitor")
                _sct = st.session_state.get("monitor_scan_time","—")
                _stf = st.session_state.get("monitor_scan_tf","15m")
                st.caption(
                    f"Watching {len(_mlist)} stocks from {_sct} scan. "
                    f"Click Refresh to update status."
                )
                _mc1, _mc2 = st.columns([3,1])
                with _mc2:
                    _do_mon = st.button(
                        "🔄 Refresh Monitor",
                        key="monitor_refresh",
                        use_container_width=True
                    )

                _still_v = []; _weaken = []; _expd = []

                for _ms in _mlist:
                    _ms_lp = live_price(_ms["sym"])
                    _ms_cp = _ms_lp["p"] if _ms_lp["ok"] else _ms["entry"]
                    _ms_sig = None
                    if _do_mon:
                        try:
                            _ms_df = candles(_ms["sym"], _stf)
                            if _ms_df is not None and len(_ms_df)>=55:
                                _ms_sig = compute_all(_ms_df, _ms_lp)
                                st.session_state[f"msig_{_ms['stock']}"] = _ms_sig
                        except Exception:
                            pass
                    _ms_sig = _ms_sig or st.session_state.get(f"msig_{_ms['stock']}")

                    _is_ce = _ms["dir"] == "UPTREND"
                    _sl_hit  = (_ms_cp<=_ms["sl"]) if _is_ce else (_ms_cp>=_ms["sl"])
                    _tgt_hit = (_ms_cp>=_ms["t1"]) if _is_ce else (_ms_cp<=_ms["t1"])
                    _pnl = (_ms_cp-_ms["entry"]) if _is_ce else (_ms["entry"]-_ms_cp)
                    _pnl_col = "#16a34a" if _pnl>=0 else "#dc2626"

                    if _ms_sig:
                        _cur_dir   = _ms_sig["direction"]
                        _cur_score = max(_ms_sig["up_score"],_ms_sig["dn_score"])
                        _cur_rsi   = _ms_sig["rv"]
                        _cur_vwap  = _ms_sig["vwv"]

                        _chk = sum([
                            _cur_dir == _ms["dir"],
                            _cur_score >= 7,
                            (55<=_cur_rsi<=75) if _is_ce else (25<=_cur_rsi<=45),
                            (_ms_cp>_cur_vwap) if _is_ce else (_ms_cp<_cur_vwap),
                        ])
                    else:
                        _chk = 3  # assume valid if no fresh data

                    if _sl_hit:
                        _st="🔴 SL HIT"; _sc="#dc2626"; _sb="#fef2f2"; _sa="Exit immediately"; _expd.append(_ms)
                    elif _tgt_hit:
                        _st="🟢 TARGET HIT"; _sc="#16a34a"; _sb="#f0fdf4"; _sa="Book profit!"; _still_v.append(_ms)
                    elif _chk>=3:
                        _st="🟢 STILL VALID"; _sc="#16a34a"; _sb="#f0fdf4"; _sa="Hold position"; _still_v.append(_ms)
                    elif _chk==2:
                        _st="🟡 WEAKENING"; _sc="#d97706"; _sb="#fffbeb"; _sa="Consider 50% exit"; _weaken.append(_ms)
                    else:
                        _st="🔴 EXPIRED"; _sc="#dc2626"; _sb="#fef2f2"; _sa="Signal no longer valid"; _expd.append(_ms)

                    _rsi_disp = f"RSI {_ms_sig['rv']:.0f}" if _ms_sig else "—"
                    st.markdown(
                        f"<div style='background:{_sb};border:1.5px solid {_sc};"
                        f"border-radius:10px;padding:10px 14px;margin-bottom:5px;"
                        f"display:flex;justify-content:space-between;"
                        f"align-items:center;flex-wrap:wrap;gap:6px'>"
                        f"<div><span style='font-size:14px;font-weight:700;"
                        f"color:#1e293b'>{_ms['stock']}</span>"
                        f"<span style='font-size:11px;color:#64748b;"
                        f"margin-left:8px'>{'BUY CE' if _is_ce else 'BUY PE'} | "
                        f"{_rsi_disp}</span></div>"
                        f"<div style='font-size:13px;font-weight:700;"
                        f"color:{_sc}'>{_st}<br>"
                        f"<span style='font-size:11px;font-weight:400;"
                        f"color:#64748b'>{_sa}</span></div>"
                        f"<div style='text-align:right;"
                        f"font-size:13px;font-weight:700;color:{_pnl_col}'>"
                        f"₹{_ms_cp:,.2f}<br>"
                        f"<span style='font-size:11px'>{_pnl:+.2f} pts</span>"
                        f"</div></div>",
                        unsafe_allow_html=True
                    )

                st.markdown(
                    f"<div style='background:#f8fafc;border-radius:8px;"
                    f"padding:8px 14px;font-size:13px;margin-top:6px'>"
                    f"🟢 Valid: <b>{len(_still_v)}</b> &nbsp;|&nbsp; "
                    f"🟡 Weakening: <b>{len(_weaken)}</b> &nbsp;|&nbsp; "
                    f"🔴 Expired/SL: <b>{len(_expd)}</b>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                if st.button("🗑️ Clear Monitor", key="monitor_clear"):
                    st.session_state.pop("monitor_stocks",None)
                    st.rerun()


            if results_sorted:
                st.markdown("---")
                import plotly.graph_objects as go_scan
                fig_scan = go_scan.Figure(go_scan.Bar(
                    x=[r["Stock"] for r in results_sorted],
                    y=[r["Combined"] for r in results_sorted],
                    marker_color=["#16a34a" if r["Direction"]=="UPTREND"
                                  else "#dc2626" for r in results_sorted],
                    text=[f"{r['Combined']}/10" for r in results_sorted],
                    textposition="outside",
                ))
                fig_scan.add_hline(y=8, line_dash="dash", line_color="#16a34a",
                                   annotation_text="Strong (8+)")
                fig_scan.add_hline(y=6, line_dash="dot", line_color="#f59e0b",
                                   annotation_text="Good (6+)")
                fig_scan.update_layout(
                    template="plotly_white", height=300,
                    yaxis_range=[0,11],
                    margin=dict(l=10,r=10,t=20,b=60),
                    xaxis_tickangle=-35, showlegend=False,
                    title="Combined Score (60% Technical + 40% Historical)"
                )
                st.plotly_chart(fig_scan, use_container_width=True)

        # Auto scan loop
        if auto_scan:
            st.info("Next scan in 5 minutes...")
            time.sleep(300)
            st.rerun()

    else:
        # Default screen
        st.markdown("""
        ### How the auto scanner works

        1. Select a stock group (up to 30 stocks)
        2. Choose timeframe — **15m** for intraday signals
        3. Click **Scan All Stocks Now**
        4. Results appear sorted by score in under 60 seconds
        5. Strong signals (8+) show full entry, SL and targets
        6. Click **Analyse** on any signal to deep-dive into that stock

        ### Telegram alerts

        Setup Telegram alerts above so your phone buzzes
        automatically when any stock hits a score of 8 or above —
        even when you are not watching the screen.

        ### Best times to scan

        | Time | What to do |
        |------|------------|
        | 9:30 AM | First scan of day — find morning setups |
        | 11:00 AM | Mid-morning scan — confirms trends |
        | 1:30 PM | Afternoon scan — fresh setups |
        | Enable auto scan | Scanner runs every 5 min automatically |
        """)


with T4:
    st.markdown("### 🤖 ML Prediction Engine")
    st.caption(
        "Trained on historical candle data using "
        "Random Forest + Gradient Boosting. "
        "Predicts next 3-candle direction."
    )

    # ── Inline stock search ───────────────────────────────
    st.markdown(
        "<div style='background:#f0f9ff;border:1px solid "
        "#bae6fd;border-radius:10px;padding:10px 14px;"
        "margin-bottom:12px'>"
        "<span style='font-size:13px;color:#0369a1;"
        "font-weight:600'>🔍 Search or browse stocks</span>"
        "</div>",
        unsafe_allow_html=True
    )
    inline_stock_search("t4")
    st.markdown("---")

    # ── Stock info bar ─────────────────────────────────
    ml_lp = live_price(stick)
    pc_ml = "#16a34a" if ml_lp["chg"] >= 0 else "#dc2626"
    arr_ml= "▲" if ml_lp["chg"] >= 0 else "▼"

    if ml_lp["ok"]:
        st.markdown(
            f"<div style='background:#1e3a5f;"
            f"border-radius:10px;padding:12px 18px;"
            f"display:flex;justify-content:space-between;"
            f"align-items:center;margin-bottom:12px'>"
            f"<span style='color:#fff;font-size:16px;"
            f"font-weight:700'>{sname}</span>"
            f"<span style='color:#fff;font-size:18px;"
            f"font-weight:700'>₹{ml_lp['p']:,.2f}</span>"
            f"<span style='color:{pc_ml};font-weight:600'>"
            f"{arr_ml}{abs(ml_lp['chg']):.2f}%</span>"
            f"<span style='color:#93c5fd;font-size:12px'>"
            f"{stick}</span>"
            f"</div>",
            unsafe_allow_html=True
        )

    # ── Controls ───────────────────────────────────────
    ctl1, ctl2, ctl3 = st.columns([2, 2, 2])

    with ctl1:
        ml_tf_opt = st.selectbox(
            "Timeframe",
            ["15m", "30m", "1h", "1d"],
            index=3,
            key="ml_tf_select",
            help="1d gives most candles = better accuracy"
        )
    with ctl2:
        run_ml = st.button(
            "🚀 Train & Predict",
            type="primary",
            key="run_ml",
            use_container_width=True
        )
    with ctl3:
        # Clear old results when stock changes
        if st.button(
            "🔄 Reset",
            key="ml_reset",
            use_container_width=True,
            help="Clear old results and start fresh"
        ):
            for k in ["ml_result","ml_model_data",
                      "ml_stock","ml_tf"]:
                st.session_state.pop(k, None)
            st.rerun()

    # Clear cached results if stock or timeframe changed
    prev_stock = st.session_state.get("ml_stock", "")
    prev_tf    = st.session_state.get("ml_tf", "")
    if prev_stock != sname or prev_tf != ml_tf_opt:
        for k in ["ml_result", "ml_model_data"]:
            st.session_state.pop(k, None)
        st.session_state["ml_stock"] = sname
        st.session_state["ml_tf"]    = ml_tf_opt

    # ── Load data and run ──────────────────────────────
    if run_ml:
        with st.spinner(
            f"Loading {sname} candle data "
            f"({ml_tf_opt} timeframe)..."
        ):
            ml_df = candles(stick, ml_tf_opt)

        if ml_df.empty or len(ml_df) < 100:
            st.error(
                f"Only {len(ml_df)} candles available for "
                f"{sname} on {ml_tf_opt}. "
                f"Need at least 100. "
                f"Switch to **1d** timeframe."
            )
        else:
            st.caption(
                f"✅ {len(ml_df)} candles loaded | "
                f"Last: "
                f"{ml_df.index[-1].strftime('%d %b %H:%M')} "
                f"IST | Timeframe: {ml_tf_opt}"
            )

            # Real-time approximation
            st.markdown("---")
            st.markdown("#### ⚡ Real-Time Approximation")
            st.caption(
                "Live price injected into indicators "
                "to bridge the 15-minute delay."
            )

            rt = approximate_realtime(
                ml_df, ml_lp["p"] if ml_lp["ok"] else 0
            )

            if rt.get("ok"):
                ra, rb, rc_ = st.columns(3)
                bc = rt["bias_color"]

                with ra:
                    bias_bg = (
                        "#f0fdf4" if rt["live_bias"]=="BULLISH"
                        else "#fef2f2"
                        if rt["live_bias"]=="BEARISH"
                        else "#fffbeb"
                    )
                    st.markdown(
                        f"<div style='background:{bias_bg};"
                        f"border-radius:10px;padding:16px;"
                        f"text-align:center'>"
                        f"<div style='font-size:11px;"
                        f"color:#64748b;text-transform:uppercase;"
                        f"letter-spacing:1px'>Live Bias</div>"
                        f"<div style='font-size:28px;"
                        f"font-weight:700;color:{bc}'>"
                        f"{rt['live_bias']}</div>"
                        f"<div style='font-size:12px;"
                        f"color:#64748b;margin-top:4px'>"
                        f"{rt['bull_count']} bull / "
                        f"{rt['bear_count']} bear</div>"
                        f"</div>",
                        unsafe_allow_html=True
                    )

                with rb:
                    sc_col = (
                        "#16a34a"
                        if rt["since_close"] >= 0
                        else "#dc2626"
                    )
                    st.markdown(
                        f"<div style='background:#f8fafc;"
                        f"border-radius:10px;padding:16px;"
                        f"text-align:center'>"
                        f"<div style='font-size:11px;"
                        f"color:#64748b;text-transform:uppercase;"
                        f"letter-spacing:1px'>"
                        f"Live vs Last Candle</div>"
                        f"<div style='font-size:28px;"
                        f"font-weight:700;color:{sc_col}'>"
                        f"{rt['since_close']:+.3f}%</div>"
                        f"<div style='font-size:12px;"
                        f"color:#64748b;margin-top:4px'>"
                        f"Candle pos: {rt['candle_pos']:.0f}% "
                        f"({rt['candle_zone']})<br>"
                        f"Micro trend: "
                        f"<b>{rt['micro_trend']}</b></div>"
                        f"</div>",
                        unsafe_allow_html=True
                    )

                with rc_:
                    st.markdown(
                        f"<div style='background:#f8fafc;"
                        f"border-radius:10px;padding:16px'>"
                        f"<div style='font-size:11px;"
                        f"color:#64748b;text-transform:uppercase;"
                        f"letter-spacing:1px;margin-bottom:8px'>"
                        f"Live Indicators</div>"
                        f"<table style='width:100%;"
                        f"font-size:13px'>"
                        f"<tr><td style='color:#64748b'>"
                        f"RSI</td><td style='text-align:right;"
                        f"font-weight:600;color:#1e293b'>"
                        f"{rt['rsi_live']}</td></tr>"
                        f"<tr><td style='color:#64748b'>"
                        f"EMA9</td><td style='text-align:right;"
                        f"font-weight:600;color:#ca8a04'>"
                        f"₹{rt['ema9_live']:,}</td></tr>"
                        f"<tr><td style='color:#64748b'>"
                        f"EMA21</td><td style='text-align:right;"
                        f"font-weight:600;color:#ea580c'>"
                        f"₹{rt['ema21_live']:,}</td></tr>"
                        f"<tr><td style='color:#64748b'>"
                        f"VWAP</td><td style='text-align:right;"
                        f"font-weight:600;color:#1e293b'>"
                        f"₹{rt['vwap_live']:,}</td></tr>"
                        f"<tr><td style='color:#64748b'>"
                        f"VWAP Dev</td>"
                        f"<td style='text-align:right;"
                        f"font-weight:600;"
                        f"color:{'#16a34a' if rt['vwap_dev']>0 else '#dc2626'}'>"
                        f"{rt['vwap_dev']:+.2f}%</td></tr>"
                        f"</table></div>",
                        unsafe_allow_html=True
                    )

                # Live signals
                st.markdown("#### 📡 Live Signal Breakdown")
                for sig_txt in rt["live_signals"]:
                    c_ = ("#f0fdf4" if "✅" in sig_txt
                          else "#fef2f2" if "❌" in sig_txt
                          else "#fffbeb")
                    st.markdown(
                        f"<div style='background:{c_};"
                        f"border-radius:6px;padding:8px 14px;"
                        f"margin:3px 0;font-size:13px;"
                        f"color:#374151'>{sig_txt}</div>",
                        unsafe_allow_html=True
                    )

            # ML training
            st.markdown("---")
            st.markdown("#### 🧠 ML Prediction")

            with st.spinner(
                "Training Random Forest + "
                "Gradient Boosting model..."
            ):
                model_data = train_model(ml_df)

            if not model_data.get("ok"):
                st.error(
                    f"Training failed: "
                    f"{model_data.get('reason','Unknown error')}"
                )
            else:
                pred = predict_next_move(ml_df, model_data)
                st.session_state["ml_result"]     = pred
                st.session_state["ml_model_data"] = model_data

                if pred and pred.get("ok"):
                    pc_   = pred["sig_color"]
                    conf  = pred["confidence"]

                    # Prediction card
                    pred_bg = (
                        "#f0fdf4" if pred["prediction"]=="UPTREND"
                        else "#fef2f2"
                        if pred["prediction"]=="DOWNTREND"
                        else "#fffbeb"
                    )
                    st.markdown(
                        f"<div style='background:{pred_bg};"
                        f"border:2px solid {pc_};"
                        f"border-radius:14px;padding:24px;"
                        f"text-align:center;margin:12px 0'>"
                        f"<div style='font-size:12px;"
                        f"color:#64748b;letter-spacing:2px;"
                        f"text-transform:uppercase'>"
                        f"ML PREDICTION — NEXT 3 CANDLES</div>"
                        f"<div style='font-size:52px;"
                        f"font-weight:700;color:{pc_};"
                        f"line-height:1.1;margin:8px 0'>"
                        f"{pred['prediction']}</div>"
                        f"<div style='font-size:26px;"
                        f"font-weight:600;color:{pc_}'>"
                        f"{pred['signal']}</div>"
                        f"<div style='font-size:15px;"
                        f"color:#64748b;margin-top:8px'>"
                        f"Confidence: {pred['reliability']} "
                        f"({conf}%)</div>"
                        f"<div style='margin-top:12px;"
                        f"display:flex;justify-content:center;"
                        f"gap:12px;flex-wrap:wrap'>"
                        f"<span style='background:#dcfce7;"
                        f"color:#16a34a;padding:4px 14px;"
                        f"border-radius:12px;font-size:13px'>"
                        f"📈 Uptrend: "
                        f"{pred['probabilities'].get('UPTREND',0):.1f}%"
                        f"</span>"
                        f"<span style='background:#fee2e2;"
                        f"color:#dc2626;padding:4px 14px;"
                        f"border-radius:12px;font-size:13px'>"
                        f"📉 Downtrend: "
                        f"{pred['probabilities'].get('DOWNTREND',0):.1f}%"
                        f"</span>"
                        f"<span style='background:#fef9c3;"
                        f"color:#854d0e;padding:4px 14px;"
                        f"border-radius:12px;font-size:13px'>"
                        f"➡️ Sideways: "
                        f"{pred['probabilities'].get('SIDEWAYS',0):.1f}%"
                        f"</span>"
                        f"</div></div>",
                        unsafe_allow_html=True
                    )

                    # Stats
                    ms1, ms2, ms3 = st.columns(3)
                    ms1.metric(
                        "Model Accuracy",
                        f"{pred['model_accuracy']}%"
                        if pred["model_accuracy"] else "N/A"
                    )
                    ms2.metric(
                        "Trained on",
                        f"{pred['n_trained']} candles"
                    )
                    ms3.metric("Timeframe", ml_tf_opt)

                    # Probability chart
                    st.markdown("#### 📊 Probability Distribution")
                    probs = pred["probabilities"]
                    import plotly.graph_objects as go_ml
                    fig_prob = go_ml.Figure(go_ml.Bar(
                        x=list(probs.keys()),
                        y=list(probs.values()),
                        marker_color=[
                            "#16a34a" if k=="UPTREND"
                            else "#dc2626" if k=="DOWNTREND"
                            else "#f59e0b"
                            for k in probs.keys()
                        ],
                        text=[f"{v:.1f}%" for v in probs.values()],
                        textposition="outside"
                    ))
                    fig_prob.update_layout(
                        template="plotly_white",
                        height=260,
                        yaxis_range=[0, 100],
                        yaxis_title="Probability %",
                        margin=dict(l=10,r=10,t=10,b=10)
                    )
                    st.plotly_chart(
                        fig_prob, use_container_width=True
                    )

                    # Top features
                    st.markdown("#### 🔬 Top Features Used")
                    for feat in pred["top_contrib"]:
                        imp = feat["importance"]
                        fc  = ("#16a34a" if imp >= 10
                               else "#f59e0b" if imp >= 5
                               else "#94a3b8")
                        st.markdown(
                            f"<div style='background:#f8fafc;"
                            f"border-radius:6px;padding:10px 14px;"
                            f"margin:4px 0'>"
                            f"<div style='display:flex;"
                            f"justify-content:space-between'>"
                            f"<span style='color:#475569;"
                            f"font-size:13px'>{feat['feature']}"
                            f"</span>"
                            f"<span style='color:{fc};"
                            f"font-size:13px;font-weight:600'>"
                            f"{imp:.1f}% importance</span></div>"
                            f"<div style='background:#e2e8f0;"
                            f"height:4px;border-radius:2px;"
                            f"margin-top:6px'>"
                            f"<div style='background:{fc};"
                            f"width:{min(imp*5,100):.0f}%;"
                            f"height:4px;border-radius:2px'>"
                            f"</div></div></div>",
                            unsafe_allow_html=True
                        )

                    # Combined signal
                    st.markdown("---")
                    st.markdown("### 🎯 Combined Signal")

                    ml_dir   = pred["prediction"]
                    tech_sig = compute_all(ml_df, ml_lp)
                    tech_dir = (tech_sig["direction"]
                                if tech_sig else "SIDEWAYS")
                    rt_bias  = (rt.get("live_bias","NEUTRAL")
                                if rt and rt.get("ok")
                                else "NEUTRAL")

                    all_bull = (
                        ml_dir   == "UPTREND" and
                        tech_dir == "UPTREND" and
                        rt_bias  == "BULLISH"
                    )
                    all_bear = (
                        ml_dir   == "DOWNTREND" and
                        tech_dir == "DOWNTREND" and
                        rt_bias  == "BEARISH"
                    )

                    if all_bull and tech_sig:
                        st.success(
                            f"🔥 ALL THREE AGREE — BULLISH\n\n"
                            f"ML: UPTREND ({conf}%) ✅ | "
                            f"Technical: {tech_sig['up_score']}/10 ✅ | "
                            f"Live: BULLISH ✅\n\n"
                            f"Highest quality CE setup. "
                            f"Enter on pullback to "
                            f"EMA9 (₹{tech_sig['e9v']:,}) "
                            f"with SL ₹{tech_sig['sl_long']:,}"
                        )
                    elif all_bear and tech_sig:
                        st.error(
                            f"🔥 ALL THREE AGREE — BEARISH\n\n"
                            f"ML: DOWNTREND ({conf}%) ✅ | "
                            f"Technical: {tech_sig['dn_score']}/10 ✅ | "
                            f"Live: BEARISH ✅\n\n"
                            f"Highest quality PE setup. "
                            f"Enter on bounce to "
                            f"EMA9 (₹{tech_sig['e9v']:,}) "
                            f"with SL ₹{tech_sig['sl_short']:,}"
                        )
                    elif ml_dir == tech_dir and ml_dir != "SIDEWAYS":
                        col_ = ("#16a34a"
                                if ml_dir == "UPTREND"
                                else "#dc2626")
                        st.markdown(
                            f"<div style='background:#f8fafc;"
                            f"border:1px solid #e2e8f0;"
                            f"border-radius:10px;"
                            f"padding:14px 18px'>"
                            f"<b style='color:{col_}'>"
                            f"⚡ ML + Technical agree: {ml_dir}"
                            f"</b><br>"
                            f"<span style='color:#64748b;"
                            f"font-size:13px'>"
                            f"Live bias is {rt_bias}. "
                            f"Wait for it to confirm "
                            f"before entering.</span></div>",
                            unsafe_allow_html=True
                        )
                    else:
                        st.warning(
                            f"⚠️ Mixed signals — "
                            f"ML says {ml_dir} but "
                            f"Technical says {tech_dir}. "
                            "Do not trade when signals conflict."
                        )

    elif "ml_result" in st.session_state:
        # Show cached result with note
        st.info(
            f"Showing previous result for "
            f"{st.session_state.get('ml_stock', sname)}. "
            f"Click **Train & Predict** to get fresh prediction "
            f"for {sname}."
        )
        pred = st.session_state["ml_result"]
        if pred and pred.get("ok"):
            pc_ = pred["sig_color"]
            st.markdown(
                f"<div style='background:#f8fafc;"
                f"border:2px solid {pc_};"
                f"border-radius:12px;padding:20px;"
                f"text-align:center'>"
                f"<div style='font-size:42px;font-weight:700;"
                f"color:{pc_}'>{pred['prediction']}</div>"
                f"<div style='font-size:20px;color:{pc_}'>"
                f"{pred['signal']}</div>"
                f"<div style='font-size:14px;color:#64748b;"
                f"margin-top:6px'>"
                f"Confidence: {pred['confidence']}%</div>"
                f"</div>",
                unsafe_allow_html=True
            )
    else:
        st.info(
            f"Stock selected: **{sname}**  \n\n"
            f"Click **🚀 Train & Predict** above to run "
            f"the ML model and get a prediction."
        )

        with st.expander("📖 How ML prediction works"):
            st.markdown("""
            The model trains on YOUR stock's own historical
            data — not generic market data. It learns the
            specific patterns of that stock by analyzing
            30+ technical features.

            | Level | Condition | Action |
            |-------|-----------|--------|
            | **Best** | ML + Technical + Live agree | Enter trade |
            | **Good** | ML + Technical agree | Wait for live |
            | **Avoid** | ML and Technical disagree | No trade |

            **Use 1d timeframe** for best accuracy —
            more candles = better trained model.
            """)


with T5:
    st.markdown("### 🏦 Smart Money & Institutional Analysis")

    if "sig" not in dir() or sig is None:
        st.info("Select a stock in Tab 2 first.")
    else:
        sm1, sm2 = st.columns(2)

        with sm1:
            st.markdown("#### 📊 Money Flow Indicators")
            cmf_col = ("#00ff88" if sig["cmfv"] > 0.05
                       else "#ff4455" if sig["cmfv"] < -0.05
                       else "#ffcc00")
            st.markdown(f"""
            <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)'>
              <div style='font-size:11px;color:#64748b;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px'>CHAIKIN MONEY FLOW</div>
              <div style='font-size:40px;font-weight:700;
                          color:{cmf_col}'>
                  {sig['cmfv']:+.3f}
              </div>
              <div style='color:#64748b;font-size:13px;
                          margin-top:4px'>
                  {'🟢 Institutions BUYING' if sig['cmfv']>0.05
                   else '🔴 Institutions SELLING' if sig['cmfv']<-0.05
                   else '🟡 Neutral'}
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown(f"""
            <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)' style='margin-top:8px'>
              <table style='width:100%;font-size:13px'>
                <tr>
                  <td style='color:#555'>OBV Direction</td>
                  <td style='text-align:right;
                      color:{"#00ff88" if sig["obv_bull"] else "#ff4455"}'>
                      {"▲ Rising" if sig["obv_bull"] else "▼ Falling"}
                  </td>
                </tr>
                <tr>
                  <td style='color:#555'>Volume Ratio</td>
                  <td style='text-align:right;
                      color:{"#00ff88" if sig["vol_ratio"]>=1.5 else "#fff"}'>
                      {sig['vol_ratio']}× avg
                  </td>
                </tr>
                <tr>
                  <td style='color:#555'>VWAP Deviation</td>
                  <td style='text-align:right;color:#1e293b'>
                      {((sig['cp']-sig['vwv'])/sig['vwv']*100):+.2f}%
                  </td>
                </tr>
                <tr>
                  <td style='color:#555'>Liquidity Sweep↓</td>
                  <td style='text-align:right'>
                      {"🚀 YES — BUY" if sig['sweep_low'] else "—"}
                  </td>
                </tr>
                <tr>
                  <td style='color:#555'>Liquidity Sweep↑</td>
                  <td style='text-align:right'>
                      {"⚠️ YES — SELL" if sig['sweep_high'] else "—"}
                  </td>
                </tr>
                <tr>
                  <td style='color:#555'>BOS Bullish</td>
                  <td style='text-align:right;color:#00ff88'>
                      {"✅ YES" if sig['bos_bull'] else "—"}
                  </td>
                </tr>
                <tr>
                  <td style='color:#555'>BOS Bearish</td>
                  <td style='text-align:right;color:#ff4455'>
                      {"✅ YES" if sig['bos_bear'] else "—"}
                  </td>
                </tr>
              </table>
            </div>
            """, unsafe_allow_html=True)

        with sm2:
            st.markdown("#### 🕯️ Candlestick Patterns")
            if sig["patterns"]:
                for pname, pbias, pmeaning in sig["patterns"]:
                    pc_ = ("#00ff88" if pbias=="bullish"
                           else "#ff4455" if pbias=="bearish"
                           else "#ffcc00")
                    st.markdown(f"""
                    <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)'
                         style='border-left:3px solid {pc_};
                                margin-bottom:8px'>
                      <div style='color:{pc_};font-weight:600'>
                          {pname}
                      </div>
                      <div style='color:#6b7280;font-size:12px;
                                  margin-top:4px'>
                          {pmeaning}
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.caption("No strong patterns on current candle")

            st.markdown("#### 📈 Support & Resistance")
            st.markdown(f"""
            <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)'>
              <div style='display:flex;
                          justify-content:space-between;
                          margin-bottom:12px'>
                <div style='text-align:center'>
                  <div style='font-size:11px;color:#64748b;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px'>SUPPORT</div>
                  <div style='color:#00ff88;font-size:22px;
                              font-weight:700'>
                      ₹{sig['sup']:,}
                  </div>
                </div>
                <div style='text-align:center'>
                  <div style='font-size:11px;color:#64748b;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px'>CURRENT</div>
                  <div style='color:#1e293b;font-size:22px;
                              font-weight:700'>
                      ₹{sig['cp']:,}
                  </div>
                </div>
                <div style='text-align:center'>
                  <div style='font-size:11px;color:#64748b;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px'>RESISTANCE</div>
                  <div style='color:#ff4455;font-size:22px;
                              font-weight:700'>
                      ₹{sig['res']:,}
                  </div>
                </div>
              </div>
              <div style='font-size:13px;color:#555'>
                  Risk: ₹{sig['res']-sig['cp']:.1f} to resistance &nbsp;|&nbsp;
                  Reward: ₹{sig['cp']-sig['sup']:.1f} to support
              </div>
            </div>
            """, unsafe_allow_html=True)

        # CMF chart
        st.markdown("---")
        st.markdown("#### 📊 Money Flow Chart (CMF)")
        cmf_s = sig["cmfs"].tail(80)
        fig_c = go.Figure()
        fig_c.add_trace(go.Bar(
            x=df.tail(80).index, y=cmf_s,
            marker_color=["#00ff88" if v>=0 else "#ff4455"
                          for v in cmf_s],
            name="CMF", opacity=0.85
        ))
        fig_c.add_hline(y= 0.1, line_dash="dash",
                        line_color="#00ff88", opacity=0.5,
                        annotation_text="Inst. buying (0.1)")
        fig_c.add_hline(y=-0.1, line_dash="dash",
                        line_color="#ff4455", opacity=0.5,
                        annotation_text="Inst. selling (-0.1)")
        fig_c.add_hline(y=0, line_color="#333",
                        line_dash="dot")
        fig_c.update_layout(
            template="plotly_white", height=250,
            margin=dict(l=10,r=10,t=20,b=10),
            title="Chaikin Money Flow — "
                  "Green = institutions buying | "
                  "Red = institutions selling",
            yaxis_range=[-0.5,0.5]
        )
        st.plotly_chart(fig_c, width="stretch")

        with st.expander(
            "📖 How to read Smart Money signals"
        ):
            st.markdown("""
            | Signal | Meaning | Action |
            |--------|---------|--------|
            | **CMF > 0.1** | Big money flowing IN | 🟢 CE signal |
            | **CMF < -0.1** | Big money flowing OUT | 🔴 PE signal |
            | **OBV Rising** | Volume confirms price rise | 🟢 Bullish |
            | **OBV Falling** | Volume confirms price fall | 🔴 Bearish |
            | **Liquidity Sweep ↓** | Stop hunt below lows — institutions bought | 🚀 Strong CE |
            | **Liquidity Sweep ↑** | Stop hunt above highs — institutions sold | ⚠️ Strong PE |
            | **BOS Bullish** | Break of structure upward with volume | 🟢 CE confirmed |
            | **BOS Bearish** | Break of structure downward with volume | 🔴 PE confirmed |
            | **Vol Ratio > 2.0** | Institutional block trade | 📊 Watch direction |
            """)

# P&L Calculator tab removed
if False:
    st.markdown("### 🧮 Options P&L Calculator")
    st.caption(
        "Enter trade details to see exact profit, loss "
        "and breakeven before placing any trade"
    )

    strat = st.selectbox("Strategy", [
        "📈 Buy CE (Bullish)",
        "📉 Buy PE (Bearish)",
        "🔀 Bull Call Spread",
        "🔀 Bear Put Spread",
        "🦋 Long Straddle",
    ])

    st.markdown("---")
    pi1,pi2,pi3 = st.columns(3)
    with pi1:
        sp  = st.number_input(
            "Spot Price ₹",
            value=float(lp["p"]) if lp["ok"] else 22500.0,
            step=50.0)
        iv  = st.slider("IV %", 5, 80, 15) / 100
    with pi2:
        sk  = st.number_input(
            "Strike ₹",
            value=round((float(lp["p"])
                         if lp["ok"] else 22500)/50)*50.0,
            step=50.0)
        dte = st.number_input("Days to Expiry",0,90,7)
    with pi3:
        lots  = st.number_input("Lots",1,50,1)
        lsz   = st.number_input(
            "Lot Size",
            value=int(LOT_SIZES.get(stick,25)),
            min_value=1)

    T_yr  = max(dte/365, 0.001)
    r_r   = 0.065
    units = lots * lsz

    def pl_chart(prices, pnls, color, be,
                 spot_v, title="P&L"):
        fig_ = go.Figure()
        fig_.add_trace(go.Scatter(
            x=list(prices), y=pnls,
            line=dict(color=color,width=2),
            fill="tozeroy",
            fillcolor=f"rgba({'0,255,100' if color=='lime' else '255,60,80'},0.06)",
            name="P&L"))
        fig_.add_hline(y=0,line_dash="dash",
                       line_color="white",opacity=0.3)
        fig_.add_vline(x=be,line_dash="dot",
                       line_color="yellow",
                       annotation_text=f"BE ₹{be:,.0f}",
                       annotation_position="top right")
        fig_.add_vline(x=spot_v,line_dash="dot",
                       line_color="cyan",
                       annotation_text=f"Spot ₹{spot_v:,.0f}",
                       annotation_position="top left")
        fig_.update_layout(
            template="plotly_white",height=320,
            xaxis_title="Spot at Expiry ₹",
            yaxis_title="P&L ₹",
            margin=dict(l=10,r=10,t=20,b=10))
        return fig_

    if strat == "📈 Buy CE (Bullish)":
        prem = st.number_input(
            "CE Premium ₹",
            value=float(bs(sp,sk,T_yr,r_r,iv,"CE")),
            step=0.5)
        tc   = prem*units
        be_  = sk+prem
        g    = greeks(sp,sk,T_yr,r_r,iv,"CE")
        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Total Cost",  f"₹{tc:,.0f}")
        m2.metric("Breakeven",   f"₹{be_:,.1f}")
        m3.metric("Max Loss",    f"₹{tc:,.0f}")
        m4.metric("Max Profit",  "Unlimited ∞")
        m5.metric("Delta",       g["d"])
        g1,g2,g3,g4 = st.columns(4)
        g1.metric("Delta",      g["d"])
        g2.metric("Gamma",      g["g"])
        g3.metric("Theta/day",  f"₹{abs(g['th']*units):,.1f}")
        g4.metric("Vega/1%IV",  f"₹{g['v']*units:,.1f}")
        prices = np.linspace(sp*0.88,sp*1.12,200)
        pnls_  = [(max(p-sk,0)-prem)*units for p in prices]
        st.plotly_chart(
            pl_chart(prices,pnls_,"lime",be_,sp),
            width="stretch")
        st.info(f"""
        💡 Max risk ₹{tc:,.0f} | Need > ₹{be_:,.1f} to profit |
        Theta burns ₹{abs(g['th']*units):,.1f}/day |
        Best exit at 60–80% profit
        """)

    elif strat == "📉 Buy PE (Bearish)":
        prem = st.number_input(
            "PE Premium ₹",
            value=float(bs(sp,sk,T_yr,r_r,iv,"PE")),
            step=0.5)
        tc  = prem*units
        be_ = sk-prem
        g   = greeks(sp,sk,T_yr,r_r,iv,"PE")
        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Total Cost",  f"₹{tc:,.0f}")
        m2.metric("Breakeven",   f"₹{be_:,.1f}")
        m3.metric("Max Loss",    f"₹{tc:,.0f}")
        m4.metric("Max Profit",  f"₹{be_*units:,.0f}")
        m5.metric("Delta",       g["d"])
        g1,g2,g3,g4 = st.columns(4)
        g1.metric("Delta",      g["d"])
        g2.metric("Gamma",      g["g"])
        g3.metric("Theta/day",  f"₹{abs(g['th']*units):,.1f}")
        g4.metric("Vega/1%IV",  f"₹{g['v']*units:,.1f}")
        prices = np.linspace(sp*0.88,sp*1.12,200)
        pnls_  = [(max(sk-p,0)-prem)*units for p in prices]
        st.plotly_chart(
            pl_chart(prices,pnls_,"#ff4455",be_,sp),
            width="stretch")
        st.info(f"""
        💡 Max risk ₹{tc:,.0f} | Need < ₹{be_:,.1f} to profit |
        Theta burns ₹{abs(g['th']*units):,.1f}/day
        """)

    elif strat == "🔀 Bull Call Spread":
        ca,cb = st.columns(2)
        with ca:
            bk = st.number_input("Buy CE Strike",
                                 value=sk,step=50.0)
            bp = st.number_input(
                "Buy CE Premium",
                value=float(bs(sp,bk,T_yr,r_r,iv,"CE")),
                step=0.5)
        with cb:
            sk2= st.number_input("Sell CE Strike",
                                  value=sk+100,step=50.0)
            sp2= st.number_input(
                "Sell CE Premium",
                value=float(bs(sp,sk2,T_yr,r_r,iv,"CE")),
                step=0.5)
        net  = (bp-sp2)*units
        maxp = (sk2-bk-bp+sp2)*units
        be_  = bk+bp-sp2
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Net Cost",   f"₹{net:,.0f}")
        m2.metric("Breakeven",  f"₹{be_:,.1f}")
        m3.metric("Max Loss",   f"₹{net:,.0f}")
        m4.metric("Max Profit", f"₹{maxp:,.0f}")
        prices = np.linspace(sp*0.9,sp*1.1,200)
        pnls_  = [(max(p-bk,0)-max(p-sk2,0)
                   -bp+sp2)*units for p in prices]
        st.plotly_chart(
            pl_chart(prices,pnls_,"lime",be_,sp),
            width="stretch")
        st.success(f"Pay ₹{net:,.0f} | Max profit ₹{maxp:,.0f} above ₹{sk2:,.0f}")

    elif strat == "🔀 Bear Put Spread":
        ca,cb = st.columns(2)
        with ca:
            bk = st.number_input("Buy PE Strike",
                                 value=sk,step=50.0)
            bp = st.number_input(
                "Buy PE Premium",
                value=float(bs(sp,bk,T_yr,r_r,iv,"PE")),
                step=0.5)
        with cb:
            sk2= st.number_input("Sell PE Strike",
                                  value=sk-100,step=50.0)
            sp2= st.number_input(
                "Sell PE Premium",
                value=float(bs(sp,sk2,T_yr,r_r,iv,"PE")),
                step=0.5)
        net  = (bp-sp2)*units
        maxp = (bk-sk2-bp+sp2)*units
        be_  = bk-bp+sp2
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Net Cost",   f"₹{net:,.0f}")
        m2.metric("Breakeven",  f"₹{be_:,.1f}")
        m3.metric("Max Loss",   f"₹{net:,.0f}")
        m4.metric("Max Profit", f"₹{maxp:,.0f}")
        prices = np.linspace(sp*0.9,sp*1.1,200)
        pnls_  = [(max(bk-p,0)-max(sk2-p,0)
                   -bp+sp2)*units for p in prices]
        st.plotly_chart(
            pl_chart(prices,pnls_,"#ff4455",be_,sp),
            width="stretch")
        st.error(f"Pay ₹{net:,.0f} | Max profit ₹{maxp:,.0f} below ₹{sk2:,.0f}")

    elif strat == "🦋 Long Straddle":
        cep = st.number_input(
            "CE Premium",
            value=float(bs(sp,sk,T_yr,r_r,iv,"CE")),
            step=0.5)
        pep = st.number_input(
            "PE Premium",
            value=float(bs(sp,sk,T_yr,r_r,iv,"PE")),
            step=0.5)
        tc   = (cep+pep)*units
        ube  = sk+cep+pep
        lbe  = sk-cep-pep
        mv_n = round((ube-sp)/sp*100,1)
        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Total Cost",   f"₹{tc:,.0f}")
        m2.metric("Upper BE",     f"₹{ube:,.1f}")
        m3.metric("Lower BE",     f"₹{lbe:,.1f}")
        m4.metric("Max Loss",     f"₹{tc:,.0f}")
        m5.metric("Move Needed",  f"{mv_n}%")
        prices = np.linspace(sp*0.85,sp*1.15,200)
        pnls_  = [(max(p-sk,0)+max(sk-p,0)
                   -cep-pep)*units for p in prices]
        fig_st = go.Figure()
        fig_st.add_trace(go.Scatter(
            x=list(prices),y=pnls_,
            line=dict(color="violet",width=2),
            fill="tozeroy",
            fillcolor="rgba(150,0,200,0.06)",
            name="P&L"))
        fig_st.add_hline(y=0,line_dash="dash",
                         line_color="white",opacity=0.3)
        for bv,lb in [(ube,"Upper BE"),(lbe,"Lower BE")]:
            fig_st.add_vline(x=bv,line_dash="dot",
                             line_color="lime",
                             annotation_text=f"{lb} ₹{bv:,.0f}")
        fig_st.update_layout(
            template="plotly_white",height=300,
            xaxis_title="Spot ₹",yaxis_title="P&L ₹",
            margin=dict(l=10,r=10,t=10,b=10))
        st.plotly_chart(fig_st, width="stretch")
        st.warning(f"""
        Need {mv_n}% move either direction to profit.
        Best used BEFORE: RBI Policy | Results | Budget | US Fed.
        Exit immediately AFTER the event — IV collapses!
        """)

    # Lot size reference
    with st.expander("📋 NSE Lot Size Reference"):
        lst = pd.DataFrame([
            {"Index/Stock":"NIFTY 50","Lot":25,"Approx Margin":"₹1.2L"},
            {"Index/Stock":"BANK NIFTY","Lot":15,"Approx Margin":"₹45K"},
            {"Index/Stock":"Reliance","Lot":250,"Approx Margin":"₹60K"},
            {"Index/Stock":"TCS","Lot":150,"Approx Margin":"₹55K"},
            {"Index/Stock":"HDFC Bank","Lot":550,"Approx Margin":"₹90K"},
            {"Index/Stock":"Infosys","Lot":300,"Approx Margin":"₹50K"},
            {"Index/Stock":"SBI","Lot":1500,"Approx Margin":"₹95K"},
            {"Index/Stock":"Bajaj Finance","Lot":125,"Approx Margin":"₹85K"},
            {"Index/Stock":"ITC","Lot":3200,"Approx Margin":"₹75K"},
            {"Index/Stock":"Tata Motors DVR","Lot":1425,"Approx Margin":"₹90K"},
        ])
        st.dataframe(lst, width="stretch", hide_index=True)


# ── Risk Calculator ───────────────────────────────────────
    st.markdown("---")
    st.markdown("### 💰 Risk Calculator — Position Sizing")
    st.caption(
        "Never risk more than 2% of your capital per trade. "
        "This calculator tells you exact lot size to buy."
    )

    rc1, rc2, rc3 = st.columns(3)
    with rc1:
        rc_capital = st.number_input(
            "Your total capital (Rs)",
            value=100000,
            step=10000,
            min_value=10000,
            key="rc_capital"
        )
        rc_risk_pct = st.slider(
            "Risk per trade (%)",
            min_value=0.5,
            max_value=5.0,
            value=2.0,
            step=0.5,
            key="rc_risk_pct",
            help="Professional traders risk max 2% per trade"
        )
    with rc2:
        rc_entry = st.number_input(
            "Entry price (Rs)",
            value=float(sig["cp"]) if sig else 100.0,
            step=0.5,
            min_value=0.1,
            key="rc_entry"
        )
        rc_sl = st.number_input(
            "Stop loss price (Rs)",
            value=float(sig["sl_long"]) if sig else 90.0,
            step=0.5,
            min_value=0.1,
            key="rc_sl"
        )
    with rc3:
        rc_lot = st.number_input(
            "Lot size (shares per lot)",
            value=50,
            step=1,
            min_value=1,
            key="rc_lot",
            help="Check NSE website for current lot size"
        )
        rc_target = st.number_input(
            "Target price (Rs)",
            value=float(sig["tgt1"]) if sig else 120.0,
            step=0.5,
            key="rc_target"
        )

    if st.button(
        "Calculate Position Size",
        type="primary",
        key="rc_calc",
        use_container_width=True
    ):
        rc_risk_amt    = rc_capital * rc_risk_pct / 100
        rc_risk_per_sh = abs(rc_entry - rc_sl)
        rc_reward      = abs(rc_target - rc_entry)
        rc_rr          = round(rc_reward / (rc_risk_per_sh + 0.001), 2)

        if rc_risk_per_sh <= 0:
            st.error("Entry and Stop Loss cannot be the same price")
        else:
            rc_shares    = int(rc_risk_amt / rc_risk_per_sh)
            rc_lots      = max(1, rc_shares // rc_lot)
            rc_act_shares= rc_lots * rc_lot
            rc_invest    = rc_act_shares * rc_entry
            rc_max_loss  = rc_act_shares * rc_risk_per_sh
            rc_max_gain  = rc_act_shares * rc_reward

            rm1, rm2, rm3, rm4 = st.columns(4)
            rm1.metric(
                "Lots to buy",
                f"{rc_lots} lots",
                delta=f"{rc_act_shares} shares"
            )
            rm2.metric(
                "Capital needed",
                f"₹{rc_invest:,.0f}",
                delta=f"{rc_invest/rc_capital*100:.1f}% of capital"
            )
            rm3.metric(
                "Max loss",
                f"₹{rc_max_loss:,.0f}",
                delta=f"{rc_max_loss/rc_capital*100:.1f}% of capital",
                delta_color="inverse"
            )
            rm4.metric(
                "Max profit",
                f"₹{rc_max_gain:,.0f}",
                delta=f"R:R {rc_rr}:1"
            )

            if rc_rr >= 1.5:
                st.success(
                    f"✅ Good trade setup — R:R {rc_rr}:1 | "
                    f"Buy {rc_lots} lots ({rc_act_shares} shares) | "
                    f"Risk ₹{rc_max_loss:,.0f} "
                    f"({rc_max_loss/rc_capital*100:.1f}% of capital)"
                )
            else:
                st.warning(
                    f"⚠️ R:R {rc_rr}:1 is below 1.5 — "
                    f"consider skipping this trade or "
                    f"moving target to improve R:R"
                )

    # ── Price Alerts ──────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🔔 Price Alerts — Get Telegram notification")
    st.caption(
        "Set a price level for any stock. "
        "The terminal checks every 2 minutes "
        "and sends Telegram alert when price is hit."
    )

    if not tg_configured():
        st.warning(
            "Setup Telegram in Auto Scanner tab first "
            "to receive price alerts."
        )
    else:
        pa1, pa2, pa3, pa4 = st.columns(4)
        with pa1:
            pa_stock = st.text_input(
                "Stock name",
                value=sname if sname else "NIFTY 50",
                key="pa_stock"
            )
            pa_sym = STOCKS.get(pa_stock, "^NSEI")
        with pa2:
            pa_direction = st.selectbox(
                "Alert when price",
                ["Goes ABOVE", "Goes BELOW"],
                key="pa_direction"
            )
        with pa3:
            pa_price = st.number_input(
                "Alert price (Rs)",
                value=float(sig["cp"]) if sig else 0.0,
                step=0.5,
                min_value=0.0,
                key="pa_price"
            )
        with pa4:
            pa_note = st.text_input(
                "Note (optional)",
                placeholder="e.g. Breakout level",
                key="pa_note"
            )

        if st.button(
            "Set Alert",
            type="primary",
            key="pa_set",
            use_container_width=True
        ):
            if "price_alerts" not in st.session_state:
                st.session_state["price_alerts"] = []
            st.session_state["price_alerts"].append({
                "stock":     pa_stock,
                "sym":       pa_sym,
                "direction": pa_direction,
                "price":     pa_price,
                "note":      pa_note,
                "triggered": False
            })
            st.success(
                f"✅ Alert set — will notify when "
                f"{pa_stock} {pa_direction.lower()} "
                f"₹{pa_price:,.2f}"
            )

        # Show active alerts
        alerts = st.session_state.get("price_alerts", [])
        active = [a for a in alerts if not a["triggered"]]

        if active:
            st.markdown(f"**Active alerts ({len(active)})**")
            for ai, alert in enumerate(active):
                ac1, ac2, ac3 = st.columns([3,2,1])
                ac1.markdown(
                    f"**{alert['stock']}** — "
                    f"{alert['direction']} ₹{alert['price']:,.2f}"
                    + (f" — {alert['note']}" if alert['note'] else "")
                )
                # Check current price
                _alp = live_price(alert["sym"])
                if _alp["ok"]:
                    ac2.markdown(
                        f"Current: **₹{_alp['p']:,.2f}**"
                    )
                    # Check if triggered
                    triggered = (
                        (alert["direction"] == "Goes ABOVE" and
                         _alp["p"] >= alert["price"]) or
                        (alert["direction"] == "Goes BELOW" and
                         _alp["p"] <= alert["price"])
                    )
                    if triggered and not alert["triggered"]:
                        alert["triggered"] = True
                        tok = st.session_state.get(
                            "tg_token_saved", ""
                        )
                        cid = st.session_state.get(
                            "tg_chat_saved", ""
                        )
                        msg = (
                            f"🔔 PRICE ALERT — {alert['stock']}\n"
                            f"Price {alert['direction'].lower()} "
                            f"₹{alert['price']:,.2f}\n"
                            f"Current price: ₹{_alp['p']:,.2f}\n"
                            + (f"Note: {alert['note']}"
                               if alert['note'] else "")
                        )
                        send_telegram(tok, cid, msg)
                        st.success(
                            f"🔔 Alert triggered! "
                            f"{alert['stock']} hit ₹{alert['price']:,.2f}"
                        )
                if ac3.button(
                    "Remove",
                    key=f"pa_rm_{ai}",
                    use_container_width=True
                ):
                    st.session_state["price_alerts"].pop(ai)
                    st.rerun()
        else:
            st.caption("No active price alerts")

        # Auto-check alerts every 2 min
        if st.session_state.get("auto_rf"):
            for alert in st.session_state.get(
                "price_alerts", []
            ):
                if not alert["triggered"]:
                    _alp2 = live_price(alert["sym"])
                    if _alp2["ok"]:
                        triggered2 = (
                            (alert["direction"]=="Goes ABOVE" and
                             _alp2["p"] >= alert["price"]) or
                            (alert["direction"]=="Goes BELOW" and
                             _alp2["p"] <= alert["price"])
                        )
                        if triggered2:
                            alert["triggered"] = True
                            tok2 = st.session_state.get(
                                "tg_token_saved",""
                            )
                            cid2 = st.session_state.get(
                                "tg_chat_saved",""
                            )
                            msg2 = (
                                f"🔔 ALERT — {alert['stock']}\n"
                                f"{alert['direction']} "
                                f"₹{alert['price']:,.2f}\n"
                                f"Now at ₹{_alp2['p']:,.2f}"
                            )
                            send_telegram(tok2, cid2, msg2)


# News & Events tab removed
if False: pass
# ╔══════════════════════════════════════════════════════╗
# ║  TAB 5 — NEWS & EVENTS                              ║
# ╚══════════════════════════════════════════════════════╝
with T7:
    st.markdown("### 📰 News & Market Events")

    nt1,nt2,nt3,nt4 = st.tabs([
        "🇮🇳 Market News",
        f"📌 {sname[:12]}",
        "🌍 Global",
        "📅 Daily Checklist",
    ])

    def render_news(arts):
        if not arts:
            st.info("No news found."); return
        for a in arts:
            st.markdown(f"""
            <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)'>
              <a href='{a['link']}' target='_blank'
                 style='color:#374151;font-size:14px;
                        font-weight:500;text-decoration:none'>
                  {a['title']}
              </a>
              <div style='color:#94a3b8;font-size:11px;
                          margin-top:5px'>
                  📰 {a['src']} &nbsp;|&nbsp; 🕐 {a['date']}
              </div>
            </div>""", unsafe_allow_html=True)

    with nt1:
        with st.spinner("Loading..."):
            render_news(get_news(
                "NSE BSE Nifty Sensex India stock market"))
    with nt2:
        with st.spinner("Loading..."):
            render_news(get_news(f"{sname} NSE India stock"))
    with nt3:
        with st.spinner("Loading..."):
            render_news(get_news(
                "US Fed Dow Nasdaq global markets"))
    with nt4:
        tstate_now = best_trading_time()
        tclr_now   = {"best":"#00ff88","good":"#88ff44",
                      "ok":"#ffcc00","caution":"#ff8844",
                      "avoid":"#ff4455","closed":"#555",
                      "pre_market":"#aaa"}.get(tstate_now,"#aaa")
        tmsg_now   = {"best":"✅ BEST TIME — enter now",
                      "good":"🟡 GOOD TIME",
                      "ok":"🟡 OK — lunch hours",
                      "caution":"⚠️ CAUTION — choppy",
                      "avoid":"❌ AVOID — too volatile",
                      "closed":"⛔ MARKET CLOSED",
                      "pre_market":"🕐 PRE-MARKET — prepare"
                     }.get(tstate_now,"—")

        st.markdown(f"""
        <div style='background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.06)' style='border:1.5px solid {tclr_now};
             text-align:center;padding:20px'>
          <div style='font-size:24px;font-weight:700;
                      color:{tclr_now}'>{tmsg_now}</div>
          <div style='color:#64748b;font-size:13px;margin-top:6px'>
              {now_ist().strftime('%H:%M IST')}
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("### ☀️ Pre-Market Checklist (8:45–9:15 AM)")
        checklist_items = [
            "Check Gift Nifty direction on Google",
            "Check US market close (Dow, Nasdaq, S&P)",
            "Check Crude Oil price (affects energy stocks)",
            "Check Gold price (affects MCX traders)",
            "Look at Tab 5 for any big events today",
            "Check PCR ratio in Options Chain tab",
            "Note first 15-min candle high and low",
        ]
        for item in checklist_items:
            st.checkbox(item, key=f"pre_{item[:20]}")

        st.markdown("---")
        st.markdown("""
        ### ⏰ Trading Time Guide

        | Time | Recommendation | Why |
        |------|----------------|-----|
        | 9:15–9:30 AM | ❌ Do not trade | Opening chaos |
        | 9:30–10:30 AM | ✅ Best window | Strong trends form |
        | 10:30–12:00 PM | ✅ Good | Clear price action |
        | 12:00–1:30 PM | 🟡 OK | Slow lunch hours |
        | 1:30–2:30 PM | ✅ Good | Afternoon momentum |
        | 2:30–3:00 PM | ⚠️ Caution | Choppy |
        | 3:00–3:30 PM | ❌ Do not trade | Erratic closing |

        ### 📅 Events — Avoid trading options on these days

        | Event | Impact |
        |-------|--------|
        | RBI Monetary Policy | 🔴 Very High — avoid |
        | US Fed Meeting | 🔴 Very High — avoid |
        | Union Budget (Feb 1) | 🔴 Very High — avoid |
        | Weekly F&O Expiry (Thu) | 🟠 High — be careful |
        | Monthly Expiry (last Thu) | 🔴 Very High — avoid |
        | Quarterly Results | 🟠 High — be careful |
        | US CPI data | 🟡 Medium |
        | India CPI / WPI | 🟡 Medium |

        ### 📋 Pre-Trade Rules (check all before every trade)

        1. Uptrend/Downtrend score ≥ 7 ✅
        2. RSI in correct zone (55–68 for CE, 32–45 for PE) ✅
        3. Price above/below VWAP ✅
        4. Volume surge ≥ 1.2× ✅
        5. MACD confirming direction ✅
        6. ADX > 20 ✅
        7. CMF confirming direction ✅
        8. OBV confirming direction ✅
        9. Risk-Reward ≥ 1.5 ✅
        10. No contradicting candlestick pattern ✅
        11. Trading time is green ✅
        12. No major event today ✅
        """)


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 8 — PAPER TRADING                              ║
# ╚══════════════════════════════════════════════════════╝

# ╔══════════════════════════════════════════════════════╗
# ║  TAB 8 — MARKET PULSE                               ║
# ╚══════════════════════════════════════════════════════╝
with T6:
    st.markdown("### 📊 Market Pulse — VIX + FII/DII")
    st.caption(
        "India VIX shows market fear. "
        "FII/DII shows institutional money flow. "
        "These two together tell you the market mood."
    )

    # ── India VIX ─────────────────────────────────────────
    st.markdown("#### ⚡ India VIX — Market Fear Index")

    if st.button("Refresh VIX + FII/DII", key="pulse_refresh",
                 type="primary"):
        get_india_vix.clear()
        get_fii_dii.clear()
        st.rerun()

    vix_data = get_india_vix()

    if vix_data["ok"]:
        vix_v = vix_data["vix"]
        vix_col = (
            "#dc2626" if vix_v > 25 else
            "#ea580c" if vix_v > 20 else
            "#d97706" if vix_v > 15 else
            "#16a34a"
        )
        vix_bg = (
            "#fef2f2" if vix_v > 25 else
            "#fff7ed" if vix_v > 20 else
            "#fffbeb" if vix_v > 15 else
            "#f0fdf4"
        )

        vc1, vc2, vc3 = st.columns(3)
        with vc1:
            st.markdown(
                f"<div style='background:{vix_bg};"
                f"border:2px solid {vix_col};"
                f"border-radius:14px;padding:20px;"
                f"text-align:center'>"
                f"<div style='font-size:12px;color:#64748b;"
                f"text-transform:uppercase;letter-spacing:1px'>"
                f"India VIX</div>"
                f"<div style='font-size:52px;font-weight:700;"
                f"color:{vix_col};line-height:1'>{vix_v}</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:{vix_col};margin-top:4px'>"
                f"{vix_data['level']}</div>"
                f"<div style='font-size:12px;color:#64748b;"
                f"margin-top:4px'>"
                f"Change: {vix_data['chg']:+.2f}%</div>"
                f"</div>",
                unsafe_allow_html=True
            )
        with vc2:
            st.markdown(
                f"<div style='background:#f8fafc;"
                f"border-radius:12px;padding:16px'>"
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#374151;margin-bottom:10px'>"
                f"VIX Levels Guide</div>"
                f"<div style='font-size:12px;color:#475569;"
                f"line-height:2'>"
                f"🟢 Below 12 — Very calm, best for buying<br>"
                f"🟢 12–15 — Calm, good for CE/PE<br>"
                f"🟡 15–20 — Elevated, trade carefully<br>"
                f"🟠 20–25 — High fear, reduce size<br>"
                f"🔴 Above 25 — Extreme fear, avoid options"
                f"</div></div>",
                unsafe_allow_html=True
            )
        with vc3:
            st.markdown(
                f"<div style='background:#eff6ff;"
                f"border:1px solid #93c5fd;"
                f"border-radius:12px;padding:16px'>"
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#1d4ed8;margin-bottom:8px'>"
                f"Today's Advice</div>"
                f"<div style='font-size:13px;color:#374151'>"
                f"{vix_data['advice']}</div>"
                f"<div style='margin-top:12px;font-size:12px;"
                f"color:#64748b'>"
                f"High VIX = expensive options = "
                f"avoid buying<br>"
                f"Low VIX = cheap options = "
                f"good time to buy</div>"
                f"</div>",
                unsafe_allow_html=True
            )

        # VIX history chart
        try:
            vix_hist = yf.download(
                "^INDIAVIX", period="3mo",
                interval="1d", progress=False
            )
            if not vix_hist.empty:
                import plotly.graph_objects as go_vix
                fig_vix = go_vix.Figure()
                vix_close = vix_hist["Close"].squeeze()
                fig_vix.add_trace(go_vix.Scatter(
                    x=vix_hist.index,
                    y=vix_close,
                    fill="tozeroy",
                    line=dict(color="#3b82f6", width=2),
                    fillcolor="rgba(59,130,246,0.1)",
                    name="India VIX"
                ))
                for lvl, col_, nm in [
                    (25, "#dc2626", "Extreme Fear"),
                    (20, "#ea580c", "High Fear"),
                    (15, "#d97706", "Elevated"),
                ]:
                    fig_vix.add_hline(
                        y=lvl,
                        line_dash="dash",
                        line_color=col_,
                        opacity=0.5,
                        annotation_text=nm
                    )
                fig_vix.update_layout(
                    template="plotly_white",
                    height=250,
                    margin=dict(l=10,r=10,t=20,b=20),
                    title="India VIX — Last 3 Months",
                    yaxis_title="VIX"
                )
                st.plotly_chart(fig_vix, use_container_width=True)
        except Exception:
            pass
    else:
        st.warning(
            "VIX data unavailable. "
            "Check internet connection or try refreshing."
        )

    # ── IV Rank Section ───────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📊 IV Rank — Are Options Cheap or Expensive?")
    st.caption(
        "IV Rank tells you if options are cheap or expensive "
        "right now compared to the past year. "
        "Always check before buying CE or PE."
    )

    iv_data = get_iv_rank()
    if iv_data["ok"]:
        iv1, iv2, iv3 = st.columns(3)
        with iv1:
            st.markdown(
                f"<div style='background:{iv_data['bg']};"
                f"border:2px solid {iv_data['color']};"
                f"border-radius:14px;padding:20px;"
                f"text-align:center'>"
                f"<div style='font-size:12px;color:#64748b;"
                f"text-transform:uppercase;letter-spacing:1px'>"
                f"IV Rank</div>"
                f"<div style='font-size:52px;font-weight:700;"
                f"color:{iv_data['color']};line-height:1'>"
                f"{iv_data['iv_rank']}</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:{iv_data['color']};margin-top:4px'>"
                f"{iv_data['signal']}</div>"
                f"<div style='font-size:12px;color:#64748b;"
                f"margin-top:4px'>"
                f"Current VIX: {iv_data['current_iv']}</div>"
                f"</div>",
                unsafe_allow_html=True
            )
        with iv2:
            st.markdown(
                f"<div style='background:#f8fafc;"
                f"border-radius:12px;padding:16px'>"
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#374151;margin-bottom:10px'>"
                f"52-Week Range</div>"
                f"<div style='font-size:13px;color:#475569;"
                f"line-height:2.2'>"
                f"Current: <b>{iv_data['current_iv']}</b><br>"
                f"52W High: <b>{iv_data['high_52w']}</b><br>"
                f"52W Low: <b>{iv_data['low_52w']}</b><br>"
                f"IV Percentile: <b>{iv_data['iv_pct']}%</b>"
                f"</div></div>",
                unsafe_allow_html=True
            )
        with iv3:
            st.markdown(
                f"<div style='background:{iv_data['bg']};"
                f"border:1px solid {iv_data['color']};"
                f"border-radius:12px;padding:16px'>"
                f"<div style='font-size:13px;font-weight:700;"
                f"color:{iv_data['color']};margin-bottom:8px'>"
                f"Today's Advice</div>"
                f"<div style='font-size:13px;color:#374151'>"
                f"{iv_data['advice']}</div>"
                f"<hr style='border-color:#e2e8f0;margin:10px 0'>"
                f"<div style='font-size:12px;color:#64748b'>"
                f"IV Rank > 70 = Expensive → sell options or avoid buying<br>"
                f"IV Rank 30-70 = Fair → buy options normally<br>"
                f"IV Rank < 30 = Cheap → best time to buy CE/PE"
                f"</div></div>",
                unsafe_allow_html=True
            )
    else:
        st.info("IV Rank data unavailable. Check VIX connection.")

    # ── FII / DII Data ─────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 🏦 FII / DII — Institutional Money Flow")
    st.caption(
        "FII = Foreign Institutional Investors (foreign funds). "
        "DII = Domestic Institutional Investors (Indian mutual funds, LIC). "
        "When FII buys heavily — market usually goes up next session."
    )

    fii_data = get_fii_dii()

    if fii_data["ok"]:
        fii_net = fii_data["fii_net"]
        dii_net = fii_data["dii_net"]

        # Check if data is actually zero (NSE blocked) or real zero
        _data_valid = (
            fii_data["fii_buy"] != 0 or
            fii_data["fii_sell"] != 0
        )

        if not _data_valid:
            st.warning(
                "⚠️ FII/DII data showing ₹0 — NSE is blocking "
                "automated requests right now. "
                "This is common during market hours. "
                "Check manually at **nseindia.com** → "
                "Market Data → FII/DII Activity. "
                "Try clicking Refresh after some time."
            )
        else:
            fd1, fd2, fd3, fd4 = st.columns(4)
            fd1.metric("FII Buy",  f"₹{fii_data['fii_buy']:,.0f}Cr")
            fd2.metric("FII Sell", f"₹{fii_data['fii_sell']:,.0f}Cr")
            fd3.metric(
                "FII Net",
                f"₹{fii_net:,.0f}Cr",
                delta=f"{'Buying' if fii_net>0 else 'Selling'}",
                delta_color="normal" if fii_net > 0 else "inverse"
            )
            fd4.metric(
                "DII Net",
                f"₹{dii_net:,.0f}Cr",
                delta=f"{'Buying' if dii_net>0 else 'Selling'}",
                delta_color="normal" if dii_net > 0 else "inverse"
            )

        # Signal interpretation
        if fii_net > 1000 and dii_net > 0:
            st.success(
                f"🔥 Both FII and DII buying — "
                f"strong bullish signal for tomorrow. "
                f"FII net: ₹{fii_net:,.0f}Cr | "
                f"DII net: ₹{dii_net:,.0f}Cr"
            )
        elif fii_net > 500:
            st.success(
                f"✅ FII buying ₹{fii_net:,.0f}Cr — "
                f"bullish bias for tomorrow"
            )
        elif fii_net < -1000:
            st.error(
                f"🔴 FII selling ₹{abs(fii_net):,.0f}Cr — "
                f"bearish pressure. "
                f"{'DII countering with buying' if dii_net>0 else 'DII also selling — double bearish'}"
            )
        else:
            st.info(
                f"FII net: ₹{fii_net:,.0f}Cr | "
                f"DII net: ₹{dii_net:,.0f}Cr — "
                f"Neutral flow"
            )

        # Last 10 days table
        if fii_data.get("data"):
            with st.expander("Last 10 days FII/DII data"):
                rows = []
                for d in fii_data["data"][:10]:
                    fb = float(d.get("fiiBuy",0) or 0)
                    fs = float(d.get("fiiSell",0) or 0)
                    db = float(d.get("diiBuy",0) or 0)
                    ds = float(d.get("diiSell",0) or 0)
                    rows.append({
                        "Date":     d.get("date",""),
                        "FII Buy":  f"₹{fb:,.0f}Cr",
                        "FII Sell": f"₹{fs:,.0f}Cr",
                        "FII Net":  f"₹{fb-fs:+,.0f}Cr",
                        "DII Buy":  f"₹{db:,.0f}Cr",
                        "DII Sell": f"₹{ds:,.0f}Cr",
                        "DII Net":  f"₹{db-ds:+,.0f}Cr",
                    })
                st.dataframe(
                    pd.DataFrame(rows),
                    use_container_width=True,
                    hide_index=True
                )
    else:
        st.warning(
            "FII/DII data unavailable from NSE. "
            "NSE website may be blocking automated requests. "
            "Try during market hours (9 AM - 6 PM IST)."
        )
        st.info(
            "Manual check: Go to **nseindia.com** → "
            "Market Data → FII/DII Activity"
        )



    # ── Sector Heatmap ─────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 🌡️ Sector Heatmap — Today's Strength")
    st.caption(
        "Always trade stocks from the strongest sector. "
        "Green = bullish today, Red = bearish today."
    )

    if st.button("Load Sector Heatmap", key="sector_heatmap_btn"):
        st.session_state["load_heatmap"] = True

    if st.session_state.get("load_heatmap"):
        _sector_map = {
            "🏦 Banking":   ["HDFCBANK.NS","ICICIBANK.NS","SBIN.NS"],
            "💻 IT":        ["TCS.NS","INFY.NS","WIPRO.NS"],
            "💊 Pharma":    ["SUNPHARMA.NS","DRREDDY.NS","CIPLA.NS"],
            "🚗 Auto":      ["MARUTI.NS","TATAMOTORS.NS","BAJAJ-AUTO.NS"],
            "⚡ Energy":    ["RELIANCE.NS","ONGC.NS","NTPC.NS"],
            "🔩 Metals":    ["TATASTEEL.NS","JSWSTEEL.NS","HINDALCO.NS"],
            "🏗️ Infra":    ["LT.NS","ADANIPORTS.NS","SIEMENS.NS"],
            "🛒 FMCG":      ["HINDUNILVR.NS","ITC.NS","NESTLEIND.NS"],
            "💰 Finance":   ["BAJFINANCE.NS","BAJAJFINSV.NS","HDFCLIFE.NS"],
            "📡 Telecom":   ["BHARTIARTL.NS","TATACOMM.NS"],
        }
        _sector_results = {}
        _prog_s = st.progress(0, text="Loading sectors...")
        for _si, (_sname, _stocks) in enumerate(_sector_map.items()):
            _prog_s.progress(
                int((_si+1)/len(_sector_map)*100),
                text=f"Loading {_sname}..."
            )
            _changes = []
            for _stk in _stocks:
                try:
                    _tk = yf.Ticker(_stk)
                    _fi = _tk.fast_info
                    _p  = float(_fi.last_price or 0)
                    _pc = float(_fi.previous_close or _p)
                    if _pc > 0:
                        _changes.append(round((_p-_pc)/_pc*100,2))
                except Exception:
                    pass
            if _changes:
                _sector_results[_sname] = round(
                    sum(_changes)/len(_changes), 2
                )
        _prog_s.empty()

        if _sector_results:
            _sorted_s = sorted(
                _sector_results.items(),
                key=lambda x: x[1], reverse=True
            )
            _hcols = st.columns(2)
            for _hi, (_sname, _chg) in enumerate(_sorted_s):
                _hcol = _hcols[_hi % 2]
                if _chg >= 1.5:
                    _hbg="#f0fdf4"; _hbr="#16a34a"; _htc="#166534"; _hic="🔥"
                elif _chg >= 0.3:
                    _hbg="#f0fdf4"; _hbr="#86efac"; _htc="#166534"; _hic="✅"
                elif _chg >= -0.3:
                    _hbg="#f8fafc"; _hbr="#cbd5e1"; _htc="#475569"; _hic="➡️"
                elif _chg >= -1.5:
                    _hbg="#fef2f2"; _hbr="#fca5a5"; _htc="#991b1b"; _hic="⚠️"
                else:
                    _hbg="#fef2f2"; _hbr="#dc2626"; _htc="#991b1b"; _hic="🔴"
                _hcol.markdown(
                    f"<div style='background:{_hbg};"
                    f"border:1.5px solid {_hbr};"
                    f"border-radius:10px;padding:14px;"
                    f"margin-bottom:8px;"
                    f"display:flex;justify-content:"
                    f"space-between;align-items:center'>"
                    f"<span style='font-size:14px;"
                    f"font-weight:700;color:{_htc}'>"
                    f"{_hic} {_sname}</span>"
                    f"<span style='font-size:22px;"
                    f"font-weight:700;color:{_htc}'>"
                    f"{_chg:+.2f}%</span></div>",
                    unsafe_allow_html=True
                )
            _best_s  = _sorted_s[0]
            _worst_s = _sorted_s[-1]
            st.success(
                f"🔥 Strongest: **{_best_s[0]}** "
                f"({_best_s[1]:+.2f}%) — "
                f"Focus CE trades here"
            )
            if _worst_s[1] < -0.3:
                st.error(
                    f"🔴 Weakest: **{_worst_s[0]}** "
                    f"({_worst_s[1]:+.2f}%) — "
                    f"Consider PE trades here"
                )

    # ── Economic Calendar ─────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📅 Economic Calendar — Upcoming Events")
    st.caption(
        "High impact events can move markets 2-3% in minutes. "
        "Avoid entering new trades before HIGH impact events."
    )

    _cal_events = get_economic_calendar()
    if _cal_events:
        for _ev in _cal_events[:8]:
            _imp    = _ev["impact"]
            _days   = _ev["days_away"]
            _imp_col = (
                "#dc2626" if _imp=="HIGH"
                else "#d97706" if _imp=="MEDIUM"
                else "#16a34a"
            )
            _imp_bg = (
                "#fef2f2" if _imp=="HIGH"
                else "#fffbeb" if _imp=="MEDIUM"
                else "#f0fdf4"
            )
            _days_label = (
                "🔴 TODAY" if _days==0
                else "🟠 TOMORROW" if _days==1
                else f"In {_days} days"
            )
            st.markdown(
                f"<div style='background:{_imp_bg};"
                f"border-left:4px solid {_imp_col};"
                f"border-radius:0 8px 8px 0;"
                f"padding:10px 14px;margin:4px 0;"
                f"display:flex;justify-content:space-between;"
                f"align-items:center'>"
                f"<div>"
                f"<span style='font-size:13px;font-weight:700;"
                f"color:#1e293b'>{_ev['event']}</span>"
                f"<span style='font-size:11px;color:#64748b;"
                f"margin-left:8px'>{_ev['date']}</span>"
                f"{'<br><span style="font-size:11px;color:#64748b">' + _ev['note'] + '</span>' if _ev['note'] else ''}"
                f"</div>"
                f"<div style='text-align:right'>"
                f"<span style='background:{_imp_col};color:white;"
                f"padding:2px 8px;border-radius:10px;"
                f"font-size:11px;font-weight:700'>{_imp}</span>"
                f"<br><span style='font-size:11px;color:{_imp_col};"
                f"font-weight:600'>{_days_label}</span>"
                f"</div></div>",
                unsafe_allow_html=True
            )

        # Warning if high impact event today or tomorrow
        _urgent = [
            e for e in _cal_events
            if e["impact"]=="HIGH" and e["days_away"] <= 1
        ]
        if _urgent:
            st.error(
                f"⚠️ HIGH IMPACT EVENT {'TODAY' if _urgent[0]['days_away']==0 else 'TOMORROW'}: "
                f"{_urgent[0]['event']} — "
                f"Avoid entering new trades. "
                f"Existing trades: tighten stop loss."
            )
    else:
        st.info("No major events in next 7 days. Clear to trade.")

    # ── Market Breadth ─────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📊 Market Breadth — Is the rally broad?")
    st.caption(
        "Breadth above 2:1 means most stocks rising — strong market. "
        "Breadth below 1:1 means rally is narrow — avoid CE trades."
    )
    if st.button("Check Market Breadth", key="breadth_btn"):
        st.session_state["load_breadth"] = True

    if st.session_state.get("load_breadth"):
        _nifty50 = [
            "RELIANCE.NS","TCS.NS","HDFCBANK.NS","ICICIBANK.NS",
            "INFY.NS","HINDUNILVR.NS","ITC.NS","SBIN.NS",
            "BHARTIARTL.NS","KOTAKBANK.NS","LT.NS","AXISBANK.NS",
            "BAJFINANCE.NS","ASIANPAINT.NS","MARUTI.NS",
            "SUNPHARMA.NS","TITAN.NS","WIPRO.NS","TATASTEEL.NS",
            "JSWSTEEL.NS","TATAMOTORS.NS","HCLTECH.NS","TECHM.NS",
            "DRREDDY.NS","CIPLA.NS","NTPC.NS","ONGC.NS",
            "POWERGRID.NS","ULTRACEMCO.NS","NESTLEIND.NS",
        ]
        _adv=0; _dec=0; _unc=0
        _prog_b = st.progress(0, text="Checking breadth...")
        for _bi, _bs in enumerate(_nifty50):
            _prog_b.progress(
                int((_bi+1)/len(_nifty50)*100),
                text=f"Checking {_bs}..."
            )
            try:
                _btk = yf.Ticker(_bs)
                _bfi = _btk.fast_info
                _bp  = float(_bfi.last_price or 0)
                _bpc = float(_bfi.previous_close or _bp)
                if _bpc > 0:
                    _bchg = (_bp-_bpc)/_bpc*100
                    if _bchg > 0.2:   _adv += 1
                    elif _bchg < -0.2: _dec += 1
                    else:              _unc += 1
            except Exception:
                pass
        _prog_b.empty()

        _ratio = round(_adv/(_dec+0.001), 2)
        bb1,bb2,bb3,bb4 = st.columns(4)
        bb1.metric("Advancing", f"{_adv}")
        bb2.metric("Declining", f"{_dec}")
        bb3.metric("Unchanged", f"{_unc}")
        bb4.metric("A/D Ratio", f"{_ratio}:1")

        if _ratio >= 3:
            st.success(f"🔥 Very strong breadth {_ratio}:1 — High confidence CE day")
        elif _ratio >= 1.5:
            st.success(f"✅ Good breadth {_ratio}:1 — CE trades favoured")
        elif _ratio >= 0.7:
            st.warning(f"⚠️ Mixed breadth {_ratio}:1 — Trade smaller size")
        else:
            st.error(f"🔴 Weak breadth {_ratio}:1 — Avoid CE, consider PE")

        _total_b = _adv+_dec+_unc
        if _total_b > 0:
            _ap = int(_adv/_total_b*100)
            _dp = int(_dec/_total_b*100)
            st.markdown(
                f"<div style='height:24px;border-radius:12px;"
                f"overflow:hidden;display:flex;margin-top:8px'>"
                f"<div style='width:{_ap}%;background:#16a34a;"
                f"color:white;font-size:11px;font-weight:700;"
                f"display:flex;align-items:center;"
                f"justify-content:center'>{_ap}%</div>"
                f"<div style='width:{100-_ap-_dp}%;"
                f"background:#94a3b8'></div>"
                f"<div style='width:{_dp}%;background:#dc2626;"
                f"color:white;font-size:11px;font-weight:700;"
                f"display:flex;align-items:center;"
                f"justify-content:center'>{_dp}%</div>"
                f"</div>",
                unsafe_allow_html=True
            )



# ╔══════════════════════════════════════════════════════╗
# ║  TAB 9 — OPTIONS CHAIN                              ║
# ╚══════════════════════════════════════════════════════╝
with T7:
    st.markdown("### 🔗 Options Chain + OI Analysis")
    st.caption(
        "OI Change is the most powerful intraday options signal. "
        "Where OI is being built = where institutions are positioned."
    )

    oc1, oc2, oc3 = st.columns([2, 1, 1])
    with oc1:
        oc_symbol = st.selectbox(
            "Index",
            ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"],
            key="oc_symbol"
        )
    with oc2:
        if st.button(
            "Load Options Chain",
            type="primary",
            key="oc_load",
            use_container_width=True
        ):
            get_options_chain.clear()
            get_oi_change_data.clear()
    with oc3:
        if st.button(
            "Load OI Analysis",
            key="oc_oi_load",
            use_container_width=True
        ):
            get_oi_change_data.clear()

    # ── OI Change Analysis ─────────────────────────────────
    st.markdown("#### 📈 OI Change Analysis — Where is money flowing?")

    oi_data = get_oi_change_data(oc_symbol)
    if oi_data["ok"]:
        # Signal banner
        st.markdown(
            f"<div style='background:{oi_data['oi_color']}22;"
            f"border:2px solid {oi_data['oi_color']};"
            f"border-radius:12px;padding:14px 20px;"
            f"margin-bottom:12px'>"
            f"<span style='font-size:18px;font-weight:700;"
            f"color:{oi_data['oi_color']}'>"
            f"⚡ {oi_data['oi_signal']}</span>"
            f"<span style='font-size:13px;color:#475569;"
            f"margin-left:16px'>{oi_data['oi_advice']}</span>"
            f"</div>",
            unsafe_allow_html=True
        )

        # Key levels
        ok1, ok2, ok3, ok4 = st.columns(4)
        ok1.metric(
            "Max Call OI (Resistance)",
            f"₹{oi_data['max_call_oi_strike']:,}",
            help="Strongest resistance — institutions sold calls here"
        )
        ok2.metric(
            "Max Put OI (Support)",
            f"₹{oi_data['max_put_oi_strike']:,}",
            help="Strongest support — institutions sold puts here"
        )
        ok3.metric(
            "Max Call OI Build",
            f"₹{oi_data['max_call_chg_strike']:,}",
            help="Strike where most new call positions added today"
        )
        ok4.metric(
            "Max Put OI Build",
            f"₹{oi_data['max_put_chg_strike']:,}",
            help="Strike where most new put positions added today"
        )

        # OI change totals
        call_chg = oi_data["total_call_oi_chg"]
        put_chg  = oi_data["total_put_oi_chg"]
        cc = "#16a34a" if call_chg > 0 else "#dc2626"
        pc = "#16a34a" if put_chg  > 0 else "#dc2626"

        st.markdown(
            f"<div style='background:#f8fafc;"
            f"border-radius:10px;padding:12px 18px;"
            f"font-size:13px;margin:8px 0'>"
            f"Total Call OI Change: "
            f"<b style='color:{cc}'>{call_chg:+,}</b> contracts &nbsp;|&nbsp; "
            f"Total Put OI Change: "
            f"<b style='color:{pc}'>{put_chg:+,}</b> contracts &nbsp;|&nbsp; "
            f"Spot: ₹{oi_data['spot']:,.2f} &nbsp;|&nbsp; "
            f"ATM: ₹{oi_data['atm']:,} &nbsp;|&nbsp; "
            f"Expiry: {oi_data['expiry']}"
            f"</div>",
            unsafe_allow_html=True
        )

        # OI Change interpretation guide
        with st.expander("📖 How to read OI Change"):
            st.markdown("""
            | OI Change | Meaning | Signal |
            |-----------|---------|--------|
            | Call OI ↑ + Price ↑ | Short covering in calls | Bullish |
            | Call OI ↑ + Price ↓ | New call writing (resistance) | Bearish |
            | Put OI ↑ + Price ↓ | Short covering in puts | Bearish |
            | Put OI ↑ + Price ↑ | New put writing (support) | Bullish |
            | Call OI ↓ + Price ↑ | Call unwinding | Mildly bullish |
            | Put OI ↓ + Price ↓ | Put unwinding | Mildly bearish |

            **Key rule:** If Put OI is building at a strike BELOW current price
            = strong support at that level. Institutions are selling puts there
            = they believe price will NOT fall below that strike.
            """)

        # Near ATM OI Change table
        st.markdown("#### OI Change near ATM strikes")
        atm_val  = oi_data["atm"]
        step_oc  = 50 if oc_symbol == "NIFTY" else 100
        near_strikes = [
            s for s in oi_data["strikes"]
            if abs(s["strike"] - atm_val) <= step_oc * 8
        ]

        if near_strikes:
            import plotly.graph_objects as go_oi

            strikes_list  = [s["strike"] for s in near_strikes]
            call_oi_chg   = [s["ce_oi_chg"] for s in near_strikes]
            put_oi_chg    = [s["pe_oi_chg"] for s in near_strikes]

            fig_oi = go_oi.Figure()
            fig_oi.add_trace(go_oi.Bar(
                x=strikes_list,
                y=call_oi_chg,
                name="Call OI Change",
                marker_color=[
                    "#16a34a" if v > 0 else "#dc2626"
                    for v in call_oi_chg
                ],
                opacity=0.8
            ))
            fig_oi.add_trace(go_oi.Bar(
                x=strikes_list,
                y=[-v for v in put_oi_chg],
                name="Put OI Change (inverted)",
                marker_color=[
                    "#3b82f6" if v > 0 else "#f59e0b"
                    for v in put_oi_chg
                ],
                opacity=0.8
            ))
            fig_oi.add_vline(
                x=atm_val,
                line_dash="dash",
                line_color="#374151",
                annotation_text=f"ATM {atm_val}",
                annotation_position="top"
            )
            fig_oi.update_layout(
                template="plotly_white",
                height=350,
                barmode="overlay",
                title=(
                    f"{oc_symbol} OI Change — "
                    f"Green=Call build, Blue=Put build"
                ),
                xaxis_title="Strike Price",
                yaxis_title="OI Change (contracts)",
                margin=dict(l=10,r=10,t=40,b=40),
                legend=dict(
                    orientation="h",
                    y=1.02
                )
            )
            st.plotly_chart(fig_oi, use_container_width=True)

            # IV display near ATM
            st.markdown("#### IV (Implied Volatility) near ATM")
            iv_rows = []
            for s in near_strikes:
                iv_rows.append({
                    "Strike": s["strike"],
                    "ATM": "◀ ATM" if s["is_atm"] else "",
                    "CE LTP": f"₹{s['ce_ltp']:,.2f}",
                    "CE IV %": f"{s['ce_iv']:.1f}%",
                    "CE OI Chg": f"{s['ce_oi_chg']:+,}",
                    "PE LTP": f"₹{s['pe_ltp']:,.2f}",
                    "PE IV %": f"{s['pe_iv']:.1f}%",
                    "PE OI Chg": f"{s['pe_oi_chg']:+,}",
                })
            st.dataframe(
                pd.DataFrame(iv_rows),
                hide_index=True,
                width="stretch"
            )

            # IV Skew insight
            atm_row = next(
                (s for s in near_strikes if s["is_atm"]), None
            )
            if atm_row:
                ce_iv = atm_row["ce_iv"]
                pe_iv = atm_row["pe_iv"]
                if pe_iv > ce_iv * 1.1:
                    st.warning(
                        f"⚠️ Put IV ({pe_iv}%) > Call IV ({ce_iv}%) — "
                        f"market pricing in more downside risk. "
                        f"PE options are expensive right now."
                    )
                elif ce_iv > pe_iv * 1.1:
                    st.warning(
                        f"⚠️ Call IV ({ce_iv}%) > Put IV ({pe_iv}%) — "
                        f"market pricing in more upside risk. "
                        f"CE options are expensive right now."
                    )
                else:
                    st.success(
                        f"✅ CE IV ({ce_iv}%) ≈ PE IV ({pe_iv}%) — "
                        f"balanced market. Both CE and PE fairly priced."
                    )
    else:
        st.warning(
            "OI data unavailable from NSE. "
            "NSE may be blocking. Try during market hours."
        )

    st.markdown("---")

    oc_data = get_options_chain(oc_symbol)

    if oc_data["ok"]:
        spot = oc_data["spot"]
        expiries = oc_data["expiries"]

        st.markdown(
            f"<div style='background:#1e3a5f;color:white;"
            f"border-radius:10px;padding:10px 18px;"
            f"font-size:16px;font-weight:700;margin-bottom:12px'>"
            f"{oc_symbol} Spot: ₹{spot:,.2f}</div>",
            unsafe_allow_html=True
        )

        if expiries:
            sel_exp = st.selectbox(
                "Expiry",
                expiries[:5],
                key="oc_expiry"
            )

            # Filter data for selected expiry
            exp_data = [
                d for d in oc_data["data"]
                if d.get("expiryDate") == sel_exp
            ]

            # Calculate ATM strike
            step = 50 if oc_symbol == "NIFTY" else 100
            atm = round(spot / step) * step

            # Build chain table
            rows = []
            total_call_oi = 0
            total_put_oi  = 0

            for item in exp_data:
                strike = item.get("strikePrice", 0)
                ce = item.get("CE", {})
                pe = item.get("PE", {})
                if not ce and not pe:
                    continue

                ce_oi  = int(ce.get("openInterest", 0) or 0)
                ce_doi = int(ce.get("changeinOpenInterest", 0) or 0)
                ce_iv  = round(float(ce.get("impliedVolatility", 0) or 0), 1)
                ce_ltp = round(float(ce.get("lastPrice", 0) or 0), 2)

                pe_oi  = int(pe.get("openInterest", 0) or 0)
                pe_doi = int(pe.get("changeinOpenInterest", 0) or 0)
                pe_iv  = round(float(pe.get("impliedVolatility", 0) or 0), 1)
                pe_ltp = round(float(pe.get("lastPrice", 0) or 0), 2)

                total_call_oi += ce_oi
                total_put_oi  += pe_oi

                rows.append({
                    "CE OI":    ce_oi,
                    "CE Chg":   ce_doi,
                    "CE IV":    ce_iv,
                    "CE LTP":   ce_ltp,
                    "Strike":   strike,
                    "ATM":      "◀ ATM" if strike == atm else "",
                    "PE LTP":   pe_ltp,
                    "PE IV":    pe_iv,
                    "PE Chg":   pe_doi,
                    "PE OI":    pe_oi,
                })

            if rows:
                # ── Max Pain Calculation ───────────────────
                # Max Pain = strike where total option buyers lose most
                df_chain_full = pd.DataFrame(rows)
                max_pain_val  = spot  # default
                min_pain_loss = float("inf")

                for _mp_strike in df_chain_full["Strike"].values:
                    _total_loss = 0
                    for _, _row in df_chain_full.iterrows():
                        _s = _row["Strike"]
                        _ce_oi = _row["CE OI"]
                        _pe_oi = _row["PE OI"]
                        # CE buyer loses if strike > current
                        if _mp_strike > _s:
                            _total_loss += (_mp_strike - _s) * _ce_oi
                        # PE buyer loses if strike < current
                        if _mp_strike < _s:
                            _total_loss += (_s - _mp_strike) * _pe_oi
                    if _total_loss < min_pain_loss:
                        min_pain_loss = _total_loss
                        max_pain_val  = _mp_strike

                # Distance from spot to Max Pain
                _mp_dist  = max_pain_val - spot
                _mp_dir   = "above" if max_pain_val > spot else "below"
                _mp_pct   = round(abs(_mp_dist) / spot * 100, 2)

                # Expected Move
                # Use ATM straddle price as proxy
                _atm_row = next(
                    (r for r in rows if r["Strike"] == atm), None
                )
                _exp_move = 0
                _exp_move_pct = 0
                if _atm_row:
                    _atm_ce = _atm_row["CE LTP"]
                    _atm_pe = _atm_row["PE LTP"]
                    _exp_move = round(_atm_ce + _atm_pe, 2)
                    _exp_move_pct = round(_exp_move / spot * 100, 2)

                # PCR
                pcr = round(total_put_oi / (total_call_oi + 1), 2)
                pcr_signal = (
                    "🟢 Bullish (PCR > 1.3 = oversold)"
                    if pcr > 1.3 else
                    "🔴 Bearish (PCR < 0.7 = overbought)"
                    if pcr < 0.7 else
                    "🟡 Neutral"
                )

                pm1,pm2,pm3,pm4,pm5 = st.columns(5)
                pm1.metric("Total Call OI", f"{total_call_oi:,}")
                pm2.metric("Total Put OI",  f"{total_put_oi:,}")
                pm3.metric(
                    "PCR",
                    f"{pcr}",
                    delta=pcr_signal,
                    delta_color="normal" if pcr > 1 else "inverse"
                )
                pm4.metric(
                    "Max Pain",
                    f"₹{max_pain_val:,}",
                    delta=f"{_mp_pct:.1f}% {_mp_dir} spot",
                    help="Strike where option buyers lose most. "
                         "Price tends to move toward Max Pain on expiry."
                )
                pm5.metric(
                    "Expected Move",
                    f"₹{_exp_move:,.0f}",
                    delta=f"±{_exp_move_pct:.1f}% from spot",
                    help="ATM straddle price = how much market "
                         "expects index to move by expiry."
                )

                # Max Pain interpretation
                _mp_color = "#7c3aed"
                if abs(_mp_dist) < step * 2:
                    st.info(
                        f"📍 Max Pain at ₹{max_pain_val:,} — "
                        f"Very close to spot ₹{spot:,.0f}. "
                        f"Market likely to stay range-bound near "
                        f"this level by expiry."
                    )
                elif max_pain_val < spot:
                    st.warning(
                        f"📍 Max Pain at ₹{max_pain_val:,} — "
                        f"₹{abs(_mp_dist):,.0f} BELOW spot. "
                        f"Sellers may push market down toward "
                        f"₹{max_pain_val:,} by expiry. "
                        f"Be cautious with CE trades far above Max Pain."
                    )
                else:
                    st.success(
                        f"📍 Max Pain at ₹{max_pain_val:,} — "
                        f"₹{abs(_mp_dist):,.0f} ABOVE spot. "
                        f"Market may be supported and move up toward "
                        f"₹{max_pain_val:,} by expiry. "
                        f"Supports CE trades."
                    )

                # Expected Move interpretation
                if _exp_move > 0:
                    st.markdown(
                        f"<div style='background:#faf5ff;"
                        f"border:1px solid #c4b5fd;"
                        f"border-radius:8px;padding:10px 16px;"
                        f"font-size:13px;color:#6d28d9;"
                        f"margin-bottom:8px'>"
                        f"<b>Expected Move:</b> Market expects "
                        f"{oc_symbol} to move ±₹{_exp_move:,.0f} "
                        f"(±{_exp_move_pct:.1f}%) from current spot "
                        f"₹{spot:,.0f} by expiry {sel_exp}. "
                        f"Upper range: ₹{spot+_exp_move:,.0f} | "
                        f"Lower range: ₹{spot-_exp_move:,.0f}"
                        f"</div>",
                        unsafe_allow_html=True
                    )

                # Show chain near ATM (±10 strikes)
                df_chain = pd.DataFrame(rows)
                df_chain_near = df_chain[
                    (df_chain["Strike"] >= atm - step*10) &
                    (df_chain["Strike"] <= atm + step*10)
                ].copy()

                # Highlight max OI
                max_call_oi = df_chain["CE OI"].max()
                max_put_oi  = df_chain["PE OI"].max()

                st.markdown(
                    "**Max Call OI** (Resistance): "
                    f"**{df_chain.loc[df_chain['CE OI'].idxmax(), 'Strike']:,}** "
                    f"({max_call_oi:,} contracts)  |  "
                    "**Max Put OI** (Support): "
                    f"**{df_chain.loc[df_chain['PE OI'].idxmax(), 'Strike']:,}** "
                    f"({max_put_oi:,} contracts)"
                )

                st.dataframe(
                    df_chain_near.set_index("Strike"),
                    use_container_width=True,
                    height=500
                )
    else:
        st.warning(
            "Options chain data unavailable from NSE. "
            "NSE may be blocking automated requests."
        )
        st.info(
            "While connected to Zerodha Kite, "
            "options chain loads from live data. "
            "Login with Kite in the sidebar for live options data."
        )
        # Try Yahoo Finance options as fallback
        st.markdown("#### Yahoo Finance Options (Fallback)")
        yf_sym = "^NSEI" if oc_symbol=="NIFTY" else "^NSEBANK"
        try:
            tk_oc = yf.Ticker(yf_sym)
            exp_yf = tk_oc.options
            if exp_yf:
                sel_yf = st.selectbox(
                    "Expiry (Yahoo)",
                    exp_yf[:3],
                    key="oc_yf_exp"
                )
                chain_yf = tk_oc.option_chain(sel_yf)
                ycol1, ycol2 = st.columns(2)
                with ycol1:
                    st.markdown("**Calls**")
                    st.dataframe(
                        chain_yf.calls[[
                            "strike","lastPrice",
                            "openInterest","impliedVolatility"
                        ]].head(20),
                        use_container_width=True,
                        hide_index=True
                    )
                with ycol2:
                    st.markdown("**Puts**")
                    st.dataframe(
                        chain_yf.puts[[
                            "strike","lastPrice",
                            "openInterest","impliedVolatility"
                        ]].head(20),
                        use_container_width=True,
                        hide_index=True
                    )
        except Exception as e:
            st.caption(f"Yahoo options also unavailable: {e}")


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 10 — BACKTEST                                  ║
# ╚══════════════════════════════════════════════════════╝
with T8:
    st.markdown("### 🧪 Strategy Backtest")
    st.caption(
        "Test your signal strategy on historical data. "
        "See what the win rate and profit would have been "
        "if you had traded every signal for the past 6 months."
    )

    # ── Settings ──────────────────────────────────────────
    bc1, bc2, bc3 = st.columns(3)
    with bc1:
        bt_stock = st.text_input(
            "Stock to backtest",
            value="NIFTY 50",
            key="bt_stock"
        )
        bt_sym = STOCKS.get(bt_stock, "^NSEI")
    with bc2:
        bt_months = st.slider(
            "Months of history",
            1, 12, 6,
            key="bt_months"
        )
        bt_tf = st.selectbox(
            "Timeframe",
            ["15m","1h","1d"],
            index=2,
            key="bt_tf"
        )
    with bc3:
        bt_min_score = st.slider(
            "Min signal score",
            5, 10, 7,
            key="bt_min_score"
        )
        bt_capital = st.number_input(
            "Capital per trade (Rs)",
            value=10000,
            step=1000,
            key="bt_capital"
        )

    if st.button(
        "Run Backtest",
        type="primary",
        key="bt_run",
        use_container_width=True
    ):
        with st.spinner(
            f"Running backtest on {bt_stock} "
            f"({bt_months} months, {bt_tf})..."
        ):
            # Fetch historical data
            days_map = {"15m":30,"1h":60,"1d":bt_months*30}
            bt_days  = days_map.get(bt_tf, bt_months*30)
            end_bt   = datetime.now()
            start_bt = end_bt - timedelta(days=bt_days)

            bt_df = yf.download(
                bt_sym,
                start=start_bt,
                end=end_bt,
                interval=bt_tf,
                auto_adjust=True,
                progress=False
            )
            if isinstance(bt_df.columns, pd.MultiIndex):
                bt_df.columns = bt_df.columns.get_level_values(0)
            bt_df.dropna(inplace=True)

        if bt_df.empty or len(bt_df) < 100:
            st.error(
                f"Not enough data for {bt_stock}. "
                f"Try 1d timeframe or longer period."
            )
        else:
            # Run signals on rolling windows
            trades = []
            window_size = 55
            step_size   = 5

            prog_bt = st.progress(
                0, text="Scanning historical signals..."
            )
            total_windows = (len(bt_df)-window_size)//step_size

            for wi, start_i in enumerate(
                range(0, len(bt_df)-window_size, step_size)
            ):
                prog_bt.progress(
                    min(int(wi/total_windows*100), 99),
                    text=f"Scanning... {wi}/{total_windows}"
                )
                window_df = bt_df.iloc[start_i:start_i+window_size]
                try:
                    fake_lp = {
                        "ok": True,
                        "p": float(window_df["Close"].iloc[-1]),
                        "chg": 0, "chg_abs": 0,
                        "high": float(window_df["High"].iloc[-1]),
                        "low":  float(window_df["Low"].iloc[-1]),
                        "prev": float(window_df["Close"].iloc[-2])
                    }
                    sig_bt = compute_all(window_df, fake_lp)
                    if sig_bt is None:
                        continue

                    score = max(
                        sig_bt["up_score"],
                        sig_bt["dn_score"]
                    )
                    if score < bt_min_score:
                        continue

                    direction = sig_bt["direction"]
                    if direction not in ["UPTREND","DOWNTREND"]:
                        continue

                    entry_price = sig_bt["cp"]
                    sl = (sig_bt["sl_long"]
                          if direction=="UPTREND"
                          else sig_bt["sl_short"])
                    target = (sig_bt["tgt1"]
                              if direction=="UPTREND"
                              else sig_bt["tgt1s"])
                    signal_date = window_df.index[-1]

                    # Check outcome on next candles
                    future = bt_df.iloc[
                        start_i+window_size:
                        start_i+window_size+10
                    ]
                    if future.empty:
                        continue

                    outcome  = "OPEN"
                    exit_px  = float(future["Close"].iloc[-1])
                    exit_dt  = future.index[-1]

                    for _, row_f in future.iterrows():
                        if direction == "UPTREND":
                            if float(row_f["Low"]) <= sl:
                                outcome = "LOSS"
                                exit_px = sl
                                exit_dt = row_f.name
                                break
                            if float(row_f["High"]) >= target:
                                outcome = "WIN"
                                exit_px = target
                                exit_dt = row_f.name
                                break
                        else:
                            if float(row_f["High"]) >= sl:
                                outcome = "LOSS"
                                exit_px = sl
                                exit_dt = row_f.name
                                break
                            if float(row_f["Low"]) <= target:
                                outcome = "WIN"
                                exit_px = target
                                exit_dt = row_f.name
                                break

                    if outcome == "OPEN":
                        pnl_bt = (
                            (exit_px - entry_price)
                            if direction=="UPTREND"
                            else (entry_price - exit_px)
                        )
                        outcome = "WIN" if pnl_bt > 0 else "LOSS"
                    else:
                        pnl_bt = (
                            (exit_px - entry_price)
                            if direction=="UPTREND"
                            else (entry_price - exit_px)
                        )

                    pnl_pct = round(pnl_bt/entry_price*100, 2)

                    trades.append({
                        "Date":      signal_date.strftime("%d %b"),
                        "Direction": direction,
                        "Score":     score,
                        "Entry":     round(entry_price, 2),
                        "Exit":      round(exit_px, 2),
                        "P&L pts":   round(pnl_bt, 2),
                        "P&L %":     pnl_pct,
                        "Result":    outcome,
                    })
                except Exception:
                    continue

            prog_bt.empty()

            if not trades:
                st.warning(
                    f"No signals found with score {bt_min_score}+. "
                    "Try lowering the minimum score."
                )
            else:
                df_bt = pd.DataFrame(trades)
                wins  = df_bt[df_bt["Result"]=="WIN"]
                losses= df_bt[df_bt["Result"]=="LOSS"]
                wr_bt = round(len(wins)/len(df_bt)*100,1)
                avg_w = round(wins["P&L %"].mean(),2) if len(wins)>0 else 0
                avg_l = round(losses["P&L %"].mean(),2) if len(losses)>0 else 0
                pf_bt = round(
                    abs(wins["P&L %"].sum()) /
                    (abs(losses["P&L %"].sum())+0.001), 2
                )
                total_ret = round(df_bt["P&L %"].sum(), 2)

                # Summary metrics
                bm1,bm2,bm3,bm4,bm5 = st.columns(5)
                bm1.metric("Total Trades", len(df_bt))
                bm2.metric("Win Rate",     f"{wr_bt}%")
                bm3.metric("Profit Factor",f"{pf_bt}")
                bm4.metric("Avg Win",      f"{avg_w}%")
                bm5.metric("Avg Loss",     f"{avg_l}%")

                # Verdict
                if wr_bt >= 55 and pf_bt >= 1.5:
                    st.success(
                        f"🔥 Strong strategy! Win rate {wr_bt}% | "
                        f"Profit factor {pf_bt} | "
                        f"Total return {total_ret}% over "
                        f"{bt_months} months. "
                        f"This signal system is working well."
                    )
                elif wr_bt >= 45 and pf_bt >= 1.0:
                    st.warning(
                        f"📈 Decent strategy. Win rate {wr_bt}% | "
                        f"Profit factor {pf_bt}. "
                        f"Profitable but room to improve."
                    )
                else:
                    st.error(
                        f"⚠️ Weak results. Win rate {wr_bt}% | "
                        f"Profit factor {pf_bt}. "
                        f"Try increasing min score or "
                        f"changing timeframe."
                    )

                # Equity curve
                import plotly.graph_objects as go_bt
                cumulative = [100]
                for _, row_b in df_bt.iterrows():
                    cumulative.append(
                        cumulative[-1] * (1 + row_b["P&L %"]/100)
                    )

                fig_bt = go_bt.Figure()
                fig_bt.add_trace(go_bt.Scatter(
                    y=cumulative,
                    mode="lines",
                    line=dict(
                        color="#16a34a"
                        if cumulative[-1] >= 100
                        else "#dc2626",
                        width=2
                    ),
                    fill="tozeroy",
                    fillcolor=(
                        "rgba(22,163,74,0.1)"
                        if cumulative[-1] >= 100
                        else "rgba(220,38,38,0.1)"
                    ),
                    name="Equity Curve"
                ))
                fig_bt.add_hline(
                    y=100,
                    line_dash="dash",
                    line_color="#94a3b8",
                    annotation_text="Starting capital"
                )
                fig_bt.update_layout(
                    template="plotly_white",
                    height=280,
                    title=(
                        f"Equity Curve — "
                        f"{bt_stock} {bt_tf} "
                        f"(Score {bt_min_score}+)"
                    ),
                    yaxis_title="Portfolio value (start=100)",
                    margin=dict(l=10,r=10,t=40,b=10)
                )
                st.plotly_chart(fig_bt, use_container_width=True)

                # Win/Loss chart
                fig_wl = go_bt.Figure(go_bt.Bar(
                    x=df_bt["Date"],
                    y=df_bt["P&L %"],
                    marker_color=[
                        "#16a34a" if r=="WIN" else "#dc2626"
                        for r in df_bt["Result"]
                    ],
                    name="P&L %"
                ))
                fig_wl.update_layout(
                    template="plotly_white",
                    height=220,
                    title="Individual trade P&L %",
                    margin=dict(l=10,r=10,t=40,b=40),
                    xaxis_tickangle=-45
                )
                st.plotly_chart(fig_wl, use_container_width=True)

                # Full trades table
                with st.expander("All trades"):
                    st.dataframe(
                        df_bt,
                        width='stretch',
                        hide_index=True
                    )

    else:
        st.info(
            "Configure settings above and click "
            "**Run Backtest** to test your strategy "
            "on historical data."
        )
        st.markdown("""
        ### How backtesting works

        The backtest takes every 15m/1h/1d candle from the past
        and runs your signal engine on it. When score crosses
        your minimum threshold it records a virtual trade.

        Then it checks the next 10 candles to see if the trade
        hit the target (WIN) or stop loss (LOSS).

        **What good results look like:**

        | Metric | Target |
        |--------|--------|
        | Win Rate | 45% or above |
        | Profit Factor | 1.5 or above |
        | Avg Win > Avg Loss | Always |

        **Important:** Backtesting on daily (1d) timeframe
        is most reliable. 15m backtests have many false
        signals due to noise.
        """)


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 11 — SIGNAL HUB                                ║
# ╚══════════════════════════════════════════════════════╝
with T9:
    st.markdown("### 🎯 Signal Hub — All-in-One Analysis")
    st.caption(
        "Everything in one screen — Technical signals, "
        "OI Change and IV Rank combined. "
        "No more switching tabs."
    )

    # ── Step 1: Select Index and Stock ────────────────────
    hub_c1, hub_c2, hub_c3 = st.columns(3)
    with hub_c1:
        hub_index = st.selectbox(
            "Index for OI Analysis",
            ["NIFTY","BANKNIFTY","FINNIFTY"],
            key="hub_index"
        )
    with hub_c2:
        hub_stock = st.text_input(
            "Stock to analyse",
            value=st.session_state.get("sn","NIFTY 50"),
            key="hub_stock"
        )
        hub_sym = STOCKS.get(hub_stock, "^NSEI")
    with hub_c3:
        hub_tf = st.selectbox(
            "Timeframe",
            ["15m","30m","1h","1d"],
            index=0,
            key="hub_tf"
        )

    if st.button(
        "🚀 Run Full Analysis",
        type="primary",
        key="hub_run",
        use_container_width=True
    ):
        st.session_state["hub_run_flag"] = True

    if not st.session_state.get("hub_run_flag"):
        st.info(
            "Select your index and stock above, "
            "then click Run Full Analysis."
        )
    else:
        with st.spinner("Loading all data simultaneously..."):
            # Fetch all data in parallel
            hub_lp   = live_price(hub_sym)
            hub_df   = candles(hub_sym, hub_tf)
            hub_sig  = None
            if hub_df is not None and len(hub_df) >= 55:
                hub_sig = compute_all(hub_df, hub_lp)

            # ML
            hub_ml_dir  = "UNKNOWN"
            hub_ml_conf = 0
            hub_ml_ok   = False
            try:
                hub_df_d = candles(hub_sym, "1d")
                if hub_df_d is not None and len(hub_df_d) >= 100:
                    hub_ml_model = train_model(hub_df_d)
                    if hub_ml_model.get("ok"):
                        hub_ml_pred = predict_next_move(
                            hub_df_d, hub_ml_model
                        )
                        if hub_ml_pred and hub_ml_pred.get("ok"):
                            hub_ml_dir  = hub_ml_pred["prediction"]
                            hub_ml_conf = hub_ml_pred["confidence"]
                            hub_ml_ok   = True
            except Exception:
                pass

            # IV Rank
            hub_iv = get_iv_rank()

            # OI Change
            hub_oi = get_oi_change_data(hub_index)

            # MTF
            hub_1h_dir = "UNKNOWN"
            hub_1d_dir = "UNKNOWN"
            try:
                df_1h = candles(hub_sym, "1h")
                if df_1h is not None and len(df_1h) >= 55:
                    sig_1h = compute_all(df_1h, hub_lp)
                    if sig_1h:
                        hub_1h_dir = sig_1h["direction"]
            except Exception:
                pass
            try:
                df_1d = candles(hub_sym, "1d")
                if df_1d is not None and len(df_1d) >= 55:
                    sig_1d = compute_all(df_1d, hub_lp)
                    if sig_1d:
                        hub_1d_dir = sig_1d["direction"]
            except Exception:
                pass

        if not hub_sig:
            st.error(
                f"Not enough data for {hub_stock}. "
                "Try a different stock or timeframe."
            )
        else:
            hub_dir   = hub_sig["direction"]
            hub_score = max(
                hub_sig["up_score"],
                hub_sig["dn_score"]
            )
            hub_col = (
                "#16a34a" if hub_dir=="UPTREND"
                else "#dc2626" if hub_dir=="DOWNTREND"
                else "#f59e0b"
            )

            # ══ SECTION 1: Overall verdict ════════════════
            # Count confirmations
            confirms = []
            conflicts = []

            # Technical score
            if hub_score >= 7:
                confirms.append(
                    f"✅ Technical score {hub_score}/10"
                )
            else:
                conflicts.append(
                    f"❌ Technical score {hub_score}/10 — weak"
                )

            # ML
            if hub_ml_ok and hub_ml_dir == hub_dir:
                confirms.append(
                    f"✅ ML confirms {hub_ml_dir} "
                    f"({hub_ml_conf}%)"
                )
            elif hub_ml_ok:
                conflicts.append(
                    f"❌ ML says {hub_ml_dir} — conflicts"
                )

            # MTF
            if hub_1h_dir == hub_dir:
                confirms.append("✅ 1h timeframe agrees")
            else:
                conflicts.append(
                    f"❌ 1h says {hub_1h_dir}"
                )
            if hub_1d_dir == hub_dir:
                confirms.append("✅ 1d timeframe agrees")
            else:
                conflicts.append(
                    f"❌ 1d says {hub_1d_dir}"
                )

            # OI Change
            if hub_oi["ok"]:
                oi_sig = hub_oi["oi_signal"]
                if (hub_dir=="UPTREND" and
                        "BULLISH" in oi_sig):
                    confirms.append(
                        f"✅ OI signal: {oi_sig}"
                    )
                elif (hub_dir=="DOWNTREND" and
                        "BEARISH" in oi_sig):
                    confirms.append(
                        f"✅ OI signal: {oi_sig}"
                    )
                elif "UNWIND" in oi_sig or "MIXED" in oi_sig:
                    conflicts.append(
                        f"⚠️ OI signal: {oi_sig}"
                    )
                else:
                    conflicts.append(
                        f"❌ OI signal: {oi_sig} — conflicts"
                    )

            # IV Rank
            if hub_iv["ok"]:
                if hub_iv["iv_rank"] < 30:
                    confirms.append(
                        f"✅ IV Rank {hub_iv['iv_rank']} — "
                        f"options cheap"
                    )
                elif hub_iv["iv_rank"] > 70:
                    conflicts.append(
                        f"⚠️ IV Rank {hub_iv['iv_rank']} — "
                        f"options expensive"
                    )
                else:
                    confirms.append(
                        f"✅ IV Rank {hub_iv['iv_rank']} — "
                        f"options fairly priced"
                    )

            # Supertrend
            if hub_sig.get("st_bull") == (hub_dir=="UPTREND"):
                confirms.append("✅ Supertrend confirms")
            else:
                conflicts.append("❌ Supertrend conflicts")

            # CPR
            cpr_pos = hub_sig.get("cpr_position","INSIDE")
            if (hub_dir=="UPTREND" and cpr_pos=="ABOVE"):
                confirms.append("✅ Price above CPR")
            elif (hub_dir=="DOWNTREND" and cpr_pos=="BELOW"):
                confirms.append("✅ Price below CPR")
            elif cpr_pos == "INSIDE":
                conflicts.append("⚠️ Price inside CPR")

            # Final grade
            total   = len(confirms) + len(conflicts)
            conf_ct = len(confirms)

            if conf_ct >= 7:
                grade     = "💎 DIAMOND"
                grade_col = "#7c3aed"
                grade_bg  = "linear-gradient(135deg,#1e1b4b,#3730a3)"
                grade_txt = "#ffffff"
                action    = "ENTER TRADE"
            elif conf_ct >= 5:
                grade     = "🔥 STRONG"
                grade_col = "#16a34a"
                grade_bg  = "#f0fdf4"
                grade_txt = "#166534"
                action    = "ENTER WITH NORMAL SIZE"
            elif conf_ct >= 3:
                grade     = "⚡ MODERATE"
                grade_col = "#d97706"
                grade_bg  = "#fffbeb"
                grade_txt = "#92400e"
                action    = "ENTER WITH HALF SIZE"
            else:
                grade     = "⚠️ WEAK"
                grade_col = "#dc2626"
                grade_bg  = "#fef2f2"
                grade_txt = "#991b1b"
                action    = "DO NOT TRADE"

            # ── BIG VERDICT BANNER ─────────────────────────
            st.markdown(
                f"<div style='background:{grade_bg};"
                f"border-radius:16px;padding:24px;"
                f"text-align:center;margin-bottom:16px'>"
                f"<div style='font-size:36px;font-weight:700;"
                f"color:{grade_txt}'>{grade}</div>"
                f"<div style='font-size:18px;color:{grade_txt};"
                f"margin-top:8px;font-weight:600'>"
                f"{hub_stock} — {hub_dir}"
                f"</div>"
                f"<div style='font-size:16px;color:{grade_txt};"
                f"margin-top:6px'>{action}</div>"
                f"<div style='font-size:13px;color:{grade_txt};"
                f"opacity:0.8;margin-top:8px'>"
                f"{conf_ct} of {total} factors confirm"
                f"</div></div>",
                unsafe_allow_html=True
            )

            # ── 3 COLUMN LAYOUT ────────────────────────────
            col_tech, col_oi, col_iv = st.columns(3)

            # Column 1: Technical + ML + MTF
            with col_tech:
                st.markdown(
                    "<div style='background:#ffffff;"
                    "border:1px solid #e2e8f0;"
                    "border-radius:12px;padding:16px;"
                    "height:100%'>"
                    "<div style='font-size:13px;font-weight:700;"
                    "color:#374151;margin-bottom:12px;"
                    "text-transform:uppercase;letter-spacing:1px'>"
                    "📊 Technical Analysis</div>",
                    unsafe_allow_html=True
                )
                # Score
                st.markdown(
                    f"<div style='text-align:center;"
                    f"padding:12px;background:#f8fafc;"
                    f"border-radius:8px;margin-bottom:10px'>"
                    f"<div style='font-size:11px;color:#64748b'>"
                    f"Signal Score</div>"
                    f"<div style='font-size:36px;font-weight:700;"
                    f"color:{hub_col}'>{hub_score}/10</div>"
                    f"<div style='font-size:13px;color:{hub_col};"
                    f"font-weight:600'>{hub_dir}</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                # Key indicators
                indicators = [
                    ("RSI", f"{hub_sig['rv']:.0f}",
                     "#16a34a" if 50<hub_sig['rv']<75 else "#dc2626"),
                    ("Supertrend",
                     "BUY" if hub_sig.get("st_bull") else "SELL",
                     "#16a34a" if hub_sig.get("st_bull") else "#dc2626"),
                    ("CPR", hub_sig.get("cpr_position","—"),
                     "#16a34a" if hub_sig.get("cpr_position")=="ABOVE"
                     else "#dc2626" if hub_sig.get("cpr_position")=="BELOW"
                     else "#f59e0b"),
                    ("VWAP",
                     "Above" if hub_sig["cp"]>hub_sig["vwv"]
                     else "Below",
                     "#16a34a" if hub_sig["cp"]>hub_sig["vwv"]
                     else "#dc2626"),
                    ("ML (1d)", hub_ml_dir,
                     "#16a34a" if hub_ml_dir=="UPTREND"
                     else "#dc2626" if hub_ml_dir=="DOWNTREND"
                     else "#f59e0b"),
                    ("1h TF", hub_1h_dir,
                     "#16a34a" if hub_1h_dir==hub_dir
                     else "#dc2626"),
                    ("1d TF", hub_1d_dir,
                     "#16a34a" if hub_1d_dir==hub_dir
                     else "#dc2626"),
                ]
                for ind_name, ind_val, ind_col in indicators:
                    st.markdown(
                        f"<div style='display:flex;"
                        f"justify-content:space-between;"
                        f"padding:5px 0;border-bottom:"
                        f"1px solid #f1f5f9;font-size:12px'>"
                        f"<span style='color:#64748b'>"
                        f"{ind_name}</span>"
                        f"<span style='font-weight:600;"
                        f"color:{ind_col}'>{ind_val}</span>"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                st.markdown("</div>", unsafe_allow_html=True)

            # Column 2: OI Change
            with col_oi:
                st.markdown(
                    "<div style='background:#ffffff;"
                    "border:1px solid #e2e8f0;"
                    "border-radius:12px;padding:16px;"
                    "height:100%'>"
                    "<div style='font-size:13px;font-weight:700;"
                    "color:#374151;margin-bottom:12px;"
                    "text-transform:uppercase;letter-spacing:1px'>"
                    "⚡ OI Change Analysis</div>",
                    unsafe_allow_html=True
                )
                if hub_oi["ok"]:
                    _oic = hub_oi["oi_color"]
                    st.markdown(
                        f"<div style='text-align:center;"
                        f"padding:12px;background:{_oic}22;"
                        f"border-radius:8px;margin-bottom:10px'>"
                        f"<div style='font-size:18px;"
                        f"font-weight:700;color:{_oic}'>"
                        f"{hub_oi['oi_signal']}</div>"
                        f"<div style='font-size:11px;color:#64748b;"
                        f"margin-top:4px'>"
                        f"{hub_oi['oi_advice']}</div>"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                    oi_levels = [
                        ("Resistance (Max Call OI)",
                         f"₹{hub_oi['max_call_oi_strike']:,}",
                         "#dc2626"),
                        ("Support (Max Put OI)",
                         f"₹{hub_oi['max_put_oi_strike']:,}",
                         "#16a34a"),
                        ("Call OI Build",
                         f"₹{hub_oi['max_call_chg_strike']:,}",
                         "#dc2626"),
                        ("Put OI Build",
                         f"₹{hub_oi['max_put_chg_strike']:,}",
                         "#16a34a"),
                        ("Total Call OI Chg",
                         f"{hub_oi['total_call_oi_chg']:+,}",
                         "#16a34a" if hub_oi["total_call_oi_chg"]<0
                         else "#dc2626"),
                        ("Total Put OI Chg",
                         f"{hub_oi['total_put_oi_chg']:+,}",
                         "#16a34a" if hub_oi["total_put_oi_chg"]>0
                         else "#dc2626"),
                        ("Spot Price",
                         f"₹{hub_oi['spot']:,.0f}",
                         "#374151"),
                        ("ATM Strike",
                         f"₹{hub_oi['atm']:,}",
                         "#7c3aed"),
                    ]
                    for _n, _v, _c in oi_levels:
                        st.markdown(
                            f"<div style='display:flex;"
                            f"justify-content:space-between;"
                            f"padding:5px 0;border-bottom:"
                            f"1px solid #f1f5f9;font-size:12px'>"
                            f"<span style='color:#64748b'>{_n}"
                            f"</span>"
                            f"<span style='font-weight:600;"
                            f"color:{_c}'>{_v}</span>"
                            f"</div>",
                            unsafe_allow_html=True
                        )
                else:
                    st.warning(
                        "OI data unavailable. "
                        "Try during market hours."
                    )
                st.markdown("</div>", unsafe_allow_html=True)

            # Column 3: IV Rank + Entry levels
            with col_iv:
                st.markdown(
                    "<div style='background:#ffffff;"
                    "border:1px solid #e2e8f0;"
                    "border-radius:12px;padding:16px;"
                    "height:100%'>"
                    "<div style='font-size:13px;font-weight:700;"
                    "color:#374151;margin-bottom:12px;"
                    "text-transform:uppercase;letter-spacing:1px'>"
                    "📈 IV Rank + Entry Levels</div>",
                    unsafe_allow_html=True
                )
                if hub_iv["ok"]:
                    _ivc = hub_iv["color"]
                    _ivb = hub_iv["bg"]
                    st.markdown(
                        f"<div style='text-align:center;"
                        f"padding:12px;background:{_ivb};"
                        f"border-radius:8px;margin-bottom:10px'>"
                        f"<div style='font-size:11px;"
                        f"color:#64748b'>IV Rank</div>"
                        f"<div style='font-size:36px;"
                        f"font-weight:700;color:{_ivc}'>"
                        f"{hub_iv['iv_rank']}</div>"
                        f"<div style='font-size:13px;"
                        f"color:{_ivc};font-weight:600'>"
                        f"{hub_iv['signal']}</div>"
                        f"<div style='font-size:11px;"
                        f"color:#64748b;margin-top:4px'>"
                        f"{hub_iv['advice']}</div>"
                        f"</div>",
                        unsafe_allow_html=True
                    )

                # Entry / SL / Target
                _hub_entry = hub_sig.get(
                    "entry_long" if hub_dir=="UPTREND"
                    else "entry_short",
                    hub_sig["e9v"]
                )
                _hub_sl = (
                    hub_sig["sl_long"]
                    if hub_dir=="UPTREND"
                    else hub_sig["sl_short"]
                )
                _hub_t1 = (
                    hub_sig["tgt1"]
                    if hub_dir=="UPTREND"
                    else hub_sig["tgt1s"]
                )
                _hub_t2 = (
                    hub_sig["tgt2"]
                    if hub_dir=="UPTREND"
                    else hub_sig["tgt2s"]
                )
                _hub_rr = round(
                    abs(_hub_t1 - _hub_entry) /
                    (abs(_hub_entry - _hub_sl) + 0.001),
                    2
                )

                # VWAP Zone in Signal Hub
                _hz = hub_sig.get("vwap_zone","FAIR_VALUE")
                _hz_map = {
                    "EXTREME_OB": ("🔴 Extreme Overbought", "#dc2626"),
                    "OVERBOUGHT": ("🟠 Overbought", "#ea580c"),
                    "FAIR_VALUE": ("🟢 Fair Value", "#16a34a"),
                    "OVERSOLD":   ("🟢 Oversold — CE Entry!", "#16a34a"),
                    "EXTREME_OS": ("💎 Extreme Oversold!", "#7c3aed"),
                }
                _hz_lbl, _hz_col = _hz_map.get(
                    _hz, ("Normal", "#374151")
                )

                levels = [
                    ("Current Price",
                     f"₹{hub_sig['cp']:,.2f}",
                     "#374151"),
                    ("VWAP Zone",
                     _hz_lbl,
                     _hz_col),
                    ("Entry Zone",
                     f"₹{_hub_entry:,.2f}",
                     "#16a34a"),
                    ("Stop Loss",
                     f"₹{_hub_sl:,.2f}",
                     "#dc2626"),
                    ("Target 1",
                     f"₹{_hub_t1:,.2f}",
                     "#1d4ed8"),
                    ("Target 2",
                     f"₹{_hub_t2:,.2f}",
                     "#1d4ed8"),
                    ("R:R Ratio",
                     f"{_hub_rr}:1",
                     "#16a34a" if _hub_rr>=1.5 else "#dc2626"),
                    ("ATR",
                     f"₹{hub_sig['atrv']:,.2f}",
                     "#374151"),
                ]
                for _n, _v, _c in levels:
                    st.markdown(
                        f"<div style='display:flex;"
                        f"justify-content:space-between;"
                        f"padding:5px 0;border-bottom:"
                        f"1px solid #f1f5f9;font-size:12px'>"
                        f"<span style='color:#64748b'>{_n}</span>"
                        f"<span style='font-weight:600;"
                        f"color:{_c}'>{_v}</span>"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                st.markdown("</div>", unsafe_allow_html=True)

            # ── Expected Move Calculator ──────────────────
            st.markdown("---")
            st.markdown("#### 📐 Expected Move Calculator")
            em1, em2, em3 = st.columns(3)
            with em1:
                em_opt_price = st.number_input(
                    "Option premium you plan to pay (₹)",
                    value=100.0, step=5.0, key="em_opt_price"
                )
                em_delta = st.slider(
                    "Estimated Delta",
                    0.1, 0.9, 0.5, 0.05,
                    key="em_delta",
                    help="ATM=0.5, ITM=0.6-0.8, OTM=0.2-0.4"
                )
            with em2:
                em_lot = st.number_input(
                    "Lot size", value=50, key="em_lot"
                )
                em_lots = st.number_input(
                    "Number of lots", value=1, key="em_lots"
                )
            with em3:
                em_target_pct = st.slider(
                    "Target profit on option (%)",
                    10, 200, 50, 10, key="em_target_pct"
                )

            if em_opt_price > 0 and em_delta > 0:
                _em_stock_move = round(em_opt_price / em_delta, 2)
                _em_target_opt = round(em_opt_price * (1 + em_target_pct/100), 2)
                _em_target_move= round(_em_target_opt / em_delta, 2)
                _em_sl_opt     = round(em_opt_price * 0.5, 2)
                _em_sl_move    = round(em_opt_price * 0.5 / em_delta, 2)
                _em_total_risk = round(em_opt_price * em_lot * em_lots, 2)
                _em_total_gain = round(
                    (em_opt_price * em_target_pct/100) * em_lot * em_lots, 2
                )

                st.markdown(
                    f"<div style='background:#faf5ff;"
                    f"border:1.5px solid #7c3aed;"
                    f"border-radius:12px;padding:16px;"
                    f"margin-bottom:8px'>"
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#6d28d9;margin-bottom:10px'>"
                    f"Expected Move Analysis</div>"
                    f"<div style='display:grid;"
                    f"grid-template-columns:repeat(3,1fr);gap:10px'>"

                    f"<div style='background:white;border-radius:8px;"
                    f"padding:10px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"Stock must move</div>"
                    f"<div style='font-size:18px;font-weight:700;"
                    f"color:#dc2626'>₹{_em_stock_move:,.0f}</div>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"just to break even</div></div>"

                    f"<div style='background:white;border-radius:8px;"
                    f"padding:10px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"For {em_target_pct}% profit</div>"
                    f"<div style='font-size:18px;font-weight:700;"
                    f"color:#16a34a'>₹{_em_target_move:,.0f}</div>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"stock move needed</div></div>"

                    f"<div style='background:white;border-radius:8px;"
                    f"padding:10px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"Total investment</div>"
                    f"<div style='font-size:18px;font-weight:700;"
                    f"color:#374151'>₹{_em_total_risk:,.0f}</div>"
                    f"<div style='font-size:10px;color:#16a34a'>"
                    f"Target P&L: ₹{_em_total_gain:+,.0f}</div></div>"

                    f"</div>"
                    f"<div style='font-size:12px;color:#6d28d9;"
                    f"margin-top:10px'>"
                    f"50% SL on option = stock reverses ₹{_em_sl_move:,.0f} | "
                    f"Option SL at ₹{_em_sl_opt:,.0f}</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )

            # ── Confirmation checklist ─────────────────────
            st.markdown("---")
            chk1, chk2 = st.columns(2)
            with chk1:
                st.markdown(
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#16a34a;margin-bottom:8px'>"
                    f"✅ Confirming factors "
                    f"({len(confirms)})</div>",
                    unsafe_allow_html=True
                )
                for c in confirms:
                    st.markdown(
                        f"<div style='background:#f0fdf4;"
                        f"border-left:3px solid #86efac;"
                        f"padding:6px 12px;margin:3px 0;"
                        f"border-radius:0 6px 6px 0;"
                        f"font-size:12px;color:#166534'>"
                        f"{c}</div>",
                        unsafe_allow_html=True
                    )
            with chk2:
                st.markdown(
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#dc2626;margin-bottom:8px'>"
                    f"❌ Conflicting factors "
                    f"({len(conflicts)})</div>",
                    unsafe_allow_html=True
                )
                for c in conflicts:
                    st.markdown(
                        f"<div style='background:#fef2f2;"
                        f"border-left:3px solid #fca5a5;"
                        f"padding:6px 12px;margin:3px 0;"
                        f"border-radius:0 6px 6px 0;"
                        f"font-size:12px;color:#991b1b'>"
                        f"{c}</div>",
                        unsafe_allow_html=True
                    )

            # ── Send to Telegram ───────────────────────────
            st.markdown("---")
            if tg_configured():
                if st.button(
                    "📱 Send Full Analysis to Telegram",
                    key="hub_tg",
                    type="primary",
                    use_container_width=True
                ):
                    _tok = st.session_state.get(
                        "tg_token_saved",""
                    )
                    _cid = st.session_state.get(
                        "tg_chat_saved",""
                    )
                    _msg = (
                        f"{grade} — {hub_stock}\n"
                        f"{hub_dir} | Score {hub_score}/10\n"
                        f"ML: {hub_ml_dir} ({hub_ml_conf}%)\n"
                        f"OI: {hub_oi.get('oi_signal','—')}\n"
                        f"IV Rank: {hub_iv.get('iv_rank','—')} "
                        f"({hub_iv.get('signal','—')})\n"
                        f"Entry: Rs{_hub_entry:,.0f} | "
                        f"SL: Rs{_hub_sl:,.0f}\n"
                        f"T1: Rs{_hub_t1:,.0f} | "
                        f"R:R {_hub_rr}:1\n"
                        f"Confirms: {len(confirms)}/{total} | "
                        f"Action: {action}"
                    )
                    if send_telegram(_tok, _cid, _msg):
                        st.success("✅ Full analysis sent!")
                    else:
                        st.error("❌ Failed to send")


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 12 — TRADE MANAGER                             ║
# ╚══════════════════════════════════════════════════════╝
with T10:
    st.markdown("### 🛡️ Trade Manager — Active Trade Monitor")
    st.caption(
        "Add your active trades here. The system monitors each trade "
        "every candle and tells you HOLD / REDUCE / EXIT with full "
        "reasoning. Signals never disappear silently."
    )

    # ── Initialize trade store ─────────────────────────────
    # Load from file if session_state is empty (page refresh etc)
    if "active_trades" not in st.session_state:
        st.session_state["active_trades"] = load_trades()
    if "trade_journal" not in st.session_state:
        st.session_state["trade_journal"] = load_journal()

    # ── Add new trade ──────────────────────────────────────
    with st.expander(
        "➕ Add New Trade",
        expanded=len(st.session_state["active_trades"]) == 0
    ):
        af1, af2, af3 = st.columns(3)
        with af1:
            tm_stock = st.text_input(
                "Stock name", value="NIFTY 50", key="tm_stock"
            )
            # Case-insensitive stock lookup
            tm_sym = None
            for _k, _v in STOCKS.items():
                if _k.lower() == tm_stock.lower():
                    tm_sym = _v
                    tm_stock = _k  # use correct case
                    break
            if not tm_sym:
                tm_sym = "^NSEI"  # fallback
            tm_type = st.selectbox(
                "Signal type", ["BUY CE","BUY PE"], key="tm_type"
            )
            tm_style = st.selectbox(
                "Trade style",
                ["Intraday (exit 2:45 PM)",
                 "Swing 1 day",
                 "Swing 3 days",
                 "Swing 1 week"],
                key="tm_style"
            )
        with af2:
            tm_entry  = st.number_input(
                "Entry price (₹)", value=0.0,
                step=0.5, key="tm_entry"
            )
            tm_sl     = st.number_input(
                "Stop loss (₹)", value=0.0,
                step=0.5, key="tm_sl"
            )
            tm_target = st.number_input(
                "Target (₹)", value=0.0,
                step=0.5, key="tm_target"
            )
        with af3:
            tm_lots = st.number_input(
                "Lots", value=1, min_value=1, key="tm_lots"
            )
            tm_tf   = st.selectbox(
                "Timeframe", ["15m","1h","1d"], key="tm_tf"
            )
            tm_opt_price = st.number_input(
                "Option premium paid (₹)",
                value=0.0, step=0.5, key="tm_opt_price",
                help="Price you paid for the CE/PE option"
            )

        if st.button(
            "Add Trade to Monitor",
            type="primary", key="tm_add",
            use_container_width=True
        ):
            # Validate stock name
            if tm_sym == "^NSEI" and tm_stock != "NIFTY 50":
                st.warning(
                    f"Stock '{tm_stock}' not found in database. "
                    "Check spelling — use exact name like "
                    "'Titan Company', 'HDFC Bank', 'Reliance'"
                )
            if tm_entry > 0 and tm_sl > 0 and tm_target > 0:
                import datetime as _dt2
                st.session_state["active_trades"].append({
                    "id":           len(st.session_state["active_trades"])+1,
                    "stock":        tm_stock,
                    "sym":          tm_sym,
                    "type":         tm_type,
                    "entry":        tm_entry,
                    "sl":           tm_sl,
                    "target":       tm_target,
                    "lots":         tm_lots,
                    "lots_rem":     tm_lots,
                    "style":        tm_style,
                    "tf":           tm_tf,
                    "opt_price":    tm_opt_price,
                    "added_at":     _dt2.datetime.now().strftime("%d %b %H:%M"),
                    "status":       "ACTIVE",
                    "last_action":  "Trade added",
                })
                # Save to file immediately
                save_trades(st.session_state["active_trades"])
                st.success(
                    f"✅ {tm_stock} {tm_type} added to Trade Manager!"
                )
                st.rerun()
            else:
                st.error(
                    "Please fill Entry, Stop Loss and Target."
                )

    # ── Monitor active trades ──────────────────────────────
    trades = st.session_state.get("active_trades", [])
    active = [t for t in trades if t["status"] == "ACTIVE"]
    closed = [t for t in trades if t["status"] != "ACTIVE"]

    if not active:
        st.info(
            "No active trades being monitored. "
            "Add a trade above after entering on Zerodha."
        )
    else:
        st.markdown(f"### 📊 Monitoring {len(active)} active trade(s)")

        rc1, rc2 = st.columns([3,1])
        with rc1:
            st.caption(
                "Click Refresh to fetch latest signals. "
                "Live price updates on every page load."
            )
        with rc2:
            if st.button(
                "🔄 Refresh Signals",
                key="tm_refresh_all",
                type="primary",
                use_container_width=True
            ):
                st.session_state["tm_refresh"] = True
                for _t in active:
                    st.session_state.pop(
                        f"tm_sig_{_t['id']}", None
                    )
                st.rerun()

        # Clear refresh flag after use
        if st.session_state.get("tm_refresh"):
            st.session_state["tm_refresh"] = False

        for idx, trade in enumerate(active):
            st.markdown("---")

            # ── Fast loading — cache everything ────────
            # Live price: cached 30 seconds
            _tm_lp_key = f"tm_lp_{trade['sym']}"
            _tm_lp_ts  = f"tm_lp_ts_{trade['sym']}"
            import time as _tm_time
            _now_ts = _tm_time.time()
            _last_ts = st.session_state.get(_tm_lp_ts, 0)

            if _now_ts - _last_ts > 30:  # refresh every 30 sec
                tm_lp = live_price(trade["sym"])
                st.session_state[_tm_lp_key] = tm_lp
                st.session_state[_tm_lp_ts]  = _now_ts
            else:
                tm_lp = st.session_state.get(
                    _tm_lp_key,
                    {"ok": False, "p": trade["entry"]}
                )

            # Signal: only fetch on Refresh button click
            tm_sig = None
            if st.session_state.get("tm_refresh"):
                try:
                    tm_df = candles(trade["sym"], trade["tf"])
                    if tm_df is not None and len(tm_df) >= 55:
                        tm_sig = compute_all(tm_df, tm_lp)
                        st.session_state[
                            f"tm_sig_{trade['id']}"
                        ] = tm_sig
                except Exception:
                    pass

            # Use cached signal
            if tm_sig is None:
                tm_sig = st.session_state.get(
                    f"tm_sig_{trade['id']}", None
                )

            cp_now  = tm_lp["p"] if tm_lp["ok"] else trade["entry"]
            is_ce   = trade["type"] == "BUY CE"
            direction = "UPTREND" if is_ce else "DOWNTREND"

            # Stock price P&L
            pnl_pts = (
                (cp_now - trade["entry"]) if is_ce
                else (trade["entry"] - cp_now)
            )
            pnl_pct = round(pnl_pts / trade["entry"] * 100, 2)
            pnl_col = "#16a34a" if pnl_pts >= 0 else "#dc2626"

            # Options P&L calculation
            opt_paid = trade.get("opt_price", 0)
            opt_lot  = 50  # default NSE lot size
            # Estimate current option price using delta approximation
            # Delta ~ 0.5 for ATM, higher for ITM
            _atm_dist = abs(cp_now - trade["entry"])
            _atm_pct  = _atm_dist / (trade["entry"] + 0.001)
            _delta    = max(0.2, min(0.8, 0.5 + (pnl_pts/(trade["entry"]+0.001)) * 2))
            opt_curr_est = max(
                0.5,
                opt_paid + (pnl_pts * _delta)
            ) if opt_paid > 0 else 0

            opt_pnl_per_lot = round(
                (opt_curr_est - opt_paid) * opt_lot, 2
            ) if opt_paid > 0 else 0
            opt_pnl_total = round(
                opt_pnl_per_lot * trade.get("lots", 1), 2
            ) if opt_paid > 0 else 0
            opt_pnl_col = "#16a34a" if opt_pnl_total >= 0 else "#dc2626"

            # Theta decay warning
            # Options lose ~0.3-0.5% of premium per day for ATM options
            opt_theta_daily = round(opt_paid * 0.004, 2) if opt_paid > 0 else 0
            days_held = 1  # approximate

            # SL / Target hit (based on stock price)
            sl_hit  = (
                (cp_now <= trade["sl"]) if is_ce
                else (cp_now >= trade["sl"])
            )
            tgt_hit = (
                (cp_now >= trade["target"]) if is_ce
                else (cp_now <= trade["target"])
            )

            # Options-specific exit: if option lost > 50% of premium
            opt_stop_hit = (
                opt_paid > 0 and
                opt_curr_est < opt_paid * 0.5
            )

            # ── Strength analysis ──────────────────────────
            strength = 0
            confirms = []
            warnings = []

            if tm_sig:
                cur_dir   = tm_sig["direction"]
                cur_score = max(tm_sig["up_score"], tm_sig["dn_score"])

                # Direction
                if cur_dir == direction:
                    strength += 2
                    confirms.append(f"Trend still {direction}")
                else:
                    strength -= 2
                    warnings.append(f"Trend changed to {cur_dir}")

                # Score
                if cur_score >= 7:
                    strength += 2
                    confirms.append(f"Score {cur_score}/10 strong")
                elif cur_score >= 5:
                    strength += 1
                    confirms.append(f"Score {cur_score}/10 moderate")
                else:
                    strength -= 1
                    warnings.append(f"Score {cur_score}/10 weak")

                # RSI
                rsi = tm_sig["rv"]
                if is_ce and rsi > 50:
                    strength += 1
                    confirms.append(f"RSI bullish {rsi:.0f}")
                elif not is_ce and rsi < 50:
                    strength += 1
                    confirms.append(f"RSI bearish {rsi:.0f}")
                else:
                    strength -= 1
                    warnings.append(f"RSI against trade {rsi:.0f}")

                # VWAP
                if is_ce and cp_now > tm_sig["vwv"]:
                    strength += 1
                    confirms.append("Price above VWAP")
                elif not is_ce and cp_now < tm_sig["vwv"]:
                    strength += 1
                    confirms.append("Price below VWAP")
                else:
                    warnings.append("Price crossed VWAP against trade")

                # EMA21
                if is_ce and cp_now > tm_sig["e21v"]:
                    strength += 1
                    confirms.append("Price above EMA21")
                elif not is_ce and cp_now < tm_sig["e21v"]:
                    strength += 1
                    confirms.append("Price below EMA21")
                else:
                    warnings.append("Price crossed EMA21 against trade")

                # Supertrend
                st_ok = (
                    (is_ce and tm_sig.get("st_bull", True)) or
                    (not is_ce and not tm_sig.get("st_bull", True))
                )
                if st_ok:
                    strength += 1
                    confirms.append("Supertrend confirms")
                else:
                    warnings.append("Supertrend flipped against trade")

                # MACD — use safe .get() with fallback
                _tm_mv  = tm_sig.get("mv", 0)
                _tm_msv = tm_sig.get("msv", 0)
                macd_ok = (
                    (is_ce     and _tm_mv > _tm_msv) or
                    (not is_ce and _tm_mv < _tm_msv)
                )
                if macd_ok:
                    strength += 1
                    confirms.append("MACD confirms")
                else:
                    warnings.append("MACD turned against trade")

                # Volume
                if tm_sig["vsurge"]:
                    strength += 1
                    confirms.append(
                        f"Volume surge {tm_sig['vol_ratio']:.1f}x"
                    )

            # ── MTF check ──────────────────────────────────
            mtf_1h = "UNKNOWN"
            try:
                df_1h = candles(trade["sym"], "1h")
                if df_1h is not None and len(df_1h) >= 55:
                    sig_1h = compute_all(df_1h, tm_lp)
                    if sig_1h:
                        mtf_1h = sig_1h["direction"]
                        if mtf_1h == direction:
                            strength += 1
                            confirms.append("1h timeframe agrees")
                        else:
                            warnings.append(f"1h says {mtf_1h}")
            except Exception:
                pass

            # ── Noise vs reversal detection ─────────────────
            # Key logic: distinguish temporary weakness from real reversal
            # Real reversal = 3+ warnings AND score < 5 AND VWAP crossed
            real_reversal = (
                len(warnings) >= 3 and
                (tm_sig["up_score"] if is_ce else tm_sig["dn_score"]) < 5
                if tm_sig else False
            )
            temp_weakness = (
                len(warnings) >= 2 and not real_reversal
            )

            # ── Holding time recommendation ─────────────────
            style = trade["style"]
            if "Intraday" in style:
                hold_advice  = "Exit by 2:45 PM regardless"
                max_hold     = "Today only"
                theta_urgent = opt_theta_daily > 0
            elif "1 day" in style:
                hold_advice  = "Review tomorrow morning before 9:30 AM"
                max_hold     = "1 trading day"
                theta_urgent = opt_theta_daily > opt_paid * 0.05
            elif "3 days" in style:
                hold_advice  = "Review every morning. Exit if score drops below 5."
                max_hold     = "3 trading days"
                theta_urgent = opt_theta_daily > opt_paid * 0.03
            else:
                hold_advice  = "Review every morning. Weekly options lose value fast."
                max_hold     = "5 trading days"
                theta_urgent = opt_theta_daily > opt_paid * 0.02

            # Theta warning for swing trades
            if theta_urgent and opt_paid > 0:
                warnings.append(
                    f"⚠️ Theta eating ₹{opt_theta_daily:.0f}/day — "
                    f"option losing value even if stock stays flat"
                )

            # ── Final verdict ───────────────────────────────
            if opt_stop_hit and not sl_hit:
                verdict  = "🔴 EXIT NOW — OPTION DOWN 50%"
                v_col    = "#dc2626"
                v_bg     = "#fef2f2"
                v_action = (
                    f"Option premium dropped from "
                    f"₹{opt_paid} to ~₹{opt_curr_est:.0f}. "
                    "50% stop loss on option hit. Exit now."
                )
            elif sl_hit:
                verdict     = "🔴 EXIT NOW — STOP LOSS HIT"
                v_col       = "#dc2626"
                v_bg        = "#fef2f2"
                v_action    = "Exit immediately on Zerodha. No waiting."
                trade["status"] = "CLOSED — SL HIT"
            elif tgt_hit:
                verdict     = "🟢 EXIT NOW — TARGET HIT"
                v_col       = "#16a34a"
                v_bg        = "#f0fdf4"
                v_action    = "Book full profit. Excellent trade!"
                trade["status"] = "CLOSED — TARGET HIT"
            elif real_reversal:
                verdict     = "🔴 EXIT NOW — TREND REVERSED"
                v_col       = "#dc2626"
                v_bg        = "#fef2f2"
                v_action    = (
                    "This is a real reversal not noise. "
                    "Exit full position immediately."
                )
            elif temp_weakness:
                verdict     = "🟡 REDUCE — TEMPORARY WEAKNESS"
                v_col       = "#d97706"
                v_bg        = "#fffbeb"
                v_action    = (
                    "Exit 50% now to protect profit/reduce loss. "
                    "Hold remaining 50% with tight stop."
                )
            elif strength >= 6:
                verdict     = "🟢 HOLD — STRONG SIGNAL"
                v_col       = "#16a34a"
                v_bg        = "#f0fdf4"
                v_action    = (
                    f"Hold full position. {hold_advice}."
                )
            elif strength >= 3:
                verdict     = "🟡 HOLD WITH CAUTION"
                v_col       = "#d97706"
                v_bg        = "#fffbeb"
                v_action    = (
                    "Signal still valid but weakening. "
                    "Trail stop loss closer to current price."
                )
            else:
                verdict     = "🟠 REDUCE — SIGNAL WEAKENING"
                v_col       = "#ea580c"
                v_bg        = "#fff7ed"
                v_action    = (
                    "Exit 50% now. Move SL to breakeven for rest."
                )

            # ── Display trade card ──────────────────────────
            st.markdown(
                f"<div style='background:{v_bg};"
                f"border:2px solid {v_col};"
                f"border-radius:14px;padding:20px;"
                f"margin-bottom:8px'>"

                # Header row
                f"<div style='display:flex;justify-content:"
                f"space-between;align-items:center;"
                f"flex-wrap:wrap;gap:8px;margin-bottom:14px'>"
                f"<div>"
                f"<span style='font-size:20px;font-weight:700;"
                f"color:#1e293b'>{trade['stock']}</span>"
                f"<span style='background:#1e293b;color:white;"
                f"padding:3px 10px;border-radius:10px;"
                f"font-size:12px;margin-left:8px'>"
                f"{trade['type']}</span>"
                f"<span style='font-size:11px;color:#64748b;"
                f"margin-left:8px'>Added {trade['added_at']}"
                f" | {trade['style']}</span>"
                f"</div>"
                f"<div style='font-size:20px;font-weight:700;"
                f"color:{pnl_col}'>"
                f"{pnl_pts:+.2f} pts ({pnl_pct:+.2f}%)"
                f"</div></div>"

                # Verdict banner
                f"<div style='background:{v_col};color:white;"
                f"border-radius:8px;padding:12px 16px;"
                f"margin-bottom:12px'>"
                f"<div style='font-size:18px;font-weight:700'>"
                f"{verdict}</div>"
                f"<div style='font-size:13px;margin-top:4px'>"
                f"{v_action}</div>"
                f"</div>"

                # Price levels
                f"<div style='display:grid;"
                f"grid-template-columns:repeat(5,1fr);"
                f"gap:8px;margin-bottom:12px'>"

                f"<div style='background:white;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>Entry</div>"
                f"<div style='font-size:15px;font-weight:700;"
                f"color:#374151'>₹{trade['entry']:,.2f}</div></div>"

                f"<div style='background:white;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>"
                f"Current</div>"
                f"<div style='font-size:15px;font-weight:700;"
                f"color:{pnl_col}'>₹{cp_now:,.2f}</div></div>"

                f"<div style='background:#fef2f2;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>SL</div>"
                f"<div style='font-size:15px;font-weight:700;"
                f"color:#dc2626'>₹{trade['sl']:,.2f}</div></div>"

                f"<div style='background:#f0fdf4;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>"
                f"Target</div>"
                f"<div style='font-size:15px;font-weight:700;"
                f"color:#16a34a'>₹{trade['target']:,.2f}</div></div>"

                f"<div style='background:#eff6ff;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>"
                f"Max Hold</div>"
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#1d4ed8'>{max_hold}</div></div>"
                f"</div>"

                + (
                    f"<div style='display:grid;"
                    f"grid-template-columns:repeat(4,1fr);"
                    f"gap:8px;margin-top:8px'>"

                    f"<div style='background:#faf5ff;"
                    f"border-radius:8px;padding:10px;"
                    f"text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"Option Paid</div>"
                    f"<div style='font-size:14px;font-weight:700;"
                    f"color:#7c3aed'>₹{opt_paid:,.0f}</div></div>"

                    f"<div style='background:#faf5ff;"
                    f"border-radius:8px;padding:10px;"
                    f"text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"Est. Now</div>"
                    f"<div style='font-size:14px;font-weight:700;"
                    f"color:{opt_pnl_col}'>"
                    f"₹{opt_curr_est:,.0f}</div></div>"

                    f"<div style='background:#faf5ff;"
                    f"border-radius:8px;padding:10px;"
                    f"text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"Option P&L</div>"
                    f"<div style='font-size:14px;font-weight:700;"
                    f"color:{opt_pnl_col}'>"
                    f"₹{opt_pnl_total:+,.0f}</div></div>"

                    f"<div style='background:#faf5ff;"
                    f"border-radius:8px;padding:10px;"
                    f"text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>"
                    f"Theta/day</div>"
                    f"<div style='font-size:14px;font-weight:700;"
                    f"color:#dc2626'>-₹{opt_theta_daily:,.0f}</div>"
                    f"</div></div>"
                    if opt_paid > 0 else ""
                )

                + f"</div>",
                unsafe_allow_html=True
            )

            # Confirms and warnings
            cw1, cw2 = st.columns(2)
            with cw1:
                if confirms:
                    st.markdown(
                        f"**✅ Supporting factors ({len(confirms)})**"
                    )
                    for c in confirms:
                        st.markdown(
                            f"<div style='background:#f0fdf4;"
                            f"border-left:3px solid #86efac;"
                            f"padding:5px 10px;margin:2px 0;"
                            f"border-radius:0 6px 6px 0;"
                            f"font-size:12px;color:#166534'>"
                            f"✅ {c}</div>",
                            unsafe_allow_html=True
                        )
            with cw2:
                if warnings:
                    st.markdown(
                        f"**⚠️ Warning factors ({len(warnings)})**"
                    )
                    for w in warnings:
                        st.markdown(
                            f"<div style='background:#fef2f2;"
                            f"border-left:3px solid #fca5a5;"
                            f"padding:5px 10px;margin:2px 0;"
                            f"border-radius:0 6px 6px 0;"
                            f"font-size:12px;color:#991b1b'>"
                            f"⚠️ {w}</div>",
                            unsafe_allow_html=True
                        )

            # Noise vs reversal explanation
            if real_reversal:
                st.error(
                    "🔴 REAL REVERSAL DETECTED — "
                    "3+ factors have turned against your trade. "
                    "This is not noise. Exit now."
                )
            elif temp_weakness:
                st.warning(
                    "🟡 TEMPORARY WEAKNESS — "
                    "Some factors weakening but trend not broken. "
                    "Reduce size, tighten stop loss. Watch next 2 candles."
                )

            # Action buttons
            ab1, ab2, ab3, ab4 = st.columns(4)
            with ab1:
                if st.button(
                    "✅ Mark Closed",
                    key=f"tm_close_{idx}",
                    use_container_width=True
                ):
                    trade["status"] = "CLOSED — MANUAL"
                    save_trades(st.session_state["active_trades"])
                    st.rerun()
            with ab2:
                if st.button(
                    "📉 Reduce 50%",
                    key=f"tm_reduce_{idx}",
                    use_container_width=True
                ):
                    trade["lots_rem"] = max(
                        1, trade["lots_rem"] // 2
                    )
                    trade["last_action"] = "Reduced 50%"
                    save_trades(st.session_state["active_trades"])
                    st.success(
                        f"Reduced to {trade['lots_rem']} lots"
                    )
            with ab3:
                if st.button(
                    "🔄 Update SL",
                    key=f"tm_update_sl_{idx}",
                    use_container_width=True
                ):
                    # Trail SL to breakeven or current - ATR
                    new_sl = (
                        max(trade["entry"], cp_now - tm_sig["atrv"])
                        if tm_sig and is_ce
                        else min(trade["entry"], cp_now + tm_sig["atrv"])
                        if tm_sig
                        else trade["sl"]
                    )
                    trade["sl"] = round(new_sl, 2)
                    save_trades(st.session_state["active_trades"])
                    st.success(f"SL trailed to ₹{trade['sl']:,.2f}")
            with ab4:
                if tg_configured():
                    if st.button(
                        "📱 Send Update",
                        key=f"tm_tg_{idx}",
                        use_container_width=True
                    ):
                        _tok = st.session_state.get(
                            "tg_token_saved", ""
                        )
                        _cid = st.session_state.get(
                            "tg_chat_saved", ""
                        )
                        _msg = (
                            f"🛡️ Trade Update — {trade['stock']}\n"
                            f"{trade['type']} | {verdict}\n"
                            f"Entry ₹{trade['entry']:,.0f} | "
                            f"Now ₹{cp_now:,.0f}\n"
                            f"P&L: {pnl_pts:+.2f} pts "
                            f"({pnl_pct:+.2f}%)\n"
                            f"Action: {v_action}"
                        )
                        if send_telegram(_tok, _cid, _msg):
                            st.success("✅ Update sent!")

    # ── Trade Journal Analytics ───────────────────────────
    st.markdown("---")
    st.markdown("### 📈 Trade Journal Analytics")
    st.caption(
        "Records your trades and shows your personal win rate patterns. "
        "Add trades here after closing them to build your analytics."
    )

    # Manual trade record form
    with st.expander("📝 Record Completed Trade"):
        ja1, ja2, ja3 = st.columns(3)
        with ja1:
            j_stock  = st.text_input("Stock", key="j_stock")
            j_type   = st.selectbox("Type", ["BUY CE","BUY PE"], key="j_type")
            j_result = st.selectbox("Result", ["WIN","LOSS","BREAKEVEN"], key="j_result")
        with ja2:
            j_entry  = st.number_input("Entry ₹", value=0.0, key="j_entry")
            j_exit   = st.number_input("Exit ₹",  value=0.0, key="j_exit")
            j_score  = st.slider("Signal score", 5, 10, 7, key="j_score")
        with ja3:
            j_time   = st.selectbox(
                "Entry time",
                ["9:15-10:00 (Opening)",
                 "10:00-11:00 (Mid AM)",
                 "11:00-13:00 (Mid session)",
                 "13:00-14:30 (Afternoon)",
                 "14:30+ (Pre-close)"],
                key="j_time"
            )
            j_vix    = st.number_input("VIX at entry", value=15.0, step=0.1, key="j_vix")
            j_emotion= st.selectbox(
                "Emotional state",
                ["Calm","Anxious","Overconfident","FOMO","Disciplined"],
                key="j_emotion"
            )

        if st.button("Save Trade Record", key="j_save", type="primary"):
            if "trade_journal" not in st.session_state:
                st.session_state["trade_journal"] = []
            import datetime as _jdt
            _pnl = round(
                ((j_exit - j_entry) if j_type=="BUY CE"
                 else (j_entry - j_exit)) / (j_entry+0.001) * 100, 2
            ) if j_entry > 0 and j_exit > 0 else 0

            st.session_state["trade_journal"].append({
                "date":    _jdt.datetime.now().strftime("%d %b %Y"),
                "stock":   j_stock,
                "type":    j_type,
                "result":  j_result,
                "entry":   j_entry,
                "exit":    j_exit,
                "pnl_pct": _pnl,
                "score":   j_score,
                "time":    j_time,
                "vix":     j_vix,
                "emotion": j_emotion,
            })
            save_journal(st.session_state["trade_journal"])
            st.success("Trade recorded ✅")

    # Analytics display
    journal = st.session_state.get("trade_journal", [])
    if len(journal) >= 3:
        import pandas as _jpd
        jdf = _jpd.DataFrame(journal)
        wins  = jdf[jdf["result"]=="WIN"]
        total = len(jdf)
        wr    = round(len(wins)/total*100, 1)

        # Overall metrics
        jm1,jm2,jm3,jm4 = st.columns(4)
        jm1.metric("Total Trades", total)
        jm2.metric("Win Rate",     f"{wr}%")
        jm3.metric("Avg Win",      f"{wins['pnl_pct'].mean():.1f}%" if len(wins)>0 else "—")
        jm4.metric("Avg Loss",     f"{jdf[jdf['result']=='LOSS']['pnl_pct'].mean():.1f}%" if len(jdf[jdf['result']=='LOSS'])>0 else "—")

        # Win rate by time
        st.markdown("**Win Rate by Entry Time**")
        _t_wr = jdf.groupby("time").apply(
            lambda x: round(len(x[x["result"]=="WIN"])/len(x)*100,1)
        ).reset_index()
        _t_wr.columns = ["Time", "Win Rate %"]
        st.dataframe(_t_wr, hide_index=True, width="stretch")

        # Win rate by signal score
        st.markdown("**Win Rate by Signal Score**")
        _s_wr = jdf.groupby("score").apply(
            lambda x: round(len(x[x["result"]=="WIN"])/len(x)*100,1)
        ).reset_index()
        _s_wr.columns = ["Score", "Win Rate %"]
        st.dataframe(_s_wr, hide_index=True, width="stretch")

        # Win rate by VIX
        jdf["vix_range"] = _jpd.cut(
            jdf["vix"],
            bins=[0,13,17,20,100],
            labels=["<13 Calm","13-17 Normal","17-20 Elevated",">20 High"]
        )
        _v_wr = jdf.groupby("vix_range", observed=True).apply(
            lambda x: round(len(x[x["result"]=="WIN"])/len(x)*100,1)
        ).reset_index()
        _v_wr.columns = ["VIX Range","Win Rate %"]
        st.markdown("**Win Rate by VIX Level**")
        st.dataframe(_v_wr, hide_index=True, width="stretch")

        # Emotion analysis
        st.markdown("**Emotion vs Results**")
        _e_wr = jdf.groupby("emotion").apply(
            lambda x: round(len(x[x["result"]=="WIN"])/len(x)*100,1)
        ).reset_index()
        _e_wr.columns = ["Emotion","Win Rate %"]
        st.dataframe(_e_wr, hide_index=True, width="stretch")

        # Key insight
        if len(_t_wr) > 1:
            _best_time = _t_wr.loc[_t_wr["Win Rate %"].idxmax(),"Time"]
            st.info(
                f"💡 Your best trading time: **{_best_time}** — "
                f"focus your entries during this window"
            )

        # All trades table
        with st.expander("All recorded trades"):
            st.dataframe(
                jdf[["date","stock","type","result",
                     "pnl_pct","score","time","vix","emotion"]],
                hide_index=True,
                width="stretch"
            )
    elif len(journal) > 0:
        st.info(
            f"You have recorded {len(journal)} trade(s). "
            "Record at least 3 trades to see analytics."
        )
    else:
        st.info(
            "No trades recorded yet. "
            "Record your completed trades above to build analytics. "
            "After 20+ trades you will see powerful patterns."
        )

    # ── Closed trades ──────────────────────────────────────
    if closed:
        with st.expander(f"📋 Closed trades ({len(closed)})"):
            for t in closed:
                _fc = "#16a34a" if "TARGET" in t["status"] else "#dc2626"
                st.markdown(
                    f"<div style='background:#f8fafc;"
                    f"border-left:3px solid {_fc};"
                    f"padding:8px 14px;margin:4px 0;"
                    f"border-radius:0 8px 8px 0;font-size:12px'>"
                    f"<b>{t['stock']}</b> {t['type']} | "
                    f"Entry ₹{t['entry']:,.0f} | "
                    f"Added {t['added_at']} | "
                    f"<b style='color:{_fc}'>{t['status']}</b>"
                    f"</div>",
                    unsafe_allow_html=True
                )
            if st.button(
                "Clear closed trades",
                key="tm_clear_closed"
            ):
                st.session_state["active_trades"] = [
                    t for t in trades
                    if t["status"] == "ACTIVE"
                ]
                save_trades(st.session_state["active_trades"])
                st.rerun()


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 11 — PAPER TRADING                             ║
# ╚══════════════════════════════════════════════════════╝
with T11:
    st.markdown("### 📝 Paper Trading — Virtual Trading Simulator")
    st.caption(
        "Practice trading with virtual money before using real capital. "
        "All signals come from your live terminal. "
        "Track performance, learn from mistakes, build confidence."
    )

    # ── Initialize paper trading state ────────────────────
    if "pt_trades" not in st.session_state:
        st.session_state["pt_trades"] = load_paper_trades()
    if "pt_capital" not in st.session_state:
        st.session_state["pt_capital"] = 100000.0
    if "pt_balance" not in st.session_state:
        # Calculate balance from closed trades
        _closed_pnl = sum(
            t.get("pnl_rs", 0)
            for t in st.session_state["pt_trades"]
            if t.get("status") == "CLOSED"
        )
        st.session_state["pt_balance"] = (
            st.session_state["pt_capital"] + _closed_pnl
        )

    # ── Capital settings ───────────────────────────────────
    with st.expander("⚙️ Settings", expanded=False):
        _set1, _set2 = st.columns(2)
        with _set1:
            _new_cap = st.number_input(
                "Starting capital (₹)",
                value=st.session_state["pt_capital"],
                step=10000.0, min_value=10000.0,
                key="pt_cap_input"
            )
        with _set2:
            if st.button(
                "Reset Paper Trading",
                key="pt_reset",
                type="secondary"
            ):
                st.session_state["pt_trades"]  = []
                st.session_state["pt_capital"] = _new_cap
                st.session_state["pt_balance"] = _new_cap
                save_paper_trades([])
                st.success("Paper trading reset ✅")
                st.rerun()

    # ── Performance Dashboard ──────────────────────────────
    _pt_all    = st.session_state["pt_trades"]
    _pt_closed = [t for t in _pt_all if t.get("status")=="CLOSED"]
    _pt_open   = [t for t in _pt_all if t.get("status")=="OPEN"]

    _total_pnl = sum(t.get("pnl_rs",0) for t in _pt_closed)
    _wins      = [t for t in _pt_closed if t.get("pnl_rs",0) > 0]
    _losses    = [t for t in _pt_closed if t.get("pnl_rs",0) <= 0]
    _win_rate  = round(len(_wins)/len(_pt_closed)*100,1) if _pt_closed else 0
    _balance   = st.session_state["pt_capital"] + _total_pnl
    _ret_pct   = round(_total_pnl/st.session_state["pt_capital"]*100,2)

    # Dashboard metrics
    dm1,dm2,dm3,dm4,dm5 = st.columns(5)
    dm1.metric(
        "Virtual Capital",
        f"₹{_balance:,.0f}",
        delta=f"₹{_total_pnl:+,.0f}",
        delta_color="normal" if _total_pnl >= 0 else "inverse"
    )
    dm2.metric("Total Trades", len(_pt_closed))
    dm3.metric("Win Rate",     f"{_win_rate}%")
    dm4.metric(
        "Total P&L",
        f"₹{_total_pnl:+,.0f}",
        delta=f"{_ret_pct:+.2f}%",
        delta_color="normal" if _total_pnl >= 0 else "inverse"
    )
    dm5.metric("Open Trades", len(_pt_open))

    # Performance bar
    if _pt_closed:
        _avg_win  = round(sum(t["pnl_rs"] for t in _wins)  /max(len(_wins),1),0)
        _avg_loss = round(sum(t["pnl_rs"] for t in _losses)/max(len(_losses),1),0)
        _pf = round(abs(sum(t["pnl_rs"] for t in _wins)) / (abs(sum(t["pnl_rs"] for t in _losses))+0.001),2)

        st.markdown(
            f"<div style='background:#f8fafc;border-radius:10px;"
            f"padding:12px 16px;margin:8px 0;font-size:13px'>"
            f"Avg Win: <b style='color:#16a34a'>₹{_avg_win:+,.0f}</b> &nbsp;|&nbsp; "
            f"Avg Loss: <b style='color:#dc2626'>₹{_avg_loss:+,.0f}</b> &nbsp;|&nbsp; "
            f"Profit Factor: <b>{'🔥 ' if _pf>=1.5 else ''}{_pf}</b> &nbsp;|&nbsp; "
            f"Return: <b style='color:{'#16a34a' if _ret_pct>=0 else '#dc2626'}'>{_ret_pct:+.2f}%</b>"
            f"</div>",
            unsafe_allow_html=True
        )

    st.markdown("---")

    # ── Enter New Paper Trade ──────────────────────────────
    st.markdown("### ➕ Enter New Paper Trade")
    st.caption(
        "Get signal from Auto Scanner or Signal Hub, "
        "then enter details here to simulate the trade."
    )

    with st.form("pt_entry_form"):
        pf1, pf2, pf3 = st.columns(3)
        with pf1:
            pt_stock  = st.text_input("Stock name", value="NIFTY 50")
            pt_signal = st.selectbox("Signal", ["BUY CE","BUY PE"])
            pt_score  = st.slider("Signal score", 5, 10, 8)
        with pf2:
            pt_entry  = st.number_input("Entry price (₹)", value=0.0, step=0.5)
            pt_sl     = st.number_input("Stop loss (₹)",   value=0.0, step=0.5)
            pt_target = st.number_input("Target (₹)",      value=0.0, step=0.5)
        with pf3:
            pt_strike = st.text_input("Option strike", placeholder="e.g. 24000 CE")
            pt_premium= st.number_input("Premium paid (₹)", value=0.0, step=0.5)
            pt_lots   = st.number_input("Lots", value=1, min_value=1)
            pt_lotsize= st.number_input("Lot size", value=50, min_value=1)
            pt_style  = st.selectbox(
                "Trade style",
                ["Intraday","Swing 1 day","Swing 3 days","Swing 1 week"]
            )

        pt_submit = st.form_submit_button(
            "📝 Enter Paper Trade",
            use_container_width=True,
            type="primary"
        )

    if pt_submit:
        if pt_entry > 0 and pt_sl > 0 and pt_target > 0:
            import datetime as _ptdt
            _pt_sym = STOCKS.get(pt_stock, "^NSEI")
            _pt_invest = round(pt_premium * pt_lotsize * pt_lots, 2) if pt_premium > 0 else 0
            _new_trade = {
                "id":        len(_pt_all) + 1,
                "stock":     pt_stock,
                "sym":       _pt_sym,
                "signal":    pt_signal,
                "score":     pt_score,
                "entry":     pt_entry,
                "sl":        pt_sl,
                "target":    pt_target,
                "strike":    pt_strike,
                "premium":   pt_premium,
                "lots":      pt_lots,
                "lot_size":  pt_lotsize,
                "invested":  _pt_invest,
                "style":     pt_style,
                "date":      _ptdt.datetime.now().strftime("%d %b %Y"),
                "time":      _ptdt.datetime.now().strftime("%H:%M"),
                "status":    "OPEN",
                "exit_price": 0.0,
                "exit_premium": 0.0,
                "pnl_pts":   0.0,
                "pnl_rs":    0.0,
                "result":    "",
                "notes":     "",
            }
            st.session_state["pt_trades"].append(_new_trade)
            save_paper_trades(st.session_state["pt_trades"])
            st.success(
                f"✅ Paper trade entered — "
                f"{pt_stock} {pt_signal} at ₹{pt_entry}"
            )
            st.rerun()
        else:
            st.error("Please fill Entry, Stop Loss and Target.")

    # ── Open Paper Trades ──────────────────────────────────
    if _pt_open:
        st.markdown("---")
        st.markdown(f"### 📊 Open Paper Trades ({len(_pt_open)})")

        for _pi, _pt in enumerate(_pt_open):
            _is_ce  = _pt["signal"] == "BUY CE"
            _pt_lp  = live_price(_pt["sym"])
            _cp_now = _pt_lp["p"] if _pt_lp["ok"] else _pt["entry"]

            # Current P&L
            _pnl_pts = (
                (_cp_now - _pt["entry"]) if _is_ce
                else (_pt["entry"] - _cp_now)
            )
            _pnl_pct = round(_pnl_pts / _pt["entry"] * 100, 2)
            _pnl_col = "#16a34a" if _pnl_pts >= 0 else "#dc2626"

            # SL / Target check
            _sl_hit  = ((_cp_now <= _pt["sl"]) if _is_ce else (_cp_now >= _pt["sl"]))
            _tgt_hit = ((_cp_now >= _pt["target"]) if _is_ce else (_cp_now <= _pt["target"]))

            _border = (
                "#dc2626" if _sl_hit else
                "#16a34a" if _tgt_hit else
                "#3b82f6"
            )

            st.markdown(
                f"<div style='background:#f8fafc;"
                f"border:2px solid {_border};"
                f"border-radius:12px;padding:16px;"
                f"margin-bottom:10px'>"
                f"<div style='display:flex;justify-content:"
                f"space-between;align-items:center;"
                f"margin-bottom:10px'>"
                f"<div>"
                f"<span style='font-size:17px;font-weight:700;"
                f"color:#1e293b'>{_pt['stock']}</span>"
                f"<span style='background:#1e293b;color:white;"
                f"padding:2px 10px;border-radius:10px;"
                f"font-size:12px;margin-left:8px'>{_pt['signal']}</span>"
                f"<span style='font-size:11px;color:#64748b;"
                f"margin-left:8px'>{_pt['date']} {_pt['time']} | "
                f"{_pt['style']} | Score {_pt['score']}/10</span>"
                f"</div>"
                f"<div style='font-size:18px;font-weight:700;"
                f"color:{_pnl_col}'>{_pnl_pts:+.2f} pts "
                f"({_pnl_pct:+.2f}%)</div></div>"

                + (
                    "<div style='background:#fef2f2;color:#dc2626;"
                    "border-radius:6px;padding:6px 12px;"
                    "font-weight:700;margin-bottom:8px'>"
                    "🔴 STOP LOSS HIT — Close this trade!</div>"
                    if _sl_hit else
                    "<div style='background:#f0fdf4;color:#16a34a;"
                    "border-radius:6px;padding:6px 12px;"
                    "font-weight:700;margin-bottom:8px'>"
                    "🟢 TARGET HIT — Book profit!</div>"
                    if _tgt_hit else ""
                ) +

                f"<div style='display:grid;"
                f"grid-template-columns:repeat(5,1fr);gap:8px'>"
                f"<div style='background:white;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>Entry</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:#374151'>₹{_pt['entry']:,.2f}</div></div>"
                f"<div style='background:white;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>Current</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:{_pnl_col}'>₹{_cp_now:,.2f}</div></div>"
                f"<div style='background:#fef2f2;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>SL</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:#dc2626'>₹{_pt['sl']:,.2f}</div></div>"
                f"<div style='background:#f0fdf4;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>Target</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:#16a34a'>₹{_pt['target']:,.2f}</div></div>"
                f"<div style='background:#eff6ff;border-radius:8px;"
                f"padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#64748b'>Premium</div>"
                f"<div style='font-size:14px;font-weight:700;"
                f"color:#1d4ed8'>₹{_pt['premium']:,.0f}</div></div>"
                f"</div></div>",
                unsafe_allow_html=True
            )

            # Close trade form
            with st.expander(f"Close this trade — {_pt['stock']}"):
                _cl1, _cl2, _cl3 = st.columns(3)
                with _cl1:
                    _exit_px = st.number_input(
                        "Exit stock price (₹)",
                        value=float(_cp_now),
                        step=0.5,
                        key=f"pt_exit_px_{_pi}"
                    )
                with _cl2:
                    _exit_prem = st.number_input(
                        "Exit option premium (₹)",
                        value=float(_pt["premium"]),
                        step=0.5,
                        key=f"pt_exit_prem_{_pi}"
                    )
                with _cl3:
                    _exit_notes = st.text_input(
                        "Notes",
                        placeholder="Why did you exit?",
                        key=f"pt_notes_{_pi}"
                    )

                if st.button(
                    f"✅ Close Trade",
                    key=f"pt_close_{_pi}",
                    type="primary",
                    use_container_width=True
                ):
                    # Calculate P&L
                    _pnl_stock = (
                        (_exit_px - _pt["entry"]) if _is_ce
                        else (_pt["entry"] - _exit_px)
                    )
                    _pnl_option = (
                        (_exit_prem - _pt["premium"])
                        * _pt["lot_size"] * _pt["lots"]
                    ) if _pt["premium"] > 0 else 0

                    _final_pnl = _pnl_option if _pt["premium"] > 0 else (
                        _pnl_stock * _pt["lot_size"] * _pt["lots"]
                    )
                    _result = "WIN" if _final_pnl > 0 else "LOSS"

                    import datetime as _cdt
                    _pt["exit_price"]   = _exit_px
                    _pt["exit_premium"] = _exit_prem
                    _pt["pnl_pts"]      = round(_pnl_stock, 2)
                    _pt["pnl_rs"]       = round(_final_pnl, 2)
                    _pt["result"]       = _result
                    _pt["notes"]        = _exit_notes
                    _pt["exit_date"]    = _cdt.datetime.now().strftime("%d %b %Y %H:%M")
                    _pt["status"]       = "CLOSED"

                    save_paper_trades(st.session_state["pt_trades"])
                    st.success(
                        f"{'✅ WIN' if _result=='WIN' else '❌ LOSS'} — "
                        f"₹{_final_pnl:+,.0f} P&L recorded!"
                    )
                    st.rerun()

    # ── Trade History ──────────────────────────────────────
    if _pt_closed:
        st.markdown("---")
        st.markdown(f"### 📋 Trade History ({len(_pt_closed)} trades)")

        # Equity curve
        import plotly.graph_objects as _ptgo
        _running = [st.session_state["pt_capital"]]
        for _t in _pt_closed:
            _running.append(_running[-1] + _t["pnl_rs"])

        _fig_eq = _ptgo.Figure()
        _fig_eq.add_trace(_ptgo.Scatter(
            y=_running,
            mode="lines+markers",
            line=dict(
                color="#16a34a" if _running[-1] >= _running[0]
                else "#dc2626",
                width=2
            ),
            fill="tozeroy",
            fillcolor=(
                "rgba(22,163,74,0.1)"
                if _running[-1] >= _running[0]
                else "rgba(220,38,38,0.1)"
            ),
            name="Portfolio Value"
        ))
        _fig_eq.add_hline(
            y=st.session_state["pt_capital"],
            line_dash="dash", line_color="#94a3b8",
            annotation_text="Starting capital"
        )
        _fig_eq.update_layout(
            template="plotly_white", height=250,
            title="Paper Trading Equity Curve",
            yaxis_title="Portfolio Value (₹)",
            margin=dict(l=10,r=10,t=40,b=10)
        )
        st.plotly_chart(_fig_eq, use_container_width=True)

        # Trade history table
        _hist_rows = []
        for _t in reversed(_pt_closed):
            _hist_rows.append({
                "Date":    _t.get("date",""),
                "Stock":   _t["stock"],
                "Signal":  _t["signal"],
                "Score":   _t["score"],
                "Entry":   f"₹{_t['entry']:,.2f}",
                "Exit":    f"₹{_t.get('exit_price',0):,.2f}",
                "P&L pts": f"{_t['pnl_pts']:+.2f}",
                "P&L ₹":   f"₹{_t['pnl_rs']:+,.0f}",
                "Result":  _t["result"],
                "Notes":   _t.get("notes",""),
            })

        import pandas as _ptpd
        _hist_df = _ptpd.DataFrame(_hist_rows)
        st.dataframe(
            _hist_df, hide_index=True,
            use_container_width=True
        )

        # Export to Excel
        if st.button(
            "📥 Export Paper Trading History",
            key="pt_export",
            use_container_width=True
        ):
            import io
            from openpyxl import Workbook as _WBPT
            from openpyxl.styles import (
                PatternFill as _PFPT,
                Font as _FntPT,
                Alignment as _AlPT
            )
            _wb = _WBPT()
            _ws = _wb.active
            _ws.title = "Paper Trades"
            _pt_exp_hdrs = [
                "Date","Stock","Signal","Score","Entry","Exit",
                "P&L pts","P&L ₹","Result","Style","Notes"
            ]
            _hf = _PFPT("solid", fgColor="0f766e")
            for _ci,_h in enumerate(_pt_exp_hdrs,1):
                _c = _ws.cell(row=1,column=_ci,value=_h)
                _c.fill=_hf
                _c.font=_FntPT(color="FFFFFF",bold=True)
                _c.alignment=_AlPT(horizontal="center")
                _ws.column_dimensions[_c.column_letter].width=max(12,len(_h)+2)

            _gf=_PFPT("solid",fgColor="d1fae5")
            _rf=_PFPT("solid",fgColor="fee2e2")
            for _ri,_t in enumerate(_pt_closed,2):
                _trow=[
                    _t.get("date",""),_t["stock"],_t["signal"],
                    _t["score"],_t["entry"],_t.get("exit_price",0),
                    _t["pnl_pts"],_t["pnl_rs"],_t["result"],
                    _t["style"],_t.get("notes","")
                ]
                _tfl=_gf if _t["pnl_rs"]>0 else _rf
                for _ci,_v in enumerate(_trow,1):
                    _c=_ws.cell(row=_ri,column=_ci,value=_v)
                    _c.fill=_tfl
                    _c.alignment=_AlPT(horizontal="center")

            _buf=io.BytesIO()
            _wb.save(_buf); _buf.seek(0)
            st.download_button(
                label="📥 Download Paper Trading Excel",
                data=_buf.getvalue(),
                file_name="paper_trades.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="pt_dl"
            )


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 12 — AUTO ORDERS                               ║
# ╚══════════════════════════════════════════════════════╝
with T12:
    st.markdown("### ⚡ Auto Order Placement")

    # ── Safety gate — must be explicitly enabled ───────────
    st.warning(
        "⚠️ **This tab places REAL orders on your Zerodha account.** "
        "Money will be debited from your real trading account. "
        "Start with 1 lot minimum. Test thoroughly before scaling up."
    )

    # Master enable switch — off by default
    _ao_enabled = st.toggle(
        "Enable Auto Order Placement",
        value=st.session_state.get("ao_enabled", False),
        key="ao_toggle",
        help="Turn ON only when you are ready to place real orders"
    )
    st.session_state["ao_enabled"] = _ao_enabled

    if not _ao_enabled:
        st.info(
            "Auto orders are currently **disabled**. "
            "Toggle ON above when you are ready. "
            "We recommend paper trading for 2 weeks first "
            "before enabling this."
        )
        st.markdown("""
        ### 📋 How Auto Orders work

        **Step 1** — Enable the toggle above

        **Step 2** — Sync your Zerodha portfolio
        Click **Sync Portfolio** to see your current positions

        **Step 3** — Place order from Signal Hub
        When Signal Hub shows a Diamond or Strong signal →
        click **Place Order** → order goes to Zerodha instantly

        **Step 4** — GTT stop loss set automatically
        Immediately after entry — GTT order placed at your SL level

        **Step 5** — Trade added to Trade Manager automatically
        No manual entry needed — trade appears in Trade Manager

        ### ⚠️ Important rules
        - Always start with **1 lot only**
        - Keep GTT stop loss active at all times
        - Exit by **2:45 PM** for intraday trades
        - Never place orders during 9:15-10:00 AM opening session
        """)
    else:
        # Check Kite connection
        _kite_ao = get_kite()
        if not _kite_ao:
            st.error(
                "❌ Kite not connected. "
                "Login with Zerodha Kite in the sidebar first."
            )
        else:
            st.success("✅ Kite connected — Ready to place orders")

            # ── Portfolio Sync ─────────────────────────────
            st.markdown("---")
            st.markdown("#### 📊 Portfolio Sync")

            if st.button(
                "🔄 Sync Portfolio from Zerodha",
                key="ao_sync",
                type="primary"
            ):
                with st.spinner("Fetching positions from Zerodha..."):
                    try:
                        _positions = _kite_ao.positions()
                        _holdings  = _kite_ao.holdings()
                        st.session_state["ao_positions"] = _positions
                        st.session_state["ao_holdings"]  = _holdings
                        st.success("✅ Portfolio synced!")
                    except Exception as _e:
                        st.error(f"Failed to sync: {_e}")

            # Show positions
            _positions = st.session_state.get("ao_positions", {})
            if _positions:
                _net_pos = _positions.get("net", [])
                if _net_pos:
                    st.markdown("**Open Positions:**")
                    import pandas as _aopd
                    _pos_rows = []
                    for _p in _net_pos:
                        if _p.get("quantity", 0) != 0:
                            _pnl = _p.get("pnl", 0)
                            _pos_rows.append({
                                "Symbol":    _p.get("tradingsymbol",""),
                                "Qty":       _p.get("quantity",0),
                                "Avg Price": f"₹{_p.get('average_price',0):,.2f}",
                                "LTP":       f"₹{_p.get('last_price',0):,.2f}",
                                "P&L":       f"₹{_pnl:+,.0f}",
                                "Product":   _p.get("product",""),
                            })
                    if _pos_rows:
                        st.dataframe(
                            _aopd.DataFrame(_pos_rows),
                            hide_index=True,
                            use_container_width=True
                        )
                else:
                    st.info("No open positions currently.")

            # ── Place Order ────────────────────────────────
            st.markdown("---")
            st.markdown("#### 📝 Place New Order")
            st.caption(
                "Get signal from Scanner or Signal Hub first, "
                "then fill details here."
            )

            _ao1, _ao2, _ao3 = st.columns(3)
            with _ao1:
                _ao_stock   = st.text_input(
                    "Stock/Index", value="NIFTY 50", key="ao_stock"
                )
                _ao_signal  = st.selectbox(
                    "Signal", ["BUY CE","BUY PE"], key="ao_signal"
                )
                _ao_exchange= st.selectbox(
                    "Exchange", ["NFO","NSE","BSE"], key="ao_exch"
                )
            with _ao2:
                # Smart symbol builder
                st.caption("Option details")
                _ao_strike = st.number_input(
                    "Strike price",
                    value=0.0, step=50.0, key="ao_strike"
                )
                _ao_opt_type = st.selectbox(
                    "CE or PE",
                    ["CE","PE"],
                    index=0 if "CE" in st.session_state.get("ao_signal","BUY CE") else 1,
                    key="ao_opt_type"
                )
                # Auto-build symbol
                _ao_sym_prefix = (
                    "NIFTY" if "NIFTY" in _ao_stock.upper() and "BANK" not in _ao_stock.upper()
                    else "BANKNIFTY" if "BANK" in _ao_stock.upper()
                    else _ao_stock.upper().replace(" ","")[:10]
                )
                # Get nearest expiry from instruments if available
                _inst_map = st.session_state.get("kite_inst_map", {})
                _ao_symbol_auto = ""
                if _ao_strike > 0:
                    # Find matching symbols in instruments
                    _matches = [
                        k for k in _inst_map.keys()
                        if _ao_sym_prefix in k
                        and str(int(_ao_strike)) in k
                        and k.endswith(_ao_opt_type)
                    ]
                    if _matches:
                        _matches.sort()
                        _ao_symbol_auto = st.selectbox(
                            "Select symbol",
                            _matches[:10],
                            key="ao_sym_select"
                        )
                    else:
                        _ao_symbol_auto = st.text_input(
                            "Symbol (type manually)",
                            placeholder=f"{_ao_sym_prefix}...{int(_ao_strike)}{_ao_opt_type}",
                            key="ao_symbol"
                        )
                else:
                    _ao_symbol_auto = st.text_input(
                        "Symbol (enter strike above)",
                        placeholder="Enter strike price first",
                        key="ao_symbol",
                        disabled=True
                    )
                _ao_symbol = _ao_symbol_auto

                _ao_qty = st.number_input(
                    "Quantity (shares)", value=50,
                    min_value=1, key="ao_qty"
                )
                _ao_price = st.number_input(
                    "Limit price (₹, 0=Market)",
                    value=0.0, step=0.5, key="ao_price"
                )
            with _ao3:
                _ao_sl      = st.number_input(
                    "Stop loss (stock price ₹)",
                    value=0.0, step=0.5, key="ao_sl"
                )
                _ao_target  = st.number_input(
                    "Target (stock price ₹)",
                    value=0.0, step=0.5, key="ao_target"
                )
                _ao_product = st.selectbox(
                    "Product type",
                    ["MIS (Intraday)","NRML (Overnight)"],
                    key="ao_product"
                )

            # Order preview
            if _ao_symbol:
                _prod = "MIS" if "MIS" in _ao_product else "NRML"
                _order_type = "MARKET" if _ao_price == 0 else "LIMIT"
                _prod_note = (
                    "Intraday — auto squared off at 3:15 PM if not exited"
                    if _prod == "MIS"
                    else "Overnight — position carries to next day"
                )
                st.markdown(
                    f"<div style='background:#fffbeb;"
                    f"border:1.5px solid #f59e0b;"
                    f"border-radius:10px;padding:14px;"
                    f"margin:8px 0;font-size:13px'>"
                    f"<b>Order Preview:</b><br>"
                    f"BUY {_ao_qty} × {_ao_symbol} "
                    f"@ {'MARKET' if _ao_price==0 else f'₹{_ao_price}'} "
                    f"| <b>{_prod}</b> — {_prod_note}<br>"
                    f"GTT Stop Loss: ₹{_ao_sl} | "
                    f"Target: ₹{_ao_target}"
                    f"</div>",
                    unsafe_allow_html=True
                )

            # Confirmation checkbox — extra safety
            _ao_confirm = st.checkbox(
                "✅ I confirm this is a REAL order with REAL money",
                key="ao_confirm"
            )

            if st.button(
                "⚡ Place Order on Zerodha",
                key="ao_place",
                type="primary",
                use_container_width=True,
                disabled=not (_ao_confirm and bool(_ao_symbol))
            ):
                with st.spinner("Placing order..."):
                    try:
                        _prod = (
                            "MIS" if "MIS" in _ao_product
                            else "NRML"
                        )
                        _o_type = (
                            "MARKET" if _ao_price == 0
                            else "LIMIT"
                        )
                        _order_id = _kite_ao.place_order(
                            tradingsymbol=_ao_symbol,
                            exchange=_ao_exchange,
                            transaction_type="BUY",
                            quantity=int(_ao_qty),
                            product=_prod,
                            order_type=_o_type,
                            price=float(_ao_price) if _ao_price > 0 else None,
                            variety="regular"
                        )
                        st.success(
                            f"✅ Order placed! Order ID: {_order_id}"
                        )

                        # Place GTT stop loss if SL provided
                        if _ao_sl > 0:
                            try:
                                # Get LTP for GTT
                                _lp_ao = live_price(
                                    STOCKS.get(_ao_stock,"^NSEI")
                                )
                                _ltp_ao = _lp_ao["p"] if _lp_ao["ok"] else _ao_sl

                                _kite_ao.place_gtt(
                                    trigger_type="single",
                                    tradingsymbol=_ao_symbol,
                                    exchange=_ao_exchange,
                                    trigger_values=[float(_ao_sl)],
                                    last_price=float(_ltp_ao),
                                    orders=[{
                                        "transaction_type": "SELL",
                                        "quantity":  int(_ao_qty),
                                        "product":   _prod,
                                        "order_type":"MARKET",
                                        "price":     float(_ao_sl),
                                    }]
                                )
                                st.success(
                                    f"✅ GTT Stop Loss set at ₹{_ao_sl}"
                                )
                            except Exception as _ge:
                                st.warning(
                                    f"GTT failed: {_ge}. "
                                    "Set stop loss manually in Zerodha."
                                )

                        # Auto-add to Trade Manager
                        import datetime as _aodt
                        _new_tm = {
                            "id":        len(st.session_state.get("active_trades",[])) + 1,
                            "stock":     _ao_stock,
                            "sym":       STOCKS.get(_ao_stock,"^NSEI"),
                            "type":      _ao_signal,
                            "entry":     _lp_ao["p"] if _lp_ao["ok"] else 0,
                            "sl":        _ao_sl,
                            "target":    _ao_target,
                            "lots":      1,
                            "lots_rem":  1,
                            "style":     "Intraday (exit 2:45 PM)",
                            "tf":        "1h",
                            "opt_price": _ao_price,
                            "added_at":  _aodt.datetime.now().strftime("%d %b %H:%M"),
                            "status":    "ACTIVE",
                            "last_action": f"Auto order {_order_id}",
                        }
                        if "active_trades" not in st.session_state:
                            st.session_state["active_trades"] = []
                        st.session_state["active_trades"].append(_new_tm)
                        save_trades(st.session_state["active_trades"])
                        st.info(
                            "✅ Trade added to Trade Manager automatically."
                        )

                    except Exception as _oe:
                        st.error(
                            f"❌ Order failed: {_oe}. "
                            "Check symbol name and Kite connection."
                        )

            # ── Order History ──────────────────────────────
            st.markdown("---")
            st.markdown("#### 📋 Today's Order History")

            if st.button("Load Order History", key="ao_history"):
                try:
                    _orders = _kite_ao.orders()
                    if _orders:
                        import pandas as _ohpd
                        _oh_rows = []
                        for _o in _orders:
                            _oh_rows.append({
                                "Time":     _o.get("order_timestamp",""),
                                "Symbol":   _o.get("tradingsymbol",""),
                                "Type":     _o.get("transaction_type",""),
                                "Qty":      _o.get("quantity",0),
                                "Price":    f"₹{_o.get('average_price',0):,.2f}",
                                "Status":   _o.get("status",""),
                                "Order ID": _o.get("order_id",""),
                            })
                        st.dataframe(
                            _ohpd.DataFrame(_oh_rows),
                            hide_index=True,
                            use_container_width=True
                        )
                    else:
                        st.info("No orders placed today.")
                except Exception as _he:
                    st.error(f"Failed to load orders: {_he}")


# ╔══════════════════════════════════════════════════════╗
# ║  TAB 13 — EVENING SCAN                              ║
# ╚══════════════════════════════════════════════════════╝
with T13:
    st.markdown("### 🌙 Evening Pre-Scan — Tomorrow's Watchlist")
    st.caption(
        "Run this at 3:30 PM after market closes. "
        "Identifies tomorrow's candidates based on daily candles. "
        "⚠️ Always confirm with morning 15m scanner before entering. "
        "Overnight news can change direction completely."
    )
    st.warning(
        "⚠️ Evening Scan is a PREPARATION TOOL only — not a trade signal. "
        "Never enter a trade based on Evening Scan alone. "
        "Always wait for the 9:30 AM scanner to confirm the same direction "
        "before entering any trade."
    )

    # ── How to use guide ──────────────────────────────────
    with st.expander("📖 How to use Evening Scan", expanded=False):
        st.markdown("""
        **Step 1 — Run at 3:30 PM (after market close)**
        Click Scan Tomorrow's Candidates below.

        **Step 2 — Review the watchlist**
        Stocks are ranked by daily signal strength.
        Note the top 5-8 stocks.

        **Step 3 — Check overnight context**
        Before sleeping check:
        - US market direction (S&P 500, Nasdaq)
        - Gift Nifty futures direction
        - Any major overnight news

        **Step 4 — Next morning at 9:00 AM**
        These stocks are your priority candidates.
        Run Prepare for Trading → then Scanner.
        Confirm the evening signal with the morning 15m signal.

        **Step 5 — Enter only when both confirm**
        Evening scan said BUY CE on HDFC Bank +
        Morning 15m scanner also says BUY CE →
        Highest confidence entry.
        """)

    # ── Settings ──────────────────────────────────────────
    ev1, ev2, ev3 = st.columns(3)
    with ev1:
        ev_group = st.selectbox(
            "Stock group",
            ["Top F&O Stocks"] + list(SECTORS.keys()),
            key="ev_group"
        )
    with ev2:
        ev_min_score = st.slider(
            "Minimum daily score",
            5, 9, 6,
            key="ev_min_score",
            help="6+ for watchlist, 8+ for high conviction"
        )
    with ev3:
        ev_min_rr = st.selectbox(
            "Min R:R",
            [1.0, 1.5, 2.0],
            index=0,
            key="ev_min_rr"
        )

    # Stock universe for evening scan
    if ev_group == "Top F&O Stocks":
        ev_stocks = [
            "NIFTY 50","BANK NIFTY","Reliance","HDFC Bank",
            "ICICI Bank","TCS","Infosys","SBI","Wipro",
            "Bajaj Finance","ITC","Sun Pharma","L&T","Maruti",
            "Axis Bank","HCL Tech","ONGC","Bharti Airtel",
            "Tata Steel","JSW Steel","Kotak Bank","Titan Company",
            "Asian Paints","Nestle India","Power Grid","NTPC",
            "Bajaj Auto","Eicher Motors","Cipla","Dr Reddys",
            "Divis Lab","UltraTech Cement","Britannia"
        ]
    else:
        ev_stocks = SECTORS.get(ev_group, [])

    if st.button(
        "🌙 Scan Tomorrow's Candidates",
        type="primary",
        key="ev_scan_btn",
        use_container_width=True
    ):
        st.session_state["ev_scanning"] = True
        st.session_state["ev_results"]  = []
        st.session_state["ev_group_used"] = ev_group

    if st.session_state.get("ev_scanning"):
        _ev_results = []
        _ev_prog    = st.progress(0, text="Scanning daily candles...")
        _ev_status  = st.empty()
        _ev_total   = len(ev_stocks)

        for _ei, sname in enumerate(ev_stocks):
            sym = STOCKS.get(sname)
            if not sym:
                continue

            _ev_prog.progress(
                int((_ei+1)/_ev_total*100),
                text=f"Analysing {sname} daily chart... ({_ei+1}/{_ev_total})"
            )

            try:
                # Use daily candles for evening scan
                df_ev = candles(sym, "1d")
                if df_ev is None or len(df_ev) < 55:
                    continue

                lp_ev = live_price(sym)
                sig_ev = compute_all(df_ev, lp_ev)
                if not sig_ev:
                    continue

                cp_ev  = sig_ev["cp"]
                dir_ev = sig_ev["direction"]
                if dir_ev not in ["UPTREND","DOWNTREND"]:
                    continue

                score_ev = (
                    sig_ev["up_score"]
                    if dir_ev == "UPTREND"
                    else sig_ev["dn_score"]
                )
                if score_ev < ev_min_score:
                    continue

                # Entry / SL / Target on daily
                is_ce_ev = dir_ev == "UPTREND"
                atr_ev   = sig_ev["atrv"]
                entry_ev = round(sig_ev["e9v"], 2)
                sl_ev    = round(
                    entry_ev - atr_ev if is_ce_ev
                    else entry_ev + atr_ev, 2
                )
                t1_ev = round(
                    entry_ev + 1.5*atr_ev if is_ce_ev
                    else entry_ev - 1.5*atr_ev, 2
                )
                t2_ev = round(
                    entry_ev + 2.5*atr_ev if is_ce_ev
                    else entry_ev - 2.5*atr_ev, 2
                )
                rr_ev = round(
                    abs(t1_ev-entry_ev) /
                    (abs(entry_ev-sl_ev)+0.001), 2
                )
                if rr_ev < ev_min_rr:
                    continue

                # Daily candlestick pattern
                _patterns_ev = sig_ev.get("patterns",[])
                _bull_pats = [
                    p[0] for p in _patterns_ev
                    if p[1]=="bullish"
                ]
                _bear_pats = [
                    p[0] for p in _patterns_ev
                    if p[1]=="bearish"
                ]

                # Key level proximity
                w_pivot_ev = sig_ev.get("w_pivot", 0)
                m_pivot_ev = sig_ev.get("m_pivot", 0)
                near_weekly = (
                    abs(cp_ev - w_pivot_ev) / cp_ev < 0.01
                    if w_pivot_ev > 0 else False
                )
                near_monthly = (
                    abs(cp_ev - m_pivot_ev) / cp_ev < 0.02
                    if m_pivot_ev > 0 else False
                )

                # BB zone
                vwap_zone_ev = sig_ev.get("vwap_zone","FAIR_VALUE")

                # RSI
                rsi_ev = sig_ev["rv"]

                # Volume
                vol_surge_ev = sig_ev.get("vsurge", False)

                # ML on daily
                ml_ev_dir  = "UNKNOWN"
                ml_ev_conf = 0
                try:
                    ml_ev_model = train_model(df_ev)
                    if ml_ev_model.get("ok"):
                        ml_ev_pred = predict_next_move(
                            df_ev, ml_ev_model
                        )
                        if ml_ev_pred and ml_ev_pred.get("ok"):
                            ml_ev_dir  = ml_ev_pred["prediction"]
                            ml_ev_conf = ml_ev_pred["confidence"]
                except Exception:
                    pass

                # Conviction score
                _conviction = score_ev
                if ml_ev_dir == dir_ev:
                    _conviction += 1.5
                if near_weekly or near_monthly:
                    _conviction += 1
                if vol_surge_ev:
                    _conviction += 0.5
                if vwap_zone_ev in ["OVERSOLD","EXTREME_OS"] and is_ce_ev:
                    _conviction += 1
                if vwap_zone_ev in ["OVERBOUGHT","EXTREME_OB"] and not is_ce_ev:
                    _conviction += 1
                if _bull_pats and is_ce_ev:
                    _conviction += 0.5
                if _bear_pats and not is_ce_ev:
                    _conviction += 0.5

                _ev_results.append({
                    "Stock":       sname,
                    "Sym":         sym,
                    "Action":      "BUY CE" if is_ce_ev else "BUY PE",
                    "Direction":   dir_ev,
                    "Score":       score_ev,
                    "Conviction":  round(_conviction, 1),
                    "RSI":         round(rsi_ev, 1),
                    "RR":          rr_ev,
                    "Entry":       entry_ev,
                    "SL":          sl_ev,
                    "T1":          t1_ev,
                    "T2":          t2_ev,
                    "ATR":         round(atr_ev, 2),
                    "ML":          ml_ev_dir,
                    "ML_Conf":     ml_ev_conf,
                    "ML_Agrees":   ml_ev_dir == dir_ev,
                    "Vol_Surge":   vol_surge_ev,
                    "VWAP_Zone":   vwap_zone_ev,
                    "Near_Weekly": near_weekly,
                    "Near_Monthly":near_monthly,
                    "Bull_Pats":   ", ".join(_bull_pats[:2]),
                    "Bear_Pats":   ", ".join(_bear_pats[:2]),
                    "W_Pivot":     round(w_pivot_ev, 2),
                    "M_Pivot":     round(m_pivot_ev, 2),
                })

            except Exception:
                continue

        _ev_prog.empty()
        _ev_status.empty()

        # Sort by conviction
        _ev_results.sort(
            key=lambda x: x["Conviction"], reverse=True
        )
        st.session_state["ev_results"]  = _ev_results
        st.session_state["ev_scanning"] = False
        st.session_state["ev_scan_time"]= datetime.now().strftime(
            "%d %b %Y %H:%M"
        )
        st.rerun()

    # ── Display Evening Scan Results ──────────────────────
    _ev_res  = st.session_state.get("ev_results", [])
    _ev_time = st.session_state.get("ev_scan_time", "")

    if _ev_res:
        _ev_ce = [r for r in _ev_res if r["Direction"]=="UPTREND"]
        _ev_pe = [r for r in _ev_res if r["Direction"]=="DOWNTREND"]

        st.markdown(
            f"<div style='background:#1e1b4b;border-radius:12px;"
            f"padding:14px 20px;margin-bottom:16px'>"
            f"<span style='color:#a5b4fc;font-size:13px'>"
            f"🌙 Evening scan completed: {_ev_time} | "
            f"Scanned: {len(ev_stocks)} stocks | "
            f"Found: {len(_ev_res)} candidates | "
            f"CE: {len(_ev_ce)} | PE: {len(_ev_pe)}"
            f"</span></div>",
            unsafe_allow_html=True
        )

        # Save to watchlist button
        if st.button(
            "💾 Save as Tomorrow's Watchlist",
            key="ev_save",
            type="primary"
        ):
            st.session_state["tomorrow_watchlist"] = _ev_res
            st.success(
                f"✅ {len(_ev_res)} stocks saved as tomorrow's watchlist!"
            )

        # Tabs for CE and PE
        ev_tab1, ev_tab2 = st.tabs([
            f"📈 BUY CE ({len(_ev_ce)} stocks)",
            f"📉 BUY PE ({len(_ev_pe)} stocks)"
        ])

        def _render_ev_cards(results, is_ce):
            if not results:
                st.info("No candidates found in this direction.")
                return

            for _ri, r in enumerate(results):
                _conviction = r["Conviction"]
                if _conviction >= 9:
                    _grade    = "💎 VERY HIGH CONVICTION"
                    _gbg      = "#1e1b4b"
                    _gfg      = "#a5b4fc"
                    _border   = "#7c3aed"
                elif _conviction >= 7.5:
                    _grade    = "🔥 HIGH CONVICTION"
                    _gbg      = "#f0fdf4"
                    _gfg      = "#166534"
                    _border   = "#16a34a"
                elif _conviction >= 6:
                    _grade    = "⚡ MODERATE"
                    _gbg      = "#fffbeb"
                    _gfg      = "#92400e"
                    _border   = "#d97706"
                else:
                    _grade    = "👀 WATCH"
                    _gbg      = "#f8fafc"
                    _gfg      = "#475569"
                    _border   = "#94a3b8"

                _col = "#16a34a" if is_ce else "#dc2626"

                st.markdown(
                    f"<div style='background:{_gbg};"
                    f"border:2px solid {_border};"
                    f"border-radius:14px;padding:16px;"
                    f"margin-bottom:10px'>"

                    # Header
                    f"<div style='display:flex;"
                    f"justify-content:space-between;"
                    f"align-items:center;margin-bottom:10px'>"
                    f"<div>"
                    f"<span style='font-size:18px;font-weight:700;"
                    f"color:#1e293b'>{r['Stock']}</span>"
                    f"<span style='background:{_col};color:white;"
                    f"padding:3px 10px;border-radius:10px;"
                    f"font-size:12px;margin-left:8px'>"
                    f"{r['Action']}</span>"
                    f"<span style='font-size:12px;color:#64748b;"
                    f"margin-left:8px'>"
                    f"Daily Score {r['Score']}/10 | "
                    f"RSI {r['RSI']}</span>"
                    f"</div>"
                    f"<div style='text-align:right'>"
                    f"<div style='font-size:16px;font-weight:700;"
                    f"color:{_border}'>{_grade}</div>"
                    f"<div style='font-size:12px;color:#64748b'>"
                    f"Conviction {r['Conviction']}/11</div>"
                    f"</div></div>"

                    # Key levels
                    f"<div style='display:grid;"
                    f"grid-template-columns:repeat(5,1fr);"
                    f"gap:8px;margin-bottom:10px'>"

                    f"<div style='background:white;border-radius:8px;"
                    f"padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>Entry</div>"
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#374151'>₹{r['Entry']:,.0f}</div></div>"

                    f"<div style='background:#fef2f2;border-radius:8px;"
                    f"padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>SL</div>"
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#dc2626'>₹{r['SL']:,.0f}</div></div>"

                    f"<div style='background:#f0fdf4;border-radius:8px;"
                    f"padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>T1</div>"
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#16a34a'>₹{r['T1']:,.0f}</div></div>"

                    f"<div style='background:#f0fdf4;border-radius:8px;"
                    f"padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>T2</div>"
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#16a34a'>₹{r['T2']:,.0f}</div></div>"

                    f"<div style='background:#eff6ff;border-radius:8px;"
                    f"padding:8px;text-align:center'>"
                    f"<div style='font-size:10px;color:#64748b'>R:R</div>"
                    f"<div style='font-size:13px;font-weight:700;"
                    f"color:#1d4ed8'>{r['RR']}:1</div></div>"
                    f"</div>"

                    # Confirmation factors
                    f"<div style='display:flex;flex-wrap:wrap;gap:6px;"
                    f"margin-bottom:8px'>"
                    + (f"<span style='background:#dcfce7;color:#166534;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>✅ ML {r['ML']} {r['ML_Conf']}%"
                       f"</span>" if r["ML_Agrees"] else
                       f"<span style='background:#fee2e2;color:#991b1b;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>⚠️ ML {r['ML']}</span>")
                    + (f"<span style='background:#dcfce7;color:#166534;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>✅ Volume surge</span>"
                       if r["Vol_Surge"] else "")
                    + (f"<span style='background:#ede9fe;color:#5b21b6;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>📍 Near Weekly Pivot ₹{r['W_Pivot']:,.0f}"
                       f"</span>" if r["Near_Weekly"] else "")
                    + (f"<span style='background:#ede9fe;color:#5b21b6;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>📅 Near Monthly Pivot ₹{r['M_Pivot']:,.0f}"
                       f"</span>" if r["Near_Monthly"] else "")
                    + (f"<span style='background:#dcfce7;color:#166534;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>🕯️ {r['Bull_Pats']}</span>"
                       if r["Bull_Pats"] and is_ce else "")
                    + (f"<span style='background:#fee2e2;color:#991b1b;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>🕯️ {r['Bear_Pats']}</span>"
                       if r["Bear_Pats"] and not is_ce else "")
                    + (f"<span style='background:#f0fdf4;color:#166534;"
                       f"padding:3px 10px;border-radius:20px;"
                       f"font-size:11px'>📊 VWAP: {r['VWAP_Zone'].replace('_',' ')}"
                       f"</span>")
                    + f"</div>"

                    # Morning confirmation reminder
                    f"<div style='background:rgba(0,0,0,0.05);"
                    f"border-radius:6px;padding:6px 10px;"
                    f"font-size:11px;color:#64748b'>"
                    f"⏰ Tomorrow morning: Confirm with 15m scanner signal "
                    f"before entering. Both must agree."
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )

                # Add to watchlist button
                _ev_c1, _ev_c2 = st.columns(2)
                with _ev_c1:
                    if st.button(
                        f"📊 Open in Trade Setup",
                        key=f"ev_ts_{_ri}_{'ce' if is_ce else 'pe'}",
                        use_container_width=True
                    ):
                        st.session_state["sn"] = r["Stock"]
                        st.session_state["st"] = r["Sym"]
                        st.rerun()
                with _ev_c2:
                    if st.button(
                        f"⭐ Add to Watchlist",
                        key=f"ev_wl_{_ri}_{'ce' if is_ce else 'pe'}",
                        use_container_width=True
                    ):
                        _wl = st.session_state.get(
                            "ev_watchlist", []
                        )
                        if r["Stock"] not in _wl:
                            _wl.append(r["Stock"])
                            st.session_state["ev_watchlist"] = _wl
                            st.success(
                                f"⭐ {r['Stock']} added to "
                                f"tomorrow's watchlist!"
                            )

        with ev_tab1:
            _render_ev_cards(_ev_ce, True)
        with ev_tab2:
            _render_ev_cards(_ev_pe, False)

        # ── Tomorrow's Watchlist Summary ───────────────────
        _ev_wl = st.session_state.get("ev_watchlist", [])
        if _ev_wl:
            st.markdown("---")
            st.markdown("### ⭐ Tomorrow's Priority Watchlist")
            st.caption(
                "These stocks are pre-selected for tomorrow. "
                "Confirm with morning 15m scanner before entering."
            )
            for _wi, _wstock in enumerate(_ev_wl):
                _wr = next(
                    (r for r in _ev_res if r["Stock"]==_wstock),
                    None
                )
                if _wr:
                    _wc = "#16a34a" if _wr["Direction"]=="UPTREND" else "#dc2626"
                    st.markdown(
                        f"<div style='background:#f8fafc;"
                        f"border-left:4px solid {_wc};"
                        f"padding:8px 14px;margin:3px 0;"
                        f"border-radius:0 8px 8px 0;"
                        f"display:flex;justify-content:space-between'>"
                        f"<span style='font-weight:700'>"
                        f"{_wstock}</span>"
                        f"<span style='color:{_wc}'>"
                        f"{_wr['Action']} | "
                        f"Score {_wr['Score']}/10 | "
                        f"Entry ₹{_wr['Entry']:,.0f}"
                        f"</span></div>",
                        unsafe_allow_html=True
                    )

            if st.button(
                "🗑️ Clear Watchlist",
                key="ev_clear_wl"
            ):
                st.session_state["ev_watchlist"] = []
                st.rerun()

        # ── Export to Excel ────────────────────────────────
        st.markdown("---")
        if st.button(
            "📥 Export Evening Scan to Excel",
            key="ev_export",
            use_container_width=True
        ):
            import io
            from openpyxl import Workbook as _EVWB
            from openpyxl.styles import (
                PatternFill as _EVPF,
                Font as _EVFnt,
                Alignment as _EVAl
            )
            _evwb  = _EVWB()
            _evws  = _evwb.active
            _evws.title = "Evening Scan"

            _ev_hdrs = [
                "Stock","Action","Score","Conviction","R:R",
                "Entry","SL","T1","T2","RSI","ML","ML Conf%",
                "Vol Surge","VWAP Zone","Near Weekly","Near Monthly",
                "Bull Pattern","Bear Pattern"
            ]
            _hf = _EVPF("solid", fgColor="1e1b4b")
            for _ci, _h in enumerate(_ev_hdrs, 1):
                _c = _evws.cell(row=1, column=_ci, value=_h)
                _c.fill = _hf
                _c.font = _EVFnt(color="FFFFFF", bold=True)
                _c.alignment = _EVAl(horizontal="center")
                _evws.column_dimensions[
                    _c.column_letter
                ].width = max(12, len(_h)+2)

            _gf = _EVPF("solid", fgColor="d1fae5")
            _rf = _EVPF("solid", fgColor="fee2e2")
            for _ri, r in enumerate(_ev_res, 2):
                _row = [
                    r["Stock"], r["Action"], r["Score"],
                    r["Conviction"], r["RR"],
                    r["Entry"], r["SL"], r["T1"], r["T2"],
                    r["RSI"], r["ML"], r["ML_Conf"],
                    "Yes" if r["Vol_Surge"] else "No",
                    r["VWAP_Zone"],
                    "Yes" if r["Near_Weekly"] else "No",
                    "Yes" if r["Near_Monthly"] else "No",
                    r["Bull_Pats"], r["Bear_Pats"]
                ]
                _rfl = (
                    _gf if r["Direction"]=="UPTREND" else _rf
                )
                for _ci, _v in enumerate(_row, 1):
                    _c = _evws.cell(row=_ri, column=_ci, value=_v)
                    _c.fill = _rfl
                    _c.alignment = _EVAl(horizontal="center")

            _buf = io.BytesIO()
            _evwb.save(_buf)
            _buf.seek(0)
            _fname = (
                f"evening_scan_"
                f"{datetime.now().strftime('%d%b%Y_%H%M')}.xlsx"
            )
            st.download_button(
                label="📥 Download Evening Scan Excel",
                data=_buf.getvalue(),
                file_name=_fname,
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
                key="ev_dl"
            )

    elif not st.session_state.get("ev_scanning"):
        st.info(
            "Click **Scan Tomorrow's Candidates** above to start. "
            "Best run at 3:30 PM after market closes."
        )


# ── Auto refresh ──────────────────────────────────────────
if auto_rf:
    st.sidebar.success("🔄 Refreshing every 2 min...")
    time.sleep(120)
    st.rerun()
